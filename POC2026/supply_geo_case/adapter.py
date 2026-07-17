#!/usr/bin/env python3
"""Build a POC2026 run package from the latest supply_geo primary paths.

The adapter intentionally mirrors the etudecas output contract at a smaller
scale: data files, summaries, reports, maps and a lightweight run manifest.
"""

from __future__ import annotations

import argparse
import ast
import csv
import hashlib
import html
import importlib.util
import itertools
import json
import math
import os
import re
import subprocess
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import yaml


SCHEMA_VERSION = "poc2026.supply_geo_case.v1"
CASE_ROOT = Path(__file__).resolve().parent
REPO_ROOT = CASE_ROOT.parents[1]
DEFAULT_CONFIG = CASE_ROOT / "config" / "supply_geo_case.yml"
BW_TRISTAN_ROOT = REPO_ROOT / "bw_tristan"
BRIGHTWAY_PROJECT = "bw25-ecoinvent310"
BRIGHTWAY_REQUIRED_DATABASES = ("biosphere3", "ecoinvent-3.10-cutoff", "OPERA_siege")
_BRIGHTWAY_RUNTIME_STATUS_CACHE: dict[str, Any] | None = None
_BRIGHTWAY_EXACT_SCENARIO_CACHE: dict[tuple[float, float], list[dict[str, Any]]] = {}
_SDD_EXCHANGE_LCIA_CACHE: dict[tuple[int, str], dict[str, list[dict[str, Any]]]] = {}

EDGE_ORDER = (
    ("T4->T3", "T4", "T3"),
    ("T3->T2", "T3", "T2"),
    ("T2->T1", "T2", "T1"),
    ("T1->OEM", "T1", "OEM"),
)

ROLE_SEQUENCE = ("T4", "T3", "T2", "T1", "OEM")

FAMILY_EF_KGCO2E_PER_KG = {
    "adhesive_composite": 6.0,
    "aluminium": 9.5,
    "copper": 4.2,
    "electronics_cots": 22.0,
    "general": 5.0,
    "polymer_plastic": 3.2,
    "rubber_silicone": 3.8,
    "steel": 2.4,
    "textile_leather": 5.5,
    "titanium_carbon": 28.0,
}

MODE_EF_KGCO2E_PER_KG_KM = {
    "air": 0.00082,
    "internal": 0.0,
    "rail": 0.000035,
    "ship": 0.000015,
    "truck": 0.00012,
    "unknown": 0.00016,
}

SDD_STOCK_TARGET_MONTHS = {
    "T4": 0.6,
    "T3": 1.1,
    "T2": 1.1,
    "T1": 0.9,
    "OEM": 0.7,
}

SDD_BRIGHTWAY_ROLE_SCOPE_SHARE = {
    "T4": 0.10,
    "T3": 0.16,
    "T2": 0.22,
    "T1": 0.30,
    "OEM": 0.22,
}

SDD_BACKUP_SUPPLIER_OVERHEAD_FACTOR = 0.15

SDD_SCRAP_RECYCLING_PROFILES = {
    "adhesive_composite": {
        "recycling_rate": 0.10,
        "avoided_burden_factor": 0.10,
        "treatment_kgco2e_per_kg": 0.18,
        "replacement_factor": 1.00,
    },
    "aluminium": {
        "recycling_rate": 0.85,
        "avoided_burden_factor": 0.65,
        "treatment_kgco2e_per_kg": 0.12,
        "replacement_factor": 1.00,
    },
    "copper": {
        "recycling_rate": 0.90,
        "avoided_burden_factor": 0.75,
        "treatment_kgco2e_per_kg": 0.08,
        "replacement_factor": 1.00,
    },
    "electronics_cots": {
        "recycling_rate": 0.35,
        "avoided_burden_factor": 0.35,
        "treatment_kgco2e_per_kg": 0.45,
        "replacement_factor": 0.95,
    },
    "general": {
        "recycling_rate": 0.35,
        "avoided_burden_factor": 0.30,
        "treatment_kgco2e_per_kg": 0.20,
        "replacement_factor": 1.00,
    },
    "polymer_plastic": {
        "recycling_rate": 0.45,
        "avoided_burden_factor": 0.25,
        "treatment_kgco2e_per_kg": 0.18,
        "replacement_factor": 1.00,
    },
    "rubber_silicone": {
        "recycling_rate": 0.20,
        "avoided_burden_factor": 0.15,
        "treatment_kgco2e_per_kg": 0.22,
        "replacement_factor": 1.00,
    },
    "steel": {
        "recycling_rate": 0.80,
        "avoided_burden_factor": 0.55,
        "treatment_kgco2e_per_kg": 0.10,
        "replacement_factor": 1.00,
    },
    "textile_leather": {
        "recycling_rate": 0.15,
        "avoided_burden_factor": 0.08,
        "treatment_kgco2e_per_kg": 0.25,
        "replacement_factor": 1.00,
    },
    "titanium_carbon": {
        "recycling_rate": 0.60,
        "avoided_burden_factor": 0.45,
        "treatment_kgco2e_per_kg": 0.20,
        "replacement_factor": 1.00,
    },
}

REGION_COUNTRIES = {
    "North America": {"CA", "US", "MX"},
    "Latin America": {"AR", "BO", "BR", "CL", "CO", "CR", "EC", "GT", "HN", "PA", "PE", "PY", "UY", "VE"},
    "Europe": {
        "AT",
        "BE",
        "BG",
        "CH",
        "CZ",
        "DE",
        "DK",
        "EE",
        "ES",
        "FI",
        "FR",
        "GB",
        "GR",
        "HU",
        "IE",
        "IT",
        "LT",
        "LU",
        "LV",
        "NL",
        "NO",
        "PL",
        "PT",
        "RO",
        "SE",
        "SI",
        "SK",
    },
    "MENA": {"AE", "BH", "DZ", "EG", "IL", "IQ", "IR", "JO", "KW", "MA", "OM", "QA", "SA", "TN", "TR"},
    "Sub-Saharan Africa": {"AO", "CD", "CI", "CM", "ET", "GH", "KE", "MG", "MZ", "NG", "SN", "TZ", "ZA", "ZM"},
    "East Asia": {"CN", "HK", "JP", "KR", "MO", "MN", "TW"},
    "South Asia": {"BD", "IN", "LK", "NP", "PK"},
    "Southeast Asia": {"ID", "KH", "LA", "MM", "MY", "PH", "SG", "TH", "VN"},
    "Oceania": {"AU", "NZ"},
}

WEATHER_PROFILES = {
    "temperate_oceanic": {
        "label": "Temperate oceanic",
        "temp_offset": -0.8,
        "amplitude_multiplier": 0.85,
        "humidity_offset": 5.0,
        "precip_multiplier": 1.15,
        "storm_multiplier": 1.15,
        "cold_multiplier": 0.35,
        "hurricane_multiplier": 0.0,
    },
    "continental_cold": {
        "label": "Continental cold",
        "temp_offset": -3.5,
        "amplitude_multiplier": 1.25,
        "humidity_offset": -2.0,
        "precip_multiplier": 0.95,
        "storm_multiplier": 0.8,
        "cold_multiplier": 1.35,
        "hurricane_multiplier": 0.0,
    },
    "high_latitude_cold": {
        "label": "High latitude cold",
        "temp_offset": -6.0,
        "amplitude_multiplier": 1.45,
        "humidity_offset": 0.0,
        "precip_multiplier": 1.05,
        "storm_multiplier": 1.05,
        "cold_multiplier": 1.8,
        "hurricane_multiplier": 0.0,
    },
    "mediterranean_heat": {
        "label": "Mediterranean heat",
        "temp_offset": 2.2,
        "amplitude_multiplier": 0.95,
        "humidity_offset": -6.0,
        "precip_multiplier": 0.78,
        "storm_multiplier": 0.75,
        "cold_multiplier": 0.15,
        "hurricane_multiplier": 0.0,
        "drought_multiplier": 1.35,
    },
    "arid_heat": {
        "label": "Arid heat",
        "temp_offset": 4.5,
        "amplitude_multiplier": 1.05,
        "humidity_offset": -18.0,
        "precip_multiplier": 0.38,
        "storm_multiplier": 0.65,
        "cold_multiplier": 0.1,
        "hurricane_multiplier": 0.0,
        "drought_multiplier": 1.9,
    },
    "monsoon_hot_humid": {
        "label": "Monsoon hot humid",
        "temp_offset": 3.0,
        "amplitude_multiplier": 0.65,
        "humidity_offset": 10.0,
        "precip_multiplier": 1.45,
        "storm_multiplier": 1.35,
        "cold_multiplier": 0.0,
        "hurricane_multiplier": 0.25,
        "drought_multiplier": 0.75,
    },
    "tropical_cyclone": {
        "label": "Tropical cyclone",
        "temp_offset": 4.0,
        "amplitude_multiplier": 0.45,
        "humidity_offset": 14.0,
        "precip_multiplier": 1.65,
        "storm_multiplier": 1.55,
        "cold_multiplier": 0.0,
        "hurricane_multiplier": 1.0,
        "drought_multiplier": 0.55,
    },
    "temperate_mixed": {
        "label": "Temperate mixed",
        "temp_offset": 0.0,
        "amplitude_multiplier": 1.0,
        "humidity_offset": 0.0,
        "precip_multiplier": 1.0,
        "storm_multiplier": 1.0,
        "cold_multiplier": 0.4,
        "hurricane_multiplier": 0.0,
    },
}

DEFAULT_CLIMATE_CHANGE = {
    "enabled": True,
    "warming_deg_c_at_horizon": 2.2,
    "heatwave_factor_at_horizon": 1.75,
    "drought_factor_at_horizon": 1.55,
    "extreme_precip_factor_at_horizon": 1.35,
    "storm_wind_factor_at_horizon": 1.25,
    "hurricane_factor_at_horizon": 1.45,
    "site_cold_factor_at_horizon": 0.85,
    "maritime_storm_factor_at_horizon": 1.35,
    "maritime_hurricane_factor_at_horizon": 1.45,
    "maritime_cold_factor_at_horizon": 1.10,
}

STANDARD_DIRS = {
    "data": "data",
    "reports": "reports",
    "summaries": "summaries",
    "maps": "maps",
    "plots": "plots",
    "run": "run",
}


@dataclass(frozen=True)
class CaseBuildResult:
    output_root: Path
    summary_path: Path
    manifest_path: Path
    artifact_index_path: Path
    summary: dict[str, Any]


def clean(value: Any) -> str:
    return str(value or "").strip()


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def optional_float(value: Any) -> float | None:
    try:
        if value in (None, ""):
            return None
        number = float(value)
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def slug(value: Any) -> str:
    text = clean(value).lower()
    chars = [ch if ch.isalnum() else "_" for ch in text]
    out = "_".join(part for part in "".join(chars).split("_") if part)
    return out or "unknown"


def ascii_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.lower()


def stable_phase(value: str) -> float:
    digest = hashlib.sha256(value.encode("utf-8")).hexdigest()
    raw = int(digest[:12], 16)
    return (raw % 1000000) / 1000000.0 * 2.0 * math.pi


def load_yaml_config(config_path: Path | str = DEFAULT_CONFIG) -> dict[str, Any]:
    path = Path(config_path)
    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"Invalid supply_geo case config: {path}")
    data["_config_path"] = str(path.resolve(strict=False))
    return data


def resolve_from_config(config: dict[str, Any], raw_path: str | Path) -> Path:
    path = Path(raw_path)
    if path.is_absolute():
        return path
    config_path = Path(str(config.get("_config_path") or DEFAULT_CONFIG))
    return (config_path.parent / path).resolve(strict=False)


def ensure_standard_dirs(output_root: Path) -> dict[str, Path]:
    paths = {key: output_root / dirname for key, dirname in STANDARD_DIRS.items()}
    for path in paths.values():
        path.mkdir(parents=True, exist_ok=True)
    return paths


def load_audit_helper(config: dict[str, Any]) -> Any:
    source = config.get("source") if isinstance(config.get("source"), dict) else {}
    helper_path = resolve_from_config(config, source.get("audit_helper_script", ""))
    if not helper_path.exists():
        raise FileNotFoundError(f"supply_geo audit helper not found: {helper_path}")
    spec = importlib.util.spec_from_file_location("_poc2026_supply_geo_audit_helper", helper_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load supply_geo audit helper: {helper_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def role_config(config: dict[str, Any]) -> list[tuple[str, str]]:
    rows = config.get("roles")
    if not isinstance(rows, list) or not rows:
        return [
            ("T4", "tier4_raw_material"),
            ("T3", "tier3_first_transformation"),
            ("T2", "tier2_second_transformation"),
            ("T1", "tier1"),
            ("OEM", "oem"),
        ]
    out: list[tuple[str, str]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        code = clean(row.get("code"))
        role_hint = clean(row.get("role_hint"))
        if code and role_hint:
            out.append((code, role_hint))
    return out


def primary_entries(record: dict[str, Any], code: str, role_hint: str) -> list[dict[str, Any]]:
    if code == "OEM" or role_hint == "oem":
        candidates = [s for s in record.get("oem_sites") or [] if isinstance(s, dict)]
    else:
        candidates = [
            s
            for s in record.get("suppliers") or []
            if isinstance(s, dict) and clean(s.get("role_hint")) == role_hint
        ]
    primary = [s for s in candidates if bool(s.get("is_primary"))]
    if primary:
        return primary
    if code == "OEM" and candidates:
        return candidates[:1]
    return [
        s
        for s in candidates
        if clean(s.get("supplier_status")).startswith("baseline_primary")
    ]


def record_mass_kg(record: dict[str, Any]) -> float:
    lca = record.get("lca_traceability") if isinstance(record.get("lca_traceability"), dict) else {}
    return safe_float(
        lca.get(
            "recommended_additive_mass_kg",
            lca.get("mass_kg", record.get("mass_kg")),
        )
    )


def node_name(audit: Any, entry: dict[str, Any]) -> str:
    return clean(audit.node_label(entry)) if hasattr(audit, "node_label") else clean(entry.get("name"))


def node_status(audit: Any, entry: dict[str, Any]) -> str:
    return clean(audit.status(entry)) if hasattr(audit, "status") else clean(entry.get("supplier_status"))


def site_identity(entry: dict[str, Any]) -> str:
    supplier = clean(entry.get("supplier_id")) or slug(entry.get("name"))
    site = clean(entry.get("site_id")) or clean(entry.get("site_selection_id"))
    if not site:
        lat = clean(entry.get("lat"))
        lon = clean(entry.get("lon"))
        site = f"{lat},{lon}" if lat and lon else slug(entry.get("location") or entry.get("site_address"))
    return f"{supplier}@@{site}"


def role_node_uid(code: str, entry: dict[str, Any]) -> str:
    return f"{code}:{site_identity(entry)}"


def country_code(entry: dict[str, Any]) -> str:
    return clean(entry.get("country_code") or entry.get("site_country") or entry.get("country"))


def source_quality(audit: Any, entry: dict[str, Any]) -> str:
    if hasattr(audit, "source_quality"):
        return clean(audit.source_quality(entry))
    return clean(entry.get("source_confidence") or entry.get("site_selection_confidence") or entry.get("geocode_status"))


def max_severity(audit: Any, issues: list[dict[str, str]]) -> str:
    if hasattr(audit, "max_severity"):
        return clean(audit.max_severity(issues))
    order = {"INFO": 0, "LOW": 1, "MEDIUM": 2, "HIGH": 3, "BLOCKER": 4}
    return max((clean(row.get("severity")) for row in issues), key=lambda sev: order.get(sev, 0), default="INFO")


def scenario_match(audit: Any, record: dict[str, Any], edge: str, src: dict[str, Any], dst: dict[str, Any]) -> dict[str, Any] | None:
    scenarios = [
        s
        for s in record.get("transport_scenarios") or []
        if isinstance(s, dict) and clean(s.get("edge")) == edge
    ]
    if not scenarios:
        return None
    src_name = node_name(audit, src)
    dst_name = node_name(audit, dst)
    src_id = clean(src.get("supplier_id"))
    dst_id = clean(dst.get("supplier_id"))
    exact: list[dict[str, Any]] = []
    for scenario in scenarios:
        scenario_from = clean(scenario.get("from"))
        scenario_to = clean(scenario.get("to"))
        scenario_from_id = clean(scenario.get("from_supplier_id"))
        scenario_to_id = clean(scenario.get("to_supplier_id"))
        from_ok = scenario_from == src_name or (scenario_from_id and scenario_from_id == src_id)
        to_ok = scenario_to == dst_name or (scenario_to_id and scenario_to_id == dst_id)
        if from_ok and to_ok:
            exact.append(scenario)
    if exact:
        candidates = exact
    else:
        candidates = [
            s
            for s in scenarios
            if not clean(s.get("from"))
            and not clean(s.get("to"))
            and not clean(s.get("from_supplier_id"))
            and not clean(s.get("to_supplier_id"))
        ]
    if not candidates:
        return None
    candidates.sort(
        key=lambda s: (
            0 if "baseline" in clean(s.get("scenario_id")).lower() or "baseline" in clean(s.get("status")).lower() else 1,
            clean(s.get("scenario_id")),
        )
    )
    return candidates[0]


def path_allocation_share(role_entries: list[list[dict[str, Any]]]) -> float:
    count = 1
    for entries in role_entries:
        count *= max(1, len(entries))
    return 1.0 / count if count else 0.0


def build_primary_case_tables(
    *,
    records: list[dict[str, Any]],
    config: dict[str, Any],
    audit: Any,
) -> dict[str, Any]:
    roles = role_config(config)
    path_rows: list[dict[str, Any]] = []
    node_rows: list[dict[str, Any]] = []
    lane_rows: list[dict[str, Any]] = []
    skipped_records: list[dict[str, Any]] = []
    excluded_records: list[int] = []

    for record_index, record in enumerate(records, 1):
        if record.get("simulation_supply_usable") is False:
            excluded_records.append(record_index)
            continue

        family = audit.component_family(record)
        role_entries = [primary_entries(record, code, role_hint) for code, role_hint in roles]
        missing_roles = [roles[i][0] for i, entries in enumerate(role_entries) if not entries]
        if missing_roles:
            skipped_records.append(
                {
                    "record_index": record_index,
                    "system": record.get("system", ""),
                    "component": record.get("component", ""),
                    "missing_roles": "|".join(missing_roles),
                }
            )
            continue

        allocation = path_allocation_share(role_entries)
        component_mass = record_mass_kg(record)
        combos = list(itertools.product(*role_entries))
        for combo_index, combo in enumerate(combos, 1):
            entries_by_code = {roles[i][0]: entry for i, entry in enumerate(combo)}
            path_id = f"R{record_index:03d}-P{combo_index:03d}"
            path_mass = component_mass * allocation
            issues: list[dict[str, str]] = []
            issues.extend(audit.lca_issues(record))
            issues.extend(audit.component_model_issues(family))

            for code, entry in entries_by_code.items():
                node_issues = audit.node_generic_issues(code, entry)
                if code != "OEM":
                    node_issues.extend(audit.material_node_issues(family, code, entry))
                issues.extend(node_issues)
                node_rows.append(
                    {
                        "path_id": path_id,
                        "record_index": record_index,
                        "role": code,
                        "node_uid": role_node_uid(code, entry),
                        "site_uid": site_identity(entry),
                        "supplier_id": clean(entry.get("supplier_id")),
                        "site_id": clean(entry.get("site_id")),
                        "name": node_name(audit, entry),
                        "location": clean(entry.get("location") or entry.get("site_address")),
                        "country_code": country_code(entry),
                        "lat": entry.get("lat", ""),
                        "lon": entry.get("lon", ""),
                        "supplier_status": node_status(audit, entry),
                        "is_primary": bool(entry.get("is_primary")) or code == "OEM",
                        "allocation_share_pct": round(allocation * 100.0, 6),
                        "component_mass_kg": round(component_mass, 9),
                        "path_mass_kg": round(path_mass, 9),
                        "source_confidence": clean(entry.get("source_confidence") or entry.get("site_selection_confidence")),
                        "site_quality": source_quality(audit, entry),
                    }
                )

            t2_internal = audit.is_internalized_t2(entries_by_code["T2"], entries_by_code["T1"])
            issues.extend(audit.path_pairing_issues(entries_by_code["T2"], entries_by_code["T1"]))
            edge_audits = [
                audit.edge_audit(record, "T4", "T3", entries_by_code["T4"], entries_by_code["T3"], path_primary=True),
                audit.edge_audit(record, "T3", "T2", entries_by_code["T3"], entries_by_code["T2"], path_primary=True),
                audit.edge_audit(
                    record,
                    "T2",
                    "T1",
                    entries_by_code["T2"],
                    entries_by_code["T1"],
                    t2_internal=t2_internal,
                    path_primary=True,
                ),
                audit.edge_audit(record, "T1", "OEM", entries_by_code["T1"], entries_by_code["OEM"], path_primary=True),
            ]
            for edge_row in edge_audits:
                issues.extend(edge_row["issues"])

            severity = max_severity(audit, issues)
            readiness = audit.readiness(True, issues, list(combo))
            issue_codes = sorted({clean(issue.get("issue_code")) for issue in issues if clean(issue.get("issue_code"))})
            total_route_km = sum(safe_float(edge_row.get("distance_km")) for edge_row in edge_audits)
            modes = sorted(
                {
                    part
                    for edge_row in edge_audits
                    for part in clean(edge_row.get("modes")).split("|")
                    if part
                }
            )

            path_row = {
                "path_id": path_id,
                "record_index": record_index,
                "path_type": "primary",
                "system": record.get("system", ""),
                "component": record.get("component", ""),
                "family": family,
                "component_mass_kg": round(component_mass, 9),
                "path_allocation_share_pct": round(allocation * 100.0, 6),
                "path_mass_kg": round(path_mass, 9),
                "lca_use_class": clean((record.get("lca_traceability") or {}).get("simulation_use_class")),
                "lca_confidence": clean((record.get("lca_traceability") or {}).get("confidence")),
                "readiness": readiness,
                "max_severity": severity,
                "issue_codes": ";".join(issue_codes),
                "issue_count": len(issues),
                "total_route_km": round(total_route_km, 1),
                "allocated_kg_km": round(path_mass * total_route_km, 1),
                "modes": "|".join(modes),
                "transport_model": "lane_specific"
                if not any(edge_row.get("transport_model_status") == "missing_explicit_lane_mode" for edge_row in edge_audits)
                else "generic_phase_modes_only",
            }
            for code in ("T4", "T3", "T2", "T1", "OEM"):
                entry = entries_by_code[code]
                path_row[code.lower()] = node_name(audit, entry)
                path_row[f"{code.lower()}_site_uid"] = site_identity(entry)
                path_row[f"{code.lower()}_country_code"] = country_code(entry)
                path_row[f"{code.lower()}_status"] = node_status(audit, entry)
            for edge_row in edge_audits:
                edge_key = clean(edge_row.get("edge")).lower().replace("->", "_")
                path_row[f"{edge_key}_km"] = edge_row.get("distance_km")
                path_row[f"{edge_key}_modes"] = edge_row.get("modes")
            path_rows.append(path_row)

            for edge_row, (edge_name, code_from, code_to) in zip(edge_audits, EDGE_ORDER, strict=True):
                src = entries_by_code[code_from]
                dst = entries_by_code[code_to]
                scenario = scenario_match(audit, record, edge_name, src, dst) or {}
                distance_km = safe_float(edge_row.get("distance_km"))
                lane_rows.append(
                    {
                        "path_id": path_id,
                        "record_index": record_index,
                        "edge": edge_name,
                        "from_role": code_from,
                        "to_role": code_to,
                        "from_node_uid": role_node_uid(code_from, src),
                        "to_node_uid": role_node_uid(code_to, dst),
                        "from_site_uid": site_identity(src),
                        "to_site_uid": site_identity(dst),
                        "from_name": node_name(audit, src),
                        "to_name": node_name(audit, dst),
                        "from_country_code": country_code(src),
                        "to_country_code": country_code(dst),
                        "distance_km": round(distance_km, 1),
                        "scenario_distance_km": scenario.get("distance_km_haversine", ""),
                        "modes": clean(edge_row.get("modes")),
                        "transport_segment_used": clean(edge_row.get("transport_segment_used")),
                        "transport_model_status": clean(edge_row.get("transport_model_status")),
                        "scenario_id": clean(scenario.get("scenario_id")),
                        "scenario_status": clean(scenario.get("status")),
                        "scenario_source": clean(scenario.get("source")),
                        "path_mass_kg": round(path_mass, 9),
                        "allocated_kg_km": round(path_mass * distance_km, 1),
                    }
                )

    return {
        "paths": path_rows,
        "nodes": node_rows,
        "lanes": lane_rows,
        "skipped_records": skipped_records,
        "excluded_record_indexes": excluded_records,
    }


def unique_nodes(node_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[str, dict[str, Any]] = {}
    for row in node_rows:
        key = clean(row.get("node_uid"))
        item = by_key.setdefault(
            key,
            {
                "node_uid": key,
                "site_uid": row.get("site_uid"),
                "role": row.get("role"),
                "supplier_id": row.get("supplier_id"),
                "site_id": row.get("site_id"),
                "name": row.get("name"),
                "location": row.get("location"),
                "country_code": row.get("country_code"),
                "lat": row.get("lat"),
                "lon": row.get("lon"),
                "supplier_status": row.get("supplier_status"),
                "source_confidence": row.get("source_confidence"),
                "site_quality": row.get("site_quality"),
                "path_count": 0,
                "allocated_mass_kg": 0.0,
                "record_indexes": set(),
            },
        )
        item["path_count"] += 1
        item["allocated_mass_kg"] += safe_float(row.get("path_mass_kg"))
        item["record_indexes"].add(int(row["record_index"]))
    out: list[dict[str, Any]] = []
    for item in by_key.values():
        out.append(
            {
                **{k: v for k, v in item.items() if k != "record_indexes"},
                "allocated_mass_kg": round(item["allocated_mass_kg"], 6),
                "record_count": len(item["record_indexes"]),
            }
        )
    return sorted(out, key=lambda r: (clean(r.get("role")), clean(r.get("name")), clean(r.get("node_uid"))))


def unique_sites(node_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[str, dict[str, Any]] = {}
    for row in node_rows:
        key = clean(row.get("site_uid"))
        item = by_key.setdefault(
            key,
            {
                "site_uid": key,
                "supplier_ids": set(),
                "site_ids": set(),
                "roles": set(),
                "name": row.get("name"),
                "location": row.get("location"),
                "country_code": row.get("country_code"),
                "lat": row.get("lat"),
                "lon": row.get("lon"),
                "source_confidence": row.get("source_confidence"),
                "site_quality": row.get("site_quality"),
                "path_count": 0,
                "allocated_mass_kg": 0.0,
            },
        )
        item["supplier_ids"].add(clean(row.get("supplier_id")))
        item["site_ids"].add(clean(row.get("site_id")))
        item["roles"].add(clean(row.get("role")))
        item["path_count"] += 1
        item["allocated_mass_kg"] += safe_float(row.get("path_mass_kg"))
    out: list[dict[str, Any]] = []
    for item in by_key.values():
        out.append(
            {
                "site_uid": item["site_uid"],
                "supplier_ids": "|".join(sorted(v for v in item["supplier_ids"] if v)),
                "site_ids": "|".join(sorted(v for v in item["site_ids"] if v)),
                "roles": "|".join(sorted(v for v in item["roles"] if v)),
                "name": item["name"],
                "location": item["location"],
                "country_code": item["country_code"],
                "lat": item["lat"],
                "lon": item["lon"],
                "source_confidence": item["source_confidence"],
                "site_quality": item["site_quality"],
                "path_count": item["path_count"],
                "allocated_mass_kg": round(item["allocated_mass_kg"], 6),
            }
        )
    return sorted(out, key=lambda r: (clean(r.get("country_code")), clean(r.get("name")), clean(r.get("site_uid"))))


def unique_flows(lane_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for row in lane_rows:
        key = (
            clean(row.get("edge")),
            clean(row.get("from_node_uid")),
            clean(row.get("to_node_uid")),
            clean(row.get("modes")),
        )
        item = by_key.setdefault(
            key,
            {
                "flow_uid": hashlib.sha1("|".join(key).encode("utf-8")).hexdigest()[:16],
                "edge": row.get("edge"),
                "from_node_uid": row.get("from_node_uid"),
                "to_node_uid": row.get("to_node_uid"),
                "from_site_uid": row.get("from_site_uid"),
                "to_site_uid": row.get("to_site_uid"),
                "from_name": row.get("from_name"),
                "to_name": row.get("to_name"),
                "from_country_code": row.get("from_country_code"),
                "to_country_code": row.get("to_country_code"),
                "distance_km": safe_float(row.get("distance_km")),
                "modes": row.get("modes"),
                "path_count": 0,
                "allocated_mass_kg": 0.0,
                "allocated_kg_km": 0.0,
                "record_indexes": set(),
            },
        )
        item["path_count"] += 1
        item["allocated_mass_kg"] += safe_float(row.get("path_mass_kg"))
        item["allocated_kg_km"] += safe_float(row.get("allocated_kg_km"))
        item["record_indexes"].add(int(row["record_index"]))
    out: list[dict[str, Any]] = []
    for item in by_key.values():
        out.append(
            {
                **{k: v for k, v in item.items() if k != "record_indexes"},
                "distance_km": round(item["distance_km"], 1),
                "allocated_mass_kg": round(item["allocated_mass_kg"], 6),
                "allocated_kg_km": round(item["allocated_kg_km"], 1),
                "record_count": len(item["record_indexes"]),
            }
        )
    return sorted(out, key=lambda r: (-safe_float(r.get("allocated_kg_km")), clean(r.get("edge")), clean(r.get("from_name"))))


def world_region(country_code_value: Any, lat: float, lon: float) -> str:
    country = clean(country_code_value).upper()
    for region, countries in REGION_COUNTRIES.items():
        if country in countries:
            return region
    if -170.0 <= lon <= -30.0 and lat >= 12.0:
        return "North America"
    if -120.0 <= lon <= -30.0 and lat < 12.0:
        return "Latin America"
    if -25.0 <= lon <= 45.0 and lat >= 35.0:
        return "Europe"
    if -20.0 <= lon <= 60.0 and 12.0 <= lat < 35.0:
        return "MENA"
    if -20.0 <= lon <= 55.0 and lat < 12.0:
        return "Sub-Saharan Africa"
    if 60.0 <= lon <= 95.0 and lat >= 5.0:
        return "South Asia"
    if 95.0 <= lon <= 145.0 and lat >= 20.0:
        return "East Asia"
    if 90.0 <= lon <= 145.0 and lat < 20.0:
        return "Southeast Asia"
    if 110.0 <= lon <= 180.0 and lat < 0.0:
        return "Oceania"
    return "Other"


def weather_profile_key(country_code_value: Any, lat: float, lon: float) -> str:
    country = clean(country_code_value).upper()
    abs_lat = abs(lat)
    tropical_cyclone_countries = {
        "BD",
        "CN",
        "CU",
        "DO",
        "HK",
        "ID",
        "IN",
        "JP",
        "KR",
        "MX",
        "MY",
        "PH",
        "SG",
        "TH",
        "TW",
        "US",
        "VN",
    }
    monsoon_countries = {"BD", "ID", "IN", "LK", "MY", "PH", "SG", "TH", "VN"}
    arid_countries = {"AE", "BH", "DZ", "EG", "IL", "IQ", "IR", "JO", "KW", "MA", "OM", "QA", "SA", "TN"}
    mediterranean_countries = {"ES", "FR", "GR", "IT", "PT", "TR"}
    continental_countries = {"AT", "CA", "CH", "CN", "CZ", "DE", "HU", "KR", "PL", "SK", "US"}
    if abs_lat >= 58.0:
        return "high_latitude_cold"
    if country in tropical_cyclone_countries and 5.0 <= abs_lat <= 36.0:
        return "tropical_cyclone"
    if country in monsoon_countries or (5.0 <= abs_lat <= 28.0 and 70.0 <= lon <= 135.0):
        return "monsoon_hot_humid"
    if country in arid_countries or (15.0 <= lat <= 34.0 and -15.0 <= lon <= 65.0):
        return "arid_heat"
    if country in mediterranean_countries and 30.0 <= abs_lat <= 46.0:
        return "mediterranean_heat"
    if country in continental_countries and abs_lat >= 40.0:
        return "continental_cold"
    if abs_lat >= 43.0:
        return "temperate_oceanic"
    return "temperate_mixed"


def month_season_peak(month_index: int, peak_month: int, width: int = 3) -> float:
    month = ((month_index - 1) % 12) + 1
    distance = min(abs(month - peak_month), 12 - abs(month - peak_month))
    return clamp(1.0 - distance / max(1, width), 0.0, 1.0)


def heat_index_c(temp_c: float, humidity_pct: float) -> float:
    if temp_c < 22.0:
        return temp_c
    return temp_c + max(0.0, humidity_pct - 45.0) * 0.08


def climate_change_signal(month_index: int, horizon_months: int, climate_config: dict[str, Any] | None) -> dict[str, float]:
    config = {**DEFAULT_CLIMATE_CHANGE, **(climate_config or {})}
    if not config.get("enabled", True):
        return {
            "climate_progress": 0.0,
            "warming_delta_c": 0.0,
            "heatwave_factor": 1.0,
            "drought_factor": 1.0,
            "extreme_precip_factor": 1.0,
            "storm_wind_factor": 1.0,
            "hurricane_factor": 1.0,
            "site_cold_factor": 1.0,
            "maritime_storm_factor": 1.0,
            "maritime_hurricane_factor": 1.0,
            "maritime_cold_factor": 1.0,
            "hazard_factor": 1.0,
        }
    progress = clamp((month_index - 1) / max(1, horizon_months - 1), 0.0, 1.0)
    # Slight convexity: later years carry more of the risk growth while keeping month 1 unchanged.
    risk_progress = progress**1.15

    def end_factor(name: str) -> float:
        return safe_float(config.get(name), DEFAULT_CLIMATE_CHANGE[name])

    def grow(name: str) -> float:
        return 1.0 + (end_factor(name) - 1.0) * risk_progress

    heatwave_factor = grow("heatwave_factor_at_horizon")
    drought_factor = grow("drought_factor_at_horizon")
    extreme_precip_factor = grow("extreme_precip_factor_at_horizon")
    storm_wind_factor = grow("storm_wind_factor_at_horizon")
    hurricane_factor = grow("hurricane_factor_at_horizon")
    site_cold_factor = grow("site_cold_factor_at_horizon")
    maritime_storm_factor = grow("maritime_storm_factor_at_horizon")
    maritime_hurricane_factor = grow("maritime_hurricane_factor_at_horizon")
    maritime_cold_factor = grow("maritime_cold_factor_at_horizon")
    hazard_factor = mean(
        [
            heatwave_factor,
            drought_factor,
            extreme_precip_factor,
            storm_wind_factor,
            hurricane_factor,
            maritime_storm_factor,
            maritime_hurricane_factor,
        ]
    )
    return {
        "climate_progress": round(progress, 9),
        "warming_delta_c": safe_float(config.get("warming_deg_c_at_horizon"), DEFAULT_CLIMATE_CHANGE["warming_deg_c_at_horizon"]) * progress,
        "heatwave_factor": heatwave_factor,
        "drought_factor": drought_factor,
        "extreme_precip_factor": extreme_precip_factor,
        "storm_wind_factor": storm_wind_factor,
        "hurricane_factor": hurricane_factor,
        "site_cold_factor": site_cold_factor,
        "maritime_storm_factor": maritime_storm_factor,
        "maritime_hurricane_factor": maritime_hurricane_factor,
        "maritime_cold_factor": maritime_cold_factor,
        "hazard_factor": hazard_factor,
    }


def build_weather_row(
    site: dict[str, Any],
    month_index: int,
    horizon_months: int,
    seed: int,
    thresholds: dict[str, Any],
    climate_config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    lat = safe_float(site.get("lat"), 48.0)
    lon = safe_float(site.get("lon"), 2.0)
    site_uid = clean(site.get("site_uid"))
    region = world_region(site.get("country_code"), lat, lon)
    profile_key = weather_profile_key(site.get("country_code"), lat, lon)
    profile = WEATHER_PROFILES.get(profile_key, WEATHER_PROFILES["temperate_mixed"])
    phase = stable_phase(f"{seed}:{site_uid}")
    month_angle = 2.0 * math.pi * ((month_index - 1) % 12) / 12.0
    northern = 1.0 if lat >= 0 else -1.0
    season = math.cos(month_angle - math.pi) * northern
    abs_lat = min(abs(lat), 65.0)
    tropicality = max(0.0, 1.0 - abs_lat / 45.0)
    climate = climate_change_signal(month_index, horizon_months, climate_config)
    trend = climate["warming_delta_c"]
    anomaly = 1.6 * math.sin(month_index * 0.71 + phase) + 1.1 * math.sin(month_index * 0.17 + lon / 35.0)
    winter_peak = month_season_peak(month_index, 1 if lat >= 0 else 7, 3)
    cold_outbreak = (
        winter_peak
        * max(0.0, math.sin(month_index * 0.49 + phase * 0.7) - 0.50)
        * safe_float(profile.get("cold_multiplier"))
    )
    base_temp = 29.0 - abs_lat * 0.32 + tropicality * 2.5 + safe_float(profile.get("temp_offset"))
    amplitude = (5.5 + min(abs_lat, 52.0) * 0.11) * safe_float(profile.get("amplitude_multiplier"), 1.0)
    temp_c = base_temp + amplitude * season + trend + anomaly
    temp_c -= 7.5 * cold_outbreak

    humidity = (
        66.0
        + tropicality * 12.0
        - max(season, 0.0) * 14.0
        + 8.0 * math.sin(month_index * 0.43 + phase / 2.0)
        + safe_float(profile.get("humidity_offset"))
    )
    humidity -= max(season, 0.0) * (climate["drought_factor"] - 1.0) * 6.0
    humidity = clamp(humidity, 20.0, 98.0)

    storm_pulse = max(0.0, math.sin(month_index * 0.91 + phase) - 0.72) * climate["extreme_precip_factor"]
    dry_pulse = max(0.0, math.sin(month_index * 0.37 + phase * 1.7) - 0.62) * climate["drought_factor"]
    wet_season = max(0.0, math.sin(month_angle + phase / 4.0))
    base_precip = 35.0 + tropicality * 45.0
    hurricane_season = month_season_peak(month_index, 9 if lat >= 0 else 2, 4)
    hurricane_pulse = max(0.0, math.sin(month_index * 0.53 + phase * 1.3) - 0.36)
    hurricane = clamp(
        hurricane_season * hurricane_pulse * safe_float(profile.get("hurricane_multiplier")) * 1.45 * climate["hurricane_factor"],
        0.0,
        1.0,
    )
    drought_multiplier = safe_float(profile.get("drought_multiplier"), 1.0)
    precip = (
        base_precip
        + 38.0 * wet_season
        + 190.0 * storm_pulse * safe_float(profile.get("storm_multiplier"), 1.0)
        + 210.0 * hurricane
        - 95.0 * dry_pulse * drought_multiplier
        - 18.0 * max(season, 0.0) * climate["drought_factor"]
    )
    precip *= safe_float(profile.get("precip_multiplier"), 1.0)
    precip = max(0.0, precip)

    wind = (
        3.5
        + 2.0 * abs(math.sin(month_index * 0.29 + phase))
        + 18.0 * storm_pulse * safe_float(profile.get("storm_multiplier"), 1.0) * climate["storm_wind_factor"]
        + 19.0 * hurricane
        + 5.5 * cold_outbreak
        + 0.01 * abs(lon)
    )
    wind = clamp(wind, 0.0, 34.0)
    hi = heat_index_c(temp_c, humidity)

    heatwave = clamp(
        max(temp_c - safe_float(thresholds.get("heatwave_temp_c"), 32.0), hi - safe_float(thresholds.get("heat_index_c"), 35.0)) / 8.0,
        0.0,
        1.0,
    )
    heatwave = clamp(heatwave * climate["heatwave_factor"], 0.0, 1.0)
    drought = 0.0
    if (
        precip < safe_float(thresholds.get("drought_precip_mm"), 18.0)
        and temp_c > safe_float(thresholds.get("drought_temp_c"), 27.0)
        and humidity < safe_float(thresholds.get("drought_humidity_pct"), 55.0)
    ):
        drought = clamp(
            (safe_float(thresholds.get("drought_precip_mm"), 18.0) - precip) / 25.0
            + (temp_c - safe_float(thresholds.get("drought_temp_c"), 27.0)) / 15.0
            + (safe_float(thresholds.get("drought_humidity_pct"), 55.0) - humidity) / 60.0,
            0.0,
            1.0,
        )
    drought = clamp(drought * drought_multiplier * climate["drought_factor"], 0.0, 1.0)
    storm = clamp(
        max(
            (precip - safe_float(thresholds.get("storm_precip_mm"), 95.0)) / 120.0,
            (wind - safe_float(thresholds.get("storm_wind_ms"), 15.0)) / 14.0,
        )
        * max(climate["extreme_precip_factor"], climate["storm_wind_factor"]),
        0.0,
        1.0,
    )
    hurricane = max(
        hurricane,
        clamp(
            max(
                (precip - safe_float(thresholds.get("hurricane_precip_mm"), 155.0)) / 150.0,
                (wind - safe_float(thresholds.get("hurricane_wind_ms"), 25.0)) / 10.0,
            )
            * safe_float(profile.get("hurricane_multiplier"))
            * climate["hurricane_factor"],
            0.0,
            1.0,
        ),
    )
    cold = clamp((safe_float(thresholds.get("cold_temp_c"), -4.0) - temp_c) / 14.0, 0.0, 1.0)
    cold = clamp((cold + cold_outbreak * 0.45) * climate["site_cold_factor"], 0.0, 1.0)

    events = []
    if heatwave > 0:
        events.append("heatwave")
    if drought > 0:
        events.append("drought")
    if storm > 0:
        events.append("storm")
    if hurricane > 0:
        events.append("hurricane")
    if cold > 0:
        events.append("cold")

    return {
        "site_uid": site_uid,
        "world_region": region,
        "weather_profile": profile_key,
        "roles": site.get("roles", ""),
        "supplier": site.get("name", ""),
        "country_code": site.get("country_code", ""),
        "lat": site.get("lat", ""),
        "lon": site.get("lon", ""),
        "month_index": month_index,
        "climate_progress": round(climate["climate_progress"], 6),
        "warming_delta_c": round(climate["warming_delta_c"], 4),
        "hazard_intensification_factor": round(climate["hazard_factor"], 4),
        "heatwave_factor": round(climate["heatwave_factor"], 4),
        "drought_factor": round(climate["drought_factor"], 4),
        "storm_factor": round(max(climate["extreme_precip_factor"], climate["storm_wind_factor"]), 4),
        "hurricane_factor": round(climate["hurricane_factor"], 4),
        "cold_factor": round(climate["site_cold_factor"], 4),
        "temp_c": round(temp_c, 2),
        "humidity_pct": round(humidity, 2),
        "precip_mm": round(precip, 2),
        "wind_ms": round(wind, 2),
        "heat_index_c": round(hi, 2),
        "heatwave": round(heatwave, 4),
        "drought": round(drought, 4),
        "storm_stress": round(storm, 4),
        "hurricane": round(hurricane, 4),
        "cold_stress": round(cold, 4),
        "event_label": "|".join(events) if events else "none",
    }


def event_seed_rows(weather_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    specs = [
        ("heatwave", "heatwave", "temp_c", 0.12, 0.06, 0.18),
        ("drought", "drought", "precip_mm", 0.08, 0.10, 0.08),
        ("storm", "storm_stress", "wind_ms", 0.20, 0.38, 0.04),
        ("hurricane", "hurricane", "wind_ms", 0.32, 0.65, 0.08),
        ("cold", "cold_stress", "temp_c", 0.07, 0.12, 0.03),
    ]
    rows: list[dict[str, Any]] = []
    for row in weather_rows:
        for event_type, intensity_column, source_column, capacity_loss, lead_gain, scrap_gain in specs:
            intensity = safe_float(row.get(intensity_column))
            if intensity <= 0.0:
                continue
            event_id = f"{row['site_uid']}:{int(row['month_index']):03d}:{event_type}"
            rows.append(
                {
                    "event_id": event_id,
                    "site_uid": row.get("site_uid"),
                    "world_region": row.get("world_region"),
                    "weather_profile": row.get("weather_profile"),
                    "roles": row.get("roles"),
                    "supplier": row.get("supplier"),
                    "country_code": row.get("country_code"),
                    "month_index": row.get("month_index"),
                    "event_type": event_type,
                    "intensity": round(intensity, 4),
                    "climate_progress": row.get("climate_progress"),
                    "warming_delta_c": row.get("warming_delta_c"),
                    "hazard_intensification_factor": row.get("hazard_intensification_factor"),
                    "source_weather_column": source_column,
                    "capacity_multiplier": round(max(0.0, 1.0 - capacity_loss * intensity), 4),
                    "lead_time_multiplier": round(1.0 + lead_gain * intensity, 4),
                    "scrap_multiplier": round(1.0 + scrap_gain * intensity, 4),
                }
            )
    return rows


def build_weather_tables(config: dict[str, Any], sites: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    weather_config = config.get("weather_driver") if isinstance(config.get("weather_driver"), dict) else {}
    if not weather_config.get("enabled", True):
        return [], []
    horizon = int(weather_config.get("horizon_months") or 240)
    seed = int(weather_config.get("seed") or 2026)
    thresholds = weather_config.get("thresholds") if isinstance(weather_config.get("thresholds"), dict) else {}
    climate_config = weather_config.get("climate_change") if isinstance(weather_config.get("climate_change"), dict) else {}
    rows = [
        build_weather_row(site, month_index, horizon, seed, thresholds, climate_config)
        for site in sites
        for month_index in range(1, horizon + 1)
    ]
    return rows, event_seed_rows(rows)


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(dict.fromkeys(key for row in rows for key in row.keys()))
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def rel(path: Path, root: Path) -> str:
    try:
        return str(path.resolve(strict=False).relative_to(root.resolve(strict=False))).replace("\\", "/")
    except ValueError:
        return str(path.resolve(strict=False)).replace("\\", "/")


def browser_rel(path: Path, from_dir: Path) -> str:
    return os.path.relpath(path.resolve(strict=False), from_dir.resolve(strict=False)).replace("\\", "/")


def csv_profile(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"exists": False, "row_count": 0, "columns": []}
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        count = sum(1 for _ in reader)
    return {"exists": True, "row_count": count, "columns": columns}


def json_scalar(value: Any) -> Any:
    text = clean(value)
    if text == "":
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    if not math.isfinite(number):
        return text
    if number.is_integer() and "." not in text and "e" not in text.lower():
        return int(number)
    return round(number, 6)


def read_csv_rows(path: Path, columns: list[str] | None = None) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        selected = columns or list(reader.fieldnames or [])
        for row in reader:
            rows.append({column: json_scalar(row.get(column)) for column in selected})
    return rows


def read_csv_rows_preserve(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def artifact_record(output_root: Path, path: Path, *, group: str, domain: str, grain: str, required: bool) -> dict[str, Any]:
    record = {
        "name": path.name,
        "group": group,
        "domain": domain,
        "grain": grain,
        "required": required,
        "path": rel(path, output_root),
        "format": path.suffix.lstrip(".") or "file",
    }
    if path.suffix.lower() == ".csv":
        record.update(csv_profile(path))
    else:
        record.update({"exists": path.exists(), "size_bytes": path.stat().st_size if path.exists() else 0})
    return record


def brightway_runtime_status() -> dict[str, Any]:
    global _BRIGHTWAY_RUNTIME_STATUS_CACHE
    if _BRIGHTWAY_RUNTIME_STATUS_CACHE is not None:
        return dict(_BRIGHTWAY_RUNTIME_STATUS_CACHE)

    modules = ["brightway25", "bw2data", "bw2io", "bw2calc", "lca_algebraic"]
    availability = {name: importlib.util.find_spec(name) is not None for name in modules}
    current_status: dict[str, Any] = {}
    external_status: dict[str, Any] = {}

    if all(availability.values()):
        try:
            import bw2data as bd  # type: ignore[import-not-found]

            bd.projects.set_current(BRIGHTWAY_PROJECT)
            current_databases = sorted(list(bd.databases))
            current_status = {
                "available": True,
                "project": str(bd.projects.current),
                "databases": current_databases,
                "methods_count": len(list(bd.methods)),
                "required_databases_present": all(name in bd.databases for name in BRIGHTWAY_REQUIRED_DATABASES),
            }
        except Exception as exc:  # pragma: no cover - runtime optional
            current_status = {"available": False, "error": f"{type(exc).__name__}: {exc}"}

    external_python = os.environ.get("BRIGHTWAY_PYTHON")
    if not external_python:
        candidate = REPO_ROOT / ".venv-brightway" / "Scripts" / "python.exe"
        if not candidate.exists():
            candidate = REPO_ROOT / ".venv-brightway" / "bin" / "python"
        external_python = str(candidate) if candidate.exists() else ""

    if external_python:
        probe = (
            "import json, sys\n"
            "mods=['brightway25','bw2data','bw2io','bw2calc','lca_algebraic']\n"
            "status={'python':sys.executable,'modules':{},'available':False}\n"
            "for name in mods:\n"
            "    try:\n"
            "        __import__(name); status['modules'][name]=True\n"
            "    except Exception as exc:\n"
            "        status['modules'][name]=False; status.setdefault('module_errors',{})[name]=type(exc).__name__+': '+str(exc)\n"
            "try:\n"
            "    import bw2data as bd\n"
            f"    bd.projects.set_current({BRIGHTWAY_PROJECT!r})\n"
            "    dbs=sorted(list(bd.databases))\n"
            "    status.update({'available':all(status['modules'].values()),'project':str(bd.projects.current),'databases':dbs,'methods_count':len(list(bd.methods))})\n"
            f"    status['required_databases_present']=all(name in bd.databases for name in {list(BRIGHTWAY_REQUIRED_DATABASES)!r})\n"
            "except Exception as exc:\n"
            "    status['error']=type(exc).__name__+': '+str(exc)\n"
            "print(json.dumps(status, ensure_ascii=True))\n"
        )
        try:
            completed = subprocess.run(
                [external_python, "-c", probe],
                check=False,
                capture_output=True,
                text=True,
                timeout=45,
            )
            stdout = completed.stdout.strip().splitlines()
            external_status = json.loads(stdout[-1]) if stdout else {"available": False}
            if completed.returncode != 0:
                external_status["returncode"] = completed.returncode
                external_status["stderr"] = completed.stderr.strip()[-1000:]
        except Exception as exc:  # pragma: no cover - runtime optional
            external_status = {
                "available": False,
                "python": external_python,
                "error": f"{type(exc).__name__}: {exc}",
            }

    current_ready = bool(
        all(availability.values())
        and current_status.get("required_databases_present")
        and int(current_status.get("methods_count") or 0) > 0
    )
    external_ready = bool(
        external_status.get("available")
        and external_status.get("required_databases_present")
        and int(external_status.get("methods_count") or 0) > 0
    )
    status = {
        "can_execute_brightway": current_ready or external_ready,
        "modules": availability,
        "project": BRIGHTWAY_PROJECT,
        "required_databases": list(BRIGHTWAY_REQUIRED_DATABASES),
        "current_python": current_status,
        "external_python": external_status,
        "runtime_mode": "current_python" if current_ready else "external_python" if external_ready else "excel_fallback",
        "note": "Brightway runtime is optional here; Excel exports from bw_tristan are used when the runtime is unavailable.",
    }
    _BRIGHTWAY_RUNTIME_STATUS_CACHE = dict(status)
    return status


def brightway_python_for_runtime(runtime: dict[str, Any]) -> str:
    external = runtime.get("external_python", {}) if isinstance(runtime.get("external_python"), dict) else {}
    current = runtime.get("current_python", {}) if isinstance(runtime.get("current_python"), dict) else {}
    if external.get("available") and clean(external.get("python")):
        return clean(external.get("python"))
    if current.get("available"):
        return os.sys.executable
    return ""


def climate_normalization_factor(indicator_unit_views: list[dict[str, Any]]) -> float:
    for row in indicator_unit_views:
        if clean(row.get("short_label")) == "Climate Change - total":
            factor = safe_float(row.get("normalization_factor_per_person_year"))
            if factor > 0.0:
                return factor
    return EF30_NORMALIZATION_FACTORS["Climate change"][0]


def run_brightway_exact_scenarios(runtime: dict[str, Any], normalization_factor: float, excel_use_phase_pe: float) -> list[dict[str, Any]]:
    cache_key = (round(normalization_factor, 9), round(excel_use_phase_pe, 9))
    if cache_key in _BRIGHTWAY_EXACT_SCENARIO_CACHE:
        return [dict(row) for row in _BRIGHTWAY_EXACT_SCENARIO_CACHE[cache_key]]
    if not runtime.get("can_execute_brightway"):
        return []
    python_path = brightway_python_for_runtime(runtime)
    if not python_path:
        return []
    script = CASE_ROOT / "tools" / "run_brightway_exact_scenarios.py"
    if not script.exists():
        return []
    try:
        completed = subprocess.run(
            [
                python_path,
                str(script),
                "--normalization-factor",
                str(normalization_factor),
                "--excel-use-phase-pe",
                str(excel_use_phase_pe),
                "--json",
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=900,
            cwd=str(REPO_ROOT),
        )
    except Exception as exc:
        return [
            {
                "scenario_id": "runtime_error",
                "label": "Erreur recalcul Brightway exact",
                "root_activity_id": "",
                "root_activity_name": "",
                "validation_status": "runtime_error",
                "error": f"{type(exc).__name__}: {exc}",
            }
        ]
    stdout_lines = [line.strip() for line in completed.stdout.splitlines() if line.strip()]
    json_line = stdout_lines[-1] if stdout_lines else ""
    if completed.returncode != 0 or not json_line:
        return [
            {
                "scenario_id": "runtime_error",
                "label": "Erreur recalcul Brightway exact",
                "root_activity_id": "",
                "root_activity_name": "",
                "validation_status": "runtime_error",
                "error": (completed.stderr.strip() or completed.stdout.strip())[-1000:],
            }
        ]
    try:
        rows = json.loads(json_line)
    except json.JSONDecodeError as exc:
        return [
            {
                "scenario_id": "runtime_error",
                "label": "Erreur parsing recalcul Brightway exact",
                "root_activity_id": "",
                "root_activity_name": "",
                "validation_status": "runtime_error",
                "error": f"{type(exc).__name__}: {exc}",
            }
        ]
    result = rows if isinstance(rows, list) else []
    _BRIGHTWAY_EXACT_SCENARIO_CACHE[cache_key] = [dict(row) for row in result]
    return result


def empty_sdd_exchange_lcia_outputs(status: str, note: str) -> dict[str, list[dict[str, Any]]]:
    return {
        "exchange_lcia": [],
        "exchange_lcia_factors": [],
        "exchange_lcia_category_totals": [],
        "exchange_lcia_monthly": [],
        "exchange_lcia_top": [],
        "exchange_lcia_status": [
            {
                "status": status,
                "input_rows": 0,
                "output_rows": 0,
                "unique_exchange_signatures": 0,
                "exact_factor_count": 0,
                "exact_factor_coverage_pct": "",
                "method": "",
                "note": note,
            }
        ],
    }


def run_sdd_exchange_brightway_lcia(
    runtime: dict[str, Any],
    exchange_delta_path: Path,
    normalization_factor: float,
) -> dict[str, list[dict[str, Any]]]:
    if not exchange_delta_path.exists():
        return empty_sdd_exchange_lcia_outputs("missing_input", f"Missing exchange delta CSV: {exchange_delta_path}")
    if not runtime.get("can_execute_brightway"):
        return empty_sdd_exchange_lcia_outputs("runtime_unavailable", "Brightway runtime is not executable.")
    python_path = brightway_python_for_runtime(runtime)
    if not python_path:
        return empty_sdd_exchange_lcia_outputs("runtime_unavailable", "No Brightway Python interpreter found.")
    script = CASE_ROOT / "tools" / "run_sdd_exchange_lcia.py"
    if not script.exists():
        return empty_sdd_exchange_lcia_outputs("script_missing", f"Missing script: {script}")

    stat = exchange_delta_path.stat()
    cache_key = (stat.st_size, clean(runtime.get("runtime_mode")) or python_path)
    if cache_key in _SDD_EXCHANGE_LCIA_CACHE:
        return {
            key: [dict(row) for row in rows]
            for key, rows in _SDD_EXCHANGE_LCIA_CACHE[cache_key].items()
        }

    with tempfile.TemporaryDirectory(prefix="poc2026_sdd_exchange_lcia_") as tmp:
        tmp_dir = Path(tmp)
        output_csv = tmp_dir / "sdd_brightway_exchange_lcia.csv"
        factor_csv = tmp_dir / "sdd_brightway_exchange_lcia_factors.csv"
        category_csv = tmp_dir / "sdd_brightway_exchange_lcia_category_totals.csv"
        monthly_csv = tmp_dir / "sdd_brightway_exchange_lcia_monthly.csv"
        top_csv = tmp_dir / "sdd_brightway_exchange_lcia_top.csv"
        status_csv = tmp_dir / "sdd_brightway_exchange_lcia_status.csv"
        try:
            completed = subprocess.run(
                [
                    python_path,
                    str(script),
                    "--input-csv",
                    str(exchange_delta_path),
                    "--output-csv",
                    str(output_csv),
                    "--factor-csv",
                    str(factor_csv),
                    "--category-csv",
                    str(category_csv),
                    "--monthly-csv",
                    str(monthly_csv),
                    "--top-csv",
                    str(top_csv),
                    "--status-csv",
                    str(status_csv),
                    "--normalization-factor",
                    str(normalization_factor),
                    "--json",
                ],
                check=False,
                capture_output=True,
                text=True,
                timeout=2700,
                cwd=str(REPO_ROOT),
            )
        except Exception as exc:
            return empty_sdd_exchange_lcia_outputs("runtime_error", f"{type(exc).__name__}: {exc}")
        if completed.returncode != 0:
            message = (completed.stderr.strip() or completed.stdout.strip())[-2000:]
            return empty_sdd_exchange_lcia_outputs("runtime_error", message)
        outputs = {
            "exchange_lcia": read_csv_rows_preserve(output_csv),
            "exchange_lcia_factors": read_csv_rows_preserve(factor_csv),
            "exchange_lcia_category_totals": read_csv_rows_preserve(category_csv),
            "exchange_lcia_monthly": read_csv_rows_preserve(monthly_csv),
            "exchange_lcia_top": read_csv_rows_preserve(top_csv),
            "exchange_lcia_status": read_csv_rows_preserve(status_csv),
        }
    _SDD_EXCHANGE_LCIA_CACHE[cache_key] = {
        key: [dict(row) for row in rows]
        for key, rows in outputs.items()
    }
    return outputs


def build_excel_runtime_comparison(
    reference_person_equivalent: list[dict[str, Any]],
    exact_scenario_lcia: list[dict[str, Any]],
    normalization_factor: float,
) -> list[dict[str, Any]]:
    climate_reference = next(
        (row for row in reference_person_equivalent if clean(row.get("short_label")) == "Climate Change - total"),
        {},
    )
    if not climate_reference:
        return []
    exact_by_root = {
        clean(row.get("root_activity_id")): row
        for row in exact_scenario_lcia
        if clean(row.get("scenario_id")) == "current_export"
    }
    rows: list[dict[str, Any]] = []
    comparisons = [
        (
            "production_without_use",
            "Production/hors usage",
            "impact_without_use_person_equivalent",
            exact_by_root.get("production"),
            "Production Brightway OPERA vs Excel hors usage.",
        ),
        (
            "lifecycle_total",
            "Cycle de vie complet",
            "impact_total_person_equivalent",
            exact_by_root.get("lifecycle"),
            "Cycle complet non aligne: l'echange usage Brightway ne reproduit pas l'usage Excel.",
        ),
        (
            "lifecycle_excel_aligned",
            "Cycle complet corrige usage STELIA",
            "impact_total_person_equivalent",
            exact_by_root.get("lifecycle_excel_aligned"),
            "Cycle complet avec usage STELIA calibre et branche tkm brute retiree.",
        ),
    ]
    for scope_id, label, excel_key, exact_row, note in comparisons:
        if not exact_row:
            continue
        excel_pe = safe_float(climate_reference.get(excel_key))
        excel_kg = excel_pe * normalization_factor
        runtime_kg = safe_float(exact_row.get("score_kgco2e"))
        delta_kg = runtime_kg - excel_kg
        relative_delta = abs(delta_kg / excel_kg) if excel_kg else math.inf
        if scope_id == "lifecycle_total":
            alignment_status = "not_aligned"
        elif relative_delta < 0.02:
            alignment_status = "aligned"
        elif relative_delta < 0.2:
            alignment_status = "close_enough_for_poc"
        else:
            alignment_status = "not_aligned"
        rows.append(
            {
                "scope_id": scope_id,
                "label": label,
                "excel_person_equivalent": round(excel_pe, 9),
                "excel_kgco2e": round(excel_kg, 9),
                "runtime_kgco2e": round(runtime_kg, 9),
                "runtime_person_equivalent": round(runtime_kg / normalization_factor, 9) if normalization_factor else "",
                "delta_kgco2e": round(delta_kg, 9),
                "delta_person_equivalent": round(delta_kg / normalization_factor, 9) if normalization_factor else "",
                "relative_delta_pct": round(100.0 * delta_kg / excel_kg, 6) if excel_kg else "",
                "alignment_status": alignment_status,
                "note": note,
            }
        )
    return rows


def parameter_by_name(parameters: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {clean(row.get("name")): row for row in parameters}


def build_usage_calibration_rows(
    reference_use_phase_components: list[dict[str, Any]],
    parameters: list[dict[str, Any]],
    exact_scenario_lcia: list[dict[str, Any]],
    normalization_factor: float,
) -> list[dict[str, Any]]:
    params = parameter_by_name(parameters)
    current_aligned = next(
        (
            row
            for row in exact_scenario_lcia
            if clean(row.get("scenario_id")) == "current_export"
            and clean(row.get("root_activity_id")) == "lifecycle_excel_aligned"
        ),
        {},
    )
    current_raw = next(
        (
            row
            for row in exact_scenario_lcia
            if clean(row.get("scenario_id")) == "current_export"
            and clean(row.get("root_activity_id")) == "lifecycle"
        ),
        {},
    )
    rows: list[dict[str, Any]] = []
    for component in reference_use_phase_components:
        label = clean(component.get("label"))
        component_key = ascii_key(label)
        parameter_name = ""
        physical_amount: float | str = ""
        physical_unit = ""
        calibration_status = "excel_component_calibrated"
        interpretation = ""
        confidence = "medium"
        if "kerosene" in component_key or "jet_a1" in component_key:
            parameter_name = "fauteuil_kero_conso_passive"
            parameter = params.get(parameter_name, {})
            physical_amount = rounded_or_blank(optional_float(parameter.get("amount")), 6)
            physical_unit = "kg kerosene allocated to seat over 7 years"
            interpretation = "Passenger fuel allocation due to seat mass, from OPERA formula."
            confidence = "high"
        elif "cargo" in component_key or "plane" in component_key:
            parameter_name = "fauteuil_cargo_conso_passive"
            parameter = params.get(parameter_name, {})
            physical_amount = rounded_or_blank(optional_float(parameter.get("amount")), 6)
            physical_unit = "unresolved OPERA cargo proxy unit"
            interpretation = "Likely an aircraft mass/use proxy from Sphera cargo plane, not supplier inbound logistics."
            confidence = "low"
        elif "nettoyage" in component_key or "clean" in component_key:
            parameter_name = "fauteuil_nettoyage_savon"
            parameter = params.get(parameter_name, {})
            annual_amount = optional_float(parameter.get("amount"))
            physical_amount = round(annual_amount * 7.0, 6) if annual_amount is not None else ""
            physical_unit = "kg soap proxy over 7 years"
            interpretation = "Marginal cleaning/maintenance proxy; climate contribution is negligible."
            confidence = "medium"
        else:
            parameter = {}
            interpretation = "Excel STELIA use-phase component without explicit OPERA mapping."
            confidence = "low"

        pe = safe_float(component.get("climate_person_equivalent"))
        kgco2e = pe * normalization_factor
        rows.append(
            {
                "usage_component_id": slug(label),
                "label": label,
                "system": clean(component.get("system")),
                "component": clean(component.get("component")),
                "excel_person_equivalent": round(pe, 9),
                "excel_kgco2e": round(kgco2e, 9),
                "person_equivalent_unit": "person eq.",
                "raw_unit": "kgCO2e",
                "parameter_name": parameter_name,
                "physical_amount": physical_amount,
                "physical_unit": physical_unit,
                "parameter_formula": clean(parameter.get("formula")),
                "calibration_status": calibration_status,
                "interpretation": interpretation,
                "confidence": confidence,
                "source_sheet": clean(component.get("source_sheet")),
                "source_row": component.get("source_row"),
            }
        )

    total_use_kg = sum(safe_float(row.get("excel_kgco2e")) for row in rows)
    for row in rows:
        row["share_of_excel_use_pct"] = round(100.0 * safe_float(row.get("excel_kgco2e")) / total_use_kg, 4) if total_use_kg else ""
    if current_aligned:
        rows.append(
            {
                "usage_component_id": "lifecycle_excel_aligned_total",
                "label": "Cycle complet corrige usage STELIA",
                "system": "Cycle de vie",
                "component": "Total",
                "excel_person_equivalent": round(safe_float(current_aligned.get("score_person_equivalent")), 9),
                "excel_kgco2e": round(safe_float(current_aligned.get("score_kgco2e")), 9),
                "person_equivalent_unit": "person eq.",
                "raw_unit": "kgCO2e",
                "parameter_name": "",
                "physical_amount": "",
                "physical_unit": "",
                "parameter_formula": "",
                "calibration_status": "brightway_lifecycle_excel_aligned",
                "interpretation": "Raw tkm use branch removed; Excel STELIA use phase added to Brightway non-use lifecycle.",
                "confidence": "medium",
                "source_sheet": "Brightway runtime",
                "source_row": "",
                "share_of_excel_use_pct": "",
            }
        )
    if current_raw:
        raw_context = current_aligned or current_raw
        rows.append(
            {
                "usage_component_id": "lifecycle_raw_tkm_export",
                "label": "Cycle complet export OPERA brut",
                "system": "Cycle de vie",
                "component": "Total brut",
                "excel_person_equivalent": round(safe_float(current_raw.get("score_person_equivalent")), 9),
                "excel_kgco2e": round(safe_float(current_raw.get("score_kgco2e")), 9),
                "person_equivalent_unit": "person eq.",
                "raw_unit": "kgCO2e",
                "parameter_name": "fauteuil_distance_moyenne_vol*fauteuil_duree_vie*fauteuil_AR_an*fauteuil_masse_totale/1000",
                "physical_amount": raw_context.get("raw_passive_usage_amount", ""),
                "physical_unit": "consommation passive siege tkm units",
                "parameter_formula": "",
                "calibration_status": "raw_export_not_aligned",
                "interpretation": "Raw OPERA export routes seat mass-distance through tkm kerosene estimate and overstates lifecycle climate.",
                "confidence": "high",
                "source_sheet": "Brightway runtime",
                "source_row": "",
                "share_of_excel_use_pct": "",
            }
        )
    return rows


def indicator_unit(indicator: str) -> str:
    start = indicator.rfind("[")
    end = indicator.rfind("]")
    if start >= 0 and end > start:
        return indicator[start + 1 : end].strip()
    return ""


def load_workbook_rows(path: Path, sheet_name: str) -> list[tuple[Any, ...]]:
    if not path.exists() or importlib.util.find_spec("openpyxl") is None:
        return []
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        return [tuple(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def load_brightway_component_impacts(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows = load_workbook_rows(path, "Master")
    if len(rows) < 4:
        return [], []

    max_cols = max(len(row) for row in rows)
    families: list[str] = []
    systems: list[str] = []
    components: list[str] = []
    family = ""
    system = ""
    for col in range(1, max_cols):
        family = clean(rows[0][col] if col < len(rows[0]) else "") or family
        system = clean(rows[1][col] if col < len(rows[1]) else "") or system
        component = clean(rows[2][col] if col < len(rows[2]) else "") or system
        families.append(family)
        systems.append(system)
        components.append(component)

    impact_rows: list[dict[str, Any]] = []
    climate_rows: list[dict[str, Any]] = []
    for row_index in range(3, len(rows)):
        row = rows[row_index]
        indicator = clean(row[0] if row else "")
        if not indicator:
            continue
        unit = indicator_unit(indicator)
        is_climate_total = "climate change - total" in indicator.lower()
        for col in range(1, max_cols):
            value = optional_float(row[col] if col < len(row) else None)
            if value is None:
                continue
            item = {
                "source_file": rel(path, REPO_ROOT),
                "indicator": indicator,
                "unit": unit,
                "family": families[col - 1] if col - 1 < len(families) else "",
                "system": systems[col - 1] if col - 1 < len(systems) else "",
                "component": components[col - 1] if col - 1 < len(components) else "",
                "value": round(value, 9),
            }
            impact_rows.append(item)
            if is_climate_total:
                climate_rows.append(
                    {
                        "source_file": rel(path, REPO_ROOT),
                        "family": item["family"],
                        "system": item["system"],
                        "component": item["component"],
                        "indicator": indicator,
                        "unit": unit,
                        "climate_kgco2e": round(value, 9),
                    }
                )
    return impact_rows, climate_rows


def load_brightway_inventory(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    rows = load_workbook_rows(path, "OPERA")
    if not rows:
        return [], [], []

    parameters: list[dict[str, Any]] = []
    activities: list[dict[str, Any]] = []
    exchanges: list[dict[str, Any]] = []
    mode = ""
    current_activity: dict[str, Any] | None = None
    exchange_header: list[str] = []

    for row in rows:
        first = clean(row[0] if len(row) > 0 else "")
        second = clean(row[1] if len(row) > 1 else "")
        if first == "Project parameters":
            mode = "project_header"
            continue
        if mode == "project_header":
            mode = "project"
            continue
        if mode == "project":
            if not first:
                mode = ""
                continue
            if first in {"Activity", "Parameters", "Exchanges", "name"}:
                mode = ""
            else:
                parameters.append(
                    {
                        "name": first,
                        "amount": optional_float(row[1] if len(row) > 1 else None),
                        "formula": clean(row[2] if len(row) > 2 else ""),
                        "parameter_family": first.split("_", 1)[0] if "_" in first else first,
                    }
                )
                continue

        if first == "Activity":
            current_activity = {"activity_name": second}
            activities.append(current_activity)
            mode = "activity"
            continue
        if current_activity is not None and first in {"code", "comment", "location", "reference product", "type", "unit"}:
            current_activity[slug(first)] = second
            continue
        if first == "Exchanges" and current_activity is not None:
            mode = "exchange_header"
            continue
        if mode == "exchange_header":
            exchange_header = [slug(cell) for cell in row]
            mode = "exchanges"
            continue
        if mode == "exchanges":
            if not any(clean(cell) for cell in row[:8]):
                mode = ""
                exchange_header = []
                continue
            item = {"activity_name": current_activity.get("activity_name", "") if current_activity else ""}
            for idx, header in enumerate(exchange_header):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else ""
                item[header] = optional_float(value) if header in {"amount", "original_amount"} else clean(value)
            exchanges.append(item)

    return parameters, activities, exchanges


def indicator_summary_rows(impact_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_indicator: dict[str, list[float]] = defaultdict(list)
    units: dict[str, str] = {}
    for row in impact_rows:
        indicator = clean(row.get("indicator"))
        value = optional_float(row.get("value"))
        if indicator and value is not None:
            by_indicator[indicator].append(value)
            units[indicator] = clean(row.get("unit"))
    out = []
    for indicator, values in by_indicator.items():
        out.append(
            {
                "indicator": indicator,
                "unit": units.get(indicator, ""),
                "component_count": len(values),
                "sum_value": round(sum(values), 9),
                "mean_value": round(mean(values), 9),
                "max_value": round(max(values), 9),
            }
        )
    return sorted(out, key=lambda row: (-safe_float(row.get("sum_value")), row["indicator"]))


def short_indicator_label(indicator: str) -> str:
    label = re.sub(r"^EF\s*3\.0\s*", "", clean(indicator), flags=re.IGNORECASE)
    label = re.sub(r"\s*\[[^\]]+\]\s*$", "", label).strip()
    return label or clean(indicator)


def ef30_indicator_category(indicator: str) -> tuple[str, str]:
    text = slug(indicator)
    if "climate_change" in text:
        return "Climate change", "aggregate" if "total" in text else "subindicator"
    if "ozone_depletion" in text:
        return "Ozone depletion", "aggregate"
    if "human_toxicity_cancer" in text:
        return "Human toxicity, cancer", "aggregate" if "total" in text else "subindicator"
    if "human_toxicity_non_cancer" in text:
        return "Human toxicity, non-cancer", "aggregate" if "total" in text else "subindicator"
    if "particulate_matter" in text:
        return "Particulate matter", "aggregate"
    if "ionising_radiation" in text:
        return "Ionising radiation", "aggregate"
    if "photochemical_ozone_formation" in text:
        return "Photochemical ozone formation", "aggregate"
    if "acidification" in text:
        return "Acidification", "aggregate"
    if "eutrophication_terrestrial" in text:
        return "Terrestrial eutrophication", "aggregate"
    if "eutrophication_freshwater" in text:
        return "Freshwater eutrophication", "aggregate"
    if "eutrophication_marine" in text:
        return "Marine eutrophication", "aggregate"
    if "land_use" in text:
        return "Land use", "aggregate"
    if "ecotoxicity_freshwater" in text:
        return "Ecotoxicity freshwater", "aggregate" if "total" in text else "subindicator"
    if "water_use" in text:
        return "Water use", "aggregate"
    if "resource_use_fossils" in text:
        return "Resource depletion, fossils", "aggregate"
    if "resource_use_mineral" in text or "resource_use_minerals" in text:
        return "Resource depletion, minerals and metals", "aggregate"
    return "", "unknown"


def build_indicator_unit_views(indicator_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in indicator_rows:
        indicator = clean(row.get("indicator"))
        category, scope = ef30_indicator_category(indicator)
        factor_row = EF30_NORMALIZATION_FACTORS.get(category, {})
        factor = optional_float(factor_row.get("factor"))
        raw_value = safe_float(row.get("sum_value"))
        include_pe = bool(factor is not None and scope == "aggregate")
        if include_pe:
            pe_value: float | str = round(raw_value / factor, 9)
            status = "normalized_ef30_person_equivalent"
        elif factor is not None and scope == "subindicator":
            pe_value = ""
            status = "subindicator_excluded_to_avoid_double_counting"
        else:
            pe_value = ""
            status = "missing_normalization_factor"

        raw_unit = clean(row.get("unit"))
        short_label = short_indicator_label(indicator)
        out.append(
            {
                "indicator": indicator,
                "short_label": short_label,
                "ef30_category": category,
                "indicator_scope": scope,
                "raw_unit": raw_unit,
                "raw_sum_value": round(raw_value, 9),
                "raw_mean_value": row.get("mean_value", ""),
                "raw_max_value": row.get("max_value", ""),
                "component_count": row.get("component_count", ""),
                "normalization_factor_per_person_year": round(factor, 12) if factor is not None else "",
                "normalization_factor_unit": clean(factor_row.get("unit")),
                "person_equivalent_value": pe_value,
                "person_equivalent_unit": "person eq." if include_pe else "",
                "include_in_person_equivalent": include_pe,
                "normalization_status": status,
                "normalization_source": EF30_NORMALIZATION_SOURCE if factor is not None else "",
                "raw_plot_label": f"{short_label} [{raw_unit or 'unit unknown'}]",
                "pe_plot_label": short_label,
            }
        )
    return sorted(
        out,
        key=lambda row: (
            0 if row["include_in_person_equivalent"] else 1,
            -safe_float(row.get("person_equivalent_value"), -1.0),
            -safe_float(row.get("raw_sum_value")),
            row["short_label"],
        ),
    )


def workbook_sheet_by_tokens(workbook: Any, *tokens: str, starts_with: str | None = None) -> str:
    token_keys = [ascii_key(token) for token in tokens]
    start_key = ascii_key(starts_with) if starts_with else ""
    for name in workbook.sheetnames:
        key = ascii_key(name)
        if start_key and not key.startswith(start_key):
            continue
        if all(token in key for token in token_keys):
            return name
    return ""


def cell_number(sheet: Any, row: int, col: int) -> float | None:
    return optional_float(sheet.cell(row, col).value)


def rounded_or_blank(value: float | None, digits: int = 9) -> float | str:
    return round(value, digits) if value is not None else ""


def load_stelia_reference_workbook(path: Path) -> dict[str, list[dict[str, Any]]]:
    empty = {
        "reference_person_equivalent_results": [],
        "reference_weighted_results": [],
        "reference_phase_breakdown": [],
        "reference_scenarios": [],
        "reference_weighting_factors": [],
        "reference_climate_contributors": [],
        "reference_use_phase_components": [],
    }
    if not path.exists() or importlib.util.find_spec("openpyxl") is None:
        return empty

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        perseq_name = workbook_sheet_by_tokens(workbook, "graphes", "perseq")
        transpo_perseq_name = workbook_sheet_by_tokens(workbook, "transpo", "perseq")
        weighted_name = workbook_sheet_by_tokens(workbook, "graphes", "pond")
        weighting_name = workbook_sheet_by_tokens(workbook, starts_with="pond")
        scenarii_name = workbook_sheet_by_tokens(workbook, "scenarii")

        out = {key: list(value) for key, value in empty.items()}

        if perseq_name:
            sheet = workbook[perseq_name]
            for row_idx in range(23, 39):
                indicator = clean(sheet.cell(row_idx, 1).value)
                if not indicator:
                    continue
                out["reference_person_equivalent_results"].append(
                    {
                        "indicator": indicator,
                        "short_label": short_indicator_label(indicator),
                        "impact_total_person_equivalent": rounded_or_blank(cell_number(sheet, row_idx, 2)),
                        "impact_without_use_person_equivalent": rounded_or_blank(cell_number(sheet, row_idx, 3)),
                        "use_phase_person_equivalent": rounded_or_blank(cell_number(sheet, row_idx, 4)),
                        "person_equivalent_unit": "person eq.",
                        "source_sheet": perseq_name,
                        "source_row": row_idx,
                    }
                )

            phases = [clean(sheet.cell(44, col).value) for col in range(2, 7)]
            for row_idx in range(45, 54):
                indicator = clean(sheet.cell(row_idx, 1).value)
                if not indicator:
                    continue
                for offset, phase in enumerate(phases, start=2):
                    value = cell_number(sheet, row_idx, offset)
                    if not phase or value is None:
                        continue
                    out["reference_phase_breakdown"].append(
                        {
                            "result_view": "person_equivalent",
                            "end_of_life_variant": "landfill",
                            "indicator": indicator,
                            "short_label": short_indicator_label(indicator),
                            "phase": phase,
                            "value": round(value, 9),
                            "unit": "person eq.",
                            "source_sheet": perseq_name,
                            "source_row": row_idx,
                        }
                    )

            for scope, start_row, end_row in (
                ("system", 88, 97),
                ("component", 103, 108),
            ):
                for row_idx in range(start_row, end_row + 1):
                    label = clean(sheet.cell(row_idx, 1).value)
                    value = cell_number(sheet, row_idx, 2)
                    share = cell_number(sheet, row_idx, 3)
                    if not label or value is None:
                        continue
                    out["reference_climate_contributors"].append(
                        {
                            "contributor_scope": scope,
                            "label": label,
                            "climate_person_equivalent": round(value, 9),
                            "share_of_non_use_climate_pct": round(100.0 * share, 4) if share is not None else "",
                            "source_sheet": perseq_name,
                            "source_row": row_idx,
                        }
                    )

        if transpo_perseq_name:
            sheet = workbook[transpo_perseq_name]
            climate_col = 0
            for col_idx in range(1, sheet.max_column + 1):
                if short_indicator_label(clean(sheet.cell(1, col_idx).value)) == "Climate Change - total":
                    climate_col = col_idx
                    break
            if climate_col:
                for row_idx in range(2, sheet.max_row + 1):
                    phase = clean(sheet.cell(row_idx, 1).value)
                    system = clean(sheet.cell(row_idx, 2).value)
                    component = clean(sheet.cell(row_idx, 3).value)
                    value = cell_number(sheet, row_idx, climate_col)
                    if phase != "Utilisation 7 ans" or value is None:
                        continue
                    out["reference_use_phase_components"].append(
                        {
                            "phase": phase,
                            "system": system,
                            "component": component,
                            "label": f"{system} - {component}" if system and component else component or system,
                            "climate_person_equivalent": round(value, 9),
                            "person_equivalent_unit": "person eq.",
                            "source_sheet": transpo_perseq_name,
                            "source_row": row_idx,
                        }
                    )

        if weighted_name:
            sheet = workbook[weighted_name]
            for row_idx in range(23, 39):
                indicator = clean(sheet.cell(row_idx, 1).value)
                if not indicator:
                    continue
                out["reference_weighted_results"].append(
                    {
                        "indicator": indicator,
                        "short_label": short_indicator_label(indicator),
                        "impact_total_weighted_score": rounded_or_blank(cell_number(sheet, row_idx, 2)),
                        "per_cabin_16_seats_weighted_score": rounded_or_blank(cell_number(sheet, row_idx, 3)),
                        "impact_without_use_weighted_score": rounded_or_blank(cell_number(sheet, row_idx, 4)),
                        "use_phase_weighted_score": rounded_or_blank(cell_number(sheet, row_idx, 5)),
                        "weighted_unit": "weighted person eq.",
                        "source_sheet": weighted_name,
                        "source_row": row_idx,
                    }
                )

            for header_row, start_row, end_row, variant in (
                (44, 45, 53, "landfill"),
                (56, 57, 65, "recycling"),
            ):
                phases = [clean(sheet.cell(header_row, col).value) for col in range(2, 7)]
                for row_idx in range(start_row, end_row + 1):
                    indicator = clean(sheet.cell(row_idx, 1).value)
                    if not indicator:
                        continue
                    for offset, phase in enumerate(phases, start=2):
                        value = cell_number(sheet, row_idx, offset)
                        if not phase or value is None:
                            continue
                        out["reference_phase_breakdown"].append(
                            {
                                "result_view": "weighted",
                                "end_of_life_variant": variant,
                                "indicator": indicator,
                                "short_label": short_indicator_label(indicator),
                                "phase": phase,
                                "value": round(value, 9),
                                "unit": "weighted person eq.",
                                "source_sheet": weighted_name,
                                "source_row": row_idx,
                            }
                        )

        if weighting_name:
            sheet = workbook[weighting_name]
            for row_idx in range(2, 18):
                category = clean(sheet.cell(row_idx, 1).value)
                if not category:
                    continue
                equal_weight = cell_number(sheet, row_idx, 2)
                ef30_weight = cell_number(sheet, row_idx, 3)
                factor = cell_number(sheet, row_idx, 4)
                out["reference_weighting_factors"].append(
                    {
                        "category": category,
                        "equal_weight": rounded_or_blank(equal_weight, 12),
                        "ef30_weight_pct": round(100.0 * ef30_weight, 6) if ef30_weight is not None else "",
                        "weighting_factor_vs_equal_weight": rounded_or_blank(factor, 9),
                        "source_sheet": weighting_name,
                        "source_row": row_idx,
                    }
                )

        if scenarii_name:
            sheet = workbook[scenarii_name]

            def add_scenario_rows(scenario_id: str, scenario_label: str, rows: range, mode: str) -> None:
                for row_idx in rows:
                    phase = clean(sheet.cell(row_idx, 1).value)
                    if not phase:
                        continue
                    if mode == "sans_ife":
                        lifecycle_reduction = cell_number(sheet, row_idx, 2)
                        relative_reduction = cell_number(sheet, row_idx, 3)
                        production_reduction = cell_number(sheet, row_idx, 4)
                        baseline = cell_number(sheet, row_idx, 5)
                        scenario = cell_number(sheet, row_idx, 6)
                    else:
                        production_reduction = cell_number(sheet, row_idx, 2)
                        lifecycle_reduction = cell_number(sheet, row_idx, 3)
                        relative_reduction = cell_number(sheet, row_idx, 4)
                        baseline = cell_number(sheet, row_idx, 5)
                        scenario = cell_number(sheet, row_idx, 6)
                    out["reference_scenarios"].append(
                        {
                            "scenario_id": scenario_id,
                            "scenario_label": scenario_label,
                            "phase": phase,
                            "baseline_climate_weighted_score": rounded_or_blank(baseline),
                            "scenario_climate_weighted_score": rounded_or_blank(scenario),
                            "lifecycle_climate_reduction_pct": round(100.0 * lifecycle_reduction, 6) if lifecycle_reduction is not None else "",
                            "relative_climate_reduction_pct": round(100.0 * relative_reduction, 6) if relative_reduction is not None else "",
                            "production_only_climate_reduction_pct": round(100.0 * production_reduction, 6) if production_reduction is not None else "",
                            "weighted_unit": "weighted person eq.",
                            "source_sheet": scenarii_name,
                            "source_row": row_idx,
                        }
                    )

            add_scenario_rows("without_ife", "Sans IFE", range(4, 11), "sans_ife")
            add_scenario_rows("all_fr", "Fabrication 100% sur site / mix FR", range(26, 33), "all_fr")

        for key in out:
            out[key].sort(key=lambda row: (clean(row.get("scenario_id")), clean(row.get("result_view")), -safe_float(row.get("impact_total_person_equivalent"), -safe_float(row.get("impact_total_weighted_score"), -safe_float(row.get("value")))), clean(row.get("label") or row.get("indicator") or row.get("phase"))))
        return out
    finally:
        workbook.close()


def load_masterboard_bom_summaries(path: Path) -> dict[str, list[dict[str, Any]]]:
    empty = {"masterboard_equipment_summary": [], "masterboard_material_summary": []}
    if not path.exists() or importlib.util.find_spec("openpyxl") is None:
        return empty

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = workbook_sheet_by_tokens(workbook, starts_with="bom")
        if not sheet_name:
            return empty
        sheet = workbook[sheet_name]
        equipment: dict[str, dict[str, Any]] = {}
        material: dict[str, dict[str, Any]] = {}

        def add(group: dict[str, dict[str, Any]], key: str, qty: float, packaging: float, chips: float, raw: float) -> None:
            row = group.setdefault(
                key or "unknown",
                {
                    "label": key or "unknown",
                    "row_count": 0,
                    "quantity_per_seat": 0.0,
                    "packaging_mass": 0.0,
                    "chips_mass": 0.0,
                    "raw_mass": 0.0,
                },
            )
            row["row_count"] += 1
            row["quantity_per_seat"] += qty
            row["packaging_mass"] += packaging
            row["chips_mass"] += chips
            row["raw_mass"] += raw

        for row_idx in range(2, sheet.max_row + 1):
            equipment_label = clean(sheet.cell(row_idx, 1).value)
            material_label = clean(sheet.cell(row_idx, 4).value)
            qty = safe_float(sheet.cell(row_idx, 7).value)
            packaging = safe_float(sheet.cell(row_idx, 9).value)
            chips = safe_float(sheet.cell(row_idx, 10).value)
            raw = safe_float(sheet.cell(row_idx, 12).value)
            if not any([equipment_label, material_label, qty, packaging, chips, raw]):
                continue
            add(equipment, equipment_label, qty, packaging, chips, raw)
            add(material, material_label, qty, packaging, chips, raw)

        def finalize(group: dict[str, dict[str, Any]], group_type: str) -> list[dict[str, Any]]:
            rows: list[dict[str, Any]] = []
            for row in group.values():
                qty = safe_float(row["quantity_per_seat"])
                rows.append(
                    {
                        "group_type": group_type,
                        "label": row["label"],
                        "row_count": row["row_count"],
                        "quantity_per_seat": round(qty, 9),
                        "packaging_mass": round(safe_float(row["packaging_mass"]), 9),
                        "chips_mass": round(safe_float(row["chips_mass"]), 9),
                        "raw_mass": round(safe_float(row["raw_mass"]), 9),
                        "material_domain": "use_or_energy" if qty > 100.0 else "component_or_packaging",
                        "source_file": rel(path, REPO_ROOT),
                        "source_sheet": sheet_name,
                    }
                )
            return sorted(rows, key=lambda item: (-safe_float(item.get("quantity_per_seat")), item["label"]))[:80]

        return {
            "masterboard_equipment_summary": finalize(equipment, "equipment"),
            "masterboard_material_summary": finalize(material, "material"),
        }
    finally:
        workbook.close()


def match_brightway_component(system: str, component: str, climate_rows: list[dict[str, Any]]) -> tuple[dict[str, Any] | None, str]:
    def bw_match_slug(value: Any) -> str:
        text = slug(value)
        if "padding" in text:
            return "padding"
        if "screen_display" in text or "display_liquid_crystal" in text:
            return "ecran"
        return text

    system_slug = bw_match_slug(system)
    component_slug = bw_match_slug(component)
    exact: dict[tuple[str, str], dict[str, Any]] = {}
    by_component: dict[str, dict[str, Any]] = {}
    for row in climate_rows:
        bw_system_slug = bw_match_slug(row.get("system"))
        bw_component_slug = bw_match_slug(row.get("component"))
        exact[(bw_system_slug, bw_component_slug)] = row
        by_component.setdefault(bw_component_slug, row)
    if (system_slug, component_slug) in exact:
        return exact[(system_slug, component_slug)], "exact_system_component"
    if component_slug in by_component:
        return by_component[component_slug], "component_exact"
    for row in climate_rows:
        bw_component_slug = slug(row.get("component"))
        bw_system_slug = slug(row.get("system"))
        if component_slug and len(component_slug) >= 4 and (component_slug in bw_component_slug or bw_component_slug in component_slug):
            return row, "component_fuzzy"
        if system_slug and len(system_slug) >= 4 and system_slug == bw_system_slug:
            return row, "system_only"
    return None, "unmatched"


def build_brightway_supply_alignment(path_rows: list[dict[str, Any]], climate_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in path_rows:
        match, level = match_brightway_component(clean(row.get("system")), clean(row.get("component")), climate_rows)
        if match is None and safe_float(row.get("path_mass_kg")) <= 0.0:
            level = "not_required_zero_mass"
        out.append(
            {
                "path_id": row.get("path_id", ""),
                "record_index": row.get("record_index", ""),
                "system": row.get("system", ""),
                "component": row.get("component", ""),
                "family": row.get("family", ""),
                "path_mass_kg": row.get("path_mass_kg", 0.0),
                "match_level": level,
                "brightway_system": match.get("system", "") if match else "",
                "brightway_component": match.get("component", "") if match else "",
                "brightway_climate_kgco2e": match.get("climate_kgco2e", "") if match else "",
                "brightway_unit": match.get("unit", "") if match else "",
            }
        )
    return out


def eval_parametric_formula(expr: str, values: dict[str, float]) -> float:
    tree = ast.parse(expr.replace("^", "**"), mode="eval")

    def visit(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return visit(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.Num):
            return float(node.n)
        if isinstance(node, ast.Name):
            if node.id not in values:
                raise KeyError(node.id)
            return values[node.id]
        if isinstance(node, ast.UnaryOp):
            value = visit(node.operand)
            if isinstance(node.op, ast.USub):
                return -value
            if isinstance(node.op, ast.UAdd):
                return value
        if isinstance(node, ast.BinOp):
            left = visit(node.left)
            right = visit(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                return left / right
            if isinstance(node.op, ast.Pow):
                return left**right
        raise ValueError(f"Unsupported formula node: {type(node).__name__}")

    value = visit(tree)
    if not math.isfinite(value):
        raise ValueError("non-finite formula result")
    return value


def formula_names(expr: str) -> set[str]:
    try:
        tree = ast.parse(expr.replace("^", "**"), mode="eval")
    except SyntaxError:
        return set()
    return {node.id for node in ast.walk(tree) if isinstance(node, ast.Name)}


def parameter_tokens(name: str) -> set[str]:
    return {part for part in re.split(r"[^a-z0-9]+", clean(name).lower()) if part}


def param_is_aluminium_material(name: str) -> bool:
    tokens = parameter_tokens(name)
    has_al = any(token.startswith("al") or token.startswith("alu") for token in tokens)
    return has_al and "elec" not in tokens and "eol" not in tokens and "inc" not in tokens and "recy" not in tokens


def param_is_electricity(name: str) -> bool:
    return "elec" in parameter_tokens(name)


def param_is_transport(name: str) -> bool:
    tokens = parameter_tokens(name)
    return bool(tokens & {"truck", "tkm", "transport", "boat", "ship", "plane", "freight"})


def param_is_packaging(name: str) -> bool:
    tokens = parameter_tokens(name)
    return bool(tokens & {"pack", "box", "carton", "ply", "pe", "film"})


def param_is_end_of_life(name: str) -> bool:
    tokens = parameter_tokens(name)
    return bool(tokens & {"eol", "inc", "valo", "recy", "scrap", "waste", "landfill"})


def param_is_steel(name: str) -> bool:
    tokens = parameter_tokens(name)
    return bool(tokens & {"acier", "steel", "inox", "30ncd", "30ncd6", "35nc", "30nc", "4140", "z10", "z15"})


def param_is_polymer_textile(name: str) -> bool:
    tokens = parameter_tokens(name)
    return bool(tokens & {"nylon", "pa6", "pa66", "pu", "pvc", "kydex", "frmc", "tissu", "cuir", "velour", "velours", "silicone", "caoutchouc", "ertalon", "pc", "pether"})


PARAMETRIC_LEVERS = (
    ("aluminium_material", "Aluminium matiere +10%", param_is_aluminium_material),
    ("electricity", "Electricite process +10%", param_is_electricity),
    ("transport", "Transport foreground +10%", param_is_transport),
    ("packaging", "Packaging +10%", param_is_packaging),
    ("end_of_life", "Fin de vie / recyclage +10%", param_is_end_of_life),
    ("steel_inox", "Acier / inox +10%", param_is_steel),
    ("polymer_textile", "Polymere / textile +10%", param_is_polymer_textile),
)

EF30_NORMALIZATION_SOURCE = "https://eplca.jrc.ec.europa.eu/permalink/Normalisation_Weighting_Factors_EF_3.0.xlsx"

EF30_NORMALIZATION_FACTORS = {
    "Climate change": {"factor": 8095.525063944057, "unit": "kg CO2 eq./person"},
    "Ozone depletion": {"factor": 0.05364799056726336, "unit": "kg CFC-11 eq./person"},
    "Human toxicity, cancer": {"factor": 1.689950739575603e-05, "unit": "CTUh/person"},
    "Human toxicity, non-cancer": {"factor": 0.0002296592158999324, "unit": "CTUh/person"},
    "Particulate matter": {"factor": 0.000595386937135986, "unit": "disease incidences/person"},
    "Ionising radiation": {"factor": 4220.15981253385, "unit": "kBq U-235 eq./person"},
    "Photochemical ozone formation": {"factor": 40.601397461454425, "unit": "kg NMVOC eq./person"},
    "Acidification": {"factor": 55.569541230602006, "unit": "mol H+ eq./person"},
    "Terrestrial eutrophication": {"factor": 176.754999788942, "unit": "mol N eq./person"},
    "Freshwater eutrophication": {"factor": 1.6068521282881312, "unit": "kg P eq./person"},
    "Marine eutrophication": {"factor": 19.54518155191913, "unit": "kg N eq./person"},
    "Land use": {"factor": 819498.1829230306, "unit": "pt/person"},
    "Ecotoxicity freshwater": {"factor": 42683.16186559794, "unit": "CTUe/person"},
    "Water use": {"factor": 11468.708640759718, "unit": "m3 water eq of deprived water/person"},
    "Resource depletion, fossils": {"factor": 65004.259664016674, "unit": "MJ/person"},
    "Resource depletion, minerals and metals": {"factor": 0.06364027822595558, "unit": "kg Sb eq./person"},
}

SUPPLIER_ROLES_FOR_LOCALIZATION = ("t4", "t3", "t2", "t1")

REGIONALIZATION_SCENARIOS = (
    {
        "scenario_id": "current_export",
        "label": "Baseline export Brightway",
        "target_scope": "current",
        "description": "Etat exporte depuis bw_tristan et la supply_geo source.",
        "elec_switch_param": "",
        "al_switch_param": "",
        "transport_policy": "current routes",
        "transport_amount_factor": 1.0,
        "local_content_target_pct": "",
    },
    {
        "scenario_id": "france_first",
        "label": "100% francais si disponible",
        "target_scope": "france",
        "description": "Electricite FR, fournisseurs FR des que possible, transport localise. Aluminium force en EU car le switch FR n'existe pas encore dans bw_tristan.",
        "elec_switch_param": "fr",
        "al_switch_param": "eu",
        "transport_policy": "short_france",
        "transport_amount_factor": 0.45,
        "local_content_target_pct": 100.0,
    },
    {
        "scenario_id": "europe_first",
        "label": "100% europeen si disponible",
        "target_scope": "europe",
        "description": "Electricite EU, aluminium EU, fournisseurs europeens et distances intercontinentales reduites.",
        "elec_switch_param": "eu",
        "al_switch_param": "eu",
        "transport_policy": "europeanized",
        "transport_amount_factor": 0.70,
        "local_content_target_pct": 100.0,
    },
    {
        "scenario_id": "fully_globalized",
        "label": "Totalement mondialise",
        "target_scope": "world",
        "description": "Mix electrique et aluminium rest-of-world, chaines longues et transports intercontinentaux amplifies.",
        "elec_switch_param": "cn",
        "al_switch_param": "row",
        "transport_policy": "long_global",
        "transport_amount_factor": 1.35,
        "local_content_target_pct": 0.0,
    },
)


def build_parametric_switches(exchanges: list[dict[str, Any]]) -> list[dict[str, Any]]:
    electricity_count = sum(1 for row in exchanges if "electricity" in clean(row.get("name")).lower() or "electric" in clean(row.get("name")).lower())
    aluminium_count = sum(1 for row in exchanges if "aluminium" in clean(row.get("name")).lower())
    return [
        {
            "switch_id": "elec_switch_param",
            "label": "Mix electricite regional",
            "values": "fr|eu|us|cn",
            "affected_exchange_count": electricity_count,
            "requires_brightway_runtime": True,
            "status": "scripted_in_bw_tristan",
        },
        {
            "switch_id": "al_switch_param",
            "label": "Mix aluminium primaire",
            "values": "eu|us|cn|row",
            "affected_exchange_count": aluminium_count,
            "requires_brightway_runtime": True,
            "status": "scripted_in_bw_tristan",
        },
    ]


def country_in_target_scope(country_code: str, target_scope: str) -> bool:
    code = clean(country_code).upper()
    if target_scope == "france":
        return code == "FR"
    if target_scope == "europe":
        return code in REGION_COUNTRIES["Europe"]
    if target_scope == "world":
        return True
    return False


def supply_localization_share(path_rows: list[dict[str, Any]], target_scope: str) -> dict[str, float | str]:
    if target_scope not in {"france", "europe", "world"}:
        return {
            "current_role_mass_already_target_pct": "",
            "current_path_mass_already_target_pct": "",
            "current_non_target_role_mass_pct": "",
        }

    total_path_mass = sum(safe_float(row.get("path_mass_kg")) for row in path_rows)
    role_denominator = total_path_mass * len(SUPPLIER_ROLES_FOR_LOCALIZATION)
    role_target_mass = 0.0
    path_target_mass = 0.0
    for row in path_rows:
        path_mass = safe_float(row.get("path_mass_kg"))
        role_matches = [
            country_in_target_scope(clean(row.get(f"{role}_country_code")), target_scope)
            for role in SUPPLIER_ROLES_FOR_LOCALIZATION
        ]
        role_target_mass += path_mass * sum(1 for matched in role_matches if matched)
        if role_matches and all(role_matches):
            path_target_mass += path_mass

    role_pct = 100.0 * role_target_mass / role_denominator if role_denominator else 0.0
    path_pct = 100.0 * path_target_mass / total_path_mass if total_path_mass else 0.0
    return {
        "current_role_mass_already_target_pct": round(role_pct, 4),
        "current_path_mass_already_target_pct": round(path_pct, 4),
        "current_non_target_role_mass_pct": round(max(0.0, 100.0 - role_pct), 4),
    }


def scenario_formula_delta(
    evaluated: list[dict[str, Any]],
    parameter_values: dict[str, float],
    scaled_params: set[str],
    scale: float,
) -> dict[str, float | int]:
    if not scaled_params or math.isclose(scale, 1.0):
        return {
            "affected_exchange_count": 0,
            "baseline_abs_amount_sum": 0.0,
            "scenario_abs_amount_sum": 0.0,
            "signed_delta_amount_sum": 0.0,
            "abs_delta_amount_sum": 0.0,
            "foreground_amount_index": 1.0,
        }

    scenario_values = dict(parameter_values)
    for name in scaled_params:
        scenario_values[name] = scenario_values[name] * scale

    affected_count = 0
    baseline_abs = 0.0
    scenario_abs = 0.0
    signed_delta = 0.0
    abs_delta = 0.0
    for item in evaluated:
        if not item["names"].intersection(scaled_params):
            continue
        try:
            scenario = eval_parametric_formula(item["formula"], scenario_values)
        except Exception:
            continue
        baseline = safe_float(item["baseline"])
        delta = scenario - baseline
        if abs(delta) <= 1e-12:
            continue
        affected_count += 1
        baseline_abs += abs(baseline)
        scenario_abs += abs(scenario)
        signed_delta += delta
        abs_delta += abs(delta)

    return {
        "affected_exchange_count": affected_count,
        "baseline_abs_amount_sum": round(baseline_abs, 9),
        "scenario_abs_amount_sum": round(scenario_abs, 9),
        "signed_delta_amount_sum": round(signed_delta, 9),
        "abs_delta_amount_sum": round(abs_delta, 9),
        "foreground_amount_index": round(scenario_abs / baseline_abs, 6) if baseline_abs else 1.0,
    }


def build_regionalization_scenarios(
    *,
    parameter_values: dict[str, float],
    evaluated: list[dict[str, Any]],
    exchanges: list[dict[str, Any]],
    path_rows: list[dict[str, Any]],
    runtime_available: bool,
) -> list[dict[str, Any]]:
    transport_params = {name for name in parameter_values if param_is_transport(name)}
    electricity_count = sum(1 for row in exchanges if "electricity" in clean(row.get("name")).lower() or "electric" in clean(row.get("name")).lower())
    aluminium_count = sum(1 for row in exchanges if "aluminium" in clean(row.get("name")).lower())
    total_kg_km = sum(safe_float(row.get("allocated_kg_km")) for row in path_rows)

    rows: list[dict[str, Any]] = []
    for scenario in REGIONALIZATION_SCENARIOS:
        transport_factor = safe_float(scenario.get("transport_amount_factor"), 1.0)
        delta = scenario_formula_delta(evaluated, parameter_values, transport_params, transport_factor)
        localization = supply_localization_share(path_rows, clean(scenario.get("target_scope")))
        switch_count = 0
        if clean(scenario.get("elec_switch_param")):
            switch_count += electricity_count
        if clean(scenario.get("al_switch_param")):
            switch_count += aluminium_count
        requires_runtime = bool(switch_count)
        rows.append(
            {
                "scenario_id": scenario["scenario_id"],
                "label": scenario["label"],
                "description": scenario["description"],
                "target_scope": scenario["target_scope"],
                "elec_switch_param": scenario["elec_switch_param"],
                "al_switch_param": scenario["al_switch_param"],
                "transport_policy": scenario["transport_policy"],
                "transport_amount_factor": transport_factor,
                "local_content_target_pct": scenario["local_content_target_pct"],
                "current_role_mass_already_target_pct": localization["current_role_mass_already_target_pct"],
                "current_path_mass_already_target_pct": localization["current_path_mass_already_target_pct"],
                "current_non_target_role_mass_pct": localization["current_non_target_role_mass_pct"],
                "brightway_switch_affected_exchange_count": switch_count,
                "transport_formula_affected_exchange_count": delta["affected_exchange_count"],
                "affected_exchange_count": int(switch_count + safe_float(delta["affected_exchange_count"])),
                "baseline_transport_proxy_kg_km": round(total_kg_km, 3),
                "scenario_transport_proxy_kg_km": round(total_kg_km * transport_factor, 3),
                "transport_proxy_delta_kg_km": round(total_kg_km * (transport_factor - 1.0), 3),
                "baseline_abs_foreground_amount_sum": delta["baseline_abs_amount_sum"],
                "scenario_abs_foreground_amount_sum": delta["scenario_abs_amount_sum"],
                "signed_delta_amount_sum": delta["signed_delta_amount_sum"],
                "abs_delta_amount_sum": delta["abs_delta_amount_sum"],
                "foreground_amount_index": delta["foreground_amount_index"],
                "requires_brightway_runtime": requires_runtime,
                "can_execute_exact_lcia_now": bool(runtime_available) if requires_runtime else True,
                "proxy_available_now": True,
                "unit_note": "Transport kg.km is a supply_geo proxy; foreground amount sums mix exchange units. Exact LCIA scenario deltas require Brightway runtime.",
            }
        )
    return rows


def build_parametric_scenarios(parameters: list[dict[str, Any]], exchanges: list[dict[str, Any]], path_rows: list[dict[str, Any]], runtime_available: bool) -> dict[str, Any]:
    parameter_values = {
        clean(row.get("name")): safe_float(row.get("amount"))
        for row in parameters
        if clean(row.get("name")) and math.isfinite(safe_float(row.get("amount"), float("nan")))
    }
    formula_rows = [row for row in exchanges if clean(row.get("formula"))]
    evaluated: list[dict[str, Any]] = []
    formula_failures = 0
    for row in formula_rows:
        formula = clean(row.get("formula"))
        try:
            baseline = eval_parametric_formula(formula, parameter_values)
        except Exception:
            formula_failures += 1
            continue
        names = formula_names(formula)
        evaluated.append({"row": row, "formula": formula, "baseline": baseline, "names": names})

    sensitivity_rows: list[dict[str, Any]] = []
    lever_rows: list[dict[str, Any]] = []
    scale = 1.10
    for lever_id, label, predicate in PARAMETRIC_LEVERS:
        lever_params = {name for name in parameter_values if predicate(name)}
        if not lever_params:
            continue
        scenario_values = dict(parameter_values)
        for name in lever_params:
            scenario_values[name] = scenario_values[name] * scale

        lever_details: list[dict[str, Any]] = []
        for item in evaluated:
            if not item["names"].intersection(lever_params):
                continue
            try:
                scenario = eval_parametric_formula(item["formula"], scenario_values)
            except Exception:
                continue
            baseline = safe_float(item["baseline"])
            delta = scenario - baseline
            if abs(delta) <= 1e-12:
                continue
            source = item["row"]
            detail = {
                "lever_id": lever_id,
                "lever_label": label,
                "activity_name": source.get("activity_name", ""),
                "exchange_name": source.get("name", ""),
                "exchange_type": source.get("type", ""),
                "database": source.get("database", ""),
                "location": source.get("location", ""),
                "unit": source.get("unit", ""),
                "formula": item["formula"],
                "baseline_amount": round(baseline, 12),
                "scenario_amount": round(scenario, 12),
                "delta_amount": round(delta, 12),
                "abs_delta_amount": round(abs(delta), 12),
                "relative_delta_pct": round(100.0 * delta / baseline, 4) if baseline else "",
                "parameter_count_in_formula": len(item["names"].intersection(lever_params)),
                "label": f"{source.get('activity_name', '')} / {source.get('name', '')}",
            }
            lever_details.append(detail)
            sensitivity_rows.append(detail)
        lever_rows.append(
            {
                "lever_id": lever_id,
                "lever_label": label,
                "parameter_count": len(lever_params),
                "affected_exchange_count": len(lever_details),
                "signed_delta_amount_sum": round(sum(safe_float(row.get("delta_amount")) for row in lever_details), 9),
                "abs_delta_amount_sum": round(sum(safe_float(row.get("abs_delta_amount")) for row in lever_details), 9),
                "mean_abs_relative_delta_pct": round(mean(abs(safe_float(row.get("relative_delta_pct"))) for row in lever_details if row.get("relative_delta_pct") != ""), 4),
            }
        )

    sensitivity_rows.sort(key=lambda row: (-safe_float(row.get("abs_delta_amount")), row["lever_id"], row["activity_name"]))
    lever_rows.sort(key=lambda row: (-safe_float(row.get("abs_delta_amount_sum")), row["lever_label"]))
    switch_rows = build_parametric_switches(exchanges)
    return {
        "parametric_levers": lever_rows,
        "parametric_sensitivity": sensitivity_rows,
        "parametric_switches": switch_rows,
        "parametric_regional_scenarios": build_regionalization_scenarios(
            parameter_values=parameter_values,
            evaluated=evaluated,
            exchanges=exchanges,
            path_rows=path_rows,
            runtime_available=runtime_available,
        ),
        "formula_count": len(formula_rows),
        "formula_evaluated_count": len(evaluated),
        "formula_failure_count": formula_failures,
    }


def build_brightway_model_payload(path_rows: list[dict[str, Any]]) -> dict[str, Any]:
    root = BW_TRISTAN_ROOT
    impacts_path = root / "STELIALCASEATS.xlsx"
    reference_results_path = root / "STELIA LCA SEATS v14022022v2.xlsx"
    masterboard_path = root / "STELIA Masterboard LCA SEATS 6.0.xlsx"
    inventory_path = root / "opera_bw2 - inventaire.xlsx"
    package_path = root / "OPERA_siege.bw2package"
    impact_rows, climate_rows = load_brightway_component_impacts(impacts_path)
    reference_results = load_stelia_reference_workbook(reference_results_path)
    masterboard_summaries = load_masterboard_bom_summaries(masterboard_path)
    parameters, activities, exchanges = load_brightway_inventory(inventory_path)
    alignment = build_brightway_supply_alignment(path_rows, climate_rows)
    runtime = brightway_runtime_status()
    indicator_summary = indicator_summary_rows(impact_rows)
    indicator_unit_views = build_indicator_unit_views(indicator_summary)
    normalisation_factor = climate_normalization_factor(indicator_unit_views)
    climate_reference = next(
        (
            row
            for row in reference_results["reference_person_equivalent_results"]
            if clean(row.get("short_label")) == "Climate Change - total"
        ),
        {},
    )
    excel_use_phase_pe = safe_float(climate_reference.get("use_phase_person_equivalent"))
    if excel_use_phase_pe <= 0.0:
        excel_use_phase_pe = 56.92001
    exact_scenario_lcia = run_brightway_exact_scenarios(runtime, normalisation_factor, excel_use_phase_pe)
    excel_runtime_comparison = build_excel_runtime_comparison(
        reference_results["reference_person_equivalent_results"],
        exact_scenario_lcia,
        normalisation_factor,
    )
    usage_calibration = build_usage_calibration_rows(
        reference_results["reference_use_phase_components"],
        parameters,
        exact_scenario_lcia,
        normalisation_factor,
    )
    parametric = build_parametric_scenarios(parameters, exchanges, path_rows, bool(runtime.get("can_execute_brightway")))
    matched = [row for row in alignment if row["match_level"] != "unmatched"]
    contributive = [row for row in alignment if safe_float(row.get("path_mass_kg")) > 0.0]
    contributive_matched = [row for row in contributive if row["match_level"] not in {"unmatched", "not_required_zero_mass"}]
    zero_mass_not_required = [row for row in alignment if row["match_level"] == "not_required_zero_mass"]
    return {
        "schema_version": "poc2026.supply_geo_case.brightway_model.v1",
        "available": impacts_path.exists() or inventory_path.exists() or package_path.exists(),
        "runtime": runtime,
        "source_files": {
            "impact_results_xlsx": rel(impacts_path, REPO_ROOT),
            "reference_results_xlsx": rel(reference_results_path, REPO_ROOT),
            "masterboard_xlsx": rel(masterboard_path, REPO_ROOT),
            "inventory_xlsx": rel(inventory_path, REPO_ROOT),
            "bw2package": rel(package_path, REPO_ROOT),
            "ef30_normalization_factors": EF30_NORMALIZATION_SOURCE,
            "brightway_exact_scenario_script": rel(CASE_ROOT / "tools" / "run_brightway_exact_scenarios.py", REPO_ROOT),
        },
        "counts": {
            "impact_rows": len(impact_rows),
            "climate_component_rows": len(climate_rows),
            "indicator_unit_views": len(indicator_unit_views),
            "person_equivalent_indicators": sum(1 for row in indicator_unit_views if row["include_in_person_equivalent"]),
            "reference_person_equivalent_results": len(reference_results["reference_person_equivalent_results"]),
            "reference_weighted_results": len(reference_results["reference_weighted_results"]),
            "reference_phase_breakdown_rows": len(reference_results["reference_phase_breakdown"]),
            "reference_use_phase_components": len(reference_results["reference_use_phase_components"]),
            "reference_scenarios": len(reference_results["reference_scenarios"]),
            "masterboard_equipment_summary_rows": len(masterboard_summaries["masterboard_equipment_summary"]),
            "masterboard_material_summary_rows": len(masterboard_summaries["masterboard_material_summary"]),
            "parameters": len(parameters),
            "activities": len(activities),
            "exchanges": len(exchanges),
            "parametric_formulas": parametric["formula_count"],
            "parametric_formulas_evaluated": parametric["formula_evaluated_count"],
            "parametric_formula_failures": parametric["formula_failure_count"],
            "parametric_levers": len(parametric["parametric_levers"]),
            "parametric_sensitivity_rows": len(parametric["parametric_sensitivity"]),
            "parametric_switches": len(parametric["parametric_switches"]),
            "parametric_regional_scenarios": len(parametric["parametric_regional_scenarios"]),
            "exact_scenario_lcia": len(exact_scenario_lcia),
            "excel_runtime_comparison": len(excel_runtime_comparison),
            "usage_calibration": len(usage_calibration),
            "supply_alignment_rows": len(alignment),
            "supply_alignment_matched_rows": len(matched),
            "supply_alignment_contributive_rows": len(contributive),
            "supply_alignment_contributive_matched_rows": len(contributive_matched),
            "supply_alignment_zero_mass_not_required_rows": len(zero_mass_not_required),
        },
        "impact_rows": impact_rows,
        "component_impacts": climate_rows,
        "indicator_summary": indicator_summary,
        "indicator_unit_views": indicator_unit_views,
        "reference_person_equivalent_results": reference_results["reference_person_equivalent_results"],
        "reference_weighted_results": reference_results["reference_weighted_results"],
        "reference_phase_breakdown": reference_results["reference_phase_breakdown"],
        "reference_use_phase_components": reference_results["reference_use_phase_components"],
        "reference_scenarios": reference_results["reference_scenarios"],
        "reference_weighting_factors": reference_results["reference_weighting_factors"],
        "reference_climate_contributors": reference_results["reference_climate_contributors"],
        "masterboard_equipment_summary": masterboard_summaries["masterboard_equipment_summary"],
        "masterboard_material_summary": masterboard_summaries["masterboard_material_summary"],
        "parameters": parameters,
        "activities": activities,
        "exchanges": exchanges,
        "supply_alignment": alignment,
        "parametric_levers": parametric["parametric_levers"],
        "parametric_sensitivity": parametric["parametric_sensitivity"],
        "parametric_switches": parametric["parametric_switches"],
        "parametric_regional_scenarios": parametric["parametric_regional_scenarios"],
        "exact_scenario_lcia": exact_scenario_lcia,
        "excel_runtime_comparison": excel_runtime_comparison,
        "usage_calibration": usage_calibration,
        "top_climate_components": sorted(climate_rows, key=lambda row: -safe_float(row.get("climate_kgco2e")))[:20],
        "top_parameter_amounts": sorted(parameters, key=lambda row: -abs(safe_float(row.get("amount"))))[:20],
    }


def build_summary(
    *,
    source_json: Path,
    source_map: Path,
    root_map: Path,
    source_records: list[dict[str, Any]],
    tables: dict[str, Any],
    unique_node_rows: list[dict[str, Any]],
    unique_site_rows: list[dict[str, Any]],
    unique_flow_rows: list[dict[str, Any]],
    weather_rows: list[dict[str, Any]],
    event_rows: list[dict[str, Any]],
    transport_weather_rows: list[dict[str, Any]],
    node_operational_rows: list[dict[str, Any]],
    operational_event_rows: list[dict[str, Any]],
    sdd_results: dict[str, list[dict[str, Any]]],
    sdd_brightway: dict[str, list[dict[str, Any]]] | None = None,
    brightway_model: dict[str, Any] | None = None,
) -> dict[str, Any]:
    path_rows = tables["paths"]
    lane_rows = tables["lanes"]
    usable_records = [r for r in source_records if r.get("simulation_supply_usable") is not False]
    record_mass = sum(record_mass_kg(record) for record in usable_records)
    allocated_path_mass = sum(safe_float(row.get("path_mass_kg")) for row in path_rows)
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source": {
            "json_path": str(source_json.resolve(strict=False)),
            "map_html": str(source_map.resolve(strict=False)),
            "root_map_html": str(root_map.resolve(strict=False)),
        },
        "counts": {
            "source_records": len(source_records),
            "usable_records": len(usable_records),
            "excluded_records": len(tables["excluded_record_indexes"]),
            "skipped_records": len(tables["skipped_records"]),
            "primary_paths": len(path_rows),
            "path_node_rows": len(tables["nodes"]),
            "primary_lane_rows": len(lane_rows),
            "unique_role_nodes": len(unique_node_rows),
            "unique_sites": len(unique_site_rows),
            "unique_flows": len(unique_flow_rows),
            "weather_rows": len(weather_rows),
            "event_seed_rows": len(event_rows),
            "transport_weather_rows": len(transport_weather_rows),
            "node_operational_rows": len(node_operational_rows),
            "operational_event_rows": len(operational_event_rows),
            "sdd_node_state_rows": len(sdd_results.get("sdd_node_state", [])),
            "sdd_lane_state_rows": len(sdd_results.get("sdd_lane_state", [])),
            "sdd_flow_state_rows": len(sdd_results.get("sdd_flow_state", [])),
            "sdd_event_ledger_rows": len(sdd_results.get("sdd_event_ledger", [])),
            "sdd_brightway_inventory_delta_rows": len((sdd_brightway or {}).get("inventory_delta", [])),
            "sdd_brightway_exchange_delta_rows": len((sdd_brightway or {}).get("exchange_delta", [])),
            "sdd_brightway_exchange_lcia_rows": len((sdd_brightway or {}).get("exchange_lcia", [])),
            "sdd_brightway_exchange_lcia_factor_rows": len((sdd_brightway or {}).get("exchange_lcia_factors", [])),
            "sdd_brightway_monthly_rows": len((sdd_brightway or {}).get("monthly", [])),
            "brightway_component_impacts": (brightway_model or {}).get("counts", {}).get("climate_component_rows", 0),
            "brightway_parameters": (brightway_model or {}).get("counts", {}).get("parameters", 0),
            "brightway_supply_alignment_rows": (brightway_model or {}).get("counts", {}).get("supply_alignment_rows", 0),
            "brightway_person_equivalent_indicators": (brightway_model or {}).get("counts", {}).get("person_equivalent_indicators", 0),
            "brightway_reference_person_equivalent_results": (brightway_model or {}).get("counts", {}).get("reference_person_equivalent_results", 0),
            "brightway_reference_use_phase_components": (brightway_model or {}).get("counts", {}).get("reference_use_phase_components", 0),
            "brightway_parametric_regional_scenarios": (brightway_model or {}).get("counts", {}).get("parametric_regional_scenarios", 0),
            "brightway_exact_scenario_lcia": (brightway_model or {}).get("counts", {}).get("exact_scenario_lcia", 0),
            "brightway_excel_runtime_comparison": (brightway_model or {}).get("counts", {}).get("excel_runtime_comparison", 0),
            "brightway_usage_calibration": (brightway_model or {}).get("counts", {}).get("usage_calibration", 0),
        },
        "mass": {
            "usable_record_mass_kg": round(record_mass, 6),
            "allocated_primary_path_mass_kg": round(allocated_path_mass, 6),
            "allocation_gap_kg": round(allocated_path_mass - record_mass, 9),
        },
        "excluded_record_indexes": tables["excluded_record_indexes"],
        "readiness_counts": dict(Counter(row.get("readiness") for row in path_rows)),
        "transport_model_counts": dict(Counter(row.get("transport_model") for row in path_rows)),
        "lca_use_class_counts": dict(Counter(row.get("lca_use_class") for row in path_rows)),
        "role_node_counts": dict(Counter(row.get("role") for row in tables["nodes"])),
        "edge_counts": dict(Counter(row.get("edge") for row in lane_rows)),
        "mode_counts": dict(Counter(row.get("modes") for row in lane_rows)),
        "event_type_counts": dict(Counter(row.get("event_type") for row in event_rows)),
        "brightway_model": {
            "available": bool((brightway_model or {}).get("available")),
            "runtime": (brightway_model or {}).get("runtime", {}),
            "counts": (brightway_model or {}).get("counts", {}),
            "source_files": (brightway_model or {}).get("source_files", {}),
        },
        "sdd_brightway_coupling": {
            "available": bool((sdd_brightway or {}).get("inventory_delta")),
            "counts": {
                "inventory_delta_rows": len((sdd_brightway or {}).get("inventory_delta", [])),
                "exchange_delta_rows": len((sdd_brightway or {}).get("exchange_delta", [])),
                "exchange_lcia_rows": len((sdd_brightway or {}).get("exchange_lcia", [])),
                "exchange_lcia_factor_rows": len((sdd_brightway or {}).get("exchange_lcia_factors", [])),
                "monthly_rows": len((sdd_brightway or {}).get("monthly", [])),
                "mechanisms": len((sdd_brightway or {}).get("mechanism_totals", [])),
            },
        },
    }


def write_report(path: Path, summary: dict[str, Any]) -> None:
    counts = summary["counts"]
    mass = summary["mass"]
    lines = [
        "# POC2026 supply_geo primary case",
        "",
        f"- Schema: `{summary['schema_version']}`",
        f"- Source JSON: `{summary['source']['json_path']}`",
        f"- Source map: `{summary['source']['map_html']}`",
        f"- Source records: **{counts['source_records']}**",
        f"- Usable records: **{counts['usable_records']}**",
        f"- Excluded records: **{counts['excluded_records']}**",
        f"- Primary paths: **{counts['primary_paths']}**",
        f"- Primary lane rows: **{counts['primary_lane_rows']}**",
        f"- Unique sites: **{counts['unique_sites']}**",
        f"- Weather rows: **{counts['weather_rows']}**",
        f"- Event seed rows: **{counts['event_seed_rows']}**",
        f"- Transport weather rows: **{counts.get('transport_weather_rows', 0)}**",
        f"- Node operational rows: **{counts.get('node_operational_rows', 0)}**",
        f"- Operational event rows: **{counts.get('operational_event_rows', 0)}**",
        f"- SDD node-state rows: **{counts.get('sdd_node_state_rows', 0)}**",
        f"- SDD flow-state rows: **{counts.get('sdd_flow_state_rows', 0)}**",
        f"- Brightway component impacts: **{counts.get('brightway_component_impacts', 0)}**",
        f"- Brightway parameters: **{counts.get('brightway_parameters', 0)}**",
        "",
        "## Mass allocation",
        "",
        f"- Usable record mass kg: **{mass['usable_record_mass_kg']}**",
        f"- Allocated primary path mass kg: **{mass['allocated_primary_path_mass_kg']}**",
        f"- Allocation gap kg: **{mass['allocation_gap_kg']}**",
        "",
        "## Readiness",
        "",
    ]
    for key, value in summary.get("readiness_counts", {}).items():
        lines.append(f"- `{key}`: **{value}**")
    lines.extend(["", "## Event Seeds", ""])
    for key, value in summary.get("event_type_counts", {}).items():
        lines.append(f"- `{key}`: **{value}**")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def grouped_sum(rows: Iterable[dict[str, Any]], key: str, value: str) -> list[dict[str, Any]]:
    stats: dict[str, float] = defaultdict(float)
    for row in rows:
        label = clean(row.get(key)) or "unknown"
        stats[label] += safe_float(row.get(value))
    return [
        {"label": label, "value": round(total, 6)}
        for label, total in sorted(stats.items(), key=lambda item: (-item[1], item[0]))
    ]


def grouped_count(rows: Iterable[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    counts = Counter(clean(row.get(key)) or "unknown" for row in rows)
    return [
        {"label": label, "value": count}
        for label, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    ]


def top_rows(rows: Iterable[dict[str, Any]], *, label_key: str, value_key: str, limit: int = 15) -> list[dict[str, Any]]:
    out = [
        {
            "label": clean(row.get(label_key)) or "unknown",
            "value": round(safe_float(row.get(value_key)), 6),
            "meta": clean(row.get("roles") or row.get("country_code") or row.get("family")),
        }
        for row in rows
    ]
    out.sort(key=lambda row: (-safe_float(row.get("value")), row["label"]))
    return out[:limit]


def percentile(values: Iterable[float], q: float) -> float:
    nums = sorted(value for value in values if math.isfinite(value))
    if not nums:
        return 0.0
    position = (len(nums) - 1) * clamp(q, 0.0, 1.0)
    low = math.floor(position)
    high = math.ceil(position)
    if low == high:
        return nums[low]
    return nums[low] * (high - position) + nums[high] * (position - low)


def mean(values: Iterable[float]) -> float:
    nums = [value for value in values if math.isfinite(value)]
    return sum(nums) / len(nums) if nums else 0.0


def build_weather_month_payload(weather_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_month: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in weather_rows:
        month = int(safe_float(row.get("month_index")))
        if month > 0:
            by_month[month].append(row)

    rows: list[dict[str, Any]] = []
    for month in range(1, max(by_month.keys(), default=0) + 1):
        month_rows = by_month.get(month, [])

        def values(key: str) -> list[float]:
            return [safe_float(row.get(key)) for row in month_rows]

        temp = values("temp_c")
        humidity = values("humidity_pct")
        precip = values("precip_mm")
        wind = values("wind_ms")
        heat_index = values("heat_index_c")
        heatwave = values("heatwave")
        drought = values("drought")
        storm = values("storm_stress")
        hurricane = values("hurricane")
        cold = values("cold_stress")
        rows.append(
            {
                "month_index": month,
                "site_count": len(month_rows),
                "avg_climate_progress": round(mean(values("climate_progress")), 4),
                "avg_warming_delta_c": round(mean(values("warming_delta_c")), 4),
                "avg_hazard_intensification_factor": round(mean(values("hazard_intensification_factor")), 4),
                "avg_temp_c": round(mean(temp), 2),
                "p90_temp_c": round(percentile(temp, 0.9), 2),
                "max_temp_c": round(max(temp) if temp else 0.0, 2),
                "avg_heat_index_c": round(mean(heat_index), 2),
                "p90_heat_index_c": round(percentile(heat_index, 0.9), 2),
                "avg_humidity_pct": round(mean(humidity), 2),
                "p10_humidity_pct": round(percentile(humidity, 0.1), 2),
                "avg_precip_mm": round(mean(precip), 2),
                "p90_precip_mm": round(percentile(precip, 0.9), 2),
                "avg_wind_ms": round(mean(wind), 2),
                "max_wind_ms": round(max(wind) if wind else 0.0, 2),
                "avg_heatwave": round(mean(heatwave), 4),
                "avg_drought": round(mean(drought), 4),
                "avg_storm_stress": round(mean(storm), 4),
                "avg_hurricane": round(mean(hurricane), 4),
                "avg_cold_stress": round(mean(cold), 4),
                "heatwave_site_count": sum(1 for value in heatwave if value > 0.0),
                "drought_site_count": sum(1 for value in drought if value > 0.0),
                "storm_site_count": sum(1 for value in storm if value > 0.0),
                "hurricane_site_count": sum(1 for value in hurricane if value > 0.0),
                "cold_site_count": sum(1 for value in cold if value > 0.0),
            }
        )
    return rows


def build_ops_month_payload(event_rows: list[dict[str, Any]], horizon_months: int) -> list[dict[str, Any]]:
    by_month: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in event_rows:
        month = int(safe_float(row.get("month_index")))
        if month > 0:
            by_month[month].append(row)

    rows: list[dict[str, Any]] = []
    for month in range(1, horizon_months + 1):
        month_rows = by_month.get(month, [])
        capacity = [safe_float(row.get("capacity_multiplier"), 1.0) for row in month_rows]
        lead_time = [safe_float(row.get("lead_time_multiplier"), 1.0) for row in month_rows]
        scrap = [safe_float(row.get("scrap_multiplier"), 1.0) for row in month_rows]
        intensity = [safe_float(row.get("intensity")) for row in month_rows]
        type_counts = Counter(clean(row.get("event_type")) for row in month_rows)
        capacity_min = min(capacity) if capacity else 1.0
        rows.append(
            {
                "month_index": month,
                "event_count": len(month_rows),
                "event_intensity_sum": round(sum(intensity), 4),
                "capacity_multiplier_min": round(capacity_min, 4),
                "capacity_pressure_pct": round((1.0 - capacity_min) * 100.0, 2),
                "lead_time_multiplier_max": round(max(lead_time) if lead_time else 1.0, 4),
                "scrap_multiplier_max": round(max(scrap) if scrap else 1.0, 4),
                "heatwave": type_counts.get("heatwave", 0),
                "drought": type_counts.get("drought", 0),
                "storm": type_counts.get("storm", 0),
                "hurricane": type_counts.get("hurricane", 0),
                "cold": type_counts.get("cold", 0),
            }
        )
    return rows


def weighted_mean(rows: Iterable[dict[str, Any]], value_key: str, weight_key: str) -> float:
    total_weight = 0.0
    total_value = 0.0
    for row in rows:
        weight = safe_float(row.get(weight_key), 1.0)
        total_weight += weight
        total_value += safe_float(row.get(value_key)) * weight
    return total_value / total_weight if total_weight else 0.0


def weighted_or_mean(rows: list[dict[str, Any]], value_key: str, weight_key: str, default: float = 0.0) -> float:
    total_weight = sum(safe_float(row.get(weight_key)) for row in rows)
    if total_weight > 0.0:
        return weighted_mean(rows, value_key, weight_key)
    values = [safe_float(row.get(value_key), default) for row in rows]
    return mean(values) if values else default


def event_risk_index(row: dict[str, Any]) -> float:
    return max(
        safe_float(row.get("heatwave")),
        safe_float(row.get("drought")),
        safe_float(row.get("storm_stress")),
        safe_float(row.get("hurricane")),
        safe_float(row.get("cold_stress")),
    )


def build_weather_group_payload(
    weather_rows: list[dict[str, Any]],
    site_rows: list[dict[str, Any]],
    group_key: str,
) -> list[dict[str, Any]]:
    site_mass = {clean(row.get("site_uid")): safe_float(row.get("allocated_mass_kg")) for row in site_rows}
    by_group: dict[str, list[dict[str, Any]]] = defaultdict(list)
    sites_by_group: dict[str, set[str]] = defaultdict(set)
    for row in weather_rows:
        label = clean(row.get(group_key)) or "unknown"
        by_group[label].append(row)
        sites_by_group[label].add(clean(row.get("site_uid")))

    out: list[dict[str, Any]] = []
    for label, rows in by_group.items():
        site_ids = sites_by_group[label]
        out.append(
            {
                "label": label,
                "site_count": len(site_ids),
                "allocated_mass_kg": round(sum(site_mass.get(site_uid, 0.0) for site_uid in site_ids), 6),
                "avg_climate_progress": round(mean(safe_float(row.get("climate_progress")) for row in rows), 4),
                "avg_warming_delta_c": round(mean(safe_float(row.get("warming_delta_c")) for row in rows), 4),
                "avg_hazard_intensification_factor": round(
                    mean(safe_float(row.get("hazard_intensification_factor")) for row in rows),
                    4,
                ),
                "avg_temp_c": round(mean(safe_float(row.get("temp_c")) for row in rows), 2),
                "p90_temp_c": round(percentile((safe_float(row.get("temp_c")) for row in rows), 0.9), 2),
                "avg_precip_mm": round(mean(safe_float(row.get("precip_mm")) for row in rows), 2),
                "avg_heatwave": round(mean(safe_float(row.get("heatwave")) for row in rows), 4),
                "avg_drought": round(mean(safe_float(row.get("drought")) for row in rows), 4),
                "avg_storm_stress": round(mean(safe_float(row.get("storm_stress")) for row in rows), 4),
                "avg_hurricane": round(mean(safe_float(row.get("hurricane")) for row in rows), 4),
                "avg_cold_stress": round(mean(safe_float(row.get("cold_stress")) for row in rows), 4),
                "risk_index": round(mean(event_risk_index(row) for row in rows), 4),
            }
        )
    return sorted(out, key=lambda row: (-safe_float(row.get("risk_index")), row["label"]))


def build_weather_region_month_payload(weather_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in weather_rows:
        month = int(safe_float(row.get("month_index")))
        if month > 0:
            by_key[(clean(row.get("world_region")) or "unknown", month)].append(row)

    out: list[dict[str, Any]] = []
    for (region, month), rows in sorted(by_key.items()):
        out.append(
            {
                "world_region": region,
                "month_index": month,
                "site_count": len({clean(row.get("site_uid")) for row in rows}),
                "avg_climate_progress": round(mean(safe_float(row.get("climate_progress")) for row in rows), 4),
                "avg_warming_delta_c": round(mean(safe_float(row.get("warming_delta_c")) for row in rows), 4),
                "avg_hazard_intensification_factor": round(
                    mean(safe_float(row.get("hazard_intensification_factor")) for row in rows),
                    4,
                ),
                "avg_temp_c": round(mean(safe_float(row.get("temp_c")) for row in rows), 2),
                "risk_index": round(mean(event_risk_index(row) for row in rows), 4),
                "avg_heatwave": round(mean(safe_float(row.get("heatwave")) for row in rows), 4),
                "avg_drought": round(mean(safe_float(row.get("drought")) for row in rows), 4),
                "avg_storm_stress": round(mean(safe_float(row.get("storm_stress")) for row in rows), 4),
                "avg_hurricane": round(mean(safe_float(row.get("hurricane")) for row in rows), 4),
                "avg_cold_stress": round(mean(safe_float(row.get("cold_stress")) for row in rows), 4),
            }
        )
    return out


def maritime_route_region(from_region: str, to_region: str, from_lat: float, from_lon: float, to_lat: float, to_lon: float) -> str:
    regions = {from_region, to_region}
    mid_lat = (from_lat + to_lat) / 2.0
    if max(abs(from_lat), abs(to_lat), abs(mid_lat)) >= 58.0:
        return "High latitude shipping"
    if "Europe" in regions and "North America" in regions:
        return "North Atlantic"
    if "North America" in regions and ("East Asia" in regions or "Southeast Asia" in regions):
        return "North Pacific"
    if "Europe" in regions and ("East Asia" in regions or "Southeast Asia" in regions or "South Asia" in regions):
        return "Indian Ocean / Suez"
    if "Europe" in regions and ("MENA" in regions or "Sub-Saharan Africa" in regions):
        return "Mediterranean / Africa"
    if "Oceania" in regions or "Southeast Asia" in regions:
        return "Tropical Pacific"
    if "Latin America" in regions or ("North America" in regions and abs(mid_lat) < 35.0):
        return "Tropical Atlantic"
    if abs(mid_lat) < 30.0:
        return "Tropical maritime"
    return "Global ocean"


def maritime_profile(route_region: str) -> dict[str, float]:
    profiles = {
        "High latitude shipping": {"storm": 1.05, "hurricane": 0.0, "cold": 1.45, "monsoon": 0.0},
        "North Atlantic": {"storm": 1.25, "hurricane": 0.35, "cold": 0.95, "monsoon": 0.0},
        "North Pacific": {"storm": 1.15, "hurricane": 0.45, "cold": 0.65, "monsoon": 0.0},
        "Indian Ocean / Suez": {"storm": 0.9, "hurricane": 0.55, "cold": 0.05, "monsoon": 0.85},
        "Mediterranean / Africa": {"storm": 0.75, "hurricane": 0.05, "cold": 0.2, "monsoon": 0.2},
        "Tropical Pacific": {"storm": 1.0, "hurricane": 0.75, "cold": 0.0, "monsoon": 0.55},
        "Tropical Atlantic": {"storm": 0.95, "hurricane": 0.85, "cold": 0.05, "monsoon": 0.2},
        "Tropical maritime": {"storm": 0.9, "hurricane": 0.5, "cold": 0.0, "monsoon": 0.45},
        "Global ocean": {"storm": 0.75, "hurricane": 0.15, "cold": 0.25, "monsoon": 0.0},
    }
    return profiles.get(route_region, profiles["Global ocean"])


def build_transport_weather_rows(
    config: dict[str, Any],
    flow_rows: list[dict[str, Any]],
    site_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    weather_config = config.get("weather_driver") if isinstance(config.get("weather_driver"), dict) else {}
    if not weather_config.get("enabled", True):
        return []
    horizon = int(weather_config.get("horizon_months") or 240)
    seed = int(weather_config.get("seed") or 2026)
    climate_config = weather_config.get("climate_change") if isinstance(weather_config.get("climate_change"), dict) else {}
    sites = {clean(row.get("site_uid")): row for row in site_rows}
    rows: list[dict[str, Any]] = []
    for flow in flow_rows:
        modes = {part for part in clean(flow.get("modes")).split("|") if part}
        if "ship" not in modes:
            continue
        from_site = sites.get(clean(flow.get("from_site_uid")), {})
        to_site = sites.get(clean(flow.get("to_site_uid")), {})
        from_lat = safe_float(from_site.get("lat"), 0.0)
        from_lon = safe_float(from_site.get("lon"), 0.0)
        to_lat = safe_float(to_site.get("lat"), 0.0)
        to_lon = safe_float(to_site.get("lon"), 0.0)
        from_region = world_region(flow.get("from_country_code"), from_lat, from_lon)
        to_region = world_region(flow.get("to_country_code"), to_lat, to_lon)
        route_region = maritime_route_region(from_region, to_region, from_lat, from_lon, to_lat, to_lon)
        profile = maritime_profile(route_region)
        phase = stable_phase(f"{seed}:{flow.get('flow_uid')}:{route_region}")
        northern = 1.0 if (from_lat + to_lat) / 2.0 >= 0.0 else -1.0
        for month_index in range(1, horizon + 1):
            climate = climate_change_signal(month_index, horizon, climate_config)
            winter = month_season_peak(month_index, 1 if northern >= 0.0 else 7, 3)
            cyclone_season = month_season_peak(month_index, 9 if northern >= 0.0 else 2, 4)
            monsoon_season = month_season_peak(month_index, 7 if northern >= 0.0 else 1, 3)
            storm_pulse = max(0.0, math.sin(month_index * 0.74 + phase) - 0.35) * climate["maritime_storm_factor"]
            cyclone_pulse = max(0.0, math.sin(month_index * 0.51 + phase * 1.4) - 0.30) * climate["maritime_hurricane_factor"]
            cold_pulse = max(0.0, math.sin(month_index * 0.63 + phase * 0.8) - 0.28) * climate["maritime_cold_factor"]
            monsoon_pulse = max(0.0, math.sin(month_index * 0.47 + phase * 1.1) - 0.40) * climate["extreme_precip_factor"]
            hurricane = clamp(cyclone_season * cyclone_pulse * profile["hurricane"] * 1.35, 0.0, 1.0)
            cold = clamp(winter * cold_pulse * profile["cold"] * 1.1, 0.0, 1.0)
            monsoon = clamp(monsoon_season * monsoon_pulse * profile["monsoon"], 0.0, 1.0)
            storm = clamp(storm_pulse * profile["storm"] + hurricane * 0.65 + monsoon * 0.35 + cold * 0.28, 0.0, 1.0)
            risk = clamp(max(storm, hurricane, cold, monsoon), 0.0, 1.0)
            delay_multiplier = 1.0 + 0.28 * storm + 0.55 * hurricane + 0.22 * cold + 0.16 * monsoon
            capacity_multiplier = clamp(1.0 - 0.13 * storm - 0.22 * hurricane - 0.08 * cold - 0.05 * monsoon, 0.55, 1.0)
            rows.append(
                {
                    "flow_uid": flow.get("flow_uid"),
                    "edge": flow.get("edge"),
                    "from_site_uid": flow.get("from_site_uid"),
                    "to_site_uid": flow.get("to_site_uid"),
                    "from_name": flow.get("from_name"),
                    "to_name": flow.get("to_name"),
                    "from_country_code": flow.get("from_country_code"),
                    "to_country_code": flow.get("to_country_code"),
                    "from_region": from_region,
                    "to_region": to_region,
                    "route_region": route_region,
                    "distance_km": flow.get("distance_km"),
                    "allocated_mass_kg": flow.get("allocated_mass_kg"),
                    "allocated_kg_km": flow.get("allocated_kg_km"),
                    "month_index": month_index,
                    "climate_progress": round(climate["climate_progress"], 6),
                    "warming_delta_c": round(climate["warming_delta_c"], 4),
                    "maritime_hazard_intensification_factor": round(
                        mean(
                            [
                                climate["maritime_storm_factor"],
                                climate["maritime_hurricane_factor"],
                                climate["maritime_cold_factor"],
                                climate["extreme_precip_factor"],
                            ]
                        ),
                        4,
                    ),
                    "maritime_storm_factor": round(climate["maritime_storm_factor"], 4),
                    "maritime_hurricane_factor": round(climate["maritime_hurricane_factor"], 4),
                    "maritime_cold_factor": round(climate["maritime_cold_factor"], 4),
                    "maritime_storm": round(storm, 4),
                    "maritime_hurricane": round(hurricane, 4),
                    "maritime_cold": round(cold, 4),
                    "maritime_monsoon": round(monsoon, 4),
                    "maritime_risk_index": round(risk, 4),
                    "delay_multiplier": round(delay_multiplier, 4),
                    "capacity_multiplier": round(capacity_multiplier, 4),
                }
            )
    return rows


def build_maritime_month_payload(transport_weather_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_month: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in transport_weather_rows:
        month = int(safe_float(row.get("month_index")))
        if month > 0:
            by_month[month].append(row)
    out: list[dict[str, Any]] = []
    for month in range(1, max(by_month.keys(), default=0) + 1):
        rows = by_month.get(month, [])
        out.append(
            {
                "month_index": month,
                "flow_count": len({clean(row.get("flow_uid")) for row in rows}),
                "exposed_kg_km": round(sum(safe_float(row.get("allocated_kg_km")) for row in rows), 1),
                "avg_climate_progress": round(weighted_mean(rows, "climate_progress", "allocated_kg_km"), 4),
                "avg_warming_delta_c": round(weighted_mean(rows, "warming_delta_c", "allocated_kg_km"), 4),
                "avg_maritime_hazard_intensification_factor": round(
                    weighted_mean(rows, "maritime_hazard_intensification_factor", "allocated_kg_km"),
                    4,
                ),
                "risk_index": round(weighted_mean(rows, "maritime_risk_index", "allocated_kg_km"), 4),
                "storm": round(weighted_mean(rows, "maritime_storm", "allocated_kg_km"), 4),
                "hurricane": round(weighted_mean(rows, "maritime_hurricane", "allocated_kg_km"), 4),
                "cold": round(weighted_mean(rows, "maritime_cold", "allocated_kg_km"), 4),
                "monsoon": round(weighted_mean(rows, "maritime_monsoon", "allocated_kg_km"), 4),
                "delay_multiplier": round(weighted_mean(rows, "delay_multiplier", "allocated_kg_km"), 4),
                "capacity_multiplier": round(weighted_mean(rows, "capacity_multiplier", "allocated_kg_km"), 4),
            }
        )
    return out


def build_maritime_region_payload(transport_weather_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_region: dict[str, list[dict[str, Any]]] = defaultdict(list)
    flow_weight_by_region: dict[str, dict[str, float]] = defaultdict(dict)
    for row in transport_weather_rows:
        region = clean(row.get("route_region")) or "unknown"
        by_region[region].append(row)
        flow_weight_by_region[region][clean(row.get("flow_uid"))] = safe_float(row.get("allocated_kg_km"))
    out: list[dict[str, Any]] = []
    for region, rows in by_region.items():
        out.append(
            {
                "label": region,
                "flow_count": len(flow_weight_by_region[region]),
                "allocated_kg_km": round(sum(flow_weight_by_region[region].values()), 1),
                "avg_climate_progress": round(weighted_mean(rows, "climate_progress", "allocated_kg_km"), 4),
                "avg_warming_delta_c": round(weighted_mean(rows, "warming_delta_c", "allocated_kg_km"), 4),
                "avg_maritime_hazard_intensification_factor": round(
                    weighted_mean(rows, "maritime_hazard_intensification_factor", "allocated_kg_km"),
                    4,
                ),
                "risk_index": round(weighted_mean(rows, "maritime_risk_index", "allocated_kg_km"), 4),
                "storm": round(weighted_mean(rows, "maritime_storm", "allocated_kg_km"), 4),
                "hurricane": round(weighted_mean(rows, "maritime_hurricane", "allocated_kg_km"), 4),
                "cold": round(weighted_mean(rows, "maritime_cold", "allocated_kg_km"), 4),
                "monsoon": round(weighted_mean(rows, "maritime_monsoon", "allocated_kg_km"), 4),
                "delay_multiplier": round(weighted_mean(rows, "delay_multiplier", "allocated_kg_km"), 4),
            }
        )
    return sorted(out, key=lambda row: (-safe_float(row.get("risk_index")), row["label"]))


def transport_event_types(row: dict[str, Any], threshold: float = 0.12) -> list[str]:
    specs = [
        ("maritime_storm", "maritime_storm"),
        ("maritime_hurricane", "maritime_hurricane"),
        ("maritime_cold", "maritime_cold"),
        ("maritime_monsoon", "maritime_monsoon"),
    ]
    return [label for key, label in specs if safe_float(row.get(key)) > threshold]


def node_operational_labels(
    *,
    env_types: set[str],
    inbound_transport_types: set[str],
    outbound_transport_types: set[str],
    capacity_applied: float,
    lead_time_multiplier: float,
    scrap_multiplier: float,
    weather_row: dict[str, Any],
    inbound_risk: float,
    outbound_risk: float,
) -> list[str]:
    labels: set[str] = set()
    if capacity_applied < 0.96:
        labels.add("capacite_meteo_degradee")
    if capacity_applied < 0.90:
        labels.add("capacite_appoint")
    if lead_time_multiplier > 1.05:
        labels.add("retard_approvisionnement")
    if scrap_multiplier > 1.025:
        labels.add("recalage_qualite")
    if inbound_risk > 0.18 or inbound_transport_types:
        labels.add("congestion_logistique_inbound")
    if outbound_risk > 0.18 or outbound_transport_types:
        labels.add("congestion_logistique_outbound")
    if "hurricane" in env_types or "storm" in env_types or "maritime_hurricane" in inbound_transport_types | outbound_transport_types:
        labels.add("maintenance_corrective_meteo")
    if "heatwave" in env_types or "cold" in env_types or safe_float(weather_row.get("heatwave")) > 0.0 or safe_float(weather_row.get("cold_stress")) > 0.0:
        labels.add("overtime_energetique")
    if "cold" in env_types or "maritime_cold" in inbound_transport_types | outbound_transport_types:
        labels.add("froid_transport_ou_site")
    return sorted(labels)


def build_node_operational_tables(
    site_rows: list[dict[str, Any]],
    weather_rows: list[dict[str, Any]],
    event_rows: list[dict[str, Any]],
    transport_weather_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    weather_by_key = {
        (clean(row.get("site_uid")), int(safe_float(row.get("month_index")))): row
        for row in weather_rows
    }
    events_by_key: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in event_rows:
        events_by_key[(clean(row.get("site_uid")), int(safe_float(row.get("month_index"))))].append(row)

    inbound_by_key: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    outbound_by_key: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in transport_weather_rows:
        month = int(safe_float(row.get("month_index")))
        inbound_by_key[(clean(row.get("to_site_uid")), month)].append(row)
        outbound_by_key[(clean(row.get("from_site_uid")), month)].append(row)

    horizon = max((int(safe_float(row.get("month_index"))) for row in weather_rows), default=0)
    state_rows: list[dict[str, Any]] = []
    op_event_rows: list[dict[str, Any]] = []

    for site in site_rows:
        site_uid = clean(site.get("site_uid"))
        lat = safe_float(site.get("lat"), 0.0)
        lon = safe_float(site.get("lon"), 0.0)
        region = world_region(site.get("country_code"), lat, lon)
        profile = weather_profile_key(site.get("country_code"), lat, lon)
        for month in range(1, horizon + 1):
            weather = weather_by_key.get((site_uid, month), {})
            env_events = events_by_key.get((site_uid, month), [])
            inbound = inbound_by_key.get((site_uid, month), [])
            outbound = outbound_by_key.get((site_uid, month), [])

            env_types = {clean(row.get("event_type")) for row in env_events if clean(row.get("event_type"))}
            env_capacity = min((safe_float(row.get("capacity_multiplier"), 1.0) for row in env_events), default=1.0)
            env_lead = max((safe_float(row.get("lead_time_multiplier"), 1.0) for row in env_events), default=1.0)
            env_scrap = max((safe_float(row.get("scrap_multiplier"), 1.0) for row in env_events), default=1.0)

            inbound_capacity = min((safe_float(row.get("capacity_multiplier"), 1.0) for row in inbound), default=1.0)
            outbound_capacity = min((safe_float(row.get("capacity_multiplier"), 1.0) for row in outbound), default=1.0)
            inbound_delay = max((safe_float(row.get("delay_multiplier"), 1.0) for row in inbound), default=1.0)
            outbound_delay = max((safe_float(row.get("delay_multiplier"), 1.0) for row in outbound), default=1.0)
            inbound_risk = max((safe_float(row.get("maritime_risk_index")) for row in inbound), default=0.0)
            outbound_risk = max((safe_float(row.get("maritime_risk_index")) for row in outbound), default=0.0)
            inbound_transport_types = {event_type for row in inbound for event_type in transport_event_types(row)}
            outbound_transport_types = {event_type for row in outbound for event_type in transport_event_types(row)}

            capacity_applied = min(env_capacity, inbound_capacity, outbound_capacity)
            lead_time_multiplier = max(env_lead, inbound_delay, outbound_delay)
            scrap_multiplier = env_scrap
            disruption_index = clamp(
                max(
                    1.0 - capacity_applied,
                    (lead_time_multiplier - 1.0) / 0.75,
                    (scrap_multiplier - 1.0) / 0.28,
                    inbound_risk,
                    outbound_risk,
                    event_risk_index(weather),
                ),
                0.0,
                1.0,
            )
            service_proxy_pct = clamp(100.0 * capacity_applied / max(1.0, lead_time_multiplier) / max(1.0, scrap_multiplier), 0.0, 100.0)
            labels = node_operational_labels(
                env_types=env_types,
                inbound_transport_types=inbound_transport_types,
                outbound_transport_types=outbound_transport_types,
                capacity_applied=capacity_applied,
                lead_time_multiplier=lead_time_multiplier,
                scrap_multiplier=scrap_multiplier,
                weather_row=weather,
                inbound_risk=inbound_risk,
                outbound_risk=outbound_risk,
            )
            source_transport_flow_uids = sorted(
                {
                    clean(row.get("flow_uid"))
                    for row in inbound + outbound
                    if safe_float(row.get("maritime_risk_index")) > 0.0 and clean(row.get("flow_uid"))
                }
            )
            source_driver_types = sorted(
                {f"env:{event_type}" for event_type in env_types}
                | {f"inbound:{event_type}" for event_type in inbound_transport_types}
                | {f"outbound:{event_type}" for event_type in outbound_transport_types}
            )
            if labels and not source_driver_types:
                fallback_drivers: set[str] = set()
                weather_stresses = {
                    "env:heatwave": safe_float(weather.get("heatwave")),
                    "env:drought": safe_float(weather.get("drought")),
                    "env:storm": safe_float(weather.get("storm_stress")),
                    "env:hurricane": safe_float(weather.get("hurricane")),
                    "env:cold": safe_float(weather.get("cold_stress")),
                }
                dominant_weather, dominant_value = max(weather_stresses.items(), key=lambda item: item[1])
                if dominant_value > 0.0:
                    fallback_drivers.add(dominant_weather)
                if inbound_risk > 0.0:
                    fallback_drivers.add("inbound:maritime_risk")
                if outbound_risk > 0.0:
                    fallback_drivers.add("outbound:maritime_risk")
                source_driver_types = sorted(fallback_drivers)
            state_row = {
                "site_uid": site_uid,
                "supplier": site.get("name", ""),
                "roles": site.get("roles", ""),
                "country_code": site.get("country_code", ""),
                "world_region": clean(weather.get("world_region")) or region,
                "weather_profile": clean(weather.get("weather_profile")) or profile,
                "month_index": month,
                "allocated_mass_kg": site.get("allocated_mass_kg", 0.0),
                "env_event_types": "|".join(sorted(env_types)) if env_types else "none",
                "transport_event_types": "|".join(sorted(inbound_transport_types | outbound_transport_types)) if inbound_transport_types or outbound_transport_types else "none",
                "source_driver_types": "|".join(source_driver_types) if source_driver_types else "none",
                "source_environmental_event_ids": "|".join(sorted(clean(row.get("event_id")) for row in env_events if clean(row.get("event_id")))),
                "source_transport_flow_uids": "|".join(source_transport_flow_uids),
                "capacity_applied": round(capacity_applied, 4),
                "lead_time_multiplier": round(lead_time_multiplier, 4),
                "scrap_multiplier": round(scrap_multiplier, 4),
                "inbound_delay_multiplier": round(inbound_delay, 4),
                "outbound_delay_multiplier": round(outbound_delay, 4),
                "inbound_maritime_risk": round(inbound_risk, 4),
                "outbound_maritime_risk": round(outbound_risk, 4),
                "disruption_index": round(disruption_index, 4),
                "service_proxy_pct": round(service_proxy_pct, 2),
                "operational_event_labels": "|".join(labels) if labels else "nominal",
                "source_link_count": len(env_events) + len(source_transport_flow_uids),
            }
            state_rows.append(state_row)
            for label in labels:
                op_event_rows.append(
                    {
                        "operational_event_id": f"ops:{site_uid}:{month:03d}:{slug(label)}",
                        "site_uid": site_uid,
                        "supplier": site.get("name", ""),
                        "roles": site.get("roles", ""),
                        "country_code": site.get("country_code", ""),
                        "world_region": state_row["world_region"],
                        "month_index": month,
                        "operational_event_type": label,
                        "source_driver_types": state_row["source_driver_types"],
                        "source_environmental_event_ids": state_row["source_environmental_event_ids"],
                        "source_transport_flow_uids": state_row["source_transport_flow_uids"],
                        "capacity_applied": state_row["capacity_applied"],
                        "lead_time_multiplier": state_row["lead_time_multiplier"],
                        "scrap_multiplier": state_row["scrap_multiplier"],
                        "disruption_index": state_row["disruption_index"],
                    }
                )
    return state_rows, op_event_rows


def build_node_ops_month_payload(
    node_operational_rows: list[dict[str, Any]],
    operational_event_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_month: dict[int, list[dict[str, Any]]] = defaultdict(list)
    event_counts: dict[int, Counter[str]] = defaultdict(Counter)
    event_types: set[str] = set()
    for row in node_operational_rows:
        month = int(safe_float(row.get("month_index")))
        if month > 0:
            by_month[month].append(row)
    for row in operational_event_rows:
        month = int(safe_float(row.get("month_index")))
        event_type = clean(row.get("operational_event_type"))
        if month > 0 and event_type:
            event_counts[month][event_type] += 1
            event_types.add(event_type)

    out: list[dict[str, Any]] = []
    for month in range(1, max(by_month.keys(), default=0) + 1):
        rows = by_month.get(month, [])
        event_counter = event_counts.get(month, Counter())
        item = {
            "month_index": month,
            "site_count": len(rows),
            "affected_site_count": sum(1 for row in rows if clean(row.get("operational_event_labels")) != "nominal"),
            "avg_capacity_applied": round(weighted_mean(rows, "capacity_applied", "allocated_mass_kg"), 4),
            "min_capacity_applied": round(min((safe_float(row.get("capacity_applied"), 1.0) for row in rows), default=1.0), 4),
            "avg_lead_time_multiplier": round(weighted_mean(rows, "lead_time_multiplier", "allocated_mass_kg"), 4),
            "max_lead_time_multiplier": round(max((safe_float(row.get("lead_time_multiplier"), 1.0) for row in rows), default=1.0), 4),
            "avg_scrap_multiplier": round(weighted_mean(rows, "scrap_multiplier", "allocated_mass_kg"), 4),
            "avg_disruption_index": round(weighted_mean(rows, "disruption_index", "allocated_mass_kg"), 4),
            "avg_service_proxy_pct": round(weighted_mean(rows, "service_proxy_pct", "allocated_mass_kg"), 2),
            "operational_event_count": sum(event_counter.values()),
        }
        for event_type in sorted(event_types):
            item[event_type] = event_counter.get(event_type, 0)
        out.append(item)
    return out


def build_node_ops_region_payload(node_operational_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_region: dict[str, list[dict[str, Any]]] = defaultdict(list)
    sites_by_region: dict[str, set[str]] = defaultdict(set)
    for row in node_operational_rows:
        region = clean(row.get("world_region")) or "unknown"
        by_region[region].append(row)
        sites_by_region[region].add(clean(row.get("site_uid")))
    out: list[dict[str, Any]] = []
    for region, rows in by_region.items():
        out.append(
            {
                "label": region,
                "site_count": len(sites_by_region[region]),
                "affected_site_months": sum(1 for row in rows if clean(row.get("operational_event_labels")) != "nominal"),
                "avg_capacity_applied": round(weighted_mean(rows, "capacity_applied", "allocated_mass_kg"), 4),
                "avg_lead_time_multiplier": round(weighted_mean(rows, "lead_time_multiplier", "allocated_mass_kg"), 4),
                "avg_disruption_index": round(weighted_mean(rows, "disruption_index", "allocated_mass_kg"), 4),
                "avg_service_proxy_pct": round(weighted_mean(rows, "service_proxy_pct", "allocated_mass_kg"), 2),
            }
        )
    return sorted(out, key=lambda row: (-safe_float(row.get("avg_disruption_index")), row["label"]))


def build_node_ops_lineage_payload(operational_event_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    links: dict[tuple[str, str], dict[str, Any]] = {}
    for row in operational_event_rows:
        target = clean(row.get("operational_event_type"))
        drivers = [driver for driver in clean(row.get("source_driver_types")).split("|") if driver and driver != "none"]
        if not target or not drivers:
            continue
        weight = max(0.01, safe_float(row.get("disruption_index")))
        for driver in drivers:
            key = (driver, target)
            item = links.setdefault(key, {"source": driver, "target": target, "count": 0, "weight": 0.0})
            item["count"] += 1
            item["weight"] += weight
    out = [
        {
            "source": source,
            "target": target,
            "label": f"{source} -> {target}",
            "count": item["count"],
            "weight": round(item["weight"], 3),
            "value": round(item["weight"], 3),
        }
        for (source, target), item in links.items()
    ]
    return sorted(out, key=lambda row: (-safe_float(row.get("weight")), row["source"], row["target"]))


def family_ef(family: Any) -> float:
    return FAMILY_EF_KGCO2E_PER_KG.get(clean(family), FAMILY_EF_KGCO2E_PER_KG["general"])


def scrap_recycling_profile(family: Any) -> dict[str, float]:
    profile = SDD_SCRAP_RECYCLING_PROFILES.get(clean(family), SDD_SCRAP_RECYCLING_PROFILES["general"])
    return {key: safe_float(value) for key, value in profile.items()}


def mode_ef(modes: Any) -> float:
    parts = [part for part in clean(modes).split("|") if part] or ["unknown"]
    return sum(MODE_EF_KGCO2E_PER_KG_KM.get(part, MODE_EF_KGCO2E_PER_KG_KM["unknown"]) for part in parts) / len(parts)


def lane_key(row: dict[str, Any]) -> tuple[str, str, int]:
    return (
        clean(row.get("from_site_uid")),
        clean(row.get("to_site_uid")),
        int(safe_float(row.get("month_index"))),
    )


def build_transport_lookup(transport_weather_rows: list[dict[str, Any]]) -> dict[tuple[str, str, int], dict[str, Any]]:
    grouped: dict[tuple[str, str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in transport_weather_rows:
        grouped[lane_key(row)].append(row)

    out: dict[tuple[str, str, int], dict[str, Any]] = {}
    for key, rows in grouped.items():
        out[key] = {
            "maritime_risk_index": max(safe_float(row.get("maritime_risk_index")) for row in rows),
            "delay_multiplier": max(safe_float(row.get("delay_multiplier"), 1.0) for row in rows),
            "capacity_multiplier": min(safe_float(row.get("capacity_multiplier"), 1.0) for row in rows),
            "source_transport_flow_uids": "|".join(sorted({clean(row.get("flow_uid")) for row in rows if clean(row.get("flow_uid"))})),
            "route_region": "|".join(sorted({clean(row.get("route_region")) for row in rows if clean(row.get("route_region"))})),
        }
    return out


def build_path_lane_index(lane_rows: list[dict[str, Any]]) -> dict[str, dict[str, dict[str, Any]]]:
    by_path: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for row in lane_rows:
        by_path[clean(row.get("path_id"))][clean(row.get("edge"))] = row
    return by_path


def build_site_month_index(rows: list[dict[str, Any]]) -> dict[tuple[str, int], dict[str, Any]]:
    return {
        (clean(row.get("site_uid")), int(safe_float(row.get("month_index")))): row
        for row in rows
    }


def baseline_path_impact(path_row: dict[str, Any], lane_rows: list[dict[str, Any]]) -> dict[str, float]:
    mass = safe_float(path_row.get("path_mass_kg"))
    material = mass * family_ef(path_row.get("family"))
    transport = sum(safe_float(row.get("path_mass_kg")) * safe_float(row.get("distance_km")) * mode_ef(row.get("modes")) for row in lane_rows)
    production_energy = mass * 0.55
    return {
        "material": material,
        "transport": transport,
        "production_energy": production_energy,
        "total": material + transport + production_energy,
    }


def nominal_node_row(site_uid: str, month_index: int) -> dict[str, Any]:
    return {
        "site_uid": site_uid,
        "month_index": month_index,
        "capacity_applied": 1.0,
        "lead_time_multiplier": 1.0,
        "scrap_multiplier": 1.0,
        "disruption_index": 0.0,
        "service_proxy_pct": 100.0,
        "operational_event_labels": "nominal",
        "source_driver_types": "none",
        "source_environmental_event_ids": "",
        "source_transport_flow_uids": "",
    }


def simulate_sdd_supply(
    *,
    path_rows: list[dict[str, Any]],
    lane_rows: list[dict[str, Any]],
    node_operational_rows: list[dict[str, Any]],
    transport_weather_rows: list[dict[str, Any]],
    horizon_months: int,
) -> dict[str, list[dict[str, Any]]]:
    node_lookup = build_site_month_index(node_operational_rows)
    transport_lookup = build_transport_lookup(transport_weather_rows)
    lanes_by_path = build_path_lane_index(lane_rows)

    stock_state: dict[tuple[str, str], float] = {}
    backlog_state: dict[tuple[str, str], float] = defaultdict(float)
    prev_output_service: dict[tuple[str, str], float] = defaultdict(lambda: 1.0)
    node_state_rows: list[dict[str, Any]] = []
    lane_state_rows: list[dict[str, Any]] = []
    flow_state_rows: list[dict[str, Any]] = []
    event_ledger_rows: list[dict[str, Any]] = []
    monthly_impacts: list[dict[str, Any]] = []

    path_rows_sorted = sorted(path_rows, key=lambda row: clean(row.get("path_id")))
    for path in path_rows_sorted:
        path_id = clean(path.get("path_id"))
        mass = max(safe_float(path.get("path_mass_kg")), 0.000001)
        for role in ROLE_SEQUENCE:
            stock_state[(path_id, role)] = mass * SDD_STOCK_TARGET_MONTHS.get(role, 0.8)

    for month in range(1, horizon_months + 1):
        current_output_service: dict[tuple[str, str], float] = {}
        month_totals = defaultdict(float)
        path_service_weight = 0.0
        path_service_total = 0.0
        affected_path_count = 0
        max_path_backlog = 0.0

        for path in path_rows_sorted:
            path_id = clean(path.get("path_id"))
            family = clean(path.get("family"))
            mass = max(safe_float(path.get("path_mass_kg")), 0.000001)
            path_lanes = lanes_by_path.get(path_id, {})
            ordered_lanes = [path_lanes.get(edge_name) for edge_name, _, _ in EDGE_ORDER]
            baseline = baseline_path_impact(path, [row for row in ordered_lanes if row])
            path_surimpact = defaultdict(float)
            path_disruption = 0.0
            path_backup_mass = 0.0
            path_premium_mass = 0.0
            path_service = 1.0
            path_backlog = 0.0

            lane_state_by_edge: dict[str, dict[str, Any]] = {}
            for lane in [row for row in ordered_lanes if row]:
                key = (
                    clean(lane.get("from_site_uid")),
                    clean(lane.get("to_site_uid")),
                    month,
                )
                weather = transport_lookup.get(key, {})
                risk = safe_float(weather.get("maritime_risk_index"))
                delay = safe_float(weather.get("delay_multiplier"), 1.0)
                capacity = safe_float(weather.get("capacity_multiplier"), 1.0)
                premium_mode_share = 0.0
                decision = "nominal"
                if risk > 0.35 or delay > 1.18:
                    premium_mode_share = min(0.35, risk * 0.45)
                    delay = max(1.0, delay * (1.0 - premium_mode_share * 0.35))
                    capacity = min(1.0, capacity + premium_mode_share * 0.18)
                    decision = "premium_transport"
                edge = clean(lane.get("edge"))
                lane_state = {
                    "path_id": path_id,
                    "record_index": path.get("record_index"),
                    "month_index": month,
                    "edge": edge,
                    "from_site_uid": lane.get("from_site_uid"),
                    "to_site_uid": lane.get("to_site_uid"),
                    "from_name": lane.get("from_name"),
                    "to_name": lane.get("to_name"),
                    "modes": lane.get("modes"),
                    "distance_km": lane.get("distance_km"),
                    "path_mass_kg": lane.get("path_mass_kg"),
                    "baseline_kg_km": lane.get("allocated_kg_km"),
                    "route_region": weather.get("route_region", ""),
                    "transport_risk_index": round(risk, 4),
                    "delay_multiplier": round(delay, 4),
                    "capacity_multiplier": round(capacity, 4),
                    "premium_mode_share": round(premium_mode_share, 4),
                    "transport_decision": decision,
                    "source_transport_flow_uids": weather.get("source_transport_flow_uids", ""),
                }
                lane_state_by_edge[edge] = lane_state
                lane_state_rows.append(lane_state)

            for role_index, role in enumerate(ROLE_SEQUENCE):
                site_uid = clean(path.get(f"{role.lower()}_site_uid"))
                node = node_lookup.get((site_uid, month), nominal_node_row(site_uid, month))
                upstream_role = ROLE_SEQUENCE[role_index - 1] if role_index > 0 else ""
                inbound_lane = lane_state_by_edge.get(EDGE_ORDER[role_index - 1][0]) if role_index > 0 else None
                upstream_service = prev_output_service[(path_id, upstream_role)] if upstream_role else 1.0
                lane_capacity = safe_float(inbound_lane.get("capacity_multiplier"), 1.0) if inbound_lane else 1.0
                lane_delay = safe_float(inbound_lane.get("delay_multiplier"), 1.0) if inbound_lane else 1.0
                lane_risk = safe_float(inbound_lane.get("transport_risk_index")) if inbound_lane else 0.0
                inbound_service = clamp(upstream_service * lane_capacity / max(1.0, lane_delay), 0.0, 1.15)

                capacity_applied = safe_float(node.get("capacity_applied"), 1.0)
                lead_time = safe_float(node.get("lead_time_multiplier"), 1.0)
                scrap_multiplier = safe_float(node.get("scrap_multiplier"), 1.0)
                node_disruption = safe_float(node.get("disruption_index"))
                disruption = clamp(max(node_disruption, lane_risk, 1.0 - capacity_applied, (lead_time - 1.0) / 0.75), 0.0, 1.0)
                path_disruption = max(path_disruption, disruption)

                required = mass + backlog_state[(path_id, role)]
                inbound_available = mass * inbound_service
                stock_start = stock_state[(path_id, role)]
                stock_draw = min(stock_start, max(0.0, required - inbound_available))
                available_input = inbound_available + stock_draw
                capacity_limit = mass * capacity_applied
                process_input = min(required, available_input, capacity_limit)
                scrap_loss_rate = clamp(scrap_multiplier - 1.0, 0.0, 0.35)
                good_output = process_input * (1.0 - scrap_loss_rate)
                gap = max(0.0, required - good_output)

                decisions: list[str] = []
                capacity_boost_output = 0.0
                backup_output = 0.0
                premium_output = 0.0
                if gap > 0.0 and (capacity_applied < 0.90 or disruption > 0.35):
                    capacity_boost_output = min(gap, mass * 0.12)
                    gap -= capacity_boost_output
                    decisions.append("capacity_boost")
                if gap > 0.0 and (disruption > 0.25 or backlog_state[(path_id, role)] > 0.0 or good_output / max(required, 0.000001) < 0.95):
                    backup_limit = mass * (0.18 if role == "OEM" else 0.32)
                    backup_output = min(gap, backup_limit)
                    gap -= backup_output
                    decisions.append("backup_supplier")
                if gap > 0.0 and inbound_lane and (lane_delay > 1.08 or backlog_state[(path_id, role)] > 0.0):
                    premium_output = min(gap, mass * 0.20)
                    gap -= premium_output
                    decisions.append("premium_transport")

                final_output = good_output + capacity_boost_output + backup_output + premium_output
                service_level = clamp(final_output / max(required, 0.000001), 0.0, 1.15)
                backlog_end = max(0.0, required - final_output)
                stock_after_draw = stock_start - stock_draw
                excess_inbound = max(0.0, inbound_available - process_input)
                stock_target = mass * SDD_STOCK_TARGET_MONTHS.get(role, 0.8)
                recovery_fill = mass * max(0.0, 1.0 - disruption) * 0.04 if backlog_end <= 0.0 else 0.0
                stock_end = clamp(stock_after_draw + excess_inbound * 0.35 + recovery_fill, 0.0, stock_target)
                backlog_state[(path_id, role)] = backlog_end
                stock_state[(path_id, role)] = stock_end
                current_output_service[(path_id, role)] = clamp(final_output / mass, 0.0, 1.15)

                scrap_mass = max(0.0, process_input - good_output)
                inbound_distance = safe_float(inbound_lane.get("distance_km")) if inbound_lane else 0.0
                ef = family_ef(family)
                path_surimpact["backup_material"] += backup_output * ef * 1.45
                path_surimpact["premium_transport"] += premium_output * inbound_distance * MODE_EF_KGCO2E_PER_KG_KM["air"]
                path_surimpact["scrap_rework"] += scrap_mass * ef
                path_surimpact["capacity_energy"] += capacity_boost_output * 0.85 + mass * disruption * 0.10
                path_surimpact["maintenance"] += mass * disruption * 0.08
                path_surimpact["backlog_penalty"] += backlog_end * 0.05
                path_backup_mass += backup_output
                path_premium_mass += premium_output

                source_drivers = clean(node.get("source_driver_types"))
                event_labels = clean(node.get("operational_event_labels"))
                node_state_rows.append(
                    {
                        "path_id": path_id,
                        "record_index": path.get("record_index"),
                        "month_index": month,
                        "role": role,
                        "site_uid": site_uid,
                        "supplier": path.get(role.lower(), ""),
                        "family": family,
                        "path_mass_kg": round(mass, 9),
                        "required_mass_kg": round(required, 6),
                        "inbound_service": round(inbound_service, 4),
                        "capacity_applied": round(capacity_applied, 4),
                        "lead_time_multiplier": round(lead_time, 4),
                        "scrap_multiplier": round(scrap_multiplier, 4),
                        "stock_start_kg": round(stock_start, 6),
                        "stock_draw_kg": round(stock_draw, 6),
                        "stock_end_kg": round(stock_end, 6),
                        "backlog_start_kg": round(required - mass, 6),
                        "backlog_end_kg": round(backlog_end, 6),
                        "good_output_kg": round(good_output, 6),
                        "backup_output_kg": round(backup_output, 6),
                        "capacity_boost_output_kg": round(capacity_boost_output, 6),
                        "premium_output_kg": round(premium_output, 6),
                        "final_output_kg": round(final_output, 6),
                        "service_level": round(service_level, 4),
                        "disruption_index": round(disruption, 4),
                        "decisions": "|".join(decisions) if decisions else "none",
                        "operational_event_labels": event_labels or "nominal",
                        "source_driver_types": source_drivers or "none",
                        "source_environmental_event_ids": node.get("source_environmental_event_ids", ""),
                        "source_transport_flow_uids": node.get("source_transport_flow_uids", ""),
                    }
                )
                if decisions or (event_labels and event_labels != "nominal") or backlog_end > 0.0:
                    event_ledger_rows.append(
                        {
                            "sdd_event_id": f"sdd:{path_id}:{month:03d}:{role}",
                            "path_id": path_id,
                            "record_index": path.get("record_index"),
                            "month_index": month,
                            "role": role,
                            "site_uid": site_uid,
                            "supplier": path.get(role.lower(), ""),
                            "event_labels": event_labels or "nominal",
                            "decisions": "|".join(decisions) if decisions else "none",
                            "source_driver_types": source_drivers or "none",
                            "source_environmental_event_ids": node.get("source_environmental_event_ids", ""),
                            "source_transport_flow_uids": node.get("source_transport_flow_uids", ""),
                            "service_level": round(service_level, 4),
                            "backlog_end_kg": round(backlog_end, 6),
                            "surimpact_kgCO2e": round(sum(path_surimpact.values()), 6),
                        }
                    )

                path_service = service_level if role == "OEM" else path_service
                path_backlog = backlog_end if role == "OEM" else path_backlog

            td_multiplier = 1.0 + path_disruption * 0.025
            td_total = baseline["total"] * td_multiplier
            surimpact_total = sum(path_surimpact.values())
            sdd_total = td_total + surimpact_total
            month_totals["classical_total"] += baseline["total"]
            month_totals["td_dlca_total"] += td_total
            month_totals["sdd_total"] += sdd_total
            for key, value in path_surimpact.items():
                month_totals[key] += value
            month_totals["surimpact_total"] += surimpact_total
            month_totals["backup_output_kg"] += path_backup_mass
            month_totals["premium_output_kg"] += path_premium_mass
            month_totals["oem_backlog_kg"] += path_backlog
            month_totals["path_disruption_weighted"] += path_disruption * mass
            path_service_total += path_service * mass
            path_service_weight += mass
            if path_disruption > 0.0 or path_backlog > 0.0:
                affected_path_count += 1
            max_path_backlog = max(max_path_backlog, path_backlog)
            flow_state_rows.append(
                {
                    "path_id": path_id,
                    "record_index": path.get("record_index"),
                    "month_index": month,
                    "system": path.get("system", ""),
                    "component": path.get("component", ""),
                    "family": family,
                    "path_mass_kg": round(mass, 9),
                    "oem_service_level": round(path_service, 4),
                    "oem_backlog_kg": round(path_backlog, 6),
                    "path_disruption_index": round(path_disruption, 4),
                    "backup_output_kg": round(path_backup_mass, 6),
                    "premium_output_kg": round(path_premium_mass, 6),
                    "classical_kgCO2e": round(baseline["total"], 6),
                    "td_dlca_kgCO2e": round(td_total, 6),
                    "sdd_kgCO2e": round(sdd_total, 6),
                    "surimpact_kgCO2e": round(surimpact_total, 6),
                }
            )

        prev_output_service = current_output_service
        monthly_impacts.append(
            {
                "month_index": month,
                "classical_kgCO2e": round(month_totals["classical_total"], 6),
                "td_dlca_kgCO2e": round(month_totals["td_dlca_total"], 6),
                "sdd_kgCO2e": round(month_totals["sdd_total"], 6),
                "surimpact_total": round(month_totals["surimpact_total"], 6),
                "backup_material": round(month_totals["backup_material"], 6),
                "premium_transport": round(month_totals["premium_transport"], 6),
                "scrap_rework": round(month_totals["scrap_rework"], 6),
                "capacity_energy": round(month_totals["capacity_energy"], 6),
                "maintenance": round(month_totals["maintenance"], 6),
                "backlog_penalty": round(month_totals["backlog_penalty"], 6),
                "backup_output_kg": round(month_totals["backup_output_kg"], 6),
                "premium_output_kg": round(month_totals["premium_output_kg"], 6),
                "oem_backlog_kg": round(month_totals["oem_backlog_kg"], 6),
                "max_path_backlog_kg": round(max_path_backlog, 6),
                "avg_oem_service_level": round(path_service_total / path_service_weight if path_service_weight else 1.0, 4),
                "avg_path_disruption_index": round(month_totals["path_disruption_weighted"] / path_service_weight if path_service_weight else 0.0, 4),
                "affected_path_count": affected_path_count,
            }
        )

    cumulative_rows: list[dict[str, Any]] = []
    running = defaultdict(float)
    for row in monthly_impacts:
        for key in ["classical_kgCO2e", "td_dlca_kgCO2e", "sdd_kgCO2e", "surimpact_total"]:
            running[key] += safe_float(row.get(key))
        cumulative_rows.append(
            {
                "month_index": row["month_index"],
                "classical_cumulative": round(running["classical_kgCO2e"], 6),
                "td_dlca_cumulative": round(running["td_dlca_kgCO2e"], 6),
                "sdd_cumulative": round(running["sdd_kgCO2e"], 6),
                "surimpact_cumulative": round(running["surimpact_total"], 6),
            }
        )

    method_comparison = [
        {"method": "LCA classique", "total_kgCO2e": cumulative_rows[-1]["classical_cumulative"] if cumulative_rows else 0.0},
        {"method": "TD-DLCA", "total_kgCO2e": cumulative_rows[-1]["td_dlca_cumulative"] if cumulative_rows else 0.0},
        {"method": "SDD supply_geo", "total_kgCO2e": cumulative_rows[-1]["sdd_cumulative"] if cumulative_rows else 0.0},
    ]
    return {
        "sdd_node_state": node_state_rows,
        "sdd_lane_state": lane_state_rows,
        "sdd_flow_state": flow_state_rows,
        "sdd_event_ledger": event_ledger_rows,
        "sdd_monthly_impacts": monthly_impacts,
        "sdd_cumulative_impacts": cumulative_rows,
        "sdd_method_comparison": method_comparison,
        "sdd_tier_month": build_sdd_tier_month_payload(node_state_rows),
    }


def brightway_exact_score(brightway_model: dict[str, Any], root_activity_id: str) -> float:
    for row in brightway_model.get("exact_scenario_lcia", []):
        if clean(row.get("scenario_id")) == "current_export" and clean(row.get("root_activity_id")) == root_activity_id:
            score = safe_float(row.get("score_kgco2e"))
            if score > 0.0:
                return score
    return 0.0


def path_context_by_id(path_rows: list[dict[str, Any]], brightway_model: dict[str, Any]) -> dict[str, dict[str, Any]]:
    alignment_by_path = {clean(row.get("path_id")): row for row in brightway_model.get("supply_alignment", [])}
    raw_scores: dict[str, float] = {}
    for path in path_rows:
        path_id = clean(path.get("path_id"))
        alignment = alignment_by_path.get(path_id, {})
        matched_score = safe_float(alignment.get("brightway_climate_kgco2e"))
        fallback_score = safe_float(path.get("path_mass_kg")) * family_ef(path.get("family"))
        raw_scores[path_id] = matched_score if matched_score > 0.0 else fallback_score

    production_score = brightway_exact_score(brightway_model, "production")
    raw_total = sum(raw_scores.values())
    scale = production_score / raw_total if production_score > 0.0 and raw_total > 0.0 else 1.0

    out: dict[str, dict[str, Any]] = {}
    for path in path_rows:
        path_id = clean(path.get("path_id"))
        mass = max(safe_float(path.get("path_mass_kg")), 0.000001)
        alignment = alignment_by_path.get(path_id, {})
        static_score = raw_scores.get(path_id, 0.0) * scale
        out[path_id] = {
            "path": path,
            "alignment": alignment,
            "static_production_kgco2e": static_score,
            "component_intensity_kgco2e_per_kg": static_score / mass,
            "calibration_scale_to_exact_production": scale,
            "calibration_source": "brightway_supply_alignment_scaled_to_exact_production"
            if safe_float(alignment.get("brightway_climate_kgco2e")) > 0.0
            else "family_proxy_scaled_to_exact_production",
            "match_level": clean(alignment.get("match_level")) or "family_proxy",
            "brightway_system": clean(alignment.get("brightway_system")),
            "brightway_component": clean(alignment.get("brightway_component")),
        }
    return out


def lane_state_index(lane_state_rows: list[dict[str, Any]]) -> dict[tuple[str, int, str], dict[str, Any]]:
    return {
        (clean(row.get("path_id")), int(safe_float(row.get("month_index"))), clean(row.get("edge"))): row
        for row in lane_state_rows
    }


def role_inbound_edge(role: str) -> str:
    if role not in ROLE_SEQUENCE:
        return ""
    role_index = ROLE_SEQUENCE.index(role)
    return EDGE_ORDER[role_index - 1][0] if role_index > 0 else ""


def sdd_brightway_role_scope_share(role: str) -> float:
    share = safe_float(SDD_BRIGHTWAY_ROLE_SCOPE_SHARE.get(role))
    return share if share > 0.0 else 1.0 / len(ROLE_SEQUENCE)


def brightway_key(value: Any) -> str:
    text = clean(value)
    return slug(ascii_key(text)) if text else ""


BRIGHTWAY_ACTIVITY_ALIASES = {
    "accoudoir_allee": "accoudoir_allee",
    "ceinture_de_securite": "ceinture_de_securite",
    "commande_actionnement_ecu": "commande_actionnement_ecu",
    "ecran": "ecran_seat",
    "ens_coque": "ensemble_coque",
    "ens_coussin_assise": "ensemble_coussin_assise",
    "ens_coussin_dossier": "ensemble_coussin_dossier",
    "ens_coussin_dossier_version_tetiere": "ensemble_coussin_dossier_version_tetiere",
    "ens_equipements_lateraux": "ensemble_equipement_lateral",
    "ens_palette_optimisee": "ensemble_palette_optimisee",
    "ens_stowage_lateral": "ensemble_stowage_lateral",
    "ens_structure_fauteuil": "ens_structure_fauteuil",
    "ens_structure_fixe": "ensemble_structure_fixe",
    "ens_structure_porte": "ensemble_porte",
    "ens_tablette_cocktail": "ensemble_tablette_cocktail",
    "ens_tablette_repas": "ensemble_tablette_repas",
    "ens_tetiere": "ensemble_tetiere",
    "manchette_acc_mobile": "manchette_acc_mobile",
    "powerbox": "seat_power",
    "sfcu_seat_function_control_unit": "sfcu",
    "support_clamps": "support_et_clamps",
    "support_equipe": "support_manchette_equipee",
    "system_ife": "systeme_ife_boitier",
    "system_ife_boitier": "systeme_ife_boitier",
}


def brightway_activity_candidates(row: dict[str, Any]) -> list[str]:
    raw_values = [
        row.get("brightway_system"),
        row.get("system"),
        row.get("brightway_component"),
        row.get("component"),
        row.get("brightway_exchange_proxy"),
    ]
    candidates: list[str] = []
    for value in raw_values:
        key = brightway_key(value)
        if not key:
            continue
        variants = [key, key.replace("equipements_lateraux", "equipement_lateral")]
        if key.startswith("ens_"):
            variants.append(f"ensemble_{key[4:]}")
        for variant in variants:
            alias = BRIGHTWAY_ACTIVITY_ALIASES.get(variant, variant)
            for candidate in (variant, alias):
                if candidate and candidate not in candidates:
                    candidates.append(candidate)
    return candidates


def brightway_exchanges_by_activity(exchanges: list[dict[str, Any]]) -> tuple[dict[str, list[dict[str, Any]]], dict[str, str]]:
    by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    names_by_key: dict[str, str] = {}
    for row in exchanges:
        activity_name = clean(row.get("activity_name"))
        key = brightway_key(activity_name)
        if not key:
            continue
        by_key[key].append(row)
        names_by_key.setdefault(key, activity_name)
    return by_key, names_by_key


def resolve_brightway_activity_for_delta(
    row: dict[str, Any],
    activity_exchanges: dict[str, list[dict[str, Any]]],
    activity_names: dict[str, str],
) -> tuple[str, list[dict[str, Any]], str]:
    candidates = brightway_activity_candidates(row)
    for candidate in candidates:
        if candidate in activity_exchanges:
            return activity_names.get(candidate, candidate), activity_exchanges[candidate], "exact_or_alias"
    for candidate in candidates:
        if len(candidate) < 6:
            continue
        matches = [
            key
            for key in activity_exchanges
            if key == candidate or (candidate in key and len(candidate) >= 8) or (key in candidate and len(key) >= 8)
        ]
        if len(matches) == 1:
            match = matches[0]
            return activity_names.get(match, match), activity_exchanges[match], "fuzzy_unique"
    return "", [], "activity_unmatched"


def brightway_exchange_category(row: dict[str, Any]) -> str:
    if clean(row.get("type")) == "production":
        return "production"
    key = brightway_key(
        " ".join(
            [
                clean(row.get("name")),
                clean(row.get("reference_product")),
                clean(row.get("unit")),
                clean(row.get("categories")),
            ]
        )
    )
    if any(token in key for token in ("transport", "freight", "lorry", "aircraft", "cargo_plane")):
        return "transport"
    if any(token in key for token in ("electricity", "heat", "compressed_air", "steam", "natural_gas", "energy")):
        return "energy"
    if any(token in key for token in ("lubricating_oil", "cleaning", "nettoyage", "detergent", "maintenance")):
        return "maintenance_consumable"
    if any(token in key for token in ("scrap", "waste", "treatment", "recycling", "recyclage", "wastewater")):
        return "waste_or_scrap"
    if "water" in key or "eau" in key:
        return "water"
    return "material"


def top_exchange_candidates(rows: list[dict[str, Any]], limit: int) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: -abs(safe_float(row.get("amount"))))[:limit]


def sdd_virtual_exchange(name: str, unit: str, reference_product: str, category: str) -> dict[str, Any]:
    return {
        "activity_name": "production du siege",
        "name": name,
        "amount": 1.0,
        "database": "SDD proxy",
        "location": "case",
        "unit": unit,
        "type": "technosphere",
        "reference_product": reference_product,
        "exchange_category": category,
    }


def selected_brightway_exchanges_for_mechanism(exchanges: list[dict[str, Any]], mechanism: str) -> tuple[list[dict[str, Any]], str]:
    usable: list[dict[str, Any]] = []
    for row in exchanges:
        category = brightway_exchange_category(row)
        if category == "production":
            continue
        if abs(safe_float(row.get("amount"))) <= 1e-12:
            continue
        copy = dict(row)
        copy["exchange_category"] = category
        usable.append(copy)

    mechanism = clean(mechanism)
    if mechanism == "premium_transport":
        return [
            {
                "activity_name": "transport premium SDD",
                "name": "market for transport, freight, aircraft, medium haul (virtual SDD)",
                "amount": 1.0,
                "database": "SDD proxy",
                "location": "GLO",
                "unit": "ton kilometer",
                "type": "technosphere",
                "reference_product": "transport, freight, aircraft",
                "exchange_category": "transport",
            }
        ], "virtual_air_freight_proxy"

    by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in usable:
        by_category[row["exchange_category"]].append(row)

    if mechanism in {"capacity_energy", "disruption_energy_overhead"}:
        selected = top_exchange_candidates(by_category.get("energy", []), 4)
        return selected, "energy_exchange_mapping" if selected else "no_energy_exchange_on_activity"

    if mechanism == "maintenance":
        selected = top_exchange_candidates(by_category.get("maintenance_consumable", []), 3)
        selected.extend(top_exchange_candidates(by_category.get("water", []), 2))
        if selected:
            return selected, "maintenance_consumable_mapping"
        fallback = top_exchange_candidates(by_category.get("energy", []), 3)
        return fallback, "maintenance_energy_fallback" if fallback else "no_maintenance_exchange_on_activity"

    if mechanism == "scrap_rework":
        return [
            sdd_virtual_exchange(
                "SDD scrap replacement production calibrated by component intensity",
                "kg replacement component output",
                "scrap replacement production",
                "scrap_rework",
            )
        ], "scrap_replacement_component_proxy"

    if mechanism == "scrap_treatment":
        return [
            sdd_virtual_exchange(
                "SDD scrap sorting and treatment calibrated by family profile",
                "kg scrap treated",
                "scrap sorting and treatment",
                "scrap_treatment",
            )
        ], "scrap_treatment_family_proxy"

    if mechanism == "recycling_credit":
        return [
            sdd_virtual_exchange(
                "SDD avoided virgin material credit calibrated by family profile",
                "kg scrap recovered",
                "avoided virgin material credit",
                "recycling_credit",
            )
        ], "recycling_avoided_burden_family_proxy"

    if mechanism == "backup_material":
        return [
            sdd_virtual_exchange(
                "SDD backup supplier overhead calibrated by component intensity",
                "kg overhead-equivalent backup output",
                "backup supplier overhead",
                "supplier_overhead",
            )
        ], "backup_supplier_overhead_proxy"

    selected = top_exchange_candidates(usable, 8)
    return selected, "generic_foreground_mapping" if selected else "no_exchange_on_activity"


def sdd_exchange_delta_amount(row: dict[str, Any], exchange: dict[str, Any]) -> float:
    mechanism = clean(row.get("mechanism"))
    amount_delta = safe_float(row.get("amount_delta"))
    if mechanism == "premium_transport":
        return amount_delta / 1000.0
    if clean(exchange.get("database")) == "SDD proxy":
        if mechanism == "backup_material":
            return amount_delta * safe_float(row.get("operational_multiplier"), 1.0)
        return amount_delta
    path_mass = max(safe_float(row.get("path_mass_kg")), 1e-9)
    ratio = amount_delta / path_mass
    if mechanism == "backup_material":
        ratio *= safe_float(row.get("operational_multiplier"), 1.0)
    return safe_float(exchange.get("amount")) * ratio


def exchange_delta_proxy_amount(row: dict[str, Any]) -> float:
    amount = safe_float(row.get("amount_delta"))
    if clean(row.get("mechanism")) == "backup_material":
        amount *= safe_float(row.get("operational_multiplier"), 1.0)
    if clean(row.get("mechanism")) == "premium_transport":
        return amount / 1000.0
    return amount


def append_exchange_delta_group(
    groups: dict[tuple[Any, ...], dict[str, Any]],
    *,
    source_row: dict[str, Any],
    activity_name: str,
    activity_match_status: str,
    exchange: dict[str, Any],
    exchange_category: str,
    mapping_status: str,
    mapping_rule_id: str,
    lcia_allocation_method: str,
    delta_amount: float,
    delta_kgco2e: float,
) -> None:
    key = (
        int(safe_float(source_row.get("month_index"))),
        clean(source_row.get("mechanism")),
        clean(source_row.get("inventory_delta_type")),
        activity_name,
        clean(exchange.get("name")),
        clean(exchange.get("reference_product")),
        exchange_category,
        clean(exchange.get("unit")),
        clean(exchange.get("database")),
        clean(exchange.get("location")),
        mapping_status,
        activity_match_status,
        clean(source_row.get("confidence")),
    )
    group = groups.setdefault(
        key,
        {
            "month_index": key[0],
            "mechanism": key[1],
            "inventory_delta_type": key[2],
            "activity_name": key[3],
            "exchange_name": key[4],
            "exchange_reference_product": key[5],
            "exchange_category": key[6],
            "exchange_unit": key[7],
            "exchange_database": key[8],
            "exchange_location": key[9],
            "mapping_status": key[10],
            "activity_match_status": key[11],
            "confidence": key[12],
            "delta_amount": 0.0,
            "delta_kgco2e": 0.0,
            "source_inventory_delta_kgco2e": 0.0,
            "source_amount_delta": 0.0,
            "row_count": 0,
            "site_uids": set(),
            "suppliers": set(),
            "roles": set(),
            "path_ids": set(),
            "components": set(),
            "systems": set(),
            "families": set(),
            "decision_labels": Counter(),
            "driver_types": Counter(),
            "mapping_rules": Counter(),
            "lcia_allocation_methods": Counter(),
        },
    )
    group["delta_amount"] += delta_amount
    group["delta_kgco2e"] += delta_kgco2e
    group["source_inventory_delta_kgco2e"] += safe_float(source_row.get("delta_kgco2e"))
    group["source_amount_delta"] += safe_float(source_row.get("amount_delta"))
    group["row_count"] += 1
    for field, target in (
        ("site_uid", "site_uids"),
        ("supplier", "suppliers"),
        ("role", "roles"),
        ("path_id", "path_ids"),
        ("component", "components"),
        ("system", "systems"),
        ("family", "families"),
    ):
        value = clean(source_row.get(field))
        if value:
            group[target].add(value)
    group["decision_labels"].update(split_tokens(source_row.get("decision_labels")))
    group["driver_types"].update(split_tokens(source_row.get("source_driver_types")))
    group["mapping_rules"][mapping_rule_id] += 1
    group["lcia_allocation_methods"][lcia_allocation_method] += 1


def build_sdd_brightway_exchange_delta(
    ledger_rows: list[dict[str, Any]],
    brightway_model: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    activity_exchanges, activity_names = brightway_exchanges_by_activity(brightway_model.get("exchanges", []))
    selection_cache: dict[tuple[str, str], tuple[list[dict[str, Any]], str]] = {}
    activity_cache: dict[tuple[str, str, str, str], tuple[str, list[dict[str, Any]], str]] = {}
    groups: dict[tuple[Any, ...], dict[str, Any]] = {}

    for row in ledger_rows:
        if not row.get("include_in_dynamic_acv"):
            continue
        row_delta = safe_float(row.get("delta_kgco2e"))
        if abs(row_delta) < 1e-12:
            continue
        activity_cache_key = (
            clean(row.get("brightway_system")),
            clean(row.get("system")),
            clean(row.get("brightway_component")),
            clean(row.get("component")),
        )
        activity_name, exchanges, activity_match_status = activity_cache.get(activity_cache_key, ("", [], ""))
        if not activity_match_status:
            activity_name, exchanges, activity_match_status = resolve_brightway_activity_for_delta(row, activity_exchanges, activity_names)
            activity_cache[activity_cache_key] = (activity_name, exchanges, activity_match_status)

        mechanism = clean(row.get("mechanism"))
        selection_key = (activity_name, mechanism)
        selected_exchanges, mapping_rule_id = selection_cache.get(selection_key, ([], ""))
        if not mapping_rule_id:
            selected_exchanges, mapping_rule_id = selected_brightway_exchanges_for_mechanism(exchanges, mechanism)
            selection_cache[selection_key] = (selected_exchanges, mapping_rule_id)

        if not selected_exchanges:
            proxy_unit = "ton kilometer" if mechanism == "premium_transport" else clean(row.get("amount_unit"))
            append_exchange_delta_group(
                groups,
                source_row=row,
                activity_name=activity_name or clean(row.get("brightway_activity_proxy")) or "production du siege",
                activity_match_status=activity_match_status or "activity_unmatched",
                exchange={
                    "name": clean(row.get("brightway_exchange_proxy")) or f"{mechanism} proxy",
                    "unit": proxy_unit,
                    "database": "SDD proxy",
                    "location": clean(row.get("route_region")) or "case",
                },
                exchange_category="proxy_unmapped",
                mapping_status="proxy_unmapped",
                mapping_rule_id=mapping_rule_id or "no_exchange_mapping",
                lcia_allocation_method="source_delta_kept_as_unmapped_proxy",
                delta_amount=exchange_delta_proxy_amount(row),
                delta_kgco2e=row_delta,
            )
            continue

        exchange_amounts = [sdd_exchange_delta_amount(row, exchange) for exchange in selected_exchanges]
        total_weight = sum(abs(value) for value in exchange_amounts)
        weights = [1.0 / len(selected_exchanges)] * len(selected_exchanges) if total_weight <= 1e-12 else [abs(value) / total_weight for value in exchange_amounts]
        for exchange, delta_amount, weight in zip(selected_exchanges, exchange_amounts, weights):
            mapping_status = "virtual_exchange_proxy" if clean(exchange.get("database")) == "SDD proxy" else "mapped_exchange"
            append_exchange_delta_group(
                groups,
                source_row=row,
                activity_name=activity_name or clean(exchange.get("activity_name")) or clean(row.get("brightway_activity_proxy")) or "production du siege",
                activity_match_status=activity_match_status or "virtual_proxy",
                exchange=exchange,
                exchange_category=clean(exchange.get("exchange_category")) or brightway_exchange_category(exchange),
                mapping_status=mapping_status,
                mapping_rule_id=mapping_rule_id,
                lcia_allocation_method="allocated_from_sdd_brightway_delta_pending_exact_exchange_lcia",
                delta_amount=delta_amount,
                delta_kgco2e=row_delta * weight,
            )

    exchange_rows: list[dict[str, Any]] = []
    for group in groups.values():
        site_uids = sorted(group["site_uids"])
        suppliers = sorted(group["suppliers"])
        roles = sorted(group["roles"], key=lambda role: ROLE_SEQUENCE.index(role) if role in ROLE_SEQUENCE else 99)
        paths = sorted(group["path_ids"])
        components = sorted(group["components"])
        systems = sorted(group["systems"])
        families = sorted(group["families"])
        rule_ids = [label for label, _ in group["mapping_rules"].most_common(4)]
        allocation_methods = [label for label, _ in group["lcia_allocation_methods"].most_common(3)]
        key_text = "|".join(
            clean(group.get(field))
            for field in ("month_index", "mechanism", "activity_name", "exchange_name", "mapping_status")
        )
        exchange_rows.append(
            {
                "exchange_delta_id": "sddbwx:" + hashlib.sha1(key_text.encode("utf-8")).hexdigest()[:14],
                "month_index": group["month_index"],
                "site_uid": site_uids[0] if len(site_uids) == 1 else "multi_site",
                "supplier": suppliers[0] if len(suppliers) == 1 else "multi_supplier",
                "role": roles[0] if len(roles) == 1 else "multi_role",
                "mechanism": group["mechanism"],
                "inventory_delta_type": group["inventory_delta_type"],
                "activity_name": group["activity_name"],
                "exchange_name": group["exchange_name"],
                "exchange_reference_product": group["exchange_reference_product"],
                "exchange_category": group["exchange_category"],
                "exchange_unit": group["exchange_unit"],
                "exchange_database": group["exchange_database"],
                "exchange_location": group["exchange_location"],
                "delta_amount": round(group["delta_amount"], 12),
                "delta_kgco2e": round(group["delta_kgco2e"], 9),
                "source_inventory_delta_kgco2e": round(group["source_inventory_delta_kgco2e"], 9),
                "source_amount_delta": round(group["source_amount_delta"], 9),
                "row_count": group["row_count"],
                "site_count": len(site_uids),
                "role_count": len(roles),
                "path_count": len(paths),
                "component_count": len(components),
                "system_count": len(systems),
                "site_uid_sample": "|".join(site_uids[:5]),
                "supplier_sample": " | ".join(suppliers[:3]),
                "role_sample": "|".join(roles),
                "path_id_sample": "|".join(paths[:5]),
                "component_sample": " | ".join(components[:3]),
                "system_sample": " | ".join(systems[:3]),
                "family_sample": " | ".join(families[:3]),
                "decision_label_sample": top_counter_labels(group["decision_labels"], 4),
                "driver_type_sample": top_counter_labels(group["driver_types"], 4),
                "mapping_status": group["mapping_status"],
                "activity_match_status": group["activity_match_status"],
                "mapping_rule_id": " | ".join(rule_ids),
                "lcia_allocation_method": " | ".join(allocation_methods),
                "confidence": group["confidence"],
            }
        )

    exchange_rows.sort(
        key=lambda row: (
            int(safe_float(row.get("month_index"))),
            clean(row.get("site_uid")),
            clean(row.get("role")),
            clean(row.get("mechanism")),
            -abs(safe_float(row.get("delta_kgco2e"))),
        )
    )

    category_groups: dict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: defaultdict(float))
    top_groups: dict[tuple[str, str, str], dict[str, Any]] = defaultdict(lambda: defaultdict(float))
    for row in exchange_rows:
        cat_key = (clean(row.get("exchange_category")), clean(row.get("mapping_status")))
        category_groups[cat_key]["delta_kgco2e"] += safe_float(row.get("delta_kgco2e"))
        category_groups[cat_key]["abs_delta_amount"] += abs(safe_float(row.get("delta_amount")))
        category_groups[cat_key]["row_count"] += safe_float(row.get("row_count"))
        category_groups[cat_key]["exchange_count"] += 1
        top_key = (clean(row.get("activity_name")), clean(row.get("exchange_name")), clean(row.get("exchange_category")))
        top_groups[top_key]["delta_kgco2e"] += safe_float(row.get("delta_kgco2e"))
        top_groups[top_key]["abs_delta_amount"] += abs(safe_float(row.get("delta_amount")))
        top_groups[top_key]["row_count"] += safe_float(row.get("row_count"))
        top_groups[top_key]["exchange_unit"] = clean(row.get("exchange_unit"))
        top_groups[top_key]["mapping_status"] = clean(row.get("mapping_status"))

    category_rows = [
        {
            "exchange_category": category,
            "mapping_status": mapping_status,
            "label": f"{category} / {mapping_status}",
            "value": round(safe_float(values.get("delta_kgco2e")), 9),
            "delta_kgco2e": round(safe_float(values.get("delta_kgco2e")), 9),
            "abs_delta_amount": round(safe_float(values.get("abs_delta_amount")), 9),
            "row_count": int(safe_float(values.get("row_count"))),
            "exchange_row_count": int(safe_float(values.get("exchange_count"))),
        }
        for (category, mapping_status), values in category_groups.items()
    ]
    category_rows.sort(key=lambda row: -abs(safe_float(row.get("delta_kgco2e"))))

    top_exchange_rows = [
        {
            "activity_name": activity_name,
            "exchange_name": exchange_name,
            "exchange_category": exchange_category,
            "label": f"{activity_name} -> {exchange_name}",
            "value": round(safe_float(values.get("delta_kgco2e")), 9),
            "delta_kgco2e": round(safe_float(values.get("delta_kgco2e")), 9),
            "abs_delta_amount": round(safe_float(values.get("abs_delta_amount")), 9),
            "exchange_unit": clean(values.get("exchange_unit")),
            "mapping_status": clean(values.get("mapping_status")),
            "row_count": int(safe_float(values.get("row_count"))),
        }
        for (activity_name, exchange_name, exchange_category), values in top_groups.items()
    ]
    top_exchange_rows.sort(key=lambda row: -abs(safe_float(row.get("delta_kgco2e"))))
    return exchange_rows, category_rows, top_exchange_rows[:30]


def compact_sdd_brightway_inventory_delta(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = (
            int(safe_float(row.get("month_index"))),
            clean(row.get("role")),
            clean(row.get("site_uid")),
            clean(row.get("supplier")),
            clean(row.get("mechanism")),
            clean(row.get("inventory_delta_type")),
            clean(row.get("decision_labels")),
            clean(row.get("source_driver_types")),
            clean(row.get("edge")),
            clean(row.get("route_region")),
            clean(row.get("calibration_source")),
            clean(row.get("confidence")),
            bool(row.get("include_in_dynamic_acv")),
        )
        if key not in grouped:
            grouped[key] = {
                "month_index": key[0],
                "role": key[1],
                "site_uid": key[2],
                "supplier": key[3],
                "mechanism": key[4],
                "inventory_delta_type": key[5],
                "decision_labels": key[6],
                "source_driver_types": key[7],
                "edge": key[8],
                "route_region": key[9],
                "calibration_source": key[10],
                "confidence": key[11],
                "include_in_dynamic_acv": key[12],
                "row_count": 0,
                "path_ids": set(),
                "record_indexes": set(),
                "components": set(),
                "systems": set(),
                "sdd_event_join_keys": set(),
                "source_environmental_event_samples": set(),
                "source_transport_flow_samples": set(),
                "amount_units": set(),
                "intensity_units": set(),
                "calculation_notes": set(),
                "mapping_rule_ids": set(),
                "amount_delta": 0.0,
                "delta_kgco2e": 0.0,
                "baseline_path_production_kgco2e": 0.0,
                "path_mass_kg": 0.0,
                "source_environmental_event_count": 0,
                "source_transport_flow_count": 0,
                "role_scope_weighted": 0.0,
                "role_scope_weight": 0.0,
            }
        group = grouped[key]
        amount = safe_float(row.get("amount_delta"))
        group["row_count"] += 1
        group["amount_delta"] += amount
        group["delta_kgco2e"] += safe_float(row.get("delta_kgco2e"))
        group["baseline_path_production_kgco2e"] += safe_float(row.get("baseline_path_production_kgco2e"))
        group["path_mass_kg"] += safe_float(row.get("path_mass_kg"))
        group["source_environmental_event_count"] += int(safe_float(row.get("source_environmental_event_count")))
        group["source_transport_flow_count"] += int(safe_float(row.get("source_transport_flow_count")))
        role_scope_weight = abs(amount) if abs(amount) > 0.0 else 1.0
        group["role_scope_weighted"] += safe_float(row.get("role_scope_share")) * role_scope_weight
        group["role_scope_weight"] += role_scope_weight
        for source_key, target_key in [
            ("path_id", "path_ids"),
            ("record_index", "record_indexes"),
            ("component", "components"),
            ("system", "systems"),
            ("sdd_event_join_key", "sdd_event_join_keys"),
            ("source_environmental_event_sample", "source_environmental_event_samples"),
            ("source_transport_flow_sample", "source_transport_flow_samples"),
            ("amount_unit", "amount_units"),
            ("intensity_unit", "intensity_units"),
            ("calculation_note", "calculation_notes"),
            ("mapping_rule_id", "mapping_rule_ids"),
        ]:
            for value in [part for part in clean(row.get(source_key)).split("|") if part]:
                group[target_key].add(value)

    out: list[dict[str, Any]] = []
    for key, group in grouped.items():
        key_text = "|".join(str(part) for part in key)
        path_ids = sorted(group["path_ids"])
        components = sorted(group["components"])
        systems = sorted(group["systems"])
        record_indexes = sorted(group["record_indexes"])
        join_keys = sorted(group["sdd_event_join_keys"])
        env_event_samples = sorted(group["source_environmental_event_samples"])
        transport_flow_samples = sorted(group["source_transport_flow_samples"])
        amount_units = sorted(group["amount_units"])
        intensity_units = sorted(group["intensity_units"])
        calculation_notes = sorted(group["calculation_notes"])
        mapping_rule_ids = sorted(group["mapping_rule_ids"])
        out.append(
            {
                "delta_id": "sddbwagg:" + hashlib.sha1(key_text.encode("utf-8")).hexdigest()[:14],
                "month_index": group["month_index"],
                "role": group["role"],
                "site_uid": group["site_uid"],
                "supplier": group["supplier"],
                "mechanism": group["mechanism"],
                "inventory_delta_type": group["inventory_delta_type"],
                "decision_labels": group["decision_labels"],
                "source_driver_types": group["source_driver_types"],
                "edge": group["edge"],
                "route_region": group["route_region"],
                "amount_delta": round(group["amount_delta"], 9),
                "amount_unit": " | ".join(amount_units[:3]),
                "delta_kgco2e": round(group["delta_kgco2e"], 9),
                "include_in_dynamic_acv": group["include_in_dynamic_acv"],
                "baseline_path_production_kgco2e": round(group["baseline_path_production_kgco2e"], 9),
                "path_mass_kg": round(group["path_mass_kg"], 9),
                "role_scope_share_avg": round(group["role_scope_weighted"] / group["role_scope_weight"], 6) if group["role_scope_weight"] else "",
                "row_count": group["row_count"],
                "path_count": len(path_ids),
                "record_count": len(record_indexes),
                "component_count": len(components),
                "system_count": len(systems),
                "source_environmental_event_count": group["source_environmental_event_count"],
                "source_environmental_event_sample": "|".join(env_event_samples[:5]),
                "source_transport_flow_count": group["source_transport_flow_count"],
                "source_transport_flow_sample": "|".join(transport_flow_samples[:5]),
                "path_id_sample": "|".join(path_ids[:5]),
                "record_index_sample": "|".join(record_indexes[:5]),
                "component_sample": " | ".join(components[:3]),
                "system_sample": " | ".join(systems[:3]),
                "sdd_event_join_key_sample": "|".join(join_keys[:5]),
                "intensity_unit_sample": " | ".join(intensity_units[:2]),
                "calibration_source": group["calibration_source"],
                "confidence": group["confidence"],
                "calculation_note_sample": " | ".join(calculation_notes[:2]),
                "mapping_rule_id": " | ".join(mapping_rule_ids[:3]),
            }
        )
    out.sort(
        key=lambda row: (
            int(safe_float(row.get("month_index"))),
            clean(row.get("site_uid")),
            clean(row.get("role")),
            clean(row.get("mechanism")),
        )
    )
    return out


def build_sdd_brightway_coupling(
    *,
    path_rows: list[dict[str, Any]],
    sdd_results: dict[str, list[dict[str, Any]]],
    brightway_model: dict[str, Any],
) -> dict[str, list[dict[str, Any]]]:
    path_context = path_context_by_id(path_rows, brightway_model)
    lanes = lane_state_index(sdd_results.get("sdd_lane_state", []))
    production_score = brightway_exact_score(brightway_model, "production")
    lifecycle_score = brightway_exact_score(brightway_model, "lifecycle_excel_aligned")
    usage_delta = max(0.0, lifecycle_score - production_score) if lifecycle_score and production_score else 0.0
    seats_per_month = 1.0

    ledger_rows: list[dict[str, Any]] = []

    def compact_token_summary(value: Any, limit: int = 3) -> tuple[int, str]:
        tokens = [part for part in clean(value).split("|") if part and part != "none"]
        return len(tokens), "|".join(tokens[:limit])

    def add_delta(
        *,
        node: dict[str, Any],
        mechanism: str,
        inventory_delta_type: str,
        amount_delta: float,
        amount_unit: str,
        intensity: float,
        intensity_unit: str,
        multiplier: float,
        role_scope_share: float,
        include_in_dynamic_acv: bool,
        calibration_source: str,
        confidence: str,
        brightway_exchange_proxy: str,
        note: str,
        lane: dict[str, Any] | None = None,
        model_parameters: dict[str, Any] | None = None,
    ) -> None:
        if amount_delta <= 0.0 and include_in_dynamic_acv:
            return
        path_id = clean(node.get("path_id"))
        context = path_context.get(path_id, {})
        path = context.get("path", {})
        delta = amount_delta * intensity * multiplier if include_in_dynamic_acv else 0.0
        if abs(delta) < 1e-9 and include_in_dynamic_acv:
            return
        month = int(safe_float(node.get("month_index")))
        role = clean(node.get("role"))
        env_event_count, env_event_sample = compact_token_summary(node.get("source_environmental_event_ids"))
        transport_flow_count, transport_flow_sample = compact_token_summary(node.get("source_transport_flow_uids"))
        ledger_row = {
            "delta_id": f"sddbw:{path_id}:{month:03d}:{role}:{mechanism}:{len(ledger_rows) + 1}",
            "sdd_event_join_key": f"sdd:{path_id}:{month:03d}:{role}",
            "month_index": month,
            "path_id": path_id,
            "record_index": node.get("record_index"),
            "role": role,
            "site_uid": clean(node.get("site_uid")),
            "supplier": clean(node.get("supplier")),
            "system": clean(path.get("system")),
            "component": clean(path.get("component")),
            "family": clean(node.get("family")),
            "mechanism": mechanism,
            "inventory_delta_type": inventory_delta_type,
            "decision_labels": clean(node.get("decisions")),
            "source_driver_types": clean(node.get("source_driver_types")),
            "source_environmental_event_count": env_event_count,
            "source_environmental_event_sample": env_event_sample,
            "source_transport_flow_count": transport_flow_count,
            "source_transport_flow_sample": transport_flow_sample,
            "edge": clean((lane or {}).get("edge")),
            "route_region": clean((lane or {}).get("route_region")),
            "amount_delta": round(amount_delta, 9),
            "amount_unit": amount_unit,
            "intensity_kgco2e_per_unit": round(intensity, 12),
            "intensity_unit": intensity_unit,
            "operational_multiplier": round(multiplier, 6),
            "role_scope_share": round(role_scope_share, 6),
            "delta_kgco2e": round(delta, 9),
            "include_in_dynamic_acv": include_in_dynamic_acv,
            "baseline_path_production_kgco2e": round(safe_float(context.get("static_production_kgco2e")), 9),
            "path_mass_kg": round(safe_float(node.get("path_mass_kg")), 9),
            "brightway_activity_proxy": "production du siege",
            "brightway_exchange_proxy": brightway_exchange_proxy,
            "brightway_match_level": clean(context.get("match_level")),
            "brightway_system": clean(context.get("brightway_system")),
            "brightway_component": clean(context.get("brightway_component")),
            "calibration_source": calibration_source,
            "confidence": confidence,
            "calculation_note": note,
            "mapping_rule_id": f"{mechanism}:{inventory_delta_type}",
        }
        for key, value in (model_parameters or {}).items():
            ledger_row[f"model_{key}"] = round(value, 9) if isinstance(value, (int, float)) else clean(value)
        ledger_rows.append(ledger_row)

    for node in sdd_results.get("sdd_node_state", []):
        path_id = clean(node.get("path_id"))
        context = path_context.get(path_id)
        if not context:
            continue
        role = clean(node.get("role"))
        month = int(safe_float(node.get("month_index")))
        inbound_edge = role_inbound_edge(role)
        inbound_lane = lanes.get((path_id, month, inbound_edge)) if inbound_edge else None
        component_intensity = safe_float(context.get("component_intensity_kgco2e_per_kg"))
        role_share = sdd_brightway_role_scope_share(role)
        role_local_component_intensity = component_intensity * role_share
        calibration_source = clean(context.get("calibration_source"))
        exchange_proxy = clean(context.get("brightway_component")) or clean(node.get("family")) or "component production proxy"

        backup_output = safe_float(node.get("backup_output_kg"))
        add_delta(
            node=node,
            mechanism="backup_material",
            inventory_delta_type="supplier_substitution_overhead",
            amount_delta=backup_output,
            amount_unit="kg backup component output",
            intensity=role_local_component_intensity,
            intensity_unit="kgCO2e/kg output, role-local share of path-scaled Brightway production",
            multiplier=SDD_BACKUP_SUPPLIER_OVERHEAD_FACTOR,
            role_scope_share=role_share,
            include_in_dynamic_acv=True,
            calibration_source=calibration_source,
            confidence="medium",
            brightway_exchange_proxy=exchange_proxy,
            note="Backup supplier output modelled as the differential sourcing overhead, not a full duplicate of component production.",
            lane=inbound_lane,
        )

        scrap_rate = clamp(safe_float(node.get("scrap_multiplier")) - 1.0, 0.0, 0.35)
        good_output = safe_float(node.get("good_output_kg"))
        process_input_estimate = good_output / max(1.0 - scrap_rate, 0.000001) if scrap_rate > 0.0 else good_output
        scrap_mass = max(0.0, process_input_estimate - good_output)
        scrap_profile = scrap_recycling_profile(node.get("family"))
        scrap_model_parameters = {
            "scrap_recycling_rate": scrap_profile["recycling_rate"],
            "scrap_avoided_burden_factor": scrap_profile["avoided_burden_factor"],
            "scrap_treatment_kgco2e_per_kg": scrap_profile["treatment_kgco2e_per_kg"],
            "scrap_replacement_factor": scrap_profile["replacement_factor"],
        }
        add_delta(
            node=node,
            mechanism="scrap_rework",
            inventory_delta_type="yield_loss_replacement_production",
            amount_delta=scrap_mass * scrap_profile["replacement_factor"],
            amount_unit="kg replacement component output",
            intensity=role_local_component_intensity,
            intensity_unit="kgCO2e/kg replacement output, role-local share of path-scaled Brightway production",
            multiplier=1.0,
            role_scope_share=role_share,
            include_in_dynamic_acv=True,
            calibration_source=calibration_source,
            confidence="medium",
            brightway_exchange_proxy=exchange_proxy,
            note="Scrap creates replacement production for the same role-local component scope.",
            model_parameters=scrap_model_parameters,
            lane=inbound_lane,
        )
        add_delta(
            node=node,
            mechanism="scrap_treatment",
            inventory_delta_type="scrap_sorting_treatment_burden",
            amount_delta=scrap_mass,
            amount_unit="kg scrap treated",
            intensity=scrap_profile["treatment_kgco2e_per_kg"],
            intensity_unit="kgCO2e/kg scrap treatment and sorting proxy",
            multiplier=1.0,
            role_scope_share=role_share,
            include_in_dynamic_acv=True,
            calibration_source="family_scrap_recycling_profile_v1",
            confidence="medium",
            brightway_exchange_proxy=f"{exchange_proxy} scrap treatment proxy",
            note="Additional sorting, collection and treatment burden for operation-driven scrap.",
            model_parameters=scrap_model_parameters,
            lane=inbound_lane,
        )
        add_delta(
            node=node,
            mechanism="recycling_credit",
            inventory_delta_type="avoided_virgin_material_credit",
            amount_delta=scrap_mass * scrap_profile["recycling_rate"],
            amount_unit="kg scrap recovered",
            intensity=-role_local_component_intensity * scrap_profile["avoided_burden_factor"],
            intensity_unit="kgCO2e/kg recovered scrap, avoided burden credit against role-local component intensity",
            multiplier=1.0,
            role_scope_share=role_share,
            include_in_dynamic_acv=True,
            calibration_source="family_scrap_recycling_profile_v1",
            confidence="medium",
            brightway_exchange_proxy=f"{exchange_proxy} recycling credit proxy",
            note="Avoided-burden recycling credit; kept explicit so cut-off or no-credit variants can be compared.",
            model_parameters=scrap_model_parameters,
            lane=inbound_lane,
        )

        capacity_boost = safe_float(node.get("capacity_boost_output_kg"))
        add_delta(
            node=node,
            mechanism="capacity_energy",
            inventory_delta_type="auxiliary_capacity_energy",
            amount_delta=capacity_boost,
            amount_unit="kg output supported by capacity boost",
            intensity=0.85,
            intensity_unit="kgCO2e/kg boosted output proxy",
            multiplier=1.0,
            role_scope_share=role_share,
            include_in_dynamic_acv=True,
            calibration_source="sdd_energy_proxy_pending_site_energy_brightway_mapping",
            confidence="low",
            brightway_exchange_proxy="site electricity/gas capacity proxy",
            note="Temporary capacity support energy proxy; next step is mapping to regional electricity/gas exchanges.",
            lane=inbound_lane,
        )

        disruption_energy_amount = safe_float(node.get("path_mass_kg")) * safe_float(node.get("disruption_index")) * role_share
        add_delta(
            node=node,
            mechanism="disruption_energy_overhead",
            inventory_delta_type="weather_disruption_energy_overhead",
            amount_delta=disruption_energy_amount,
            amount_unit="kg component-month disruption exposure",
            intensity=0.10,
            intensity_unit="kgCO2e/kg disruption exposure proxy",
            multiplier=1.0,
            role_scope_share=role_share,
            include_in_dynamic_acv=True,
            calibration_source="sdd_energy_proxy_pending_site_energy_brightway_mapping",
            confidence="low",
            brightway_exchange_proxy="site disruption energy proxy",
            note="Residual energy overhead for degraded operating state.",
            lane=inbound_lane,
        )

        add_delta(
            node=node,
            mechanism="maintenance",
            inventory_delta_type="corrective_maintenance_overhead",
            amount_delta=disruption_energy_amount,
            amount_unit="kg component-month disruption exposure",
            intensity=0.08,
            intensity_unit="kgCO2e/kg disruption exposure proxy",
            multiplier=1.0,
            role_scope_share=role_share,
            include_in_dynamic_acv=True,
            calibration_source="sdd_maintenance_proxy_pending_brightway_process_mapping",
            confidence="low",
            brightway_exchange_proxy="maintenance corrective proxy",
            note="Corrective maintenance proxy induced by environmental and logistics stress.",
            lane=inbound_lane,
        )

        premium_output = safe_float(node.get("premium_output_kg"))
        distance = safe_float((inbound_lane or {}).get("distance_km"))
        add_delta(
            node=node,
            mechanism="premium_transport",
            inventory_delta_type="transport_mode_switch_to_air_proxy",
            amount_delta=premium_output * distance,
            amount_unit="kg.km premium transport",
            intensity=MODE_EF_KGCO2E_PER_KG_KM["air"],
            intensity_unit="kgCO2e/kg.km air freight proxy",
            multiplier=1.0,
            role_scope_share=role_share,
            include_in_dynamic_acv=True,
            calibration_source="transport_mode_proxy_air_kgkm",
            confidence="medium",
            brightway_exchange_proxy="market for transport, freight, aircraft proxy",
            note="Premium transport for unresolved gap/backlog, using air freight kg.km proxy.",
            lane=inbound_lane,
        )

    monthly_totals: dict[int, dict[str, Any]] = defaultdict(lambda: defaultdict(float))
    mechanism_totals: dict[str, dict[str, Any]] = defaultdict(lambda: defaultdict(float))
    site_totals: dict[str, dict[str, Any]] = defaultdict(lambda: defaultdict(float))
    component_totals: dict[str, dict[str, Any]] = defaultdict(lambda: defaultdict(float))

    for row in ledger_rows:
        if not row.get("include_in_dynamic_acv"):
            continue
        month = int(safe_float(row.get("month_index")))
        mechanism = clean(row.get("mechanism"))
        delta = safe_float(row.get("delta_kgco2e"))
        monthly_totals[month]["sdd_inventory_delta_kgco2e"] += delta
        monthly_totals[month][mechanism] += delta
        mechanism_totals[mechanism]["delta_kgco2e"] += delta
        mechanism_totals[mechanism]["row_count"] += 1
        mechanism_totals[mechanism]["amount_delta"] += safe_float(row.get("amount_delta"))

        site_key = clean(row.get("site_uid")) or clean(row.get("supplier"))
        site_totals[site_key]["delta_kgco2e"] += delta
        site_totals[site_key]["row_count"] += 1
        site_totals[site_key]["label"] = clean(row.get("supplier")) or site_key
        site_totals[site_key]["country_or_role"] = clean(row.get("role"))

        component_key = f"{clean(row.get('system'))} / {clean(row.get('component'))}"
        component_totals[component_key]["delta_kgco2e"] += delta
        component_totals[component_key]["row_count"] += 1
        component_totals[component_key]["label"] = component_key
        component_totals[component_key]["family"] = clean(row.get("family"))

    legacy_monthly = {
        int(safe_float(row.get("month_index"))): row
        for row in sdd_results.get("sdd_monthly_impacts", [])
    }
    max_month = max(
        list(monthly_totals.keys()) + [int(safe_float(row.get("month_index"))) for row in sdd_results.get("sdd_monthly_impacts", [])],
        default=0,
    )
    monthly_rows: list[dict[str, Any]] = []
    cumulative_rows: list[dict[str, Any]] = []
    running_static = 0.0
    running_delta = 0.0
    running_dynamic = 0.0
    running_cycle = 0.0
    mechanism_keys = sorted(mechanism_totals.keys())
    for month in range(1, max_month + 1):
        totals = monthly_totals.get(month, {})
        delta = safe_float(totals.get("sdd_inventory_delta_kgco2e"))
        static = production_score * seats_per_month
        dynamic = static + delta
        cycle_dynamic = dynamic + usage_delta * seats_per_month if usage_delta else 0.0
        legacy = legacy_monthly.get(month, {})
        row = {
            "month_index": month,
            "seat_equivalent_volume": seats_per_month,
            "production_static_kgco2e": round(static, 9),
            "sdd_inventory_delta_kgco2e": round(delta, 9),
            "production_dynamic_kgco2e": round(dynamic, 9),
            "cycle_dynamic_with_stelia_usage_kgco2e": round(cycle_dynamic, 9) if cycle_dynamic else "",
            "delta_vs_static_pct": round(100.0 * delta / static, 6) if static else "",
            "legacy_sdd_surimpact_proxy_kgco2e": safe_float(legacy.get("surimpact_total")),
            "legacy_sdd_kgco2e": safe_float(legacy.get("sdd_kgCO2e")),
        }
        for mechanism in mechanism_keys:
            row[mechanism] = round(safe_float(totals.get(mechanism)), 9)
        monthly_rows.append(row)
        running_static += static
        running_delta += delta
        running_dynamic += dynamic
        running_cycle += cycle_dynamic if cycle_dynamic else 0.0
        cumulative_rows.append(
            {
                "month_index": month,
                "production_static_cumulative": round(running_static, 9),
                "sdd_inventory_delta_cumulative": round(running_delta, 9),
                "production_dynamic_cumulative": round(running_dynamic, 9),
                "cycle_dynamic_with_stelia_usage_cumulative": round(running_cycle, 9) if running_cycle else "",
                "delta_vs_static_pct": round(100.0 * running_delta / running_static, 6) if running_static else "",
            }
        )

    mechanism_rows = [
        {
            "mechanism": mechanism,
            "label": mechanism.replace("_", " "),
            "delta_kgco2e": round(safe_float(values.get("delta_kgco2e")), 9),
            "amount_delta": round(safe_float(values.get("amount_delta")), 9),
            "row_count": int(safe_float(values.get("row_count"))),
            "share_of_delta_pct": round(100.0 * safe_float(values.get("delta_kgco2e")) / max(sum(safe_float(v.get("delta_kgco2e")) for v in mechanism_totals.values()), 0.000001), 6),
        }
        for mechanism, values in mechanism_totals.items()
    ]
    mechanism_rows.sort(key=lambda row: -safe_float(row.get("delta_kgco2e")))

    top_sites = [
        {
            "site_uid": site_uid,
            "label": values["label"],
            "meta": values["country_or_role"],
            "value": round(safe_float(values.get("delta_kgco2e")), 9),
            "row_count": int(safe_float(values.get("row_count"))),
        }
        for site_uid, values in site_totals.items()
    ]
    top_components = [
        {"label": values["label"], "meta": values["family"], "value": round(safe_float(values.get("delta_kgco2e")), 9), "row_count": int(safe_float(values.get("row_count")))}
        for values in component_totals.values()
    ]
    top_sites.sort(key=lambda row: -safe_float(row.get("value")))
    top_components.sort(key=lambda row: -safe_float(row.get("value")))

    summary_rows = []
    if cumulative_rows:
        final = cumulative_rows[-1]
        delta = safe_float(final.get("sdd_inventory_delta_cumulative"))
        static = safe_float(final.get("production_static_cumulative"))
        summary_rows.append(
            {
                "label": "Production ACV statique programme",
                "value": round(static, 9),
                "unit": "kgCO2e",
            }
        )
        summary_rows.append(
            {
                "label": "Surimpact ACV net",
                "value": round(delta, 9),
                "unit": "kgCO2e",
            }
        )
        summary_rows.append(
            {
                "label": "Surimpact ACV par siege equivalent",
                "value": round(delta / max(max_month * seats_per_month, 0.000001), 9),
                "unit": "kgCO2e/seat eq.",
            }
        )
        summary_rows.append(
            {
                "label": "Surimpact / production statique",
                "value": round(100.0 * delta / static, 6) if static else 0.0,
                "unit": "%",
            }
        )

    exchange_delta_rows, exchange_category_rows, top_exchange_rows = build_sdd_brightway_exchange_delta(ledger_rows, brightway_model)

    return {
        "inventory_delta": compact_sdd_brightway_inventory_delta(ledger_rows),
        "exchange_delta": exchange_delta_rows,
        "exchange_category_totals": exchange_category_rows,
        "top_exchanges": top_exchange_rows,
        "monthly": monthly_rows,
        "cumulative": cumulative_rows,
        "mechanism_totals": mechanism_rows,
        "site_impacts": top_sites,
        "top_sites": top_sites[:20],
        "top_components": top_components[:20],
        "summary": summary_rows,
    }


def build_sdd_tier_month_payload(node_state_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in node_state_rows:
        by_key[(clean(row.get("role")), int(safe_float(row.get("month_index"))))].append(row)
    out: list[dict[str, Any]] = []
    for (role, month), rows in sorted(by_key.items(), key=lambda item: (item[0][1], ROLE_SEQUENCE.index(item[0][0]) if item[0][0] in ROLE_SEQUENCE else 99)):
        out.append(
            {
                "role": role,
                "month_index": month,
                "avg_service_level": round(weighted_mean(rows, "service_level", "path_mass_kg"), 4),
                "avg_disruption_index": round(weighted_mean(rows, "disruption_index", "path_mass_kg"), 4),
                "backlog_kg": round(sum(safe_float(row.get("backlog_end_kg")) for row in rows), 6),
                "stock_draw_kg": round(sum(safe_float(row.get("stock_draw_kg")) for row in rows), 6),
                "backup_output_kg": round(sum(safe_float(row.get("backup_output_kg")) for row in rows), 6),
                "premium_output_kg": round(sum(safe_float(row.get("premium_output_kg")) for row in rows), 6),
            }
        )
    return out


def split_tokens(value: Any) -> list[str]:
    return [part for part in clean(value).split("|") if part and part != "none"]


def top_counter_labels(counter: Counter[str], limit: int = 5) -> str:
    return " | ".join(f"{label} ({count})" for label, count in counter.most_common(limit)) or "none"


def build_sdd_site_map_payload(
    site_rows: list[dict[str, Any]],
    sdd_node_state: list[dict[str, Any]],
    *,
    event_exposure_rows: list[dict[str, Any]] | None = None,
    sdd_brightway_site_impacts: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    by_site: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in sdd_node_state:
        by_site[clean(row.get("site_uid"))].append(row)

    exposure_by_label_country: dict[tuple[str, str], dict[str, Any]] = {}
    exposure_by_label: dict[str, dict[str, Any]] = {}
    for row in event_exposure_rows or []:
        label = clean(row.get("label"))
        country = clean(row.get("meta"))
        if label:
            exposure_by_label_country[(label, country)] = row
            if label not in exposure_by_label or safe_float(row.get("value")) > safe_float(exposure_by_label[label].get("value")):
                exposure_by_label[label] = row

    acv_by_site_uid = {
        clean(row.get("site_uid")): row
        for row in sdd_brightway_site_impacts or []
        if clean(row.get("site_uid"))
    }

    out: list[dict[str, Any]] = []
    for site in site_rows:
        site_uid = clean(site.get("site_uid"))
        lat = safe_float(site.get("lat"), float("nan"))
        lon = safe_float(site.get("lon"), float("nan"))
        if not math.isfinite(lat) or not math.isfinite(lon):
            continue
        rows = by_site.get(site_uid, [])
        decision_counts: Counter[str] = Counter()
        driver_counts: Counter[str] = Counter()
        affected_months = set()
        for row in rows:
            decision_counts.update(split_tokens(row.get("decisions")))
            driver_counts.update(split_tokens(row.get("source_driver_types")))
            if safe_float(row.get("disruption_index")) > 0.01 or safe_float(row.get("backlog_end_kg")) > 0.0:
                affected_months.add(int(safe_float(row.get("month_index"))))
        name = clean(site.get("name"))
        country = clean(site.get("country_code"))
        exposure = exposure_by_label_country.get((name, country), exposure_by_label.get(name, {}))
        acv_impact = acv_by_site_uid.get(site_uid, {})
        out.append(
            {
                "site_uid": site_uid,
                "name": site.get("name", ""),
                "roles": site.get("roles", ""),
                "country_code": site.get("country_code", ""),
                "lat": round(lat, 6),
                "lon": round(lon, 6),
                "allocated_mass_kg": round(safe_float(site.get("allocated_mass_kg")), 6),
                "path_count": int(safe_float(site.get("path_count"))),
                "avg_service_level": round(weighted_mean(rows, "service_level", "path_mass_kg"), 4) if rows else 1.0,
                "min_service_level": round(min((safe_float(row.get("service_level"), 1.0) for row in rows), default=1.0), 4),
                "avg_disruption_index": round(weighted_mean(rows, "disruption_index", "path_mass_kg"), 4) if rows else 0.0,
                "max_disruption_index": round(max((safe_float(row.get("disruption_index")) for row in rows), default=0.0), 4),
                "peak_backlog_kg": round(max((safe_float(row.get("backlog_end_kg")) for row in rows), default=0.0), 6),
                "total_backup_output_kg": round(sum(safe_float(row.get("backup_output_kg")) for row in rows), 6),
                "total_premium_output_kg": round(sum(safe_float(row.get("premium_output_kg")) for row in rows), 6),
                "affected_month_count": len(affected_months),
                "top_decisions": top_counter_labels(decision_counts),
                "top_drivers": top_counter_labels(driver_counts),
                "weather_exposure_index": round(safe_float(exposure.get("value")), 6),
                "weather_event_count": int(safe_float(exposure.get("events"))),
                "sdd_acv_delta_kgco2e": round(safe_float(acv_impact.get("value")), 9),
                "sdd_acv_row_count": int(safe_float(acv_impact.get("row_count"))),
            }
        )
    return sorted(out, key=lambda row: (-safe_float(row.get("avg_disruption_index")), row["name"]))


def build_sdd_lane_map_payload(site_rows: list[dict[str, Any]], sdd_lane_state: list[dict[str, Any]]) -> list[dict[str, Any]]:
    site_lookup = {clean(row.get("site_uid")): row for row in site_rows}
    by_lane: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in sdd_lane_state:
        key = (clean(row.get("from_site_uid")), clean(row.get("to_site_uid")), clean(row.get("edge")))
        by_lane[key].append(row)

    out: list[dict[str, Any]] = []
    for (from_uid, to_uid, edge), rows in by_lane.items():
        from_site = site_lookup.get(from_uid, {})
        to_site = site_lookup.get(to_uid, {})
        from_lat = safe_float(from_site.get("lat"), float("nan"))
        from_lon = safe_float(from_site.get("lon"), float("nan"))
        to_lat = safe_float(to_site.get("lat"), float("nan"))
        to_lon = safe_float(to_site.get("lon"), float("nan"))
        if not all(math.isfinite(value) for value in [from_lat, from_lon, to_lat, to_lon]):
            continue
        decision_counts = Counter(clean(row.get("transport_decision")) for row in rows if clean(row.get("transport_decision")) != "nominal")
        kgkm_by_path = {
            clean(row.get("path_id")): safe_float(row.get("baseline_kg_km"))
            for row in rows
            if clean(row.get("path_id"))
        }
        out.append(
            {
                "edge": edge,
                "from_site_uid": from_uid,
                "to_site_uid": to_uid,
                "from_name": rows[0].get("from_name", ""),
                "to_name": rows[0].get("to_name", ""),
                "from_lat": round(from_lat, 6),
                "from_lon": round(from_lon, 6),
                "to_lat": round(to_lat, 6),
                "to_lon": round(to_lon, 6),
                "modes": rows[0].get("modes", ""),
                "distance_km": rows[0].get("distance_km", ""),
                "allocated_kg_km": round(sum(kgkm_by_path.values()), 1),
                "avg_transport_risk": round(weighted_or_mean(rows, "transport_risk_index", "path_mass_kg"), 4),
                "max_transport_risk": round(max((safe_float(row.get("transport_risk_index")) for row in rows), default=0.0), 4),
                "avg_delay_multiplier": round(weighted_or_mean(rows, "delay_multiplier", "path_mass_kg", 1.0), 4),
                "avg_capacity_multiplier": round(weighted_or_mean(rows, "capacity_multiplier", "path_mass_kg", 1.0), 4),
                "avg_premium_mode_share": round(weighted_or_mean(rows, "premium_mode_share", "path_mass_kg"), 4),
                "route_region": rows[0].get("route_region", ""),
                "top_decisions": top_counter_labels(decision_counts),
            }
        )
    return sorted(out, key=lambda row: (-safe_float(row.get("avg_transport_risk")), -safe_float(row.get("allocated_kg_km"))))


def build_horizon_adaptation_payload() -> dict[str, Any]:
    csv_dir = REPO_ROOT / "POC2026" / "horizon-adaptation" / "outputs" / "csv"
    state_columns = [
        "policy_label",
        "month_index",
        "warming",
        "scarcity",
        "main_supply_availability_applied",
        "primary_supply_availability_applied",
        "capacity_applied",
        "backlog_end",
        "good_output_backup_units",
        "outbound_shipments",
        "grid_energy_kwh",
        "solar_used_kwh",
        "biomass_used_kwh",
        "battery_discharge_kwh",
        "battery_soc_kwh",
        "battery_soh",
        "hvac_kwh",
        "biomass_transition_level",
        "biomass_transition_cap",
        "biomass_resource_stock_ratio",
        "operational_feedback_event",
    ]
    monthly_states = read_csv_rows(csv_dir / "ha_monthly_states.csv", state_columns)
    policy_labels = list(dict.fromkeys(clean(row.get("policy_label")) for row in monthly_states if clean(row.get("policy_label"))))
    preferred_policy = "Adaptation integree" if "Adaptation integree" in policy_labels else (policy_labels[0] if policy_labels else "")
    selected_states = [row for row in monthly_states if clean(row.get("policy_label")) == preferred_policy]
    return {
        "schema_version": "poc2026.horizon_adaptation.dashboard_subset.v1",
        "available": csv_dir.exists(),
        "source_dir": rel(csv_dir, REPO_ROOT),
        "policy_label": preferred_policy,
        "weather_driver": read_csv_rows(csv_dir / "ha_weather_driver.csv"),
        "event_timeline": read_csv_rows(csv_dir / "ha_event_timeline.csv"),
        "reference_cumulative": read_csv_rows(csv_dir / "ha_reference_cumulative.csv"),
        "event_impact": read_csv_rows(csv_dir / "ha_event_impact_breakdown.csv"),
        "monthly_state": selected_states,
        "policy_labels": policy_labels,
    }


def build_general_kpi_payload(
    *,
    summary: dict[str, Any],
    path_rows: list[dict[str, Any]],
    lane_rows: list[dict[str, Any]],
    site_rows: list[dict[str, Any]],
    weather_rows: list[dict[str, Any]],
    event_rows: list[dict[str, Any]],
    transport_weather_rows: list[dict[str, Any]],
    node_operational_rows: list[dict[str, Any]],
    operational_event_rows: list[dict[str, Any]],
    sdd_results: dict[str, list[dict[str, Any]]],
    sdd_brightway: dict[str, list[dict[str, Any]]],
    brightway_model: dict[str, Any],
    map_src: str,
) -> dict[str, Any]:
    total_kg_km = sum(safe_float(row.get("allocated_kg_km")) for row in path_rows)
    total_path_mass = sum(safe_float(row.get("path_mass_kg")) for row in path_rows)
    avg_route_km = total_kg_km / total_path_mass if total_path_mass else 0.0
    lane_specific = sum(1 for row in path_rows if row.get("transport_model") == "lane_specific")
    lane_specific_pct = 100.0 * lane_specific / len(path_rows) if path_rows else 0.0

    mode_kg_km: dict[str, float] = defaultdict(float)
    for row in lane_rows:
        parts = [part for part in clean(row.get("modes")).split("|") if part] or ["unknown"]
        split_value = safe_float(row.get("allocated_kg_km")) / len(parts)
        for part in parts:
            mode_kg_km[part] += split_value
    mode_mix = [
        {"label": label, "value": round(value, 3)}
        for label, value in sorted(mode_kg_km.items(), key=lambda item: (-item[1], item[0]))
    ]

    event_month: dict[int, dict[str, int]] = defaultdict(lambda: {"heatwave": 0, "drought": 0, "storm": 0, "hurricane": 0, "cold": 0})
    event_site_intensity: dict[str, dict[str, Any]] = {}
    for row in event_rows:
        month = int(safe_float(row.get("month_index")))
        event_type = clean(row.get("event_type"))
        if event_type:
            event_month[month][event_type] = event_month[month].get(event_type, 0) + 1
        site_uid = clean(row.get("site_uid"))
        item = event_site_intensity.setdefault(
            site_uid,
            {
                "label": clean(row.get("supplier")) or site_uid,
                "meta": clean(row.get("country_code")),
                "value": 0.0,
                "events": 0,
            },
        )
        item["value"] += safe_float(row.get("intensity"))
        item["events"] += 1
    weather_month = build_weather_month_payload(weather_rows)
    climate_horizon = weather_month[-1] if weather_month else {}
    weather_horizon = max((int(row["month_index"]) for row in weather_month), default=0)
    event_horizon = max(event_month.keys(), default=0)
    horizon_months = max(weather_horizon, event_horizon)
    event_month_rows = [
        {"month_index": month, **event_month.get(month, {"heatwave": 0, "drought": 0, "storm": 0, "hurricane": 0, "cold": 0})}
        for month in range(1, horizon_months + 1)
    ]
    ops_month = build_ops_month_payload(event_rows, horizon_months)
    weather_region = build_weather_group_payload(weather_rows, site_rows, "world_region")
    weather_profile = build_weather_group_payload(weather_rows, site_rows, "weather_profile")
    weather_region_month = build_weather_region_month_payload(weather_rows)
    maritime_month = build_maritime_month_payload(transport_weather_rows)
    maritime_region = build_maritime_region_payload(transport_weather_rows)
    node_ops_month = build_node_ops_month_payload(node_operational_rows, operational_event_rows)
    node_ops_region = build_node_ops_region_payload(node_operational_rows)
    node_ops_lineage = build_node_ops_lineage_payload(operational_event_rows)
    event_exposure = sorted(event_site_intensity.values(), key=lambda row: (-safe_float(row["value"]), row["label"]))[:15]
    for row in event_exposure:
        row["value"] = round(safe_float(row["value"]), 3)

    scatter_rows = [
        {
            "path_id": row.get("path_id"),
            "system": row.get("system"),
            "component": row.get("component"),
            "family": row.get("family"),
            "x": safe_float(row.get("total_route_km")),
            "y": safe_float(row.get("path_mass_kg")),
            "kg_km": safe_float(row.get("allocated_kg_km")),
            "readiness": row.get("readiness"),
        }
        for row in path_rows
    ]

    counts = summary.get("counts", {})
    mass = summary.get("mass", {})
    brightway_counts = brightway_model.get("counts", {})
    brightway_runtime = brightway_model.get("runtime", {})
    sdd_bw_summary = {clean(row.get("label")): row for row in sdd_brightway.get("summary", [])}
    sdd_bw_delta = safe_float(sdd_bw_summary.get("Surimpact ACV net", {}).get("value"))
    sdd_bw_delta_pct = safe_float(sdd_bw_summary.get("Surimpact / production statique", {}).get("value"))
    sdd_bw_mechanisms = {clean(row.get("mechanism")): row for row in sdd_brightway.get("mechanism_totals", [])}
    sdd_bw_recycling_credit = safe_float(sdd_bw_mechanisms.get("recycling_credit", {}).get("delta_kgco2e"))
    sdd_bw_exchange_lcia_monthly = sdd_brightway.get("exchange_lcia_monthly", [])
    sdd_bw_exchange_lcia_total = sum(safe_float(row.get("exact_delta_kgco2e")) for row in sdd_bw_exchange_lcia_monthly)
    sdd_bw_exchange_lcia_status = (sdd_brightway.get("exchange_lcia_status") or [{}])[0]
    aligned_lifecycle = next(
        (
            row
            for row in brightway_model.get("exact_scenario_lcia", [])
            if clean(row.get("scenario_id")) == "current_export"
            and clean(row.get("root_activity_id")) == "lifecycle_excel_aligned"
        ),
        {},
    )
    cards = [
        {"label": "Lignes exploitables", "value": counts.get("usable_records", 0), "unit": ""},
        {"label": "Chemins primaires", "value": counts.get("primary_paths", 0), "unit": ""},
        {"label": "Masse allouee", "value": round(safe_float(mass.get("allocated_primary_path_mass_kg")), 2), "unit": "kg"},
        {"label": "Transport alloue", "value": round(total_kg_km, 0), "unit": "kg.km"},
        {"label": "Sites uniques", "value": counts.get("unique_sites", 0), "unit": ""},
        {"label": "Evenements simules", "value": counts.get("event_seed_rows", 0), "unit": ""},
        {"label": "Liaisons renseignees", "value": round(lane_specific_pct, 1), "unit": "%"},
        {"label": "Route moy. ponderee", "value": round(avg_route_km, 1), "unit": "km"},
        {"label": "Regions meteo", "value": len(weather_region), "unit": ""},
        {"label": "Profils meteo", "value": len(weather_profile), "unit": ""},
        {"label": "Rechauffement fin horizon", "value": round(safe_float(climate_horizon.get("avg_warming_delta_c")), 2), "unit": "degC"},
        {"label": "Aleas fin horizon", "value": round(safe_float(climate_horizon.get("avg_hazard_intensification_factor")), 2), "unit": "x"},
        {"label": "Flux maritimes exposes", "value": len({clean(row.get("flow_uid")) for row in transport_weather_rows}), "unit": ""},
        {"label": "Risque maritime moy.", "value": round(mean(safe_float(row.get("risk_index")) for row in maritime_month), 3), "unit": ""},
        {"label": "Evenements operationnels", "value": len(operational_event_rows), "unit": ""},
        {"label": "Service sites moyen", "value": round(mean(safe_float(row.get("avg_service_proxy_pct")) for row in node_ops_month), 1), "unit": "%"},
        {"label": "Service SDD OEM moy.", "value": round(mean(safe_float(row.get("avg_oem_service_level")) for row in sdd_results.get("sdd_monthly_impacts", [])) * 100.0, 1), "unit": "%"},
        {"label": "Impact operationnel estime", "value": round(sum(safe_float(row.get("surimpact_total")) for row in sdd_results.get("sdd_monthly_impacts", [])), 1), "unit": "kgCO2e"},
        {"label": "Surimpact ACV net", "value": round(sdd_bw_delta / 1000.0, 1), "unit": "tCO2e"},
        {"label": "Credit recyclage ACV", "value": round(sdd_bw_recycling_credit / 1000.0, 2), "unit": "tCO2e"},
        {"label": "Recalcul Brightway partiel", "value": round(sdd_bw_exchange_lcia_total / 1000.0, 1), "unit": "tCO2e"},
        {"label": "Facteurs ACV recalcules", "value": safe_float(sdd_bw_exchange_lcia_status.get("exact_factor_count")), "unit": ""},
        {"label": "Surimpact / production", "value": round(sdd_bw_delta_pct, 1), "unit": "%"},
        {"label": "Composants ACV BW", "value": brightway_counts.get("climate_component_rows", 0), "unit": ""},
        {"label": "Parametres BW", "value": brightway_counts.get("parameters", 0), "unit": ""},
        {"label": "Indicateurs PE BW", "value": brightway_counts.get("person_equivalent_indicators", 0), "unit": ""},
        {"label": "Scenarios regionaux Brightway", "value": brightway_counts.get("parametric_regional_scenarios", 0), "unit": ""},
        {"label": "Scenarios Brightway recalcules", "value": brightway_counts.get("exact_scenario_lcia", 0), "unit": ""},
        {"label": "Cycle BW corrige", "value": round(safe_float(aligned_lifecycle.get("score_kgco2e")) / 1000.0, 1), "unit": "tCO2e"},
        {"label": "Couverture ACV chaine", "value": round(100.0 * safe_float(brightway_counts.get("supply_alignment_matched_rows")) / safe_float(brightway_counts.get("supply_alignment_rows"), 1.0), 1), "unit": "%"},
        {"label": "Brightway disponible", "value": 1 if brightway_runtime.get("can_execute_brightway") else 0, "unit": "bool"},
    ]

    return {
        "schema_version": "poc2026.supply_geo_case.kpi_dashboard.v1",
        "generated_at_utc": summary.get("generated_at_utc"),
        "map_src": map_src,
        "cards": cards,
        "readiness": [{"label": key, "value": value} for key, value in summary.get("readiness_counts", {}).items()],
        "lca_use_class": [{"label": key, "value": value} for key, value in summary.get("lca_use_class_counts", {}).items()],
        "family_mass": grouped_sum(path_rows, "family", "path_mass_kg"),
        "family_count": grouped_count(path_rows, "family"),
        "edge_kg_km": grouped_sum(lane_rows, "edge", "allocated_kg_km"),
        "mode_kg_km": mode_mix,
        "top_sites_by_mass": top_rows(
            [row for row in site_rows if "OEM" not in set(clean(row.get("roles")).split("|"))],
            label_key="name",
            value_key="allocated_mass_kg",
            limit=15,
        ),
        "weather_month": weather_month,
        "weather_region": weather_region,
        "weather_profile": weather_profile,
        "weather_region_month": weather_region_month,
        "ops_month": ops_month,
        "maritime_month": maritime_month,
        "maritime_region": maritime_region,
        "node_ops_month": node_ops_month,
        "node_ops_region": node_ops_region,
        "node_ops_lineage": node_ops_lineage,
        "sdd_monthly": sdd_results.get("sdd_monthly_impacts", []),
        "sdd_cumulative": sdd_results.get("sdd_cumulative_impacts", []),
        "sdd_method_comparison": sdd_results.get("sdd_method_comparison", []),
        "sdd_tier_month": sdd_results.get("sdd_tier_month", []),
        "sdd_brightway": {
            "summary": sdd_brightway.get("summary", []),
            "monthly": sdd_brightway.get("monthly", []),
            "cumulative": sdd_brightway.get("cumulative", []),
            "mechanism_totals": sdd_brightway.get("mechanism_totals", []),
            "exchange_category_totals": sdd_brightway.get("exchange_category_totals", []),
            "top_exchanges": sdd_brightway.get("top_exchanges", []),
            "exchange_lcia_category_totals": sdd_brightway.get("exchange_lcia_category_totals", []),
            "exchange_lcia_monthly": sdd_brightway.get("exchange_lcia_monthly", []),
            "exchange_lcia_top": sdd_brightway.get("exchange_lcia_top", []),
            "exchange_lcia_status": sdd_brightway.get("exchange_lcia_status", []),
            "site_impacts": sdd_brightway.get("site_impacts", []),
            "top_sites": sdd_brightway.get("top_sites", []),
            "top_components": sdd_brightway.get("top_components", []),
        },
        "event_month": event_month_rows,
        "event_exposure": event_exposure,
        "path_scatter": scatter_rows,
        "horizon_adaptation": build_horizon_adaptation_payload(),
        "brightway_model": {
            "schema_version": brightway_model.get("schema_version"),
            "available": brightway_model.get("available"),
            "runtime": brightway_model.get("runtime", {}),
            "source_files": brightway_model.get("source_files", {}),
            "counts": brightway_model.get("counts", {}),
            "component_impacts": brightway_model.get("component_impacts", []),
            "indicator_summary": brightway_model.get("indicator_summary", []),
            "indicator_unit_views": brightway_model.get("indicator_unit_views", []),
            "reference_person_equivalent_results": brightway_model.get("reference_person_equivalent_results", []),
            "reference_weighted_results": brightway_model.get("reference_weighted_results", []),
            "reference_phase_breakdown": brightway_model.get("reference_phase_breakdown", []),
            "reference_use_phase_components": brightway_model.get("reference_use_phase_components", []),
            "reference_scenarios": brightway_model.get("reference_scenarios", []),
            "reference_weighting_factors": brightway_model.get("reference_weighting_factors", []),
            "reference_climate_contributors": brightway_model.get("reference_climate_contributors", []),
            "masterboard_equipment_summary": brightway_model.get("masterboard_equipment_summary", []),
            "masterboard_material_summary": brightway_model.get("masterboard_material_summary", []),
            "parameters": brightway_model.get("parameters", []),
            "activities": brightway_model.get("activities", []),
            "exchanges": brightway_model.get("exchanges", []),
            "supply_alignment": brightway_model.get("supply_alignment", []),
            "parametric_levers": brightway_model.get("parametric_levers", []),
            "parametric_sensitivity": brightway_model.get("parametric_sensitivity", []),
            "parametric_switches": brightway_model.get("parametric_switches", []),
            "parametric_regional_scenarios": brightway_model.get("parametric_regional_scenarios", []),
            "exact_scenario_lcia": brightway_model.get("exact_scenario_lcia", []),
            "excel_runtime_comparison": brightway_model.get("excel_runtime_comparison", []),
            "usage_calibration": brightway_model.get("usage_calibration", []),
            "top_climate_components": brightway_model.get("top_climate_components", []),
            "top_parameter_amounts": brightway_model.get("top_parameter_amounts", []),
        },
    }


def json_for_script(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False).replace("</", "<\\/")


def write_sdd_results_map_html(
    path: Path,
    *,
    site_rows: list[dict[str, Any]],
    sdd_results: dict[str, list[dict[str, Any]]],
    source_map_src: str,
) -> None:
    sites = build_sdd_site_map_payload(site_rows, sdd_results.get("sdd_node_state", []))
    lanes = build_sdd_lane_map_payload(site_rows, sdd_results.get("sdd_lane_state", []))
    if sites:
        center_lat = sum(safe_float(row.get("lat")) for row in sites) / len(sites)
        center_lon = sum(safe_float(row.get("lon")) for row in sites) / len(sites)
    else:
        center_lat, center_lon = 20.0, 0.0
    payload = {
        "sites": sites,
        "lanes": lanes,
        "source_map_src": source_map_src,
        "center": {"lat": round(center_lat, 5), "lon": round(center_lon, 5)},
        "stats": {
            "site_count": len(sites),
            "lane_count": len(lanes),
            "avg_disruption": round(mean(safe_float(row.get("avg_disruption_index")) for row in sites), 4),
            "avg_service": round(mean(safe_float(row.get("avg_service_level")) for row in sites), 4),
        },
    }
    payload_json = json_for_script(payload)
    document = """<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>POC2026 supply_geo SDD results map</title>
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <style>
    html, body, #map { height: 100%; margin: 0; }
    body { font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif; color: #172033; }
    .panel {
      position: absolute;
      top: 14px;
      left: 14px;
      z-index: 500;
      width: 320px;
      background: rgba(255, 255, 255, 0.94);
      border: 1px solid #d8dee8;
      border-radius: 8px;
      box-shadow: 0 8px 28px rgba(23, 32, 51, 0.16);
      padding: 12px;
    }
    .panel h1 { margin: 0 0 8px; font-size: 15px; line-height: 1.2; }
    .row { display: flex; gap: 8px; align-items: center; margin-top: 8px; }
    label { font-size: 12px; font-weight: 750; color: #4b5568; }
    select, button {
      width: 100%;
      border: 1px solid #cbd5e1;
      background: #fff;
      color: #172033;
      border-radius: 6px;
      padding: 7px 8px;
      font-size: 12px;
      font-weight: 700;
    }
    button { cursor: pointer; }
    .stats { display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px; margin-top: 10px; }
    .stat { border: 1px solid #e2e8f0; border-radius: 7px; padding: 8px; background: #f8fafc; }
    .stat b { display: block; font-size: 16px; }
    .stat span { font-size: 11px; color: #64748b; font-weight: 700; }
    .legend {
      position: absolute;
      right: 14px;
      bottom: 18px;
      z-index: 500;
      background: rgba(255, 255, 255, 0.94);
      border: 1px solid #d8dee8;
      border-radius: 8px;
      padding: 10px;
      font-size: 12px;
      font-weight: 700;
      min-width: 210px;
    }
    .scale { height: 10px; border-radius: 999px; background: linear-gradient(90deg, #2ca02c, #ffbf00, #d62728); margin: 7px 0; }
    .popup h2 { margin: 0 0 6px; font-size: 14px; }
    .popup table { border-collapse: collapse; font-size: 12px; }
    .popup td { padding: 2px 6px 2px 0; vertical-align: top; }
    .popup td:first-child { color: #64748b; font-weight: 750; }
    @media (max-width: 720px) {
      .panel { width: calc(100% - 28px); }
      .legend { left: 14px; right: auto; bottom: 14px; }
    }
  </style>
</head>
<body>
  <div id="map"></div>
  <section class="panel">
    <h1>Resultats SDD supply_geo</h1>
    <label for="metric">Couleur des noeuds</label>
    <div class="row">
      <select id="metric">
        <option value="avg_disruption_index">Perturbation moyenne</option>
        <option value="avg_service_level">Service moyen</option>
        <option value="peak_backlog_kg">Retard max</option>
        <option value="affected_month_count">Mois affectes</option>
      </select>
    </div>
    <div class="row">
      <button type="button" id="toggleLanes">Masquer les liens</button>
      <button type="button" id="fitMap">Cadrer</button>
    </div>
    <div class="stats">
      <div class="stat"><b id="siteCount"></b><span>sites</span></div>
      <div class="stat"><b id="laneCount"></b><span>liens</span></div>
      <div class="stat"><b id="avgDisruption"></b><span>perturbation moy.</span></div>
      <div class="stat"><b id="avgService"></b><span>service moy.</span></div>
    </div>
    <div class="row"><button type="button" id="openSource">Ouvrir carte source</button></div>
  </section>
  <section class="legend">
    <div id="legendTitle">Perturbation moyenne</div>
    <div class="scale"></div>
    <div>Vert = faible, rouge = fort. Les liens sont colores par risque transport.</div>
  </section>
  <script>
    const MAP_PAYLOAD = __MAP_PAYLOAD__;
    const sites = MAP_PAYLOAD.sites || [];
    const lanes = MAP_PAYLOAD.lanes || [];
    const map = L.map("map", { worldCopyJump: true }).setView([MAP_PAYLOAD.center.lat, MAP_PAYLOAD.center.lon], 3);
    L.tileLayer("https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png", {
      maxZoom: 18,
      attribution: "&copy; OpenStreetMap contributors"
    }).addTo(map);

    const siteLayer = L.layerGroup().addTo(map);
    const laneLayer = L.layerGroup().addTo(map);
    let lanesVisible = true;
    let siteMarkers = [];

    function fmt(value, digits = 2) {
      const n = Number(value);
      if (!Number.isFinite(n)) return String(value ?? "");
      if (Math.abs(n) >= 1000000) return (n / 1000000).toFixed(1) + "M";
      if (Math.abs(n) >= 1000) return n.toLocaleString("fr-FR", { maximumFractionDigits: 0 });
      return n.toLocaleString("fr-FR", { maximumFractionDigits: digits });
    }

    function clamp(value, low, high) {
      return Math.max(low, Math.min(high, value));
    }

    function riskColor(value) {
      const v = clamp(Number(value) || 0, 0, 1);
      const hue = 120 - 120 * v;
      return `hsl(${hue}, 72%, 44%)`;
    }

    function serviceColor(value) {
      const v = clamp(Number(value) || 0, 0, 1);
      return riskColor(1 - v);
    }

    function backlogColor(value) {
      const v = clamp(Math.log1p(Number(value) || 0) / 4, 0, 1);
      return riskColor(v);
    }

    function metricValue(site, metric) {
      return Number(site[metric] || 0);
    }

    function metricColor(site, metric) {
      if (metric === "avg_service_level") return serviceColor(metricValue(site, metric));
      if (metric === "peak_backlog_kg" || metric === "affected_month_count") return backlogColor(metricValue(site, metric));
      return riskColor(metricValue(site, metric));
    }

    function sitePopup(site) {
      return `<div class="popup">
        <h2>${site.name || "Site"}</h2>
        <table>
          <tr><td>Roles</td><td>${site.roles || ""}</td></tr>
          <tr><td>Pays</td><td>${site.country_code || ""}</td></tr>
          <tr><td>Masse</td><td>${fmt(site.allocated_mass_kg)} kg</td></tr>
          <tr><td>Service moy.</td><td>${fmt(site.avg_service_level * 100)}%</td></tr>
          <tr><td>Service min</td><td>${fmt(site.min_service_level * 100)}%</td></tr>
          <tr><td>Perturbation</td><td>${fmt(site.avg_disruption_index)}</td></tr>
          <tr><td>Retard max</td><td>${fmt(site.peak_backlog_kg)} kg</td></tr>
          <tr><td>Mois affectes</td><td>${site.affected_month_count}</td></tr>
          <tr><td>Decisions</td><td>${site.top_decisions}</td></tr>
          <tr><td>Drivers</td><td>${site.top_drivers}</td></tr>
        </table>
      </div>`;
    }

    function lanePopup(lane) {
      return `<div class="popup">
        <h2>${lane.from_name || ""} -> ${lane.to_name || ""}</h2>
        <table>
          <tr><td>Edge</td><td>${lane.edge}</td></tr>
          <tr><td>Modes</td><td>${lane.modes || ""}</td></tr>
          <tr><td>Bassin</td><td>${lane.route_region || ""}</td></tr>
          <tr><td>kg.km</td><td>${fmt(lane.allocated_kg_km)}</td></tr>
          <tr><td>Risque moy.</td><td>${fmt(lane.avg_transport_risk)}</td></tr>
          <tr><td>Risque max</td><td>${fmt(lane.max_transport_risk)}</td></tr>
          <tr><td>Delai</td><td>${fmt(lane.avg_delay_multiplier)}</td></tr>
          <tr><td>Capacite</td><td>${fmt(lane.avg_capacity_multiplier)}</td></tr>
          <tr><td>Decisions</td><td>${lane.top_decisions}</td></tr>
        </table>
      </div>`;
    }

    function drawSites() {
      const metric = document.getElementById("metric").value;
      siteLayer.clearLayers();
      siteMarkers = sites.map(site => {
        const radius = 5 + Math.min(18, Math.sqrt(Number(site.allocated_mass_kg || 0)) * 1.5);
        const marker = L.circleMarker([site.lat, site.lon], {
          radius,
          color: "#172033",
          weight: 0.7,
          fillColor: metricColor(site, metric),
          fillOpacity: 0.82
        }).bindPopup(sitePopup(site));
        marker.addTo(siteLayer);
        return marker;
      });
      const label = document.getElementById("metric").selectedOptions[0].textContent;
      document.getElementById("legendTitle").textContent = label;
    }

    function drawLanes() {
      laneLayer.clearLayers();
      lanes.forEach(lane => {
        const weight = 1 + Math.min(7, Math.log1p(Number(lane.allocated_kg_km || 0)) / 2.4);
        L.polyline([[lane.from_lat, lane.from_lon], [lane.to_lat, lane.to_lon]], {
          color: riskColor(lane.avg_transport_risk),
          weight,
          opacity: 0.48
        }).bindPopup(lanePopup(lane)).addTo(laneLayer);
      });
    }

    function fitMap() {
      const group = L.featureGroup([...siteMarkers]);
      if (siteMarkers.length) map.fitBounds(group.getBounds().pad(0.18));
    }

    document.getElementById("siteCount").textContent = fmt(MAP_PAYLOAD.stats.site_count, 0);
    document.getElementById("laneCount").textContent = fmt(MAP_PAYLOAD.stats.lane_count, 0);
    document.getElementById("avgDisruption").textContent = fmt(MAP_PAYLOAD.stats.avg_disruption, 3);
    document.getElementById("avgService").textContent = fmt(MAP_PAYLOAD.stats.avg_service * 100, 1) + "%";
    document.getElementById("metric").addEventListener("change", drawSites);
    document.getElementById("fitMap").addEventListener("click", fitMap);
    document.getElementById("toggleLanes").addEventListener("click", event => {
      lanesVisible = !lanesVisible;
      if (lanesVisible) {
        laneLayer.addTo(map);
        event.target.textContent = "Masquer les liens";
      } else {
        map.removeLayer(laneLayer);
        event.target.textContent = "Afficher les liens";
      }
    });
    document.getElementById("openSource").addEventListener("click", () => {
      if (MAP_PAYLOAD.source_map_src) window.open(MAP_PAYLOAD.source_map_src, "_blank");
    });

    drawLanes();
    drawSites();
    setTimeout(fitMap, 120);
  </script>
</body>
</html>
"""
    document = document.replace("__MAP_PAYLOAD__", payload_json)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(document, encoding="utf-8")


def write_enriched_base_map_html(
    path: Path,
    *,
    source_map: Path,
    site_rows: list[dict[str, Any]],
    sdd_results: dict[str, list[dict[str, Any]]],
    dashboard_payload: dict[str, Any] | None = None,
) -> None:
    dashboard_payload = dashboard_payload or {}
    sdd_brightway_payload = dashboard_payload.get("sdd_brightway", {}) if isinstance(dashboard_payload.get("sdd_brightway"), dict) else {}
    sites = build_sdd_site_map_payload(
        site_rows,
        sdd_results.get("sdd_node_state", []),
        event_exposure_rows=dashboard_payload.get("event_exposure", []),
        sdd_brightway_site_impacts=sdd_brightway_payload.get("site_impacts", []),
    )
    lanes = build_sdd_lane_map_payload(site_rows, sdd_results.get("sdd_lane_state", []))
    payload = {
        "schema_version": "poc2026.supply_geo_case.base_map_sdd_overlay.v1",
        "sites": sites,
        "lanes": lanes,
        "stats": {
            "site_count": len(sites),
            "lane_count": len(lanes),
            "avg_disruption": round(mean(safe_float(row.get("avg_disruption_index")) for row in sites), 4),
            "avg_service": round(mean(safe_float(row.get("avg_service_level")) for row in sites), 4),
            "max_backlog_kg": round(max((safe_float(row.get("peak_backlog_kg")) for row in sites), default=0.0), 3),
            "high_risk_lane_count": sum(1 for row in lanes if safe_float(row.get("avg_transport_risk")) >= 0.35),
        },
    }
    payload_json = json_for_script(payload)
    compact_dashboard_payload = {
        "schema_version": "poc2026.supply_geo_case.base_map_dashboard.v1",
        "cards": dashboard_payload.get("cards", []),
        "top_sites_by_mass": dashboard_payload.get("top_sites_by_mass", []),
        "family_mass": dashboard_payload.get("family_mass", []),
        "mode_kg_km": dashboard_payload.get("mode_kg_km", []),
        "edge_kg_km": dashboard_payload.get("edge_kg_km", []),
        "event_exposure": dashboard_payload.get("event_exposure", []),
        "weather_month": dashboard_payload.get("weather_month", []),
        "weather_region": dashboard_payload.get("weather_region", []),
        "weather_profile": dashboard_payload.get("weather_profile", []),
        "weather_region_month": dashboard_payload.get("weather_region_month", []),
        "event_month": dashboard_payload.get("event_month", []),
        "ops_month": dashboard_payload.get("ops_month", []),
        "maritime_month": dashboard_payload.get("maritime_month", []),
        "maritime_region": dashboard_payload.get("maritime_region", []),
        "node_ops_month": dashboard_payload.get("node_ops_month", []),
        "node_ops_region": dashboard_payload.get("node_ops_region", []),
        "node_ops_lineage": dashboard_payload.get("node_ops_lineage", []),
        "sdd_monthly": dashboard_payload.get("sdd_monthly", []),
        "sdd_cumulative": dashboard_payload.get("sdd_cumulative", []),
        "sdd_method_comparison": dashboard_payload.get("sdd_method_comparison", []),
        "sdd_tier_month": dashboard_payload.get("sdd_tier_month", []),
        "sdd_brightway": dashboard_payload.get("sdd_brightway", {}),
        "path_scatter": dashboard_payload.get("path_scatter", []),
        "brightway_model": dashboard_payload.get("brightway_model", {}),
    }
    dashboard_json = json_for_script(compact_dashboard_payload)
    source_html = source_map.read_text(encoding="utf-8")
    injection = """
<style>
  .toolbar { align-items: flex-start; flex-wrap: wrap; }
  #chart { height: calc(100vh - 112px); min-height: 640px; }
  .sdd-view-tabs {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 6px 8px;
    border: 1px solid #d8dee8;
    border-radius: 8px;
    background: #ffffff;
  }
  .sdd-view-tabs label {
    font-size: 12px;
    font-weight: 750;
    color: #334155;
    margin-right: 2px;
  }
  .sdd-view-tabs button,
  .sdd-view-tabs select {
    border: 1px solid #cbd5e1;
    background: #f8fafc;
    color: #172033;
    border-radius: 6px;
    padding: 5px 8px;
    font-size: 12px;
    font-weight: 700;
  }
  .sdd-view-tabs button {
    cursor: pointer;
  }
  .sdd-view-tabs button.active {
    background: #172033;
    border-color: #172033;
    color: #ffffff;
  }
  .sdd-map-note {
    font-size: 12px;
    font-weight: 700;
    color: #64748b;
    padding: 6px 8px;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    background: #f8fafc;
  }
  .sdd-dashboard-panel {
    display: block;
    min-height: calc(100vh - 112px);
    padding: 16px;
    background: #f6f7f9;
    color: #172033;
    box-sizing: border-box;
    overflow: auto;
  }
  .sdd-dashboard-panel[hidden] {
    display: none !important;
  }
  .sdd-dashboard-header {
    display: flex;
    align-items: flex-end;
    justify-content: space-between;
    gap: 16px;
    margin-bottom: 12px;
  }
  .sdd-dashboard-header h1 {
    margin: 0;
    font-size: 18px;
    line-height: 1.2;
  }
  .sdd-dashboard-header p {
    margin: 4px 0 0;
    color: #64748b;
    font-size: 12px;
    font-weight: 700;
  }
  .sdd-dashboard-cards {
    display: grid;
    grid-template-columns: repeat(6, minmax(120px, 1fr));
    gap: 10px;
    margin-bottom: 12px;
  }
  .sdd-dashboard-card {
    min-height: 72px;
    border: 1px solid #d8dee8;
    border-radius: 8px;
    background: #ffffff;
    padding: 10px;
    box-sizing: border-box;
  }
  .sdd-dashboard-card span {
    display: block;
    color: #64748b;
    font-size: 11px;
    font-weight: 750;
    line-height: 1.25;
  }
  .sdd-dashboard-card b {
    display: block;
    margin-top: 8px;
    font-size: 19px;
    line-height: 1.1;
  }
  .sdd-dashboard-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(320px, 1fr));
    gap: 12px;
  }
  .sdd-dashboard-plot {
    min-height: 315px;
    border: 1px solid #d8dee8;
    border-radius: 8px;
    background: #ffffff;
  }
  @media (max-width: 900px) {
    #chart { height: calc(100vh - 180px); min-height: 560px; }
    .sdd-view-tabs { width: 100%; flex-wrap: wrap; }
    .sdd-dashboard-header { display: block; }
    .sdd-dashboard-cards { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    .sdd-dashboard-grid { grid-template-columns: 1fr; }
  }
</style>
<section id="sddDashboardPanel" class="sdd-dashboard-panel" hidden>
  <div class="sdd-dashboard-header">
    <div>
      <h1>Dashboard KPI generaux</h1>
      <p>Vue integree dans la carte de base : meteo, operations, transport et impacts SDD.</p>
    </div>
    <p id="sddDashboardStatus"></p>
  </div>
  <div id="sddDashboardCards" class="sdd-dashboard-cards"></div>
  <div class="sdd-dashboard-grid">
    <div id="baseMapKpiMethodPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiCumulativePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiWeatherPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiWeatherWaterPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiRegionalWeatherPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiOpsPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiOpsEventsPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiOpsLineagePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiMaritimePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiMaritimeRegionPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiEventPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiEventExposurePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddServicePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddTierPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddImpactPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvStaticDynamicPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvDeltaMechanismPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvExchangeCategoryPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvTopExchangesPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvExchangeExactMonthlyPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvExchangeExactTopPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvSeatEquivalentPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvTopComponentsPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiSddAcvTopSitesPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiPathScatterPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiTopSitesPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiFamilyMassPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiModeMixPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwClimatePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwIndicatorPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwRawIndicatorPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwUnitCoveragePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwReferenceWeightedPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwReferencePhasePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwReferenceScenarioPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwReferenceClimateContributorPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwAlignmentPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwParamPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwLeverPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwSensitivityPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwSwitchPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwRegionalScenarioPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwUsageBreakdownPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwExactScenarioPlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwAlignedLifecyclePlot" class="sdd-dashboard-plot"></div>
    <div id="baseMapKpiBwExcelRuntimePlot" class="sdd-dashboard-plot"></div>
  </div>
</section>
<script>
const SDD_MAP_PAYLOAD = __SDD_MAP_PAYLOAD__;
const BASE_DASHBOARD_PAYLOAD = __BASE_DASHBOARD_PAYLOAD__;
(function() {
  const payload = SDD_MAP_PAYLOAD || {};
  const dashboardPayload = BASE_DASHBOARD_PAYLOAD || {};
  const sites = payload.sites || [];
  const lanes = payload.lanes || [];
  let currentView = "source";
  let sourceState = null;

  function clone(value) {
    return JSON.parse(JSON.stringify(value));
  }

  function fmt(value, digits = 2) {
    const n = Number(value);
    if (!Number.isFinite(n)) return String(value ?? "");
    if (Math.abs(n) >= 1000000) return (n / 1000000).toFixed(1) + "M";
    if (Math.abs(n) >= 1000) {
      return n.toLocaleString("fr-FR", { maximumFractionDigits: 0 });
    }
    return n.toLocaleString("fr-FR", { maximumFractionDigits: digits });
  }

  function escapeHtml(value) {
    return String(value ?? "")
      .replaceAll("&", "&amp;")
      .replaceAll("<", "&lt;")
      .replaceAll(">", "&gt;")
      .replaceAll('"', "&quot;");
  }

  function maxOf(rows, key) {
    return Math.max(0, ...rows.map(row => Number(row[key] || 0)).filter(Number.isFinite));
  }

  function dashboardRows(key) {
    const value = dashboardPayload[key];
    return Array.isArray(value) ? value : [];
  }

  function num(row, key, fallback = 0) {
    const value = Number(row?.[key]);
    return Number.isFinite(value) ? value : fallback;
  }

  function truthy(value) {
    return value === true || value === 1 || value === "1" || value === "true" || value === "True";
  }

  function showChartCanvas() {
    const chart = document.getElementById("chart");
    const panel = document.getElementById("sddDashboardPanel");
    if (panel) panel.hidden = true;
    if (chart) chart.style.display = "";
  }

  function showDashboardCanvas() {
    const chart = document.getElementById("chart");
    const panel = document.getElementById("sddDashboardPanel");
    if (chart) chart.style.display = "none";
    if (panel) panel.hidden = false;
  }

  function dashboardLayout(title, yTitle = "") {
    return {
      title: { text: title, x: 0.02, y: 0.95, font: { size: 14 } },
      margin: { l: 58, r: 24, t: 48, b: 46 },
      xaxis: { zeroline: false, gridcolor: "#eef2f7" },
      yaxis: { title: yTitle, zeroline: false, gridcolor: "#eef2f7" },
      legend: { orientation: "h", x: 0, y: 1.12 },
      paper_bgcolor: "#ffffff",
      plot_bgcolor: "#ffffff",
      font: { family: "system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif", color: "#172033" }
    };
  }

  function renderDashboardPlot(id, traces, layout) {
    const node = document.getElementById(id);
    if (!node) return;
    Plotly.react(node, traces, layout, { displayModeBar: false, responsive: true });
  }

  function renderDashboardCards() {
    const container = document.getElementById("sddDashboardCards");
    if (!container) return;
    const cards = dashboardRows("cards");
    container.innerHTML = cards.map(card => {
      const unit = card.unit ? ` ${escapeHtml(card.unit)}` : "";
      return `<div class="sdd-dashboard-card"><span>${escapeHtml(card.label || "")}</span><b>${fmt(card.value, 1)}${unit}</b></div>`;
    }).join("");
  }

  function horizontalBarPlot(id, rows, title, valueKey = "value", labelKey = "label", color = "#2b8cbe", xTitle = "") {
    const data = (rows || []).slice(0, 16).reverse();
    const layout = dashboardLayout(title, "");
    layout.margin = { l: 170, r: 24, t: 48, b: 42 };
    layout.xaxis.title = xTitle;
    renderDashboardPlot(
      id,
      [{ type: "bar", orientation: "h", x: data.map(row => num(row, valueKey)), y: data.map(row => row[labelKey] || row.name || "n/a"), marker: { color } }],
      layout
    );
  }

  function groupedCountRows(rows, key) {
    const counts = {};
    (rows || []).forEach(row => {
      const label = String(row?.[key] || "unknown");
      counts[label] = (counts[label] || 0) + 1;
    });
    return Object.entries(counts).map(([label, value]) => ({ label, value })).sort((a, b) => b.value - a.value || a.label.localeCompare(b.label));
  }

  function topBy(rows, key, limit = 14) {
    return (rows || [])
      .filter(row => Number.isFinite(Number(row?.[key])))
      .sort((a, b) => Number(b[key]) - Number(a[key]))
      .slice(0, limit);
  }

  function phaseRows(rows, view, shortLabel) {
    return (rows || []).filter(row => row.result_view === view && row.short_label === shortLabel && row.end_of_life_variant === "landfill");
  }

  function renderDashboard() {
    currentView = "dashboard";
    setActiveButton();
    showDashboardCanvas();
    renderDashboardCards();

    const cumulative = dashboardRows("sdd_cumulative");
    const cumulativeX = cumulative.map(row => num(row, "month_index"));
    renderDashboardPlot(
      "baseMapKpiCumulativePlot",
      [
        { type: "scatter", mode: "lines", name: "LCA classique", x: cumulativeX, y: cumulative.map(row => num(row, "classical_cumulative")) },
        { type: "scatter", mode: "lines", name: "TD-DLCA", x: cumulativeX, y: cumulative.map(row => num(row, "td_dlca_cumulative")) },
        { type: "scatter", mode: "lines", name: "SDD", x: cumulativeX, y: cumulative.map(row => num(row, "sdd_cumulative")) }
      ],
      dashboardLayout("Impacts cumules", "kgCO2e")
    );

    const methods = dashboardRows("sdd_method_comparison");
    renderDashboardPlot(
      "baseMapKpiMethodPlot",
      [{ type: "bar", x: methods.map(row => row.method), y: methods.map(row => num(row, "total_kgCO2e")), marker: { color: ["#9ecae1", "#3182bd", "#f16913"] } }],
      dashboardLayout("Comparaison methodes", "kgCO2e")
    );

    const weather = dashboardRows("weather_month");
    const weatherX = weather.map(row => num(row, "month_index"));
    renderDashboardPlot(
      "baseMapKpiWeatherPlot",
      [
        { type: "scatter", mode: "lines", name: "Temp. moy.", x: weatherX, y: weather.map(row => num(row, "avg_temp_c")) },
        { type: "scatter", mode: "lines", name: "Temp. p90", x: weatherX, y: weather.map(row => num(row, "p90_temp_c")) },
        { type: "scatter", mode: "lines", name: "Rechauffement applique", x: weatherX, y: weather.map(row => num(row, "avg_warming_delta_c")), line: { dash: "dot", width: 3 } },
        { type: "scatter", mode: "lines", name: "Humidite", x: weatherX, y: weather.map(row => num(row, "avg_humidity_pct")) },
        { type: "scatter", mode: "lines", name: "Vent", x: weatherX, y: weather.map(row => num(row, "avg_wind_ms")) }
      ],
      dashboardLayout("Drivers meteo moyens", "degC / % / m/s")
    );

    const ops = dashboardRows("node_ops_month");
    const opsX = ops.map(row => num(row, "month_index"));
    renderDashboardPlot(
      "baseMapKpiOpsPlot",
      [
        { type: "scatter", mode: "lines", name: "Service estime", x: opsX, y: ops.map(row => num(row, "avg_service_proxy_pct")) },
        { type: "scatter", mode: "lines", name: "Capacite", x: opsX, y: ops.map(row => num(row, "avg_capacity_applied") * 100) },
        { type: "scatter", mode: "lines", name: "Perturbation", x: opsX, y: ops.map(row => num(row, "avg_disruption_index") * 100) },
        { type: "bar", name: "Sites affectes", x: opsX, y: ops.map(row => num(row, "affected_site_count")), opacity: 0.28 }
      ],
      dashboardLayout("Etat operationnel des sites", "% / sites")
    );

    const maritime = dashboardRows("maritime_month");
    const maritimeX = maritime.map(row => num(row, "month_index"));
    renderDashboardPlot(
      "baseMapKpiMaritimePlot",
      [
        { type: "scatter", mode: "lines", name: "Risque maritime", x: maritimeX, y: maritime.map(row => num(row, "risk_index")) },
        { type: "scatter", mode: "lines", name: "Tempete", x: maritimeX, y: maritime.map(row => num(row, "storm")) },
        { type: "scatter", mode: "lines", name: "Ouragan", x: maritimeX, y: maritime.map(row => num(row, "hurricane")) },
        { type: "scatter", mode: "lines", name: "Froid", x: maritimeX, y: maritime.map(row => num(row, "cold")) },
        { type: "scatter", mode: "lines", name: "Mousson", x: maritimeX, y: maritime.map(row => num(row, "monsoon")) }
      ],
      dashboardLayout("Risque transport maritime", "indice")
    );

    const events = dashboardRows("event_month");
    const eventX = events.map(row => num(row, "month_index"));
    const eventLayout = dashboardLayout("Evenements environnementaux", "sites/mois");
    eventLayout.barmode = "stack";
    const eventSpecs = [
      ["heatwave", "Canicule"],
      ["drought", "Secheresse"],
      ["storm", "Tempete"],
      ["hurricane", "Ouragan"],
      ["cold", "Froid"]
    ];
    renderDashboardPlot(
      "baseMapKpiEventPlot",
      eventSpecs.map(([key, label], idx) => ({
        type: "bar",
        name: label,
        x: eventX,
        y: events.map(row => num(row, key)),
        marker: { color: ["#d62728", "#e6550d", "#3182bd", "#756bb1", "#6baed6"][idx] }
      })),
      eventLayout
    );

    const topSites = dashboardRows("top_sites_by_mass").slice(0, 12).reverse();
    renderDashboardPlot(
      "baseMapKpiTopSitesPlot",
      [{ type: "bar", orientation: "h", x: topSites.map(row => num(row, "value")), y: topSites.map(row => row.label), marker: { color: "#2b8cbe" } }],
      dashboardLayout("Top sites par masse allouee", "kg")
    );

    renderDashboardPlot(
      "baseMapKpiWeatherWaterPlot",
      [
        { type: "bar", name: "Precipitations", x: weatherX, y: weather.map(row => num(row, "avg_precip_mm")), marker: { color: "#9ecae1" }, opacity: 0.58 },
        { type: "scatter", mode: "lines", name: "Humidite", x: weatherX, y: weather.map(row => num(row, "avg_humidity_pct")), line: { color: "#3182bd" } },
        { type: "scatter", mode: "lines", name: "Secheresse", x: weatherX, y: weather.map(row => num(row, "avg_drought") * 100), line: { color: "#e6550d" } },
        { type: "scatter", mode: "lines", name: "Ouragan", x: weatherX, y: weather.map(row => num(row, "avg_hurricane") * 100), line: { color: "#756bb1" } }
      ],
      dashboardLayout("Eau, humidite et risques meteo", "mm / % / indice x100")
    );

    const weatherRegion = dashboardRows("weather_region");
    renderDashboardPlot(
      "baseMapKpiRegionalWeatherPlot",
      [
        { type: "bar", name: "Temp.", x: weatherRegion.map(row => row.label), y: weatherRegion.map(row => num(row, "avg_temp_c")), marker: { color: "#fb6a4a" } },
        { type: "bar", name: "Vent", x: weatherRegion.map(row => row.label), y: weatherRegion.map(row => num(row, "avg_wind_ms")), marker: { color: "#6baed6" } },
        { type: "bar", name: "Ouragan", x: weatherRegion.map(row => row.label), y: weatherRegion.map(row => num(row, "avg_hurricane") * 100), marker: { color: "#756bb1" } }
      ],
      { ...dashboardLayout("Profils meteo par region", "valeur moyenne"), barmode: "group" }
    );

    const opsEventKeys = ["capacite_appoint", "capacite_meteo_degradee", "maintenance_corrective_meteo", "congestion_logistique_inbound", "congestion_logistique_outbound", "froid_transport_ou_site", "recalage_qualite", "retard_approvisionnement"];
    renderDashboardPlot(
      "baseMapKpiOpsEventsPlot",
      opsEventKeys.map((key, idx) => ({
        type: "scatter",
        mode: "lines",
        name: key.replaceAll("_", " "),
        x: opsX,
        y: ops.map(row => num(row, key)),
        line: { width: 1.5, color: ["#2ca02c", "#31a354", "#74c476", "#fd8d3c", "#e6550d", "#6baed6", "#636363", "#756bb1"][idx] || "#555" }
      })),
      dashboardLayout("Evenements operationnels par mois", "occurrences")
    );

    horizontalBarPlot("baseMapKpiOpsLineagePlot", dashboardRows("node_ops_lineage"), "Lineage operations -> drivers", "value", "label", "#756bb1");
    horizontalBarPlot("baseMapKpiMaritimeRegionPlot", dashboardRows("maritime_region"), "Risque maritime par bassin", "risk_index", "label", "#3182bd");
    horizontalBarPlot("baseMapKpiEventExposurePlot", dashboardRows("event_exposure"), "Top expositions meteo par site", "value", "label", "#e6550d");

    const sddMonthly = dashboardRows("sdd_monthly");
    const sddMonths = sddMonthly.map(row => num(row, "month_index"));
    renderDashboardPlot(
      "baseMapKpiSddServicePlot",
      [
        { type: "scatter", mode: "lines", name: "Service OEM", x: sddMonths, y: sddMonthly.map(row => num(row, "avg_oem_service_level") * 100), line: { color: "#2ca02c" } },
        { type: "scatter", mode: "lines", name: "TD-DLCA", x: sddMonths, y: sddMonthly.map(row => num(row, "td_dlca_kgCO2e")), yaxis: "y2", line: { color: "#3182bd" } },
        { type: "scatter", mode: "lines", name: "SDD", x: sddMonths, y: sddMonthly.map(row => num(row, "sdd_kgCO2e")), yaxis: "y2", line: { color: "#f16913" } }
      ],
      {
        ...dashboardLayout("Service supply et impact mensuel SDD", "% / kgCO2e"),
        yaxis2: { title: "kgCO2e", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      }
    );

    const tierRows = dashboardRows("sdd_tier_month");
    const roles = Array.from(new Set(tierRows.map(row => row.role))).filter(Boolean);
    renderDashboardPlot(
      "baseMapKpiSddTierPlot",
      roles.map(role => {
        const rows = tierRows.filter(row => row.role === role);
        return { type: "scatter", mode: "lines", name: role, x: rows.map(row => num(row, "month_index")), y: rows.map(row => num(row, "avg_disruption_index") * 100) };
      }),
      dashboardLayout("Perturbation SDD par rang fournisseur", "indice x100")
    );

    renderDashboardPlot(
      "baseMapKpiSddImpactPlot",
      [
        { type: "bar", name: "Surimpact", x: sddMonths, y: sddMonthly.map(row => num(row, "surimpact_total")), marker: { color: "#f16913" }, opacity: 0.62 },
        { type: "scatter", mode: "lines", name: "Surimpact cumule", x: cumulativeX, y: cumulative.map(row => num(row, "surimpact_cumulative")), yaxis: "y2", line: { color: "#d62728" } }
      ],
      {
        ...dashboardLayout("Surimpact ACV dynamique mensuel et cumule", "kgCO2e/mois"),
        yaxis2: { title: "kgCO2e cumule", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      }
    );

    const sddBw = dashboardPayload.sdd_brightway || {};
    const sddBwRows = key => Array.isArray(sddBw[key]) ? sddBw[key] : [];
    const sddBwMonthly = sddBwRows("monthly");
    const sddBwCumulative = sddBwRows("cumulative");
    const sddBwMonths = sddBwMonthly.map(row => num(row, "month_index"));
    const sddBwCumulativeMonths = sddBwCumulative.map(row => num(row, "month_index"));
    renderDashboardPlot(
      "baseMapKpiSddAcvStaticDynamicPlot",
      [
        { type: "scatter", mode: "lines", name: "Production statique BW", x: sddBwMonths, y: sddBwMonthly.map(row => num(row, "production_static_kgco2e")), line: { color: "#3182bd" } },
        { type: "scatter", mode: "lines", name: "Production dynamique SDD-BW", x: sddBwMonths, y: sddBwMonthly.map(row => num(row, "production_dynamic_kgco2e")), line: { color: "#d62728" } },
        { type: "bar", name: "Delta inventaire", x: sddBwMonths, y: sddBwMonthly.map(row => num(row, "sdd_inventory_delta_kgco2e")), yaxis: "y2", marker: { color: "#fd8d3c" }, opacity: 0.45 }
      ],
      {
        ...dashboardLayout("Couplage SDD -> ACV Brightway", "kgCO2e/mois"),
        yaxis2: { title: "delta kgCO2e", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      }
    );
    horizontalBarPlot("baseMapKpiSddAcvDeltaMechanismPlot", sddBwRows("mechanism_totals"), "Delta ACV dynamique par mecanisme SDD", "delta_kgco2e", "label", "#e6550d", "kgCO2e");
    horizontalBarPlot("baseMapKpiSddAcvExchangeCategoryPlot", sddBwRows("exchange_category_totals"), "Ledger SDD -> Brightway par categorie d'echange", "delta_kgco2e", "label", "#31a354", "kgCO2e alloues");
    horizontalBarPlot("baseMapKpiSddAcvTopExchangesPlot", sddBwRows("top_exchanges"), "Principaux postes ACV expliquant le surimpact", "delta_kgco2e", "label", "#756bb1", "kgCO2e alloues");
    const exchangeLciaMonthly = sddBwRows("exchange_lcia_monthly");
    const exchangeLciaMonths = exchangeLciaMonthly.map(row => num(row, "month_index"));
    renderDashboardPlot(
      "baseMapKpiSddAcvExchangeExactMonthlyPlot",
      [
        { type: "scatter", mode: "lines", name: "Delta alloue", x: exchangeLciaMonths, y: exchangeLciaMonthly.map(row => num(row, "allocated_delta_kgco2e")), line: { color: "#e6550d" } },
        { type: "scatter", mode: "lines", name: "Delta exact BW", x: exchangeLciaMonths, y: exchangeLciaMonthly.map(row => num(row, "exact_delta_kgco2e")), line: { color: "#3182bd" } },
        { type: "bar", name: "Ecart exact-alloue", x: exchangeLciaMonths, y: exchangeLciaMonthly.map(row => num(row, "exact_minus_allocated_kgco2e")), marker: { color: "#bdbdbd" }, opacity: 0.45 }
      ],
      dashboardLayout("Test marginal Brightway: exact vs alloue", "kgCO2e/mois")
    );
    horizontalBarPlot("baseMapKpiSddAcvExchangeExactTopPlot", sddBwRows("exchange_lcia_top"), "Principaux echanges recalcules dans Brightway", "exact_delta_kgco2e", "label", "#3182bd", "kgCO2e recalcules");
    renderDashboardPlot(
      "baseMapKpiSddAcvSeatEquivalentPlot",
      [
        { type: "scatter", mode: "lines", name: "Delta par siege eq.", x: sddBwMonths, y: sddBwMonthly.map(row => num(row, "sdd_inventory_delta_kgco2e") / Math.max(num(row, "seat_equivalent_volume", 1), 0.000001)), line: { color: "#756bb1" } },
        { type: "scatter", mode: "lines", name: "Delta cumule / prod statique", x: sddBwCumulativeMonths, y: sddBwCumulative.map(row => num(row, "delta_vs_static_pct")), yaxis: "y2", line: { color: "#2ca25f" } }
      ],
      {
        ...dashboardLayout("Intensite ACV dynamique par equivalent siege", "kgCO2e/seat eq."),
        yaxis2: { title: "% cumule", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      }
    );
    horizontalBarPlot("baseMapKpiSddAcvTopComponentsPlot", sddBwRows("top_components"), "Top composants par surimpact ACV", "value", "label", "#fb6a4a", "kgCO2e");
    horizontalBarPlot("baseMapKpiSddAcvTopSitesPlot", sddBwRows("top_sites"), "Top sites par surimpact ACV", "value", "label", "#6baed6", "kgCO2e");

    const pathScatter = dashboardRows("path_scatter");
    renderDashboardPlot(
      "baseMapKpiPathScatterPlot",
      [{
        type: "scatter",
        mode: "markers",
        name: "Chemins",
        x: pathScatter.map(row => num(row, "x")),
        y: pathScatter.map(row => num(row, "y")),
        text: pathScatter.map(row => `${escapeHtml(row.system || "")}<br>${escapeHtml(row.component || "")}<br>${fmt(row.kg_km, 0)} kg.km`),
        hoverinfo: "text",
        marker: { size: pathScatter.map(row => 6 + Math.min(18, Math.sqrt(num(row, "kg_km")) / 45)), color: pathScatter.map(row => num(row, "kg_km")), colorscale: "Viridis", showscale: true }
      }],
      dashboardLayout("Masse vs distance des chemins", "kg")
    );

    horizontalBarPlot("baseMapKpiFamilyMassPlot", dashboardRows("family_mass"), "Masse allouee par famille", "value", "label", "#31a354");
    horizontalBarPlot("baseMapKpiModeMixPlot", dashboardRows("mode_kg_km"), "Mix transport kg.km", "value", "label", "#6baed6");

    const bw = dashboardPayload.brightway_model || {};
    const bwRows = key => Array.isArray(bw[key]) ? bw[key] : [];
    const indicatorUnits = bwRows("indicator_unit_views");
    const referencePe = bwRows("reference_person_equivalent_results");
    const referenceWeighted = bwRows("reference_weighted_results");
    const peIndicators = topBy(indicatorUnits.filter(row => truthy(row.include_in_person_equivalent)), "person_equivalent_value", 16);
    horizontalBarPlot("baseMapKpiBwClimatePlot", bwRows("top_climate_components"), "Brightway: top composants climat", "climate_kgco2e", "component", "#fb6a4a", "kgCO2e");
    horizontalBarPlot("baseMapKpiBwIndicatorPlot", referencePe.length ? topBy(referencePe, "impact_total_person_equivalent", 16) : peIndicators, "Reference STELIA: indicateurs en personne eq.", referencePe.length ? "impact_total_person_equivalent" : "person_equivalent_value", referencePe.length ? "short_label" : "pe_plot_label", "#9e9ac8", "personne eq.");
    horizontalBarPlot("baseMapKpiBwRawIndicatorPlot", topBy(indicatorUnits, "raw_sum_value", 16), "Brightway: indicateurs bruts par unite", "raw_sum_value", "raw_plot_label", "#bcbddc", "valeur brute");
    horizontalBarPlot("baseMapKpiBwUnitCoveragePlot", groupedCountRows(indicatorUnits, "normalization_status"), "Brightway: statut des facteurs d'unite", "value", "label", "#969696", "indicateurs");
    horizontalBarPlot("baseMapKpiBwReferenceWeightedPlot", topBy(referenceWeighted, "impact_total_weighted_score", 16), "Reference STELIA: normalise et pondere", "impact_total_weighted_score", "short_label", "#756bb1", "personne eq. ponderee");
    horizontalBarPlot("baseMapKpiBwReferencePhasePlot", phaseRows(bwRows("reference_phase_breakdown"), "person_equivalent", "Climate change"), "Reference STELIA: climat hors usage par phase", "value", "phase", "#31a354", "personne eq.");
    horizontalBarPlot("baseMapKpiBwReferenceScenarioPlot", bwRows("reference_scenarios").filter(row => row.phase === "Total"), "Reference STELIA: reductions scenarios climat", "lifecycle_climate_reduction_pct", "scenario_label", "#2b8cbe", "% cycle de vie");
    horizontalBarPlot("baseMapKpiBwReferenceClimateContributorPlot", bwRows("reference_climate_contributors"), "Reference STELIA: contributeurs climat hors usage", "climate_person_equivalent", "label", "#fd8d3c", "personne eq.");
    horizontalBarPlot("baseMapKpiBwAlignmentPlot", groupedCountRows(bwRows("supply_alignment"), "match_level"), "Brightway: qualite du matching supply", "value", "label", "#74c476", "lignes");
    horizontalBarPlot("baseMapKpiBwParamPlot", bwRows("top_parameter_amounts"), "Brightway: principaux parametres", "amount", "name", "#fdae6b", "valeur du parametre");
    horizontalBarPlot("baseMapKpiBwLeverPlot", bwRows("parametric_levers"), "Parametrisation: effet quantite d'un +10%", "abs_delta_amount_sum", "lever_label", "#41ab5d", "delta absolu, unites mixtes");
    horizontalBarPlot("baseMapKpiBwSensitivityPlot", bwRows("parametric_sensitivity").slice(0, 14), "Parametrisation: echanges les plus sensibles", "abs_delta_amount", "label", "#fd8d3c", "delta absolu, unite de l'echange");
    horizontalBarPlot("baseMapKpiBwSwitchPlot", bwRows("parametric_switches"), "Parametrisation: switchs Brightway scriptes", "affected_exchange_count", "label", "#6baed6", "echanges affectes");
    horizontalBarPlot("baseMapKpiBwRegionalScenarioPlot", bwRows("parametric_regional_scenarios"), "Parametrisation: sourcing FR / EU / mondialise", "foreground_amount_index", "label", "#2ca25f", "indice transport foreground");
    horizontalBarPlot("baseMapKpiBwUsageBreakdownPlot", bwRows("usage_calibration").filter(row => row.system === "Consommation passive" || row.system === "Entretien"), "Usage STELIA: decomposition climat", "excel_kgco2e", "component", "#fb6a4a", "kgCO2e");
    horizontalBarPlot("baseMapKpiBwExactScenarioPlot", bwRows("exact_scenario_lcia").filter(row => row.root_activity_id === "production" && row.scenario_id !== "current_export"), "Brightway exact: delta production par scenario", "delta_kgco2e", "label", "#de2d26", "kgCO2e vs baseline");
    horizontalBarPlot("baseMapKpiBwAlignedLifecyclePlot", bwRows("exact_scenario_lcia").filter(row => row.root_activity_id === "lifecycle_excel_aligned" && row.scenario_id !== "current_export"), "Brightway exact: cycle complet corrige usage", "delta_kgco2e", "label", "#3182bd", "kgCO2e vs baseline corrige");
    horizontalBarPlot("baseMapKpiBwExcelRuntimePlot", bwRows("excel_runtime_comparison"), "Excel reference vs Brightway runtime", "relative_delta_pct", "label", "#756bb1", "% ecart");

    const status = document.getElementById("sddDashboardStatus");
    if (status) {
      const bwStatus = bw.runtime?.can_execute_brightway ? "Brightway executable" : "Brightway non executable localement, exports Excel utilises";
      status.textContent = `${fmt(weather.length, 0)} mois meteo, ${fmt(ops.length, 0)} mois operations, ${fmt(sites.length, 0)} sites - ${bwStatus}`;
    }
    setTimeout(() => {
      [
        "baseMapKpiMethodPlot",
        "baseMapKpiCumulativePlot",
        "baseMapKpiWeatherPlot",
        "baseMapKpiWeatherWaterPlot",
        "baseMapKpiRegionalWeatherPlot",
        "baseMapKpiOpsPlot",
        "baseMapKpiOpsEventsPlot",
        "baseMapKpiOpsLineagePlot",
        "baseMapKpiMaritimePlot",
        "baseMapKpiMaritimeRegionPlot",
        "baseMapKpiEventPlot",
        "baseMapKpiEventExposurePlot",
        "baseMapKpiSddServicePlot",
        "baseMapKpiSddTierPlot",
        "baseMapKpiSddImpactPlot",
        "baseMapKpiSddAcvStaticDynamicPlot",
        "baseMapKpiSddAcvDeltaMechanismPlot",
        "baseMapKpiSddAcvExchangeCategoryPlot",
        "baseMapKpiSddAcvTopExchangesPlot",
        "baseMapKpiSddAcvExchangeExactMonthlyPlot",
        "baseMapKpiSddAcvExchangeExactTopPlot",
        "baseMapKpiSddAcvSeatEquivalentPlot",
        "baseMapKpiSddAcvTopComponentsPlot",
        "baseMapKpiSddAcvTopSitesPlot",
        "baseMapKpiPathScatterPlot",
        "baseMapKpiTopSitesPlot",
        "baseMapKpiFamilyMassPlot",
        "baseMapKpiModeMixPlot",
        "baseMapKpiBwClimatePlot",
        "baseMapKpiBwIndicatorPlot",
        "baseMapKpiBwRawIndicatorPlot",
        "baseMapKpiBwUnitCoveragePlot",
        "baseMapKpiBwReferenceWeightedPlot",
        "baseMapKpiBwReferencePhasePlot",
        "baseMapKpiBwReferenceScenarioPlot",
        "baseMapKpiBwReferenceClimateContributorPlot",
        "baseMapKpiBwAlignmentPlot",
        "baseMapKpiBwParamPlot",
        "baseMapKpiBwLeverPlot",
        "baseMapKpiBwSensitivityPlot",
        "baseMapKpiBwSwitchPlot",
        "baseMapKpiBwRegionalScenarioPlot",
        "baseMapKpiBwUsageBreakdownPlot",
        "baseMapKpiBwExactScenarioPlot",
        "baseMapKpiBwAlignedLifecyclePlot",
        "baseMapKpiBwExcelRuntimePlot"
      ].forEach(id => {
        const node = document.getElementById(id);
        if (node) Plotly.Plots.resize(node);
      });
    }, 80);
  }

  function captureSourceState() {
    const chart = document.getElementById("chart");
    if (chart && chart.data && chart.layout) {
      sourceState = {
        data: clone(chart.data),
        layout: clone(chart.layout)
      };
    }
  }

  function colorScaleForMetric(metric) {
    if (metric === "avg_service_level") {
      return {
        label: "Perte de service",
        cmin: 0,
        cmax: 1,
        colorscale: [[0, "#2ca02c"], [0.5, "#ffbf00"], [1, "#d62728"]]
      };
    }
    if (metric === "weather_exposure_index") {
      return {
        label: "Exposition meteo",
        cmin: 0,
        cmax: 1,
        colorscale: [[0, "#2ca02c"], [0.5, "#ffbf00"], [1, "#d62728"]]
      };
    }
    if (metric === "sdd_acv_delta_kgco2e") {
      return {
        label: "Surimpact ACV net",
        cmin: 0,
        cmax: 1,
        colorscale: [[0, "#f7fbff"], [0.45, "#6baed6"], [1, "#d62728"]]
      };
    }
    if (metric === "allocated_mass_kg") {
      return {
        label: "Masse allouee",
        cmin: 0,
        cmax: 1,
        colorscale: [[0, "#f7fcf5"], [0.45, "#74c476"], [1, "#006d2c"]]
      };
    }
    return {
      label: metric === "peak_backlog_kg" ? "Retard max" : (metric === "affected_month_count" ? "Mois affectes" : "Perturbation"),
      cmin: 0,
      cmax: 1,
      colorscale: [[0, "#2ca02c"], [0.5, "#ffbf00"], [1, "#d62728"]]
    };
  }

  function colorValue(site, metric) {
    const raw = Number(site[metric] || 0);
    if (metric === "avg_service_level") return Math.max(0, Math.min(1, 1 - raw));
    if (metric === "peak_backlog_kg") {
      const maxBacklog = maxOf(sites, "peak_backlog_kg") || 1;
      return Math.max(0, Math.min(1, Math.log1p(raw) / Math.log1p(maxBacklog)));
    }
    if (metric === "affected_month_count") {
      const maxMonths = maxOf(sites, "affected_month_count") || 1;
      return Math.max(0, Math.min(1, raw / maxMonths));
    }
    if (metric === "weather_exposure_index") {
      const maxExposure = maxOf(sites, "weather_exposure_index") || 1;
      return Math.max(0, Math.min(1, raw / maxExposure));
    }
    if (metric === "sdd_acv_delta_kgco2e") {
      const maxDelta = maxOf(sites, "sdd_acv_delta_kgco2e") || 1;
      return Math.max(0, Math.min(1, Math.log1p(raw) / Math.log1p(maxDelta)));
    }
    if (metric === "allocated_mass_kg") {
      const maxMass = maxOf(sites, "allocated_mass_kg") || 1;
      return Math.max(0, Math.min(1, Math.log1p(raw) / Math.log1p(maxMass)));
    }
    return Math.max(0, Math.min(1, raw));
  }

  function siteHover(site) {
    return [
      `<b>${escapeHtml(site.name || "Site")}</b>`,
      `Roles: ${escapeHtml(site.roles || "")}`,
      `Pays: ${escapeHtml(site.country_code || "")}`,
      `Masse allouee: ${fmt(site.allocated_mass_kg)} kg`,
      `Service moyen: ${fmt(Number(site.avg_service_level || 0) * 100, 1)}%`,
      `Service min: ${fmt(Number(site.min_service_level || 0) * 100, 1)}%`,
      `Perturbation moyenne: ${fmt(site.avg_disruption_index, 3)}`,
      `Retard max: ${fmt(site.peak_backlog_kg)} kg`,
      `Mois affectes: ${fmt(site.affected_month_count, 0)}`,
      `Exposition meteo: ${fmt(site.weather_exposure_index, 2)} (${fmt(site.weather_event_count, 0)} evenements)`,
      `Surimpact ACV net: ${fmt(site.sdd_acv_delta_kgco2e, 1)} kgCO2e`,
      `Decisions: ${escapeHtml(site.top_decisions || "none")}`,
      `Facteurs: ${escapeHtml(site.top_drivers || "none")}`
    ].join("<br>");
  }

  function laneHover(lane) {
    return [
      `<b>${escapeHtml(lane.from_name || "")} -> ${escapeHtml(lane.to_name || "")}</b>`,
      `Lien: ${escapeHtml(lane.edge || "")}`,
      `Modes: ${escapeHtml(lane.modes || "")}`,
      `Bassin: ${escapeHtml(lane.route_region || "")}`,
      `kg.km alloues: ${fmt(lane.allocated_kg_km)}`,
      `Risque moyen: ${fmt(lane.avg_transport_risk, 3)}`,
      `Risque max: ${fmt(lane.max_transport_risk, 3)}`,
      `Delai moyen: x${fmt(lane.avg_delay_multiplier, 2)}`,
      `Capacite moyenne: ${fmt(Number(lane.avg_capacity_multiplier || 0) * 100, 1)}%`,
      `Decisions: ${escapeHtml(lane.top_decisions || "none")}`
    ].join("<br>");
  }

  function siteTrace(metric, name = "Sites SDD") {
    const maxMass = maxOf(sites, "allocated_mass_kg") || 1;
    const scale = colorScaleForMetric(metric);
    return {
      type: "scattergeo",
      mode: "markers",
      name,
      lon: sites.map(site => site.lon),
      lat: sites.map(site => site.lat),
      text: sites.map(siteHover),
      hoverinfo: "text",
      hoverlabel: { align: "left", namelength: -1 },
      marker: {
        size: sites.map(site => 7 + 24 * Math.sqrt(Math.max(0, Number(site.allocated_mass_kg || 0)) / maxMass)),
        color: sites.map(site => colorValue(site, metric)),
        cmin: scale.cmin,
        cmax: scale.cmax,
        colorscale: scale.colorscale,
        colorbar: { title: scale.label },
        line: { width: 0.8, color: "#172033" },
        opacity: 0.86
      }
    };
  }

  function laneBucketTrace(name, minRisk, maxRisk, color, width) {
    const lon = [];
    const lat = [];
    const text = [];
    lanes.forEach(lane => {
      const risk = Number(lane.avg_transport_risk || 0);
      if (risk < minRisk || risk >= maxRisk) return;
      const hover = laneHover(lane);
      lon.push(lane.from_lon, lane.to_lon, null);
      lat.push(lane.from_lat, lane.to_lat, null);
      text.push(hover, hover, null);
    });
    return {
      type: "scattergeo",
      mode: "lines",
      name,
      lon,
      lat,
      text,
      hoverinfo: "text",
      line: { color, width },
      opacity: 0.62
    };
  }

  function laneKgKmBucketTrace(name, minRatio, maxRatio, color, width) {
    const maxKgKm = maxOf(lanes, "allocated_kg_km") || 1;
    const lon = [];
    const lat = [];
    const text = [];
    lanes.forEach(lane => {
      const ratio = Number(lane.allocated_kg_km || 0) / maxKgKm;
      if (ratio < minRatio || ratio >= maxRatio) return;
      const hover = laneHover(lane);
      lon.push(lane.from_lon, lane.to_lon, null);
      lat.push(lane.from_lat, lane.to_lat, null);
      text.push(hover, hover, null);
    });
    return {
      type: "scattergeo",
      mode: "lines",
      name,
      lon,
      lat,
      text,
      hoverinfo: "text",
      line: { color, width },
      opacity: 0.58
    };
  }

  function geoLayout(title) {
    return {
      title: { text: title, x: 0.02, y: 0.98, font: { size: 16 } },
      geo: {
        scope: "world",
        projection: { type: "natural earth" },
        showland: true,
        landcolor: "#f0f0f0",
        subunitwidth: 1,
        countrywidth: 1,
        subunitcolor: "#dcdcdc",
        countrycolor: "#dcdcdc",
        showcountries: true,
        showocean: true,
        oceancolor: "#f8fafc"
      },
      margin: { l: 0, r: 0, t: 48, b: 0 },
      legend: { orientation: "h", x: 0.01, y: 1.03 },
      paper_bgcolor: "#ffffff",
      plot_bgcolor: "#ffffff"
    };
  }

  function metricSelectValue() {
    const select = document.getElementById("sddMetricSelect");
    return select ? select.value : "avg_disruption_index";
  }

  function renderPlot(traces, title) {
    showChartCanvas();
    Plotly.react("chart", traces, geoLayout(title), { displayModeBar: true, responsive: true });
  }

  function renderSource() {
    currentView = "source";
    setActiveButton();
    showChartCanvas();
    if (typeof window.draw === "function") {
      window.draw();
      setTimeout(captureSourceState, 50);
      return;
    }
    if (sourceState) {
      Plotly.react("chart", clone(sourceState.data), clone(sourceState.layout), { displayModeBar: true, responsive: true });
    }
  }

  function renderSddSites() {
    currentView = "sites";
    setActiveButton();
    renderPlot(
      [siteTrace(metricSelectValue(), "Sites fournisseurs")],
      "Resultats SDD par site fournisseur"
    );
  }

  function renderSddLanes() {
    currentView = "lanes";
    setActiveButton();
    renderPlot(
      [
        laneBucketTrace("Risque transport faible", 0, 0.20, "#2ca02c", 1.6),
        laneBucketTrace("Risque transport moyen", 0.20, 0.35, "#ff9f1c", 2.2),
        laneBucketTrace("Risque transport eleve", 0.35, 2, "#d62728", 3.0),
        siteTrace("avg_disruption_index", "Sites")
      ],
      "Resultats SDD sur les liaisons supply"
    );
  }

  function renderWeatherMap() {
    currentView = "weather";
    setActiveButton();
    renderPlot(
      [
        laneBucketTrace("Routes meteo critiques", 0.35, 2, "#d62728", 2.6),
        laneBucketTrace("Routes sous surveillance", 0.20, 0.35, "#ff9f1c", 1.8),
        siteTrace("weather_exposure_index", "Exposition meteo sites")
      ],
      "Exposition meteo regionalisee des sites et routes"
    );
  }

  function renderOperationsMap() {
    currentView = "operations";
    setActiveButton();
    renderPlot(
      [
        laneBucketTrace("Transport degrade", 0.35, 2, "#d62728", 2.8),
        siteTrace("affected_month_count", "Sites affectes par operations")
      ],
      "Operations: mois affectes, backlog et drivers"
    );
  }

  function renderSddImpact() {
    currentView = "impact";
    setActiveButton();
    renderPlot(
      [
        laneBucketTrace("Liaisons a risque eleve", 0.35, 2, "#d62728", 3.2),
        siteTrace("peak_backlog_kg", "Retards et rupture de service")
      ],
      "Impacts operationnels SDD localises"
    );
  }

  function renderAcvMap() {
    currentView = "acv";
    setActiveButton();
    renderPlot(
      [
        laneBucketTrace("Routes a risque eleve", 0.35, 2, "#756bb1", 2.5),
        siteTrace("sdd_acv_delta_kgco2e", "Delta ACV dynamique")
      ],
      "ACV dynamique: surimpact net par site"
    );
  }

  function renderCriticalityMap() {
    currentView = "criticality";
    setActiveButton();
    renderPlot(
      [
        laneKgKmBucketTrace("Flux critique", 0.55, 2, "#d62728", 3.2),
        laneKgKmBucketTrace("Flux intermediaire", 0.25, 0.55, "#f16913", 2.2),
        laneKgKmBucketTrace("Flux diffus", 0, 0.25, "#9ecae1", 1.2),
        siteTrace("allocated_mass_kg", "Masse allouee sites")
      ],
      "Criticite reseau: masse allouee et kg.km"
    );
  }

  function renderCurrentSddView() {
    if (currentView === "sites") renderSddSites();
    if (currentView === "lanes") renderSddLanes();
    if (currentView === "weather") renderWeatherMap();
    if (currentView === "operations") renderOperationsMap();
    if (currentView === "impact") renderSddImpact();
    if (currentView === "acv") renderAcvMap();
    if (currentView === "criticality") renderCriticalityMap();
    if (currentView === "dashboard") renderDashboard();
  }

  function setActiveButton() {
    document.querySelectorAll("[data-sdd-view]").forEach(button => {
      button.classList.toggle("active", button.dataset.sddView === currentView);
    });
  }

  function installControls() {
    const toolbar = document.querySelector(".toolbar");
    if (!toolbar || document.getElementById("sddViewTabs")) return;

    const tabs = document.createElement("span");
    tabs.id = "sddViewTabs";
    tabs.className = "sdd-view-tabs";
    tabs.innerHTML = `
      <label>Vues</label>
      <button type="button" data-sdd-view="source">Carte source</button>
      <button type="button" data-sdd-view="sites">SDD sites</button>
      <button type="button" data-sdd-view="lanes">SDD liaisons</button>
      <button type="button" data-sdd-view="weather">Meteo</button>
      <button type="button" data-sdd-view="operations">Operations</button>
      <button type="button" data-sdd-view="impact">SDD impact</button>
      <button type="button" data-sdd-view="acv">ACV dynamique</button>
      <button type="button" data-sdd-view="criticality">Criticite reseau</button>
      <button type="button" data-sdd-view="dashboard">Dashboard KPI</button>
      <select id="sddMetricSelect" title="Metrique de couleur des sites">
        <option value="avg_disruption_index">Perturbation</option>
        <option value="avg_service_level">Service</option>
        <option value="peak_backlog_kg">Retard max</option>
        <option value="affected_month_count">Mois affectes</option>
      </select>
    `;
    toolbar.appendChild(tabs);

    const note = document.createElement("span");
    note.className = "sdd-map-note";
    note.textContent = `${payload.stats?.site_count || 0} sites SDD, ${payload.stats?.lane_count || 0} liaisons, service moyen ${fmt(Number(payload.stats?.avg_service || 0) * 100, 1)}%`;
    toolbar.appendChild(note);

    tabs.querySelector('[data-sdd-view="source"]').addEventListener("click", renderSource);
    tabs.querySelector('[data-sdd-view="sites"]').addEventListener("click", renderSddSites);
    tabs.querySelector('[data-sdd-view="lanes"]').addEventListener("click", renderSddLanes);
    tabs.querySelector('[data-sdd-view="weather"]').addEventListener("click", renderWeatherMap);
    tabs.querySelector('[data-sdd-view="operations"]').addEventListener("click", renderOperationsMap);
    tabs.querySelector('[data-sdd-view="impact"]').addEventListener("click", renderSddImpact);
    tabs.querySelector('[data-sdd-view="acv"]').addEventListener("click", renderAcvMap);
    tabs.querySelector('[data-sdd-view="criticality"]').addEventListener("click", renderCriticalityMap);
    tabs.querySelector('[data-sdd-view="dashboard"]').addEventListener("click", renderDashboard);
    tabs.querySelector("#sddMetricSelect").addEventListener("change", () => {
      if (currentView !== "source" && currentView !== "dashboard") renderCurrentSddView();
    });
    setActiveButton();
  }

  window.renderSddSites = renderSddSites;
  window.renderSddLanes = renderSddLanes;
  window.renderWeatherMap = renderWeatherMap;
  window.renderOperationsMap = renderOperationsMap;
  window.renderSddImpact = renderSddImpact;
  window.renderAcvMap = renderAcvMap;
  window.renderCriticalityMap = renderCriticalityMap;
  window.renderBaseDashboard = renderDashboard;

  window.addEventListener("load", () => {
    setTimeout(() => {
      captureSourceState();
      installControls();
    }, 80);
  });
})();
</script>
"""
    injection = injection.replace("__SDD_MAP_PAYLOAD__", payload_json)
    injection = injection.replace("__BASE_DASHBOARD_PAYLOAD__", dashboard_json)
    insert_at = source_html.lower().rfind("</body>")
    if insert_at >= 0:
        document = source_html[:insert_at] + injection + source_html[insert_at:]
    else:
        document = source_html + injection
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(document, encoding="utf-8")


def write_results_dashboard_html(path: Path, payload: dict[str, Any]) -> None:
    payload_json = json_for_script(payload)
    title = html.escape("POC2026 supply_geo results")
    document = """<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>__TITLE__</title>
  <script src="https://cdn.plot.ly/plotly-2.32.0.min.js"></script>
  <style>
    :root {
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #172033;
      --muted: #5e6a7d;
      --line: #d8dee8;
      --blue: #1f77b4;
      --green: #2ca02c;
      --orange: #ff7f0e;
      --red: #d62728;
      --purple: #7f3c8d;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
      background: var(--bg);
      color: var(--ink);
    }
    .topbar {
      height: 56px;
      display: flex;
      align-items: center;
      gap: 18px;
      padding: 0 18px;
      border-bottom: 1px solid var(--line);
      background: var(--panel);
      position: sticky;
      top: 0;
      z-index: 20;
    }
    .brand {
      font-size: 14px;
      font-weight: 800;
      white-space: nowrap;
    }
    .tabs {
      display: inline-flex;
      border: 1px solid var(--line);
      background: #f8fafc;
    }
    .tabBtn {
      border: 0;
      background: transparent;
      color: #344256;
      font-size: 13px;
      font-weight: 750;
      padding: 8px 12px;
      cursor: pointer;
    }
    .tabBtn.active {
      background: #172033;
      color: #ffffff;
    }
    .meta {
      margin-left: auto;
      font-size: 12px;
      color: var(--muted);
      white-space: nowrap;
    }
    .tabPane { display: none; }
    .tabPane.active { display: block; }
    #mapPane { height: calc(100vh - 56px); }
    #mapFrame {
      width: 100%;
      height: 100%;
      border: 0;
      display: block;
      background: #ffffff;
    }
    .dashboard {
      max-width: 1540px;
      margin: 0 auto;
      padding: 18px;
    }
    .kpiGrid {
      display: grid;
      grid-template-columns: repeat(4, minmax(150px, 1fr));
      gap: 10px;
      margin-bottom: 12px;
    }
    .kpiCard {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
      min-height: 82px;
    }
    .kpiLabel {
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
    }
    .kpiValue {
      margin-top: 8px;
      font-size: 24px;
      font-weight: 850;
      line-height: 1.1;
    }
    .kpiUnit {
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      margin-left: 4px;
    }
    .chartGrid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
    }
    .chartPanel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      min-height: 370px;
    }
    .chartPanel.wide {
      grid-column: 1 / -1;
    }
    .plot {
      width: 100%;
      height: 340px;
    }
    .wide .plot {
      height: 390px;
    }
    @media (max-width: 980px) {
      .topbar {
        height: auto;
        min-height: 56px;
        flex-wrap: wrap;
        padding: 10px 12px;
      }
      .meta {
        width: 100%;
        margin-left: 0;
      }
      #mapPane { height: calc(100vh - 98px); }
      .kpiGrid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .chartGrid { grid-template-columns: 1fr; }
      .chartPanel.wide { grid-column: auto; }
    }
  </style>
</head>
<body>
  <header class="topbar">
    <div class="brand">POC2026 supply_geo</div>
    <nav class="tabs" aria-label="Result views">
      <button class="tabBtn active" type="button" data-tab="map">Carte</button>
      <button class="tabBtn" type="button" data-tab="kpi">KPI generaux</button>
    </nav>
    <div class="meta" id="generatedAt"></div>
  </header>

  <main>
    <section id="mapPane" class="tabPane active" role="tabpanel">
      <iframe id="mapFrame" title="Supply map"></iframe>
    </section>
    <section id="kpiPane" class="tabPane" role="tabpanel">
      <div class="dashboard">
        <div id="kpiCards" class="kpiGrid"></div>
        <div class="chartGrid">
          <section class="chartPanel wide"><div id="weatherTempPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="weatherWaterPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="weatherWindPlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="regionalWeatherHeatmapPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="regionalEventMixPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="weatherProfilePlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="maritimeRiskPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="maritimeRegionPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="maritimeEventMixPlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="nodeOpsStatePlot" class="plot"></div></section>
          <section class="chartPanel"><div id="nodeOpsEventPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="nodeOpsRegionPlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="nodeOpsLineagePlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="sddServicePlot" class="plot"></div></section>
          <section class="chartPanel"><div id="sddTierPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="sddMethodPlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="sddImpactStackPlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="sddCumulativePlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="opsProxyPlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="haEventCalendarPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="haSystemPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="haEnergyPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="haBatteryBiomassPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="haWeatherDriverPlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="haImpactCumulativePlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="haEventImpactPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="readinessPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="lcaPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="familyPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="modePlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="routeMassPlot" class="plot"></div></section>
          <section class="chartPanel"><div id="topSitePlot" class="plot"></div></section>
          <section class="chartPanel"><div id="eventExposurePlot" class="plot"></div></section>
          <section class="chartPanel wide"><div id="eventMonthPlot" class="plot"></div></section>
        </div>
      </div>
    </section>
  </main>

  <script>
    const KPI_PAYLOAD = __KPI_PAYLOAD__;
    const PLOT_CONFIG = { responsive: true, displaylogo: false };
    const COLORS = ["#1f77b4", "#2ca02c", "#ff7f0e", "#d62728", "#7f3c8d", "#17becf", "#8c564b", "#bcbd22"];

    function fmt(value) {
      const n = Number(value);
      if (!Number.isFinite(n)) return String(value ?? "");
      if (Math.abs(n) >= 1000000) return (n / 1000000).toFixed(1) + "M";
      if (Math.abs(n) >= 10000) return Math.round(n).toLocaleString("fr-FR");
      if (Math.abs(n) >= 1000) return n.toLocaleString("fr-FR", { maximumFractionDigits: 0 });
      if (Math.abs(n) >= 100) return n.toLocaleString("fr-FR", { maximumFractionDigits: 1 });
      return n.toLocaleString("fr-FR", { maximumFractionDigits: 2 });
    }

    function plotLayout(title) {
      return {
        title: { text: title, font: { size: 13 } },
        margin: { l: 54, r: 18, t: 48, b: 52 },
        paper_bgcolor: "#ffffff",
        plot_bgcolor: "#ffffff",
        font: { family: "system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif", size: 11, color: "#172033" },
        xaxis: { gridcolor: "#eef2f7", zerolinecolor: "#d8dee8" },
        yaxis: { gridcolor: "#eef2f7", zerolinecolor: "#d8dee8" },
        legend: { orientation: "h", y: -0.2 }
      };
    }

    function renderCards() {
      const root = document.getElementById("kpiCards");
      root.innerHTML = (KPI_PAYLOAD.cards || []).map(card => `
        <div class="kpiCard">
          <div class="kpiLabel">${card.label || ""}</div>
          <div class="kpiValue">${fmt(card.value)}<span class="kpiUnit">${card.unit || ""}</span></div>
        </div>
      `).join("");
    }

    function barPlot(id, rows, title, options = {}) {
      const labels = rows.map(r => r.label);
      const values = rows.map(r => Number(r.value || 0));
      const trace = {
        type: "bar",
        x: options.horizontal ? values : labels,
        y: options.horizontal ? labels : values,
        orientation: options.horizontal ? "h" : "v",
        marker: { color: rows.map((_, i) => COLORS[i % COLORS.length]) },
        hovertemplate: options.horizontal ? "%{y}<br>%{x}<extra></extra>" : "%{x}<br>%{y}<extra></extra>"
      };
      const layout = plotLayout(title);
      if (options.horizontal) {
        layout.margin.l = 170;
        layout.yaxis = { automargin: true, autorange: "reversed" };
        layout.xaxis = { gridcolor: "#eef2f7", zerolinecolor: "#d8dee8" };
      }
      Plotly.react(id, [trace], layout, PLOT_CONFIG);
    }

    function numSeries(rows, key) {
      return rows.map(r => Number(r[key] || 0));
    }

    function scaledSeries(rows, key, factor) {
      return rows.map(r => Number(r[key] || 0) * factor);
    }

    function lineTrace(name, x, y, color, options = {}) {
      return {
        type: "scatter",
        mode: options.mode || "lines",
        name,
        x,
        y,
        yaxis: options.yaxis,
        line: { color, width: options.width || 2, dash: options.dash || "solid" },
        marker: { color, size: options.markerSize || 5 },
        hovertemplate: `${name}<br>mois=%{x}<br>%{y:.2f}<extra></extra>`
      };
    }

    function barTrace(name, x, y, color, options = {}) {
      return {
        type: "bar",
        name,
        x,
        y,
        yaxis: options.yaxis,
        marker: { color, opacity: options.opacity || 0.68 },
        hovertemplate: `${name}<br>mois=%{x}<br>%{y}<extra></extra>`
      };
    }

    function renderPlots() {
      const weather = KPI_PAYLOAD.weather_month || [];
      const weatherMonths = weather.map(r => r.month_index);
      Plotly.react("weatherTempPlot", [
        lineTrace("Temperature moyenne", weatherMonths, numSeries(weather, "avg_temp_c"), "#d62728"),
        lineTrace("Temperature p90", weatherMonths, numSeries(weather, "p90_temp_c"), "#ff7f0e", { dash: "dot" }),
        lineTrace("Rechauffement applique", weatherMonths, numSeries(weather, "avg_warming_delta_c"), "#111111", { dash: "dot", width: 3 }),
        lineTrace("Heat index moyen", weatherMonths, numSeries(weather, "avg_heat_index_c"), "#7f3c8d", { dash: "dash" }),
        lineTrace("Heat index p90", weatherMonths, numSeries(weather, "p90_heat_index_c"), "#8c564b", { dash: "dashdot" })
      ], {
        ...plotLayout("Driver meteo supply_geo: temperature et heat index"),
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "degC", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const waterLayout = {
        ...plotLayout("Driver meteo supply_geo: humidite et precipitation"),
        margin: { l: 54, r: 62, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Humidite (%)", gridcolor: "#eef2f7" },
        yaxis2: { title: "Precipitations moy. (mm/mois)", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("weatherWaterPlot", [
        lineTrace("Humidite moyenne", weatherMonths, numSeries(weather, "avg_humidity_pct"), "#1f77b4"),
        lineTrace("Humidite p10", weatherMonths, numSeries(weather, "p10_humidity_pct"), "#17becf", { dash: "dot" }),
        barTrace("Precipitations moy.", weatherMonths, numSeries(weather, "avg_precip_mm"), "#9ecae1", { yaxis: "y2", opacity: 0.58 })
      ], waterLayout, PLOT_CONFIG);

      const windLayout = {
        ...plotLayout("Driver meteo supply_geo: vent et intensites d'evenements"),
        margin: { l: 54, r: 62, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Vent (m/s)", gridcolor: "#eef2f7" },
        yaxis2: { title: "Intensite moyenne", overlaying: "y", side: "right", range: [0, 1], gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("weatherWindPlot", [
        lineTrace("Vent moyen", weatherMonths, numSeries(weather, "avg_wind_ms"), "#5e6a7d"),
        lineTrace("Vent max", weatherMonths, numSeries(weather, "max_wind_ms"), "#172033", { dash: "dot" }),
        lineTrace("Canicule", weatherMonths, numSeries(weather, "avg_heatwave"), "#d62728", { yaxis: "y2" }),
        lineTrace("Secheresse", weatherMonths, numSeries(weather, "avg_drought"), "#ff7f0e", { yaxis: "y2" }),
        lineTrace("Tempete", weatherMonths, numSeries(weather, "avg_storm_stress"), "#1f77b4", { yaxis: "y2" }),
        lineTrace("Ouragan / cyclone", weatherMonths, numSeries(weather, "avg_hurricane"), "#9467bd", { yaxis: "y2" }),
        lineTrace("Froid", weatherMonths, numSeries(weather, "avg_cold_stress"), "#17becf", { yaxis: "y2" })
      ], windLayout, PLOT_CONFIG);

      const regionMonth = KPI_PAYLOAD.weather_region_month || [];
      const regions = [...new Set(regionMonth.map(r => r.world_region))];
      const regionMonths = [...new Set(regionMonth.map(r => r.month_index))].sort((a, b) => Number(a) - Number(b));
      const regionRows = regions.map(region => {
        const byMonth = new Map(regionMonth.filter(r => r.world_region === region).map(r => [Number(r.month_index), Number(r.risk_index || 0)]));
        return regionMonths.map(month => byMonth.get(Number(month)) || 0);
      });
      Plotly.react("regionalWeatherHeatmapPlot", [{
        type: "heatmap",
        x: regionMonths,
        y: regions,
        z: regionRows,
        colorscale: [[0, "#f7fbff"], [0.35, "#9ecae1"], [0.7, "#fd8d3c"], [1, "#d62728"]],
        colorbar: { title: "Risque" },
        hovertemplate: "region=%{y}<br>mois=%{x}<br>risque=%{z:.3f}<extra></extra>"
      }], {
        ...plotLayout("Risque meteo regionalise par mois"),
        margin: { l: 150, r: 60, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "", automargin: true }
      }, PLOT_CONFIG);

      const weatherRegion = KPI_PAYLOAD.weather_region || [];
      const regionLabels = weatherRegion.map(r => r.label);
      Plotly.react("regionalEventMixPlot", [
        barTrace("Canicule", regionLabels, numSeries(weatherRegion, "avg_heatwave"), "#d62728"),
        barTrace("Secheresse", regionLabels, numSeries(weatherRegion, "avg_drought"), "#ff7f0e"),
        barTrace("Tempete", regionLabels, numSeries(weatherRegion, "avg_storm_stress"), "#1f77b4"),
        barTrace("Ouragan / cyclone", regionLabels, numSeries(weatherRegion, "avg_hurricane"), "#9467bd"),
        barTrace("Froid", regionLabels, numSeries(weatherRegion, "avg_cold_stress"), "#17becf")
      ], {
        ...plotLayout("Mix d'evenements moyen par region"),
        barmode: "stack",
        xaxis: { automargin: true },
        yaxis: { title: "Intensite moyenne", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const weatherProfile = KPI_PAYLOAD.weather_profile || [];
      Plotly.react("weatherProfilePlot", [{
        type: "bar",
        x: weatherProfile.map(r => Number(r.risk_index || 0)),
        y: weatherProfile.map(r => r.label),
        orientation: "h",
        marker: { color: weatherProfile.map(r => Number(r.avg_cold_stress || 0) > Number(r.avg_heatwave || 0) ? "#17becf" : "#d62728") },
        text: weatherProfile.map(r => `${r.site_count || 0} sites`),
        hovertemplate: "%{y}<br>risque=%{x:.3f}<br>%{text}<extra></extra>"
      }], {
        ...plotLayout("Profils meteo appliques aux sites"),
        margin: { l: 180, r: 18, t: 48, b: 52 },
        xaxis: { title: "Risque moyen", gridcolor: "#eef2f7" },
        yaxis: { automargin: true, autorange: "reversed" }
      }, PLOT_CONFIG);

      const maritime = KPI_PAYLOAD.maritime_month || [];
      const maritimeMonths = maritime.map(r => r.month_index);
      const maritimeLayout = {
        ...plotLayout("Risque meteo des lanes maritimes"),
        margin: { l: 54, r: 62, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Indice / multiplicateur", gridcolor: "#eef2f7", range: [0, 1.65] },
        yaxis2: { title: "kg.km exposes", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("maritimeRiskPlot", [
        barTrace("kg.km exposes ship", maritimeMonths, numSeries(maritime, "exposed_kg_km"), "#c6dbef", { yaxis: "y2", opacity: 0.42 }),
        lineTrace("Risque maritime", maritimeMonths, numSeries(maritime, "risk_index"), "#172033", { width: 3 }),
        lineTrace("Delai transport", maritimeMonths, numSeries(maritime, "delay_multiplier"), "#9467bd"),
        lineTrace("Capacite maritime", maritimeMonths, numSeries(maritime, "capacity_multiplier"), "#2ca02c")
      ], maritimeLayout, PLOT_CONFIG);

      const maritimeRegion = KPI_PAYLOAD.maritime_region || [];
      Plotly.react("maritimeRegionPlot", [{
        type: "bar",
        x: maritimeRegion.map(r => Number(r.risk_index || 0)),
        y: maritimeRegion.map(r => r.label),
        orientation: "h",
        marker: { color: "#1f77b4" },
        text: maritimeRegion.map(r => `${fmt(r.allocated_kg_km || 0)} kg.km`),
        hovertemplate: "%{y}<br>risque=%{x:.3f}<br>%{text}<extra></extra>"
      }], {
        ...plotLayout("Risque moyen par bassin maritime"),
        margin: { l: 170, r: 18, t: 48, b: 52 },
        xaxis: { title: "Risque moyen", gridcolor: "#eef2f7" },
        yaxis: { automargin: true, autorange: "reversed" }
      }, PLOT_CONFIG);

      const maritimeRegionLabels = maritimeRegion.map(r => r.label);
      Plotly.react("maritimeEventMixPlot", [
        barTrace("Tempete", maritimeRegionLabels, numSeries(maritimeRegion, "storm"), "#1f77b4"),
        barTrace("Ouragan / typhon", maritimeRegionLabels, numSeries(maritimeRegion, "hurricane"), "#9467bd"),
        barTrace("Froid maritime", maritimeRegionLabels, numSeries(maritimeRegion, "cold"), "#17becf"),
        barTrace("Mousson", maritimeRegionLabels, numSeries(maritimeRegion, "monsoon"), "#2ca02c")
      ], {
        ...plotLayout("Mix d'evenements par bassin maritime"),
        barmode: "stack",
        xaxis: { automargin: true },
        yaxis: { title: "Intensite moyenne", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const nodeOps = KPI_PAYLOAD.node_ops_month || [];
      const nodeOpsMonths = nodeOps.map(r => r.month_index);
      const nodeOpsLayout = {
        ...plotLayout("Etat operationnel des noeuds supply_geo recalcules"),
        margin: { l: 54, r: 62, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Multiplicateur / indice", gridcolor: "#eef2f7", range: [0, 1.45] },
        yaxis2: { title: "Service estime (%)", overlaying: "y", side: "right", range: [0, 105], gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("nodeOpsStatePlot", [
        lineTrace("Capacite appliquee moyenne", nodeOpsMonths, numSeries(nodeOps, "avg_capacity_applied"), "#2ca02c"),
        lineTrace("Capacite min", nodeOpsMonths, numSeries(nodeOps, "min_capacity_applied"), "#31a354", { dash: "dot" }),
        lineTrace("Delai moyen", nodeOpsMonths, numSeries(nodeOps, "avg_lead_time_multiplier"), "#7f3c8d"),
        lineTrace("Perturbation moyenne", nodeOpsMonths, numSeries(nodeOps, "avg_disruption_index"), "#d62728", { width: 3 }),
        lineTrace("Service estime", nodeOpsMonths, numSeries(nodeOps, "avg_service_proxy_pct"), "#1f77b4", { yaxis: "y2" })
      ], nodeOpsLayout, PLOT_CONFIG);

      const opEventTypes = [
        ["capacite_meteo_degradee", "Capacite meteo degradee", "#2ca02c"],
        ["capacite_appoint", "Capacite d'appoint", "#74c476"],
        ["retard_approvisionnement", "Retard approvisionnement", "#756bb1"],
        ["congestion_logistique_inbound", "Congestion amont", "#fd8d3c"],
        ["congestion_logistique_outbound", "Congestion aval", "#fdae6b"],
        ["recalage_qualite", "Recalage qualite", "#636363"],
        ["maintenance_corrective_meteo", "Maintenance meteo", "#d62728"],
        ["overtime_energetique", "Surconsommation energetique", "#ff7f0e"],
        ["froid_transport_ou_site", "Froid transport/site", "#17becf"]
      ];
      Plotly.react("nodeOpsEventPlot", opEventTypes.map(([key, label, color]) => (
        barTrace(label, nodeOpsMonths, numSeries(nodeOps, key), color, { opacity: 0.74 })
      )), {
        ...plotLayout("Evenements operationnels recalcules depuis les drivers"),
        barmode: "stack",
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Nombre d'evenements", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const nodeOpsRegion = KPI_PAYLOAD.node_ops_region || [];
      Plotly.react("nodeOpsRegionPlot", [{
        type: "bar",
        x: nodeOpsRegion.map(r => Number(r.avg_disruption_index || 0)),
        y: nodeOpsRegion.map(r => r.label),
        orientation: "h",
        marker: { color: "#d62728" },
        text: nodeOpsRegion.map(r => `${r.site_count || 0} sites / service ${fmt(r.avg_service_proxy_pct || 0)}%`),
        hovertemplate: "%{y}<br>disruption=%{x:.3f}<br>%{text}<extra></extra>"
      }], {
        ...plotLayout("Noeuds: disruption operationnelle par region"),
        margin: { l: 150, r: 18, t: 48, b: 52 },
        xaxis: { title: "Perturbation moyenne", gridcolor: "#eef2f7" },
        yaxis: { automargin: true, autorange: "reversed" }
      }, PLOT_CONFIG);

      const lineage = (KPI_PAYLOAD.node_ops_lineage || []).slice(0, 40);
      const lineageLabels = [...new Set(lineage.flatMap(r => [r.source, r.target]))];
      const lineageIndex = new Map(lineageLabels.map((label, index) => [label, index]));
      Plotly.react("nodeOpsLineagePlot", [{
        type: "sankey",
        arrangement: "snap",
        node: {
          label: lineageLabels,
          pad: 12,
          thickness: 14,
          color: lineageLabels.map(label => label.startsWith("env:") ? "#9ecae1" : (label.includes("maritime") || label.startsWith("inbound:") || label.startsWith("outbound:") ? "#fdae6b" : "#756bb1"))
        },
        link: {
          source: lineage.map(r => lineageIndex.get(r.source)),
          target: lineage.map(r => lineageIndex.get(r.target)),
          value: lineage.map(r => Number(r.weight || r.count || 0)),
          color: "rgba(31, 119, 180, 0.22)"
        },
        hovertemplate: "%{source.label} -> %{target.label}<br>poids=%{value:.2f}<extra></extra>"
      }], {
        ...plotLayout("Tracabilite enviro / transport vers events operationnels"),
        margin: { l: 18, r: 18, t: 48, b: 20 }
      }, PLOT_CONFIG);

      const sddMonthly = KPI_PAYLOAD.sdd_monthly || [];
      const sddMonths = sddMonthly.map(r => r.month_index);
      const sddServiceLayout = {
        ...plotLayout("Full SDD supply_geo: service, backlog et chemins affectes"),
        margin: { l: 54, r: 62, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Service / disruption", gridcolor: "#eef2f7", range: [0, 1.05] },
        yaxis2: { title: "Retard kg / chemins", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("sddServicePlot", [
        lineTrace("Service OEM moyen", sddMonths, numSeries(sddMonthly, "avg_oem_service_level"), "#2ca02c", { width: 3 }),
        lineTrace("Perturbation moyenne", sddMonths, numSeries(sddMonthly, "avg_path_disruption_index"), "#d62728"),
        barTrace("Retard OEM", sddMonths, numSeries(sddMonthly, "oem_backlog_kg"), "#c6dbef", { yaxis: "y2", opacity: 0.45 }),
        lineTrace("Chemins affectes", sddMonths, numSeries(sddMonthly, "affected_path_count"), "#756bb1", { yaxis: "y2", dash: "dot" })
      ], sddServiceLayout, PLOT_CONFIG);

      const sddTier = KPI_PAYLOAD.sdd_tier_month || [];
      const tierMonths = [...new Set(sddTier.map(r => r.month_index))].sort((a, b) => Number(a) - Number(b));
      const tierRoles = ["T4", "T3", "T2", "T1", "OEM"];
      const tierColors = { T4: "#9ecae1", T3: "#3182bd", T2: "#756bb1", T1: "#ff7f0e", OEM: "#d62728" };
      Plotly.react("sddTierPlot", tierRoles.map(role => {
        const byMonth = new Map(sddTier.filter(r => r.role === role).map(r => [Number(r.month_index), Number(r.avg_service_level || 0)]));
        return lineTrace(role, tierMonths, tierMonths.map(month => byMonth.get(Number(month)) || 0), tierColors[role] || "#172033");
      }), {
        ...plotLayout("Propagation SDD: service par tier"),
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Service level", gridcolor: "#eef2f7", range: [0, 1.05] }
      }, PLOT_CONFIG);

      const sddMethods = KPI_PAYLOAD.sdd_method_comparison || [];
      Plotly.react("sddMethodPlot", [{
        type: "bar",
        x: sddMethods.map(r => r.method),
        y: sddMethods.map(r => Number(r.total_kgCO2e || 0)),
        marker: { color: ["#9ecae1", "#3182bd", "#f16913"] },
        hovertemplate: "%{x}<br>%{y:.1f} kgCO2e<extra></extra>"
      }], {
        ...plotLayout("Full SDD supply_geo: comparaison methodes"),
        xaxis: { automargin: true },
        yaxis: { title: "Total kgCO2e", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const sddImpactLayout = {
        ...plotLayout("Full SDD supply_geo: surimpacts mensuels"),
        margin: { l: 54, r: 70, t: 48, b: 52 },
        barmode: "stack",
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Surimpact mensuel kgCO2e", gridcolor: "#eef2f7" },
        yaxis2: { title: "SDD mensuel kgCO2e", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("sddImpactStackPlot", [
        barTrace("Approvisionnement alternatif", sddMonths, numSeries(sddMonthly, "backup_material"), "#9ecae1"),
        barTrace("Transport accelere", sddMonths, numSeries(sddMonthly, "premium_transport"), "#fd8d3c"),
        barTrace("Rebut et reprise", sddMonths, numSeries(sddMonthly, "scrap_rework"), "#756bb1"),
        barTrace("Energie d'appoint", sddMonths, numSeries(sddMonthly, "capacity_energy"), "#ff7f0e"),
        barTrace("Maintenance", sddMonths, numSeries(sddMonthly, "maintenance"), "#636363"),
        barTrace("Retard non absorbe", sddMonths, numSeries(sddMonthly, "backlog_penalty"), "#bdbdbd"),
        lineTrace("Impact dynamique mensuel", sddMonths, numSeries(sddMonthly, "sdd_kgCO2e"), "#d62728", { yaxis: "y2", width: 3 })
      ], sddImpactLayout, PLOT_CONFIG);

      const sddCumulative = KPI_PAYLOAD.sdd_cumulative || [];
      const sddCumMonths = sddCumulative.map(r => r.month_index);
      Plotly.react("sddCumulativePlot", [
        lineTrace("LCA classique", sddCumMonths, numSeries(sddCumulative, "classical_cumulative"), "#9ecae1"),
        lineTrace("TD-DLCA", sddCumMonths, numSeries(sddCumulative, "td_dlca_cumulative"), "#3182bd"),
        lineTrace("SDD supply_geo", sddCumMonths, numSeries(sddCumulative, "sdd_cumulative"), "#f16913", { width: 3 }),
        lineTrace("Surimpact cumule", sddCumMonths, numSeries(sddCumulative, "surimpact_cumulative"), "#d62728", { dash: "dot" })
      ], {
        ...plotLayout("Full SDD supply_geo: impacts cumules recalcules"),
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "kgCO2e cumules", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const ops = KPI_PAYLOAD.ops_month || [];
      const opsMonths = ops.map(r => r.month_index);
      const opsLayout = {
        ...plotLayout("Indicateurs operationnels derives des evenements meteo"),
        margin: { l: 54, r: 62, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Multiplicateur", gridcolor: "#eef2f7", range: [0.72, 1.45] },
        yaxis2: { title: "Nombre d'evenements", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("opsProxyPlot", [
        barTrace("Evenements simules", opsMonths, numSeries(ops, "event_count"), "#c6dbef", { yaxis: "y2", opacity: 0.52 }),
        lineTrace("Capacite appliquee min", opsMonths, numSeries(ops, "capacity_multiplier_min"), "#2ca02c"),
        lineTrace("Delai max", opsMonths, numSeries(ops, "lead_time_multiplier_max"), "#7f3c8d"),
        lineTrace("Taux de rebut max", opsMonths, numSeries(ops, "scrap_multiplier_max"), "#d62728")
      ], opsLayout, PLOT_CONFIG);

      const ha = KPI_PAYLOAD.horizon_adaptation || {};
      const haTimeline = ha.event_timeline || [];
      const haTimelineMonths = haTimeline.map(r => r.month_index);
      const calendarSpecs = [
        ["evt_canicule", "Canicule", "#d62728"],
        ["evt_secheresse", "Secheresse", "#ff7f0e"],
        ["evt_tempete_inondation", "Tempete / inondation", "#1f77b4"],
        ["evt_maintenance_corrective", "Maintenance corrective", "#2ca02c"],
        ["evt_congestion_logistique", "Congestion logistique", "#fd8d3c"],
        ["evt_overtime_energetique", "Surconsommation energetique", "#756bb1"],
        ["evt_capacite_appoint", "Capacite d'appoint", "#e377c2"],
        ["evt_recalage_qualite", "Recalage qualite", "#636363"]
      ];
      Plotly.react("haEventCalendarPlot", calendarSpecs.map(([key, label, color]) => (
        barTrace(label, haTimelineMonths, numSeries(haTimeline, key), color, { opacity: 0.72 })
      )), {
        ...plotLayout("Reference horizon-adaptation: climat et operations"),
        barmode: "stack",
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Flags mensuels", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const haState = ha.monthly_state || [];
      const haStateMonths = haState.map(r => r.month_index);
      const policyTitle = ha.policy_label ? ` (${ha.policy_label})` : "";
      Plotly.react("haSystemPlot", [
        lineTrace("Retard", haStateMonths, numSeries(haState, "backlog_end"), "#111111"),
        lineTrace("Capacite appliquee", haStateMonths, numSeries(haState, "capacity_applied"), "#2ca02c"),
        lineTrace("Production de secours", haStateMonths, numSeries(haState, "good_output_backup_units"), "#756bb1"),
        lineTrace("Disponibilite matiere principale", haStateMonths, scaledSeries(haState, "primary_supply_availability_applied", 100), "#31a354", { dash: "dot" })
      ], {
        ...plotLayout(`Reference horizon-adaptation: supply chain${policyTitle}`),
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Reponse operationnelle", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      Plotly.react("haEnergyPlot", [
        lineTrace("Solaire utilise", haStateMonths, numSeries(haState, "solar_used_kwh"), "#1f77b4"),
        lineTrace("Biomasse utilisee", haStateMonths, numSeries(haState, "biomass_used_kwh"), "#ff7f0e"),
        lineTrace("Decharge batterie", haStateMonths, numSeries(haState, "battery_discharge_kwh"), "#2ca02c"),
        lineTrace("Energie reseau", haStateMonths, numSeries(haState, "grid_energy_kwh"), "#d62728")
      ], {
        ...plotLayout(`Reference horizon-adaptation: energie${policyTitle}`),
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "kWh", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const batteryLayout = {
        ...plotLayout(`Reference horizon-adaptation: batterie, HVAC et biomasse${policyTitle}`),
        margin: { l: 54, r: 62, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "%", gridcolor: "#eef2f7" },
        yaxis2: { title: "kWh", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("haBatteryBiomassPlot", [
        lineTrace("SOH batterie", haStateMonths, scaledSeries(haState, "battery_soh", 100), "#7f3c8d"),
        lineTrace("Biomasse active", haStateMonths, scaledSeries(haState, "biomass_transition_level", 100), "#2ca02c"),
        lineTrace("Plafond biomasse", haStateMonths, scaledSeries(haState, "biomass_transition_cap", 100), "#74c476", { dash: "dash" }),
        lineTrace("SOC batterie", haStateMonths, numSeries(haState, "battery_soc_kwh"), "#1f77b4", { yaxis: "y2" }),
        lineTrace("Charge HVAC", haStateMonths, numSeries(haState, "hvac_kwh"), "#d62728", { yaxis: "y2" })
      ], batteryLayout, PLOT_CONFIG);

      const haWeather = ha.weather_driver || [];
      const haWeatherMonths = haWeather.map(r => r.month_index);
      const haWeatherLayout = {
        ...plotLayout("Reference horizon-adaptation: temperature et humidite"),
        margin: { l: 54, r: 62, t: 48, b: 52 },
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "degC", gridcolor: "#eef2f7" },
        yaxis2: { title: "% humidite", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("haWeatherDriverPlot", [
        lineTrace("Temperature", haWeatherMonths, numSeries(haWeather, "temp_c"), "#d62728"),
        lineTrace("Heat index", haWeatherMonths, numSeries(haWeather, "heat_index_c"), "#ff7f0e", { dash: "dash" }),
        lineTrace("Humidite", haWeatherMonths, numSeries(haWeather, "humidity_pct"), "#1f77b4", { yaxis: "y2" })
      ], haWeatherLayout, PLOT_CONFIG);

      const haReference = ha.reference_cumulative || [];
      const haReferenceMonths = haReference.map(r => r.month_index);
      Plotly.react("haImpactCumulativePlot", [
        lineTrace("LCA classique", haReferenceMonths, numSeries(haReference, "classical_cumulative"), "#9ecae1"),
        lineTrace("Time-Dependent DLCA", haReferenceMonths, numSeries(haReference, "td_cumulative"), "#3182bd"),
        lineTrace("SDD", haReferenceMonths, numSeries(haReference, "sdd_cumulative"), "#f16913")
      ], {
        ...plotLayout("Reference horizon-adaptation: impacts cumules"),
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Impact cumule (kgCO2e)", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      const haImpact = ha.event_impact || [];
      const haImpactMonths = haImpact.map(r => r.month_index);
      const eventImpactLayout = {
        ...plotLayout("Reference horizon-adaptation: surimpact des evenements"),
        margin: { l: 54, r: 70, t: 48, b: 52 },
        barmode: "stack",
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Surimpact mensuel", gridcolor: "#eef2f7" },
        yaxis2: { title: "Surimpact cumule", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      };
      Plotly.react("haEventImpactPlot", [
        barTrace("Approvisionnement alternatif", haImpactMonths, numSeries(haImpact, "surimpact_matiere"), "#9ecae1"),
        barTrace("Transport amont accelere", haImpactMonths, numSeries(haImpact, "surimpact_inbound"), "#6baed6"),
        barTrace("Transport aval air", haImpactMonths, numSeries(haImpact, "surimpact_transport_aval"), "#fd8d3c"),
        barTrace("Rebut / recalage", haImpactMonths, numSeries(haImpact, "surimpact_rebut"), "#756bb1"),
        lineTrace("Surimpact cumule", haImpactMonths, numSeries(haImpact, "surimpact_cumule"), "#d62728", { yaxis: "y2", width: 3 })
      ], eventImpactLayout, PLOT_CONFIG);

      barPlot("readinessPlot", KPI_PAYLOAD.readiness || [], "Readiness des chemins primaires", { horizontal: true });
      barPlot("lcaPlot", KPI_PAYLOAD.lca_use_class || [], "Classes LCA", { horizontal: true });
      barPlot("familyPlot", KPI_PAYLOAD.family_mass || [], "Masse par famille matiere", { horizontal: true });
      barPlot("modePlot", KPI_PAYLOAD.mode_kg_km || [], "Transport par mode (kg.km split)", { horizontal: false });

      const scatter = KPI_PAYLOAD.path_scatter || [];
      Plotly.react("routeMassPlot", [{
        type: "scatter",
        mode: "markers",
        x: scatter.map(r => r.x),
        y: scatter.map(r => r.y),
        text: scatter.map(r => `${r.path_id}<br>${r.system}<br>${r.component}<br>${r.family}<br>${r.readiness}`),
        marker: {
          size: scatter.map(r => Math.max(6, Math.min(28, Math.sqrt(Number(r.kg_km || 0)) / 8))),
          color: scatter.map(r => r.readiness === "primary_ready_topology" ? "#2ca02c" : "#ff7f0e"),
          opacity: 0.76,
          line: { width: 0.6, color: "#172033" }
        },
        hovertemplate: "%{text}<br>route=%{x:.1f} km<br>mass=%{y:.4f} kg<extra></extra>"
      }], {
        ...plotLayout("Masse vs distance par chemin primaire"),
        xaxis: { title: "Distance totale route (km)", gridcolor: "#eef2f7" },
        yaxis: { title: "Masse allouee (kg)", gridcolor: "#eef2f7" }
      }, PLOT_CONFIG);

      barPlot("topSitePlot", KPI_PAYLOAD.top_sites_by_mass || [], "Top sites par masse allouee", { horizontal: true });
      barPlot("eventExposurePlot", KPI_PAYLOAD.event_exposure || [], "Top expositions meteo par site", { horizontal: true });

      const months = (KPI_PAYLOAD.event_month || []).map(r => r.month_index);
      const eventTypes = ["heatwave", "drought", "storm", "hurricane", "cold"];
      const eventLabels = { heatwave: "Canicule", drought: "Secheresse", storm: "Tempete", hurricane: "Ouragan", cold: "Froid" };
      const eventColors = { heatwave: "#d62728", drought: "#ff7f0e", storm: "#1f77b4", hurricane: "#9467bd", cold: "#17becf" };
      const traces = eventTypes.map(type => ({
        type: "bar",
        name: eventLabels[type] || type,
        x: months,
        y: (KPI_PAYLOAD.event_month || []).map(r => Number(r[type] || 0)),
        marker: { color: eventColors[type] || "#5e6a7d" },
        hovertemplate: `${eventLabels[type] || type}<br>mois=%{x}<br>evenements=%{y}<extra></extra>`
      }));
      traces.push(lineTrace("Facteur d'intensification climatique", months, numSeries(weather, "avg_hazard_intensification_factor"), "#111111", { yaxis: "y2", dash: "dot", width: 3 }));
      Plotly.react("eventMonthPlot", traces, {
        ...plotLayout("Evenements meteo simules par mois"),
        barmode: "stack",
        xaxis: { title: "Mois", gridcolor: "#eef2f7" },
        yaxis: { title: "Nombre d'evenements", gridcolor: "#eef2f7" },
        yaxis2: { title: "Facteur climatique", overlaying: "y", side: "right", gridcolor: "rgba(0,0,0,0)" }
      }, PLOT_CONFIG);
    }

    function setTab(tab) {
      document.querySelectorAll(".tabBtn").forEach(btn => btn.classList.toggle("active", btn.dataset.tab === tab));
      document.getElementById("mapPane").classList.toggle("active", tab === "map");
      document.getElementById("kpiPane").classList.toggle("active", tab === "kpi");
      if (tab === "kpi") {
        setTimeout(() => {
          document.querySelectorAll(".plot").forEach(el => {
            try { Plotly.Plots.resize(el); } catch (e) {}
          });
        }, 60);
      }
    }

    function init() {
      document.getElementById("mapFrame").src = KPI_PAYLOAD.map_src || "";
      document.getElementById("generatedAt").textContent = KPI_PAYLOAD.generated_at_utc || "";
      document.querySelectorAll(".tabBtn").forEach(btn => btn.addEventListener("click", () => setTab(btn.dataset.tab || "map")));
      renderCards();
      renderPlots();
    }

    window.addEventListener("load", init);
  </script>
</body>
</html>
"""
    document = document.replace("__TITLE__", title).replace("__KPI_PAYLOAD__", payload_json)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(document, encoding="utf-8")


def write_run_package(
    *,
    output_root: Path,
    dirs: dict[str, Path],
    config: dict[str, Any],
    source_json: Path,
    source_map: Path,
    root_map: Path,
    summary: dict[str, Any],
    unique_node_rows: list[dict[str, Any]],
    unique_flow_rows: list[dict[str, Any]],
    artifact_paths: list[tuple[Path, str, str, str, bool]],
) -> tuple[Path, Path]:
    run_dir = dirs["run"]
    write_json(run_dir / "nodes.json", unique_node_rows)
    write_json(run_dir / "flows.json", unique_flow_rows)
    write_json(
        run_dir / "weather_event_policy.json",
        config.get("weather_driver") if isinstance(config.get("weather_driver"), dict) else {},
    )

    artifact_index = [
        artifact_record(output_root, path, group=group, domain=domain, grain=grain, required=required)
        for path, group, domain, grain, required in artifact_paths
    ]
    artifact_index.extend(
        [
            artifact_record(output_root, run_dir / "nodes.json", group="run", domain="nodes", grain="node", required=True),
            artifact_record(output_root, run_dir / "flows.json", group="run", domain="flows", grain="lane", required=True),
            artifact_record(
                output_root,
                run_dir / "weather_event_policy.json",
                group="run",
                domain="weather_event_policy",
                grain="case",
                required=False,
            ),
        ]
    )
    artifact_index_path = run_dir / "artifact_index.json"
    write_json(artifact_index_path, artifact_index)

    manifest = {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": summary["generated_at_utc"],
        "case_name": (config.get("case") or {}).get("name"),
        "output_dir": str(output_root.resolve(strict=False)),
        "source_json": str(source_json.resolve(strict=False)),
        "source_map_html": str(source_map.resolve(strict=False)),
        "root_map_html": str(root_map.resolve(strict=False)),
        "counts": summary["counts"],
        "capabilities": {
            "primary_supply_paths": True,
            "allocated_path_mass": True,
            "site_weather_driver": summary["counts"]["weather_rows"] > 0,
            "supplier_event_seed": summary["counts"]["event_seed_rows"] > 0,
            "transport_weather_risk": summary["counts"].get("transport_weather_rows", 0) > 0,
            "node_operational_state": summary["counts"].get("node_operational_rows", 0) > 0,
            "operational_event_lineage": summary["counts"].get("operational_event_rows", 0) > 0,
            "sdd_stateful_supply_engine": summary["counts"].get("sdd_flow_state_rows", 0) > 0,
            "sdd_brightway_inventory_delta": summary["counts"].get("sdd_brightway_inventory_delta_rows", 0) > 0,
            "sdd_brightway_exchange_delta": summary["counts"].get("sdd_brightway_exchange_delta_rows", 0) > 0,
            "sdd_brightway_exchange_lcia": summary["counts"].get("sdd_brightway_exchange_lcia_rows", 0) > 0,
            "brightway_supply_lca_source": summary["counts"].get("brightway_component_impacts", 0) > 0,
            "brightway_person_equivalent_units": summary["counts"].get("brightway_person_equivalent_indicators", 0) > 0,
            "brightway_regionalized_parametric_scenarios": summary["counts"].get("brightway_parametric_regional_scenarios", 0) > 0,
            "brightway_exact_scenario_lcia": summary["counts"].get("brightway_exact_scenario_lcia", 0) > 0,
            "brightway_excel_runtime_comparison": summary["counts"].get("brightway_excel_runtime_comparison", 0) > 0,
            "brightway_usage_calibration": summary["counts"].get("brightway_usage_calibration", 0) > 0,
            "brightway_runtime_available": bool(summary.get("brightway_model", {}).get("runtime", {}).get("can_execute_brightway")),
        },
        "entrypoints": {
            "nodes": "nodes.json",
            "flows": "flows.json",
            "artifact_index": "artifact_index.json",
            "summary": "../summaries/primary_supply_case_summary.json",
            "paths": "../data/primary_supply_paths.csv",
            "lanes": "../data/primary_supply_lanes.csv",
            "weather": "../data/site_weather_driver.csv",
            "event_seed": "../data/supplier_risk_event_seed.csv",
            "transport_weather": "../data/transport_weather_risk.csv",
            "node_operational_state": "../data/node_operational_state.csv",
            "operational_events": "../data/operational_event_seed.csv",
            "sdd_node_state": "../data/sdd_node_state.csv",
            "sdd_lane_state": "../data/sdd_lane_state.csv",
            "sdd_flow_state": "../data/sdd_flow_state.csv",
            "sdd_event_ledger": "../data/sdd_event_ledger.csv",
            "sdd_monthly_impacts": "../data/sdd_monthly_impacts.csv",
            "sdd_cumulative_impacts": "../data/sdd_cumulative_impacts.csv",
            "sdd_brightway_inventory_delta": "../data/sdd_brightway_inventory_delta.csv",
            "sdd_brightway_exchange_delta": "../data/sdd_brightway_exchange_delta.csv",
            "sdd_brightway_exchange_category_totals": "../data/sdd_brightway_exchange_category_totals.csv",
            "sdd_brightway_top_exchanges": "../data/sdd_brightway_top_exchanges.csv",
            "sdd_brightway_exchange_lcia": "../data/sdd_brightway_exchange_lcia.csv",
            "sdd_brightway_exchange_lcia_factors": "../data/sdd_brightway_exchange_lcia_factors.csv",
            "sdd_brightway_exchange_lcia_category_totals": "../data/sdd_brightway_exchange_lcia_category_totals.csv",
            "sdd_brightway_exchange_lcia_monthly": "../data/sdd_brightway_exchange_lcia_monthly.csv",
            "sdd_brightway_exchange_lcia_top": "../data/sdd_brightway_exchange_lcia_top.csv",
            "sdd_brightway_exchange_lcia_status": "../data/sdd_brightway_exchange_lcia_status.csv",
            "sdd_brightway_monthly": "../data/sdd_brightway_monthly.csv",
            "sdd_brightway_cumulative": "../data/sdd_brightway_cumulative.csv",
            "sdd_brightway_mechanism_totals": "../data/sdd_brightway_mechanism_totals.csv",
            "sdd_brightway_top_sites": "../data/sdd_brightway_top_sites.csv",
            "sdd_brightway_top_components": "../data/sdd_brightway_top_components.csv",
            "sdd_method_comparison": "../summaries/sdd_method_comparison.json",
            "brightway_component_impacts": "../data/brightway_component_impacts.csv",
            "brightway_indicator_summary": "../data/brightway_indicator_summary.csv",
            "brightway_indicator_unit_views": "../data/brightway_indicator_unit_views.csv",
            "brightway_reference_person_equivalent_results": "../data/brightway_reference_person_equivalent_results.csv",
            "brightway_reference_weighted_results": "../data/brightway_reference_weighted_results.csv",
            "brightway_reference_phase_breakdown": "../data/brightway_reference_phase_breakdown.csv",
            "brightway_reference_use_phase_components": "../data/brightway_reference_use_phase_components.csv",
            "brightway_reference_scenarios": "../data/brightway_reference_scenarios.csv",
            "brightway_reference_weighting_factors": "../data/brightway_reference_weighting_factors.csv",
            "brightway_reference_climate_contributors": "../data/brightway_reference_climate_contributors.csv",
            "brightway_masterboard_equipment_summary": "../data/brightway_masterboard_equipment_summary.csv",
            "brightway_masterboard_material_summary": "../data/brightway_masterboard_material_summary.csv",
            "brightway_parameters": "../data/brightway_parameters.csv",
            "brightway_activities": "../data/brightway_activities.csv",
            "brightway_activity_exchanges": "../data/brightway_activity_exchanges.csv",
            "brightway_supply_alignment": "../data/brightway_supply_alignment.csv",
            "brightway_parametric_levers": "../data/brightway_parametric_levers.csv",
            "brightway_parametric_sensitivity": "../data/brightway_parametric_sensitivity.csv",
            "brightway_parametric_switches": "../data/brightway_parametric_switches.csv",
            "brightway_parametric_regional_scenarios": "../data/brightway_parametric_regional_scenarios.csv",
            "brightway_exact_scenario_lcia": "../data/brightway_exact_scenario_lcia.csv",
            "brightway_excel_runtime_comparison": "../data/brightway_excel_runtime_comparison.csv",
            "brightway_usage_calibration": "../data/brightway_usage_calibration.csv",
            "brightway_model_summary": "../summaries/brightway_model_summary.json",
            "general_kpis": "../summaries/general_kpis.json",
            "base_results_map": "../maps/supply_geo_base_results_map.html",
            "dashboard": "../maps/supply_geo_base_results_map.html",
        },
    }
    manifest_path = run_dir / "run_manifest.json"
    write_json(manifest_path, manifest)
    return manifest_path, artifact_index_path


def build_supply_geo_case(
    *,
    config_path: Path | str = DEFAULT_CONFIG,
    output_dir: Path | str | None = None,
    weather_enabled: bool | None = None,
) -> CaseBuildResult:
    config = load_yaml_config(config_path)
    if weather_enabled is not None:
        weather = config.setdefault("weather_driver", {})
        if isinstance(weather, dict):
            weather["enabled"] = weather_enabled

    source_config = config.get("source") if isinstance(config.get("source"), dict) else {}
    output_config = config.get("output") if isinstance(config.get("output"), dict) else {}
    source_json = resolve_from_config(config, source_config.get("json_path", ""))
    source_map = resolve_from_config(config, source_config.get("map_html", ""))
    root_map = resolve_from_config(config, source_config.get("root_map_html", source_map))
    output_root = Path(output_dir).resolve(strict=False) if output_dir else resolve_from_config(config, output_config.get("root", "../outputs"))

    source_data = json.loads(source_json.read_text(encoding="utf-8"))
    records = source_data.get("records")
    if not isinstance(records, list):
        raise ValueError(f"Expected a records list in {source_json}")

    audit = load_audit_helper(config)
    dirs = ensure_standard_dirs(output_root)
    tables = build_primary_case_tables(records=records, config=config, audit=audit)
    unique_node_rows = unique_nodes(tables["nodes"])
    unique_site_rows = unique_sites(tables["nodes"])
    unique_flow_rows = unique_flows(tables["lanes"])
    weather_rows, event_rows = build_weather_tables(config, unique_site_rows)
    transport_weather_rows = build_transport_weather_rows(config, unique_flow_rows, unique_site_rows)
    node_operational_rows, operational_event_rows = build_node_operational_tables(
        unique_site_rows,
        weather_rows,
        event_rows,
        transport_weather_rows,
    )
    horizon_months = max((int(safe_float(row.get("month_index"))) for row in weather_rows), default=0)
    sdd_results = simulate_sdd_supply(
        path_rows=tables["paths"],
        lane_rows=tables["lanes"],
        node_operational_rows=node_operational_rows,
        transport_weather_rows=transport_weather_rows,
        horizon_months=horizon_months,
    )
    brightway_model = build_brightway_model_payload(tables["paths"])
    sdd_brightway = build_sdd_brightway_coupling(
        path_rows=tables["paths"],
        sdd_results=sdd_results,
        brightway_model=brightway_model,
    )

    data_paths = {
        "paths": dirs["data"] / "primary_supply_paths.csv",
        "nodes": dirs["data"] / "primary_supply_nodes.csv",
        "lanes": dirs["data"] / "primary_supply_lanes.csv",
        "unique_sites": dirs["data"] / "primary_supply_sites.csv",
        "weather": dirs["data"] / "site_weather_driver.csv",
        "events": dirs["data"] / "supplier_risk_event_seed.csv",
        "transport_weather": dirs["data"] / "transport_weather_risk.csv",
        "node_operational": dirs["data"] / "node_operational_state.csv",
        "operational_events": dirs["data"] / "operational_event_seed.csv",
        "sdd_node_state": dirs["data"] / "sdd_node_state.csv",
        "sdd_lane_state": dirs["data"] / "sdd_lane_state.csv",
        "sdd_flow_state": dirs["data"] / "sdd_flow_state.csv",
        "sdd_event_ledger": dirs["data"] / "sdd_event_ledger.csv",
        "sdd_monthly_impacts": dirs["data"] / "sdd_monthly_impacts.csv",
        "sdd_cumulative_impacts": dirs["data"] / "sdd_cumulative_impacts.csv",
        "sdd_brightway_inventory_delta": dirs["data"] / "sdd_brightway_inventory_delta.csv",
        "sdd_brightway_exchange_delta": dirs["data"] / "sdd_brightway_exchange_delta.csv",
        "sdd_brightway_exchange_category_totals": dirs["data"] / "sdd_brightway_exchange_category_totals.csv",
        "sdd_brightway_top_exchanges": dirs["data"] / "sdd_brightway_top_exchanges.csv",
        "sdd_brightway_exchange_lcia": dirs["data"] / "sdd_brightway_exchange_lcia.csv",
        "sdd_brightway_exchange_lcia_factors": dirs["data"] / "sdd_brightway_exchange_lcia_factors.csv",
        "sdd_brightway_exchange_lcia_category_totals": dirs["data"] / "sdd_brightway_exchange_lcia_category_totals.csv",
        "sdd_brightway_exchange_lcia_monthly": dirs["data"] / "sdd_brightway_exchange_lcia_monthly.csv",
        "sdd_brightway_exchange_lcia_top": dirs["data"] / "sdd_brightway_exchange_lcia_top.csv",
        "sdd_brightway_exchange_lcia_status": dirs["data"] / "sdd_brightway_exchange_lcia_status.csv",
        "sdd_brightway_monthly": dirs["data"] / "sdd_brightway_monthly.csv",
        "sdd_brightway_cumulative": dirs["data"] / "sdd_brightway_cumulative.csv",
        "sdd_brightway_mechanism_totals": dirs["data"] / "sdd_brightway_mechanism_totals.csv",
        "sdd_brightway_top_sites": dirs["data"] / "sdd_brightway_top_sites.csv",
        "sdd_brightway_top_components": dirs["data"] / "sdd_brightway_top_components.csv",
        "brightway_component_impacts": dirs["data"] / "brightway_component_impacts.csv",
        "brightway_indicator_summary": dirs["data"] / "brightway_indicator_summary.csv",
        "brightway_indicator_unit_views": dirs["data"] / "brightway_indicator_unit_views.csv",
        "brightway_reference_person_equivalent_results": dirs["data"] / "brightway_reference_person_equivalent_results.csv",
        "brightway_reference_weighted_results": dirs["data"] / "brightway_reference_weighted_results.csv",
        "brightway_reference_phase_breakdown": dirs["data"] / "brightway_reference_phase_breakdown.csv",
        "brightway_reference_use_phase_components": dirs["data"] / "brightway_reference_use_phase_components.csv",
        "brightway_reference_scenarios": dirs["data"] / "brightway_reference_scenarios.csv",
        "brightway_reference_weighting_factors": dirs["data"] / "brightway_reference_weighting_factors.csv",
        "brightway_reference_climate_contributors": dirs["data"] / "brightway_reference_climate_contributors.csv",
        "brightway_masterboard_equipment_summary": dirs["data"] / "brightway_masterboard_equipment_summary.csv",
        "brightway_masterboard_material_summary": dirs["data"] / "brightway_masterboard_material_summary.csv",
        "brightway_parameters": dirs["data"] / "brightway_parameters.csv",
        "brightway_activities": dirs["data"] / "brightway_activities.csv",
        "brightway_activity_exchanges": dirs["data"] / "brightway_activity_exchanges.csv",
        "brightway_supply_alignment": dirs["data"] / "brightway_supply_alignment.csv",
        "brightway_parametric_levers": dirs["data"] / "brightway_parametric_levers.csv",
        "brightway_parametric_sensitivity": dirs["data"] / "brightway_parametric_sensitivity.csv",
        "brightway_parametric_switches": dirs["data"] / "brightway_parametric_switches.csv",
        "brightway_parametric_regional_scenarios": dirs["data"] / "brightway_parametric_regional_scenarios.csv",
        "brightway_exact_scenario_lcia": dirs["data"] / "brightway_exact_scenario_lcia.csv",
        "brightway_excel_runtime_comparison": dirs["data"] / "brightway_excel_runtime_comparison.csv",
        "brightway_usage_calibration": dirs["data"] / "brightway_usage_calibration.csv",
        "skipped": dirs["data"] / "skipped_records.csv",
    }
    write_csv(data_paths["paths"], tables["paths"])
    write_csv(data_paths["nodes"], tables["nodes"])
    write_csv(data_paths["lanes"], tables["lanes"])
    write_csv(data_paths["unique_sites"], unique_site_rows)
    write_csv(data_paths["weather"], weather_rows)
    write_csv(data_paths["events"], event_rows)
    write_csv(data_paths["transport_weather"], transport_weather_rows)
    write_csv(data_paths["node_operational"], node_operational_rows)
    write_csv(data_paths["operational_events"], operational_event_rows)
    write_csv(data_paths["sdd_node_state"], sdd_results["sdd_node_state"])
    write_csv(data_paths["sdd_lane_state"], sdd_results["sdd_lane_state"])
    write_csv(data_paths["sdd_flow_state"], sdd_results["sdd_flow_state"])
    write_csv(data_paths["sdd_event_ledger"], sdd_results["sdd_event_ledger"])
    write_csv(data_paths["sdd_monthly_impacts"], sdd_results["sdd_monthly_impacts"])
    write_csv(data_paths["sdd_cumulative_impacts"], sdd_results["sdd_cumulative_impacts"])
    write_csv(data_paths["sdd_brightway_inventory_delta"], sdd_brightway["inventory_delta"])
    write_csv(data_paths["sdd_brightway_exchange_delta"], sdd_brightway["exchange_delta"])
    sdd_exchange_lcia = run_sdd_exchange_brightway_lcia(
        brightway_model.get("runtime", {}),
        data_paths["sdd_brightway_exchange_delta"],
        climate_normalization_factor(brightway_model.get("indicator_unit_views", [])),
    )
    sdd_brightway.update(sdd_exchange_lcia)
    write_csv(data_paths["sdd_brightway_exchange_category_totals"], sdd_brightway["exchange_category_totals"])
    write_csv(data_paths["sdd_brightway_top_exchanges"], sdd_brightway["top_exchanges"])
    write_csv(data_paths["sdd_brightway_exchange_lcia"], sdd_brightway["exchange_lcia"])
    write_csv(data_paths["sdd_brightway_exchange_lcia_factors"], sdd_brightway["exchange_lcia_factors"])
    write_csv(data_paths["sdd_brightway_exchange_lcia_category_totals"], sdd_brightway["exchange_lcia_category_totals"])
    write_csv(data_paths["sdd_brightway_exchange_lcia_monthly"], sdd_brightway["exchange_lcia_monthly"])
    write_csv(data_paths["sdd_brightway_exchange_lcia_top"], sdd_brightway["exchange_lcia_top"])
    write_csv(data_paths["sdd_brightway_exchange_lcia_status"], sdd_brightway["exchange_lcia_status"])
    write_csv(data_paths["sdd_brightway_monthly"], sdd_brightway["monthly"])
    write_csv(data_paths["sdd_brightway_cumulative"], sdd_brightway["cumulative"])
    write_csv(data_paths["sdd_brightway_mechanism_totals"], sdd_brightway["mechanism_totals"])
    write_csv(data_paths["sdd_brightway_top_sites"], sdd_brightway["top_sites"])
    write_csv(data_paths["sdd_brightway_top_components"], sdd_brightway["top_components"])
    write_csv(data_paths["brightway_component_impacts"], brightway_model["component_impacts"])
    write_csv(data_paths["brightway_indicator_summary"], brightway_model["indicator_summary"])
    write_csv(data_paths["brightway_indicator_unit_views"], brightway_model["indicator_unit_views"])
    write_csv(data_paths["brightway_reference_person_equivalent_results"], brightway_model["reference_person_equivalent_results"])
    write_csv(data_paths["brightway_reference_weighted_results"], brightway_model["reference_weighted_results"])
    write_csv(data_paths["brightway_reference_phase_breakdown"], brightway_model["reference_phase_breakdown"])
    write_csv(data_paths["brightway_reference_use_phase_components"], brightway_model["reference_use_phase_components"])
    write_csv(data_paths["brightway_reference_scenarios"], brightway_model["reference_scenarios"])
    write_csv(data_paths["brightway_reference_weighting_factors"], brightway_model["reference_weighting_factors"])
    write_csv(data_paths["brightway_reference_climate_contributors"], brightway_model["reference_climate_contributors"])
    write_csv(data_paths["brightway_masterboard_equipment_summary"], brightway_model["masterboard_equipment_summary"])
    write_csv(data_paths["brightway_masterboard_material_summary"], brightway_model["masterboard_material_summary"])
    write_csv(data_paths["brightway_parameters"], brightway_model["parameters"])
    write_csv(data_paths["brightway_activities"], brightway_model["activities"])
    write_csv(data_paths["brightway_activity_exchanges"], brightway_model["exchanges"])
    write_csv(data_paths["brightway_supply_alignment"], brightway_model["supply_alignment"])
    write_csv(data_paths["brightway_parametric_levers"], brightway_model["parametric_levers"])
    write_csv(data_paths["brightway_parametric_sensitivity"], brightway_model["parametric_sensitivity"])
    write_csv(data_paths["brightway_parametric_switches"], brightway_model["parametric_switches"])
    write_csv(data_paths["brightway_parametric_regional_scenarios"], brightway_model["parametric_regional_scenarios"])
    write_csv(data_paths["brightway_exact_scenario_lcia"], brightway_model["exact_scenario_lcia"])
    write_csv(data_paths["brightway_excel_runtime_comparison"], brightway_model["excel_runtime_comparison"])
    write_csv(data_paths["brightway_usage_calibration"], brightway_model["usage_calibration"])
    write_csv(data_paths["skipped"], tables["skipped_records"], fieldnames=["record_index", "system", "component", "missing_roles"])

    map_ref = {
        "source_map_html": str(source_map.resolve(strict=False)),
        "root_map_html": str(root_map.resolve(strict=False)),
        "note": "The existing supply_geo Plotly map remains the visual source of truth; this adapter injects SDD result views into an enriched copy of that base map.",
    }
    map_ref_path = dirs["maps"] / "source_map_reference.json"
    write_json(map_ref_path, map_ref)

    summary = build_summary(
        source_json=source_json,
        source_map=source_map,
        root_map=root_map,
        source_records=records,
        tables=tables,
        unique_node_rows=unique_node_rows,
        unique_site_rows=unique_site_rows,
        unique_flow_rows=unique_flow_rows,
        weather_rows=weather_rows,
        event_rows=event_rows,
        transport_weather_rows=transport_weather_rows,
        node_operational_rows=node_operational_rows,
        operational_event_rows=operational_event_rows,
        sdd_results=sdd_results,
        sdd_brightway=sdd_brightway,
        brightway_model=brightway_model,
    )
    summary_path = dirs["summaries"] / "primary_supply_case_summary.json"
    write_json(summary_path, summary)
    report_path = dirs["reports"] / "primary_supply_case_report.md"
    write_report(report_path, summary)
    sdd_method_path = dirs["summaries"] / "sdd_method_comparison.json"
    write_json(sdd_method_path, sdd_results["sdd_method_comparison"])
    brightway_summary_path = dirs["summaries"] / "brightway_model_summary.json"
    write_json(
        brightway_summary_path,
        {
            "schema_version": brightway_model.get("schema_version"),
            "available": brightway_model.get("available"),
            "runtime": brightway_model.get("runtime", {}),
            "source_files": brightway_model.get("source_files", {}),
            "counts": brightway_model.get("counts", {}),
            "parametric_levers": brightway_model.get("parametric_levers", []),
            "parametric_switches": brightway_model.get("parametric_switches", []),
            "parametric_regional_scenarios": brightway_model.get("parametric_regional_scenarios", []),
            "exact_scenario_lcia": brightway_model.get("exact_scenario_lcia", []),
            "excel_runtime_comparison": brightway_model.get("excel_runtime_comparison", []),
            "indicator_unit_views": brightway_model.get("indicator_unit_views", []),
            "reference_person_equivalent_results": brightway_model.get("reference_person_equivalent_results", []),
            "reference_weighted_results": brightway_model.get("reference_weighted_results", []),
            "reference_scenarios": brightway_model.get("reference_scenarios", []),
            "top_climate_components": brightway_model.get("top_climate_components", []),
            "top_parameter_amounts": brightway_model.get("top_parameter_amounts", []),
        },
    )
    stale_sdd_map_path = dirs["maps"] / "supply_geo_sdd_results_map.html"
    if stale_sdd_map_path.exists():
        stale_sdd_map_path.unlink()
    base_results_map_path = dirs["maps"] / "supply_geo_base_results_map.html"
    dashboard_payload = build_general_kpi_payload(
        summary=summary,
        path_rows=tables["paths"],
        lane_rows=tables["lanes"],
        site_rows=unique_site_rows,
        weather_rows=weather_rows,
        event_rows=event_rows,
        transport_weather_rows=transport_weather_rows,
        node_operational_rows=node_operational_rows,
        operational_event_rows=operational_event_rows,
        sdd_results=sdd_results,
        sdd_brightway=sdd_brightway,
        brightway_model=brightway_model,
        map_src=browser_rel(base_results_map_path, dirs["maps"]),
    )
    write_enriched_base_map_html(
        base_results_map_path,
        source_map=source_map,
        site_rows=unique_site_rows,
        sdd_results=sdd_results,
        dashboard_payload=dashboard_payload,
    )
    general_kpis_path = dirs["summaries"] / "general_kpis.json"
    write_json(general_kpis_path, dashboard_payload)
    dashboard_path = dirs["maps"] / "supply_geo_results_dashboard.html"
    if dashboard_path.exists():
        dashboard_path.unlink()

    artifact_paths = [
        (data_paths["paths"], "data", "primary_supply_paths", "path", True),
        (data_paths["nodes"], "data", "primary_supply_nodes", "path_role_node", True),
        (data_paths["lanes"], "data", "primary_supply_lanes", "path_lane", True),
        (data_paths["unique_sites"], "data", "primary_supply_sites", "site", True),
        (data_paths["weather"], "data", "site_weather_driver", "site_month", False),
        (data_paths["events"], "data", "supplier_risk_event_seed", "site_event", False),
        (data_paths["transport_weather"], "data", "transport_weather_risk", "flow_month", False),
        (data_paths["node_operational"], "data", "node_operational_state", "site_month", False),
        (data_paths["operational_events"], "data", "operational_event_seed", "site_operation_event", False),
        (data_paths["sdd_node_state"], "data", "sdd_node_state", "path_role_month", True),
        (data_paths["sdd_lane_state"], "data", "sdd_lane_state", "path_lane_month", True),
        (data_paths["sdd_flow_state"], "data", "sdd_flow_state", "path_month", True),
        (data_paths["sdd_event_ledger"], "data", "sdd_event_ledger", "sdd_event", False),
        (data_paths["sdd_monthly_impacts"], "data", "sdd_monthly_impacts", "month", True),
        (data_paths["sdd_cumulative_impacts"], "data", "sdd_cumulative_impacts", "month", True),
        (data_paths["sdd_brightway_inventory_delta"], "data", "sdd_brightway_inventory_delta", "sdd_brightway_delta", True),
        (data_paths["sdd_brightway_exchange_delta"], "data", "sdd_brightway_exchange_delta", "sdd_brightway_exchange_delta", True),
        (data_paths["sdd_brightway_exchange_category_totals"], "data", "sdd_brightway_exchange_category_totals", "exchange_category", True),
        (data_paths["sdd_brightway_top_exchanges"], "data", "sdd_brightway_top_exchanges", "exchange", True),
        (data_paths["sdd_brightway_exchange_lcia"], "data", "sdd_brightway_exchange_lcia", "sdd_brightway_exchange_lcia", True),
        (data_paths["sdd_brightway_exchange_lcia_factors"], "data", "sdd_brightway_exchange_lcia_factors", "exchange_factor", True),
        (data_paths["sdd_brightway_exchange_lcia_category_totals"], "data", "sdd_brightway_exchange_lcia_category_totals", "exchange_category", True),
        (data_paths["sdd_brightway_exchange_lcia_monthly"], "data", "sdd_brightway_exchange_lcia_monthly", "month", True),
        (data_paths["sdd_brightway_exchange_lcia_top"], "data", "sdd_brightway_exchange_lcia_top", "exchange", True),
        (data_paths["sdd_brightway_exchange_lcia_status"], "data", "sdd_brightway_exchange_lcia_status", "case", True),
        (data_paths["sdd_brightway_monthly"], "data", "sdd_brightway_monthly", "month", True),
        (data_paths["sdd_brightway_cumulative"], "data", "sdd_brightway_cumulative", "month", True),
        (data_paths["sdd_brightway_mechanism_totals"], "data", "sdd_brightway_mechanism_totals", "mechanism", True),
        (data_paths["sdd_brightway_top_sites"], "data", "sdd_brightway_top_sites", "site", True),
        (data_paths["sdd_brightway_top_components"], "data", "sdd_brightway_top_components", "component", True),
        (data_paths["brightway_component_impacts"], "data", "brightway_component_impacts", "component_indicator", True),
        (data_paths["brightway_indicator_summary"], "data", "brightway_indicator_summary", "indicator", True),
        (data_paths["brightway_indicator_unit_views"], "data", "brightway_indicator_unit_views", "indicator_unit", True),
        (data_paths["brightway_reference_person_equivalent_results"], "data", "brightway_reference_person_equivalent_results", "indicator", True),
        (data_paths["brightway_reference_weighted_results"], "data", "brightway_reference_weighted_results", "indicator", True),
        (data_paths["brightway_reference_phase_breakdown"], "data", "brightway_reference_phase_breakdown", "indicator_phase", True),
        (data_paths["brightway_reference_use_phase_components"], "data", "brightway_reference_use_phase_components", "usage_component", True),
        (data_paths["brightway_reference_scenarios"], "data", "brightway_reference_scenarios", "scenario_phase", True),
        (data_paths["brightway_reference_weighting_factors"], "data", "brightway_reference_weighting_factors", "weighting_factor", True),
        (data_paths["brightway_reference_climate_contributors"], "data", "brightway_reference_climate_contributors", "climate_contributor", True),
        (data_paths["brightway_masterboard_equipment_summary"], "data", "brightway_masterboard_equipment_summary", "equipment", True),
        (data_paths["brightway_masterboard_material_summary"], "data", "brightway_masterboard_material_summary", "material", True),
        (data_paths["brightway_parameters"], "data", "brightway_parameters", "parameter", True),
        (data_paths["brightway_activities"], "data", "brightway_activities", "activity", True),
        (data_paths["brightway_activity_exchanges"], "data", "brightway_activity_exchanges", "exchange", True),
        (data_paths["brightway_supply_alignment"], "data", "brightway_supply_alignment", "path_component", True),
        (data_paths["brightway_parametric_levers"], "data", "brightway_parametric_levers", "lever", True),
        (data_paths["brightway_parametric_sensitivity"], "data", "brightway_parametric_sensitivity", "lever_exchange", True),
        (data_paths["brightway_parametric_switches"], "data", "brightway_parametric_switches", "switch", True),
        (data_paths["brightway_parametric_regional_scenarios"], "data", "brightway_parametric_regional_scenarios", "regional_scenario", True),
        (data_paths["brightway_exact_scenario_lcia"], "data", "brightway_exact_scenario_lcia", "exact_lcia_scenario", True),
        (data_paths["brightway_excel_runtime_comparison"], "data", "brightway_excel_runtime_comparison", "runtime_reference_comparison", True),
        (data_paths["brightway_usage_calibration"], "data", "brightway_usage_calibration", "usage_calibration", True),
        (data_paths["skipped"], "data", "skipped_records", "record", False),
        (summary_path, "summaries", "primary_supply_case_summary", "case", True),
        (sdd_method_path, "summaries", "sdd_method_comparison", "case", True),
        (brightway_summary_path, "summaries", "brightway_model_summary", "case", True),
        (general_kpis_path, "summaries", "general_kpis", "case", True),
        (report_path, "reports", "primary_supply_case_report", "case", False),
        (map_ref_path, "maps", "source_map_reference", "case", False),
        (base_results_map_path, "maps", "base_results_map", "case", True),
    ]
    manifest_path, artifact_index_path = write_run_package(
        output_root=output_root,
        dirs=dirs,
        config=config,
        source_json=source_json,
        source_map=source_map,
        root_map=root_map,
        summary=summary,
        unique_node_rows=unique_node_rows,
        unique_flow_rows=unique_flow_rows,
        artifact_paths=artifact_paths,
    )
    return CaseBuildResult(
        output_root=output_root,
        summary_path=summary_path,
        manifest_path=manifest_path,
        artifact_index_path=artifact_index_path,
        summary=summary,
    )


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the POC2026 supply_geo primary case package.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--no-weather", action="store_true", help="Skip site weather/event seed generation.")
    return parser.parse_args(list(argv) if argv is not None else None)


def main(argv: Iterable[str] | None = None) -> None:
    args = parse_args(argv)
    result = build_supply_geo_case(
        config_path=args.config,
        output_dir=args.output_dir,
        weather_enabled=False if args.no_weather else None,
    )
    counts = result.summary["counts"]
    print(f"Wrote supply_geo case package: {result.output_root}")
    print(f"Primary paths: {counts['primary_paths']}")
    print(f"Unique sites: {counts['unique_sites']}")
    print(f"Event seed rows: {counts['event_seed_rows']}")
    print(f"Manifest: {result.manifest_path}")


if __name__ == "__main__":
    main()
