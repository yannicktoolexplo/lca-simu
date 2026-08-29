#!/usr/bin/env python3
"""Forensic audit of the STELIA aircraft-use model and OPERA Brightway branch."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
BW_ROOT = REPO_ROOT / "bw_tristan"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def number(value: object, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def first(rows: list[dict[str, str]], **criteria: str) -> dict[str, str]:
    for row in rows:
        if all(row.get(key) == value for key, value in criteria.items()):
            return row
    raise RuntimeError(f"Missing row for {criteria}")


def fmt(value: float, digits: int = 3) -> str:
    return f"{value:,.{digits}f}".replace(",", " ")


def workbook_path(pattern: str) -> Path:
    matches = list(BW_ROOT.glob(pattern))
    if not matches:
        raise FileNotFoundError(pattern)
    return matches[0]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "outputs",
    )
    return parser.parse_args()


def build_audit(output_root: Path) -> tuple[list[dict], list[dict], list[dict], str]:
    data = output_root / "data"
    parameters = read_csv(data / "brightway_parameters.csv")
    exchanges = read_csv(data / "brightway_activity_exchanges.csv")
    exact = read_csv(data / "brightway_exact_scenario_lcia.csv")
    comparison = read_csv(data / "brightway_excel_runtime_comparison.csv")
    component_impacts = read_csv(data / "brightway_component_impacts.csv")
    parameter_values = {row["name"]: number(row.get("amount")) for row in parameters}

    masterboard_path = workbook_path("STELIA Masterboard*.xlsx")
    stelia_path = workbook_path("STELIA LCA SEATS v14022022v2.xlsx")
    masterboard = load_workbook(masterboard_path, data_only=True, read_only=True)
    final_workbook = load_workbook(stelia_path, data_only=True, read_only=True)
    flight = masterboard["Conso Vol"]
    master = final_workbook["Master"]

    seat_kg = number(flight["B1"].value)
    mtow_kg = number(flight["B3"].value)
    stale_life_years = number(flight["B4"].value)
    distance_km = number(flight["B5"].value)
    flights_per_year = number(flight["B6"].value)
    fuel_l_per_100_pkm = number(flight["B7"].value)
    aircraft_fuel_l_per_100_km = number(flight["B8"].value)
    passenger_count = aircraft_fuel_l_per_100_km / fuel_l_per_100_pkm
    density_kg_per_l = number(flight["E13"].value) / number(flight["B13"].value)
    final_life_years = 7.0

    lifetime_distance_km = distance_km * flights_per_year * final_life_years
    seat_tkm = seat_kg / 1000 * lifetime_distance_km
    annual_seat_tkm = seat_tkm / final_life_years
    total_aircraft_fuel_l = (
        fuel_l_per_100_pkm / 100
        * passenger_count
        * lifetime_distance_km
    )
    allocated_seat_fuel_l = total_aircraft_fuel_l * seat_kg / mtow_kg
    allocated_seat_fuel_kg = allocated_seat_fuel_l * density_kg_per_l

    excel_upstream_rounded = number(master["ID5"].value)
    excel_operation_rounded = number(master["IE5"].value)
    excel_cleaning_rounded = number(master["IF5"].value)
    excel_use_rounded = excel_upstream_rounded + excel_operation_rounded + excel_cleaning_rounded
    excel_annual_use_rounded = (
        number(master["IA5"].value)
        + number(master["IB5"].value)
        + number(master["IC5"].value)
    )
    detailed_total_rounded = 0.0
    recycling_rounded = 0.0
    family_value = ""
    system_value = ""
    for column in range(2, master.max_column + 1):
        family_value = str(master.cell(1, column).value or family_value)
        system_value = str(master.cell(2, column).value or system_value)
        value = number(master.cell(5, column).value)
        if family_value == "Utilisation 1 an":
            continue
        if family_value == "Fin de vie" and system_value == "Recyclage":
            recycling_rounded += value
            continue
        detailed_total_rounded += value

    aligned = first(exact, scenario_id="current_export", root_activity_id="lifecycle_excel_aligned")
    production_exact = first(exact, scenario_id="current_export", root_activity_id="production")
    excel_use_pe_reconverted = number(aligned["excel_use_phase_kgco2e_added"])
    raw_use = number(aligned["raw_passive_usage_kgco2e_removed"])
    raw_tkm = number(aligned["raw_passive_usage_amount"])
    raw_factor_tkm = number(aligned["passive_usage_activity_score_kgco2e_per_unit"])
    non_use_brightway = number(aligned["score_kgco2e"]) - excel_use_pe_reconverted
    production_brightway = number(production_exact["score_kgco2e"])
    recycling_credit_brightway = non_use_brightway - production_brightway
    excel_total = number(first(comparison, scope_id="lifecycle_total")["excel_kgco2e"])
    climate_units = next(
        row
        for row in read_csv(data / "brightway_indicator_unit_views.csv")
        if "Climate Change - total" in row.get("indicator", "")
    )
    normalization_factor = number(climate_units["normalization_factor_per_person_year"])
    lifecycle_comparison = first(comparison, scope_id="lifecycle_total")
    excel_total_pe_reconverted = (
        number(lifecycle_comparison["excel_person_equivalent"])
        * normalization_factor
    )
    excel_climate_components = [
        row
        for row in component_impacts
        if "Climate Change - total" in row.get("indicator", "")
    ]

    def excel_family_climate(family: str) -> float:
        return sum(
            number(row.get("value"))
            for row in excel_climate_components
            if row.get("family") == family
        )

    excel_subassemblies = excel_family_climate("Sous-ensemble")
    excel_ife = excel_family_climate("IFE")
    excel_assembly = excel_family_climate("Site assemblage")
    excel_delivery = excel_family_climate("Livraison")
    excel_landfill = excel_family_climate("Fin de vie")
    excel_before_eol = excel_subassemblies + excel_ife + excel_assembly + excel_delivery

    kerosene_exchange = next(
        row
        for row in exchanges
        if row.get("activity_name") == "kerosene, production et combustion, 1tkm eq"
        and row.get("name") == "market for kerosene"
    )
    raw_fuel_kg_per_tkm = number(kerosene_exchange["amount"])
    raw_fuel_kg = raw_tkm * raw_fuel_kg_per_tkm
    raw_factor_per_kg_fuel = raw_use / raw_fuel_kg
    physical_brightway_use = allocated_seat_fuel_kg * raw_factor_per_kg_fuel

    excel_use_without_cleaning_rounded = excel_upstream_rounded + excel_operation_rounded
    excel_upstream_factor = excel_upstream_rounded / allocated_seat_fuel_kg
    excel_operation_factor = excel_operation_rounded / allocated_seat_fuel_kg
    excel_wtw_factor = excel_use_without_cleaning_rounded / allocated_seat_fuel_kg
    excel_factor_per_tkm = excel_use_without_cleaning_rounded / seat_tkm

    cargo_parameter = parameter_values["fauteuil_cargo_conso_passive"]
    kerosene_parameter = parameter_values["fauteuil_kero_conso_passive"]
    cargo_refs = sum(
        "fauteuil_cargo_conso_passive" in (row.get("formula") or "")
        for row in exchanges
    )
    kerosene_refs = sum(
        "fauteuil_kero_conso_passive" in (row.get("formula") or "")
        for row in exchanges
    )
    formula_corpus = "\n".join(row.get("formula") or "" for row in exchanges)
    disconnected_parameters = sum(
        not re.search(rf"(?<![A-Za-z0-9_]){re.escape(row['name'])}(?![A-Za-z0-9_])", formula_corpus)
        for row in parameters
    )
    passenger_exchange = next(
        row
        for row in exchanges
        if row.get("activity_name") == "consommation passive siege pkm"
        and row.get("name") == "kerosene, production et combustion, 1tkm eq"
    )
    palette_steel_exchange = next(
        row
        for row in exchanges
        if row.get("activity_name") == "Ensemble palette optimisee"
        and row.get("name") == "steel production, converter, low-alloyed"
    )
    palette_steel_formula_amount = (
        parameter_values["epo_eol_35nc6"] + parameter_values["epo_eol_acier"]
    )
    invalid_meal_table_formulas = sum(
        row.get("activity_name") == "Ensemble tablette repas"
        and "+1+" in (row.get("formula") or "")
        for row in exchanges
    )

    marginal_fuel_low_kg = seat_kg * lifetime_distance_km / 1000 * 0.02
    marginal_fuel_central_kg = seat_kg * lifetime_distance_km / 1000 * 0.025
    marginal_fuel_high_kg = seat_kg * lifetime_distance_km / 1000 * 0.03

    scenarios = [
        {
            "scenario_id": "stelia_excel_raw",
            "label": "STELIA final, impacts bruts",
            "use_kgco2e": round(excel_use_rounded, 6),
            "lifecycle_kgco2e": round(excel_total, 6),
            "fuel_kg": round(allocated_seat_fuel_kg, 6),
            "factor_kgco2e_per_kg_fuel": round(excel_use_rounded / allocated_seat_fuel_kg, 9),
            "status": "reference_historique_brute_non_independante",
            "interpretation": "Impacts bruts publies par STELIA; les facteurs Sphera sont statiques dans le classeur.",
        },
        {
            "scenario_id": "stelia_person_equivalent_reconverted",
            "label": "STELIA reconverti depuis les personnes equivalentes arrondies",
            "use_kgco2e": round(excel_use_pe_reconverted, 6),
            "lifecycle_kgco2e": round(excel_total_pe_reconverted, 6),
            "fuel_kg": round(allocated_seat_fuel_kg, 6),
            "factor_kgco2e_per_kg_fuel": round(excel_use_pe_reconverted / allocated_seat_fuel_kg, 9),
            "status": "vue_normalisee_ne_pas_utiliser_comme_reference_brute",
            "interpretation": "Les personnes equivalentes sont une vue comparative arrondie, pas la source des kgCO2e.",
        },
        {
            "scenario_id": "brightway_fuel_mass_corrected",
            "label": "Brightway ecoinvent sur la quantite de carburant STELIA, avant fin de vie",
            "use_kgco2e": round(physical_brightway_use, 6),
            "lifecycle_kgco2e": round(physical_brightway_use + production_brightway, 6),
            "fuel_kg": round(allocated_seat_fuel_kg, 6),
            "factor_kgco2e_per_kg_fuel": round(raw_factor_per_kg_fuel, 9),
            "status": "meilleure_estimation_brightway_actuelle_avant_fin_de_vie",
            "interpretation": "La masse de carburant STELIA est conservee; le fond ecoinvent et les emissions OPERA sont recalcules. Le traitement de fin de vie reste a ajouter sans credit cut-off indu.",
        },
        {
            "scenario_id": "brightway_fuel_mass_with_imported_recycling_credit",
            "label": "Brightway avec credit de recyclage OPERA importe",
            "use_kgco2e": round(physical_brightway_use, 6),
            "lifecycle_kgco2e": round(physical_brightway_use + non_use_brightway, 6),
            "fuel_kg": round(allocated_seat_fuel_kg, 6),
            "factor_kgco2e_per_kg_fuel": round(raw_factor_per_kg_fuel, 9),
            "status": "diagnostic_non_retenu_credit_incompatible_cutoff",
            "interpretation": "Ancienne valeur centrale conservee pour tracabilite; elle soustrait un credit de recyclage incompatible avec le scenario cut-off principal.",
        },
        {
            "scenario_id": "brightway_marginal_weight_central",
            "label": "Brightway, effet marginal central de la masse du siege",
            "use_kgco2e": round(marginal_fuel_central_kg * raw_factor_per_kg_fuel, 6),
            "lifecycle_kgco2e": round(marginal_fuel_central_kg * raw_factor_per_kg_fuel + production_brightway, 6),
            "fuel_kg": round(marginal_fuel_central_kg, 6),
            "factor_kgco2e_per_kg_fuel": round(raw_factor_per_kg_fuel, 9),
            "status": "sensibilite_ecoconception_a_confirmer_par_modele_mission",
            "interpretation": "Coefficient central de 0,025 kg de carburant par kg embarque et 1 000 km; pertinent pour un delta de masse, pas pour une empreinte attributionnelle.",
        },
        {
            "scenario_id": "brightway_marginal_weight_low",
            "label": "Brightway, effet marginal bas de la masse du siege",
            "use_kgco2e": round(marginal_fuel_low_kg * raw_factor_per_kg_fuel, 6),
            "lifecycle_kgco2e": round(marginal_fuel_low_kg * raw_factor_per_kg_fuel + production_brightway, 6),
            "fuel_kg": round(marginal_fuel_low_kg, 6),
            "factor_kgco2e_per_kg_fuel": round(raw_factor_per_kg_fuel, 9),
            "status": "borne_sensibilite_ecoconception",
            "interpretation": "Borne basse de 0,020 kg de carburant par kg embarque et 1 000 km.",
        },
        {
            "scenario_id": "brightway_marginal_weight_high",
            "label": "Brightway, effet marginal haut de la masse du siege",
            "use_kgco2e": round(marginal_fuel_high_kg * raw_factor_per_kg_fuel, 6),
            "lifecycle_kgco2e": round(marginal_fuel_high_kg * raw_factor_per_kg_fuel + production_brightway, 6),
            "fuel_kg": round(marginal_fuel_high_kg, 6),
            "factor_kgco2e_per_kg_fuel": round(raw_factor_per_kg_fuel, 9),
            "status": "borne_sensibilite_ecoconception",
            "interpretation": "Borne haute de 0,030 kg de carburant par kg embarque et 1 000 km.",
        },
        {
            "scenario_id": "opera_raw_air_freight",
            "label": "OPERA brut avec intensite fret aerien",
            "use_kgco2e": round(raw_use, 6),
            "lifecycle_kgco2e": round(raw_use + non_use_brightway, 6),
            "fuel_kg": round(raw_fuel_kg, 6),
            "factor_kgco2e_per_kg_fuel": round(raw_factor_per_kg_fuel, 9),
            "status": "rejete_service_de_transport_incompatible",
            "interpretation": "La tkm du siege est correcte mais elle est traitee comme une tkm de charge utile cargo.",
        },
    ]

    issues = [
        {
            "priority": 1,
            "issue_id": "person_equivalent_back_conversion",
            "object": "Comparaison POC2026 corrigee",
            "finding": "Les personnes equivalentes arrondies etaient reconverties en kgCO2e alors que les impacts bruts existent dans Master.",
            "current_value": round(excel_total, 6),
            "reference_value": round(excel_total_pe_reconverted, 6),
            "unit": "kgCO2e brut vs reconverti",
            "ratio": round(excel_total_pe_reconverted / excel_total, 9),
            "effect_kgco2e": round(excel_total_pe_reconverted - excel_total, 6),
            "confidence": "elevee",
            "source": "STELIA Master row 5; Graphes perseq!B24",
        },
        {
            "priority": 1,
            "issue_id": "fixed_one_year_double_count_import",
            "object": "Import POC2026 corrige",
            "finding": "L'import additionnait auparavant les vues d'utilisation 1 an et 7 ans; les colonnes annuelles sont maintenant exclues du total cycle de vie.",
            "current_value": 0.0,
            "reference_value": round(excel_annual_use_rounded, 6),
            "unit": "kgCO2e retires du double comptage",
            "ratio": 0.0,
            "effect_kgco2e": round(-excel_annual_use_rounded, 6),
            "confidence": "elevee",
            "source": "adapter.load_brightway_component_impacts; STELIA Master!IA5:IC5",
        },
        {
            "priority": 1,
            "issue_id": "fixed_eol_scenario_mixing",
            "object": "Import POC2026 corrige",
            "finding": "L'import additionnait auparavant l'enfouissement et le scenario alternatif de recyclage; le total principal retient maintenant l'enfouissement declare dans l'etude.",
            "current_value": 0.0,
            "reference_value": round(recycling_rounded, 6),
            "unit": "kgCO2e du scenario recyclage exclu",
            "ratio": 0.0,
            "effect_kgco2e": round(-recycling_rounded, 6),
            "confidence": "elevee",
            "source": "STELIA Master!IL5:IZ5; PDF OPERA page 10",
        },
        {
            "priority": 1,
            "issue_id": "air_freight_service_mismatch",
            "object": "OPERA Brightway brut",
            "finding": "La masse-distance du siege est correcte, mais le facteur correspond au fret utile aerien.",
            "current_value": round(raw_fuel_kg_per_tkm, 6),
            "reference_value": round(allocated_seat_fuel_kg / seat_tkm, 6),
            "unit": "kg kerosene/tkm",
            "ratio": round(raw_fuel_kg / allocated_seat_fuel_kg, 6),
            "effect_kgco2e": round(raw_use - physical_brightway_use, 6),
            "confidence": "elevee",
            "source": "brightway_activity_exchanges.csv; STELIA Masterboard/Conso Vol",
        },
        {
            "priority": 1,
            "issue_id": "sphera_to_ecoinvent_factor_gap",
            "object": "Comparaison STELIA/Brightway",
            "finding": "Le facteur cycle carburant Sphera implicite est plus faible que le facteur OPERA/eecoinvent par kg de carburant.",
            "current_value": round(raw_factor_per_kg_fuel, 6),
            "reference_value": round(excel_wtw_factor, 6),
            "unit": "kgCO2e/kg kerosene",
            "ratio": round(raw_factor_per_kg_fuel / excel_wtw_factor, 6),
            "effect_kgco2e": round(physical_brightway_use - excel_use_without_cleaning_rounded, 6),
            "confidence": "elevee",
            "source": "Brightway EF 3.0 exact; STELIA Master!ID5 et IE5",
        },
        {
            "priority": 2,
            "issue_id": "dead_cargo_parameter_undocumented",
            "object": "Parametre OPERA",
            "finding": "Le parametre cargo 42 700 vaut presque exactement un dixieme de la tkm annuelle du siege; son unite et son eventuel calibrage Sphera ne sont pas documentes, et il est deconnecte.",
            "current_value": cargo_parameter,
            "reference_value": round(annual_seat_tkm, 6),
            "unit": "tkm/an probable",
            "ratio": round(cargo_parameter / annual_seat_tkm, 6),
            "effect_kgco2e": 0.0,
            "confidence": "moyenne",
            "source": f"brightway_parameters.csv; references echanges={cargo_refs}",
        },
        {
            "priority": 1,
            "issue_id": "dead_kerosene_parameter_dimension",
            "object": "Parametre OPERA",
            "finding": "La formule remplace les 200 passagers du Masterboard par 105 kg/passager; elle ne reproduit que 52,5 % des litres alloues au siege.",
            "current_value": round(kerosene_parameter, 6),
            "reference_value": round(allocated_seat_fuel_l, 6),
            "unit": "litre probable sur 7 ans",
            "ratio": round(kerosene_parameter / allocated_seat_fuel_l, 6),
            "effect_kgco2e": 0.0,
            "confidence": "elevee",
            "source": f"opera_bw2 inventaire OPERA!B630:C639; references echanges={kerosene_refs}",
        },
        {
            "priority": 2,
            "issue_id": "hardcoded_sphera_results",
            "object": "Classeur STELIA final",
            "finding": "Les impacts annuels kerosene et operation avion sont des constantes; le classeur ne contient ni quantite ni lien vers le calcul Sphera.",
            "current_value": f"IA5={master['IA5'].value}; IB5={master['IB5'].value}",
            "reference_value": "inventaire physique et facteurs versionnes",
            "unit": "kgCO2e/an",
            "ratio": "",
            "effect_kgco2e": "",
            "confidence": "elevee",
            "source": "STELIA LCA SEATS/Master!IA5:IB5",
        },
        {
            "priority": 2,
            "issue_id": "mtow_aircraft_version",
            "object": "Hypothese avion",
            "finding": "Le Masterboard nomme A321XLR mais utilise 95 t; Airbus publie 101 a 101,5 t pour le XLR.",
            "current_value": mtow_kg,
            "reference_value": 101500.0,
            "unit": "kg MTOW",
            "ratio": round(mtow_kg / 101500, 6),
            "effect_kgco2e": round(physical_brightway_use * (mtow_kg / 101500 - 1), 6),
            "confidence": "elevee",
            "source": "STELIA Masterboard/Conso Vol!B3; Airbus A321XLR",
        },
        {
            "priority": 2,
            "issue_id": "passenger_and_load_factor_uncertainty",
            "object": "Hypothese avion",
            "finding": "Les 200 passagers et 2,5 L/100 pkm ne sont pas sources dans le dossier et aucun taux de remplissage/cargo ventral n'est explicite.",
            "current_value": passenger_count,
            "reference_value": "206-220 sieges typiques XLR, puis taux de remplissage reel",
            "unit": "passagers supposes",
            "ratio": "",
            "effect_kgco2e": "",
            "confidence": "moyenne",
            "source": "STELIA Masterboard/Conso Vol!B7:B8; Airbus A321XLR; methodologie ICAO ICEC",
        },
        {
            "priority": 3,
            "issue_id": "stale_30_year_life",
            "object": "Gouvernance Excel",
            "finding": "Conso Vol affiche 30 ans alors que l'etude finale retient 7 ans; la formule annuelle B14 annule mathematiquement cette duree.",
            "current_value": stale_life_years,
            "reference_value": final_life_years,
            "unit": "ans",
            "ratio": round(stale_life_years / final_life_years, 6),
            "effect_kgco2e": 0.0,
            "confidence": "elevee",
            "source": "STELIA Masterboard/Conso Vol!B4,B14; STELIA final/Master!ID:IF",
        },
        {
            "priority": 1,
            "issue_id": "foreground_parameters_disconnected",
            "object": "Foreground OPERA/Brightway",
            "finding": "Une part importante des parametres importes ne pilote aucun echange; les formules sont conservees comme metadonnees mais Brightway calcule sur les montants figes.",
            "current_value": disconnected_parameters,
            "reference_value": 0,
            "unit": "parametres deconnectes",
            "ratio": round(disconnected_parameters / len(parameters), 6),
            "effect_kgco2e": "non quantifie",
            "confidence": "elevee",
            "source": "brightway_parameters.csv; brightway_activity_exchanges.csv",
        },
        {
            "priority": 1,
            "issue_id": "passenger_activity_wired_to_freight",
            "object": "Activite consommation passive siege pkm",
            "finding": "Le modele nomme pkm appelle l'activite kerosene 1tkm; il ne peut pas servir de branche passager independante.",
            "current_value": passenger_exchange["name"],
            "reference_value": "kerosene, production et combustion, 1pkm eq",
            "unit": "activite Brightway",
            "ratio": "",
            "effect_kgco2e": "non quantifie",
            "confidence": "elevee",
            "source": "brightway_activity_exchanges.csv",
        },
        {
            "priority": 1,
            "issue_id": "steel_exchange_formula_factor_ten",
            "object": "Ensemble palette optimisee",
            "finding": "Le montant acier stocke est dix fois superieur a la somme des deux parametres mentionnes dans sa formule.",
            "current_value": number(palette_steel_exchange["amount"]),
            "reference_value": round(palette_steel_formula_amount, 6),
            "unit": "kg acier",
            "ratio": round(number(palette_steel_exchange["amount"]) / palette_steel_formula_amount, 6),
            "effect_kgco2e": "a recalculer",
            "confidence": "elevee",
            "source": "brightway_activity_exchanges.csv; epo_eol_35nc6 + epo_eol_acier",
        },
        {
            "priority": 1,
            "issue_id": "invalid_meal_table_formulas",
            "object": "Ensemble tablette repas",
            "finding": "Des formules contiennent des libelles d'activites inseres dans une expression arithmetique et ne sont pas evaluables.",
            "current_value": invalid_meal_table_formulas,
            "reference_value": 0,
            "unit": "formules invalides",
            "ratio": "",
            "effect_kgco2e": "non quantifie",
            "confidence": "elevee",
            "source": "brightway_activity_exchanges.csv",
        },
        {
            "priority": 1,
            "issue_id": "cutoff_recycling_credit_consistency",
            "object": "Fin de vie Brightway",
            "finding": "Les credits de substitution ne doivent pas etre melanges au scenario principal ecoinvent cut-off; enfouissement, recyclage cut-off et recyclage avec substitution sont trois scenarios distincts.",
            "current_value": "credits et traitements presents dans le foreground",
            "reference_value": "scenario de fin de vie explicite et unique",
            "unit": "regle de modelisation",
            "ratio": "",
            "effect_kgco2e": "depend du scenario",
            "confidence": "elevee",
            "source": "ecoinvent-3.10-cutoff; brightway_activity_exchanges.csv",
        },
        {
            "priority": 1,
            "issue_id": "production_scope_includes_recycling_credit",
            "object": "Resultat hors utilisation Brightway",
            "finding": "Le total hors utilisation de 2,05 tCO2e n'est pas une production: il soustrait un credit de recyclage d'environ 0,40 tCO2e a la racine production.",
            "current_value": round(non_use_brightway, 6),
            "reference_value": round(production_brightway, 6),
            "unit": "kgCO2e",
            "ratio": round(non_use_brightway / production_brightway, 6),
            "effect_kgco2e": round(recycling_credit_brightway, 6),
            "confidence": "elevee",
            "source": "brightway_exact_scenario_lcia.csv",
        },
        {
            "priority": 1,
            "issue_id": "ife_mapping_gap",
            "object": "IFE de production",
            "finding": "L'IFE explique l'essentiel de l'ecart avant fin de vie: les resultats electroniques STELIA sont figes et les proxys Brightway actuels ne reproduisent pas leur intensite.",
            "current_value": 305.3,
            "reference_value": round(excel_ife, 6),
            "unit": "kgCO2e",
            "ratio": round(305.3 / excel_ife, 6),
            "effect_kgco2e": round(305.3 - excel_ife, 6),
            "confidence": "elevee sur l'ecart, moyenne sur la cause",
            "source": "STELIA Master; recalcul Brightway exact par activite",
        },
        {
            "priority": 1,
            "issue_id": "seat_mass_versions_not_reconciled",
            "object": "Conservation de masse",
            "finding": "Le dossier melange au moins trois configurations: masse fonctionnelle 109,967 kg, fin de vie 114,320 kg et BOM Masterboard recente 123,309 kg hors emballages.",
            "current_value": 109.967,
            "reference_value": "114,320 et 123,309",
            "unit": "kg/siege",
            "ratio": "",
            "effect_kgco2e": "usage et matieres affectes",
            "confidence": "elevee",
            "source": "STELIA Master; Masterboard BOM et Conso Vol",
        },
        {
            "priority": 1,
            "issue_id": "assembly_gas_unit",
            "object": "Gaz d'assemblage Excel",
            "finding": "La consommation est libellee 217,352 kWh/siege mais l'impact historique ressemble a une interpretation en MJ; une erreur d'unite peut sous-estimer le poste d'environ 30 kgCO2e.",
            "current_value": 13.7,
            "reference_value": "environ 44 kgCO2e apres conversion de 217,352 kWh en 782 MJ",
            "unit": "kgCO2e",
            "ratio": "",
            "effect_kgco2e": "environ +30",
            "confidence": "moyenne a elevee",
            "source": "STELIA Masterboard; racine assemblage Brightway",
        },
        {
            "priority": 1,
            "issue_id": "brightway_electronics_units",
            "object": "Foreground electronique Brightway",
            "finding": "Plusieurs proxys utilisent des unites ecoinvent comme des kilogrammes: powerbox 1,5 unite pour 1,5 kg, LED 1 kg pour 0,8 kg BOM et 0,45 unite d'ecran LCD dans le support ecran.",
            "current_value": "proxys dimensionnellement heterogenes",
            "reference_value": "nombre de pieces, masse et contenu electronique documentes",
            "unit": "unites d'echange",
            "ratio": "",
            "effect_kgco2e": "au moins 170 kgCO2e suspects sur le support ecran",
            "confidence": "elevee",
            "source": "foreground OPERA; recalcul Brightway par activite",
        },
        {
            "priority": 1,
            "issue_id": "packaging_transport_double_count",
            "object": "Livraison et emballage Brightway",
            "finding": "L'emballage final de 53 kg est porte par l'activite packaging puis de nouveau par le transport vers Hambourg; le perimetre doit etre dedoublonne.",
            "current_value": 120.6,
            "reference_value": round(excel_delivery, 6),
            "unit": "kgCO2e livraison/emballage",
            "ratio": round(120.6 / excel_delivery, 6),
            "effect_kgco2e": "a recalculer apres dedoublonnage",
            "confidence": "moyenne a elevee",
            "source": "foreground OPERA; STELIA Master",
        },
    ]

    bridge = [
        {
            "step_order": 1,
            "step_id": "opera_raw",
            "label": "Utilisation OPERA brute",
            "delta_kgco2e": round(raw_use, 6),
            "running_total_kgco2e": round(raw_use, 6),
            "explanation": "La tkm du siege est traitee comme fret aerien.",
        },
        {
            "step_order": 2,
            "step_id": "correct_fuel_intensity",
            "label": "Allocation carburant a la masse totale avion",
            "delta_kgco2e": round(physical_brightway_use - raw_use, 6),
            "running_total_kgco2e": round(physical_brightway_use, 6),
            "explanation": "0,2142 kg/tkm cargo devient 0,0421 kg/tkm de masse avion.",
        },
        {
            "step_order": 3,
            "step_id": "restore_sphera_factor",
            "label": "Remplacement du facteur ecoinvent par le facteur Sphera implicite",
            "delta_kgco2e": round(excel_use_without_cleaning_rounded - physical_brightway_use, 6),
            "running_total_kgco2e": round(excel_use_without_cleaning_rounded, 6),
            "explanation": "Le fond carburant historique est plus faible que le fond ecoinvent actuel.",
        },
        {
            "step_order": 4,
            "step_id": "cleaning",
            "label": "Nettoyage et desinfection",
            "delta_kgco2e": round(excel_cleaning_rounded, 6),
            "running_total_kgco2e": round(excel_use_rounded, 6),
            "explanation": "Contribution d'entretien sur sept ans.",
        },
    ]

    report = f"""# Audit detaille STELIA - utilisation du siege en avion

## Verdict

Le total STELIA brut de **{fmt(excel_total / 1000)} tCO2e** n'est pas explique par une erreur arithmetique simple. Sa phase d'utilisation est physiquement reconstructible a partir du Masterboard : **{fmt(allocated_seat_fuel_kg / 1000)} t de kerosene** allouees au siege sur sept ans, puis facteurs Sphera amont et operation avion.

Le calcul OPERA brut a **{fmt((raw_use + non_use_brightway) / 1000)} tCO2e** utilise la bonne masse-distance, **{fmt(seat_tkm, 0)} tkm**, mais la mauvaise intensite de carburant. Il applique **{fmt(raw_fuel_kg_per_tkm, 4)} kg/tkm**, facteur de fret utile aerien, au lieu de **{fmt(allocated_seat_fuel_kg / seat_tkm, 4)} kg/tkm** pour une masse incluse dans la masse totale de l'avion. Cela multiplie le carburant par **{fmt(raw_fuel_kg / allocated_seat_fuel_kg, 3)}**.

## Reconstruction physique STELIA

- Distance cumulee : {fmt(lifetime_distance_km, 0)} km = {fmt(distance_km, 0)} km x {fmt(flights_per_year, 0)} vols/an x 7 ans.
- Consommation avion : {fmt(fuel_l_per_100_pkm, 2)} L/100 pkm x {fmt(passenger_count, 0)} passagers = {fmt(aircraft_fuel_l_per_100_km, 0)} L/100 km.
- Carburant total avion : {fmt(total_aircraft_fuel_l / 1_000_000)} millions de litres.
- Part massique du siege : {fmt(seat_kg, 3)} / {fmt(mtow_kg, 0)} = {fmt(seat_kg / mtow_kg * 100, 4)} %.
- Carburant alloue au siege : {fmt(allocated_seat_fuel_l, 0)} L, soit {fmt(allocated_seat_fuel_kg, 0)} kg avec une densite de {fmt(density_kg_per_l, 3)} kg/L.
- Facteur Sphera amont implicite : {fmt(excel_upstream_factor, 4)} kgCO2e/kg de carburant.
- Facteur operation avion implicite : {fmt(excel_operation_factor, 4)} kgCO2e/kg, proche du facteur ICAO de 3,16 kgCO2/kg.
- Utilisation brute du classeur : {fmt(excel_use_rounded / 1000)} tCO2e; reconversion incorrecte des personnes equivalentes arrondies : {fmt(excel_use_pe_reconverted / 1000)} tCO2e.

## Deux questions ACV a ne pas confondre

- **Empreinte attributionnelle du siege.** On distribue une partie de la consommation totale de l'avion au siege. C'est la logique du Masterboard par ratio masse siege / MTOW. Elle convient pour decrire une part du cycle de vie, sous reserve de documenter le partage passagers/cargo et le denominateur de masse.
- **Effet marginal d'une decision d'ecoconception.** On calcule seulement le carburant ajoute ou evite par un changement de masse. Une sensibilite publiee de 0,020 a 0,030 kg de carburant par kg embarque et 1 000 km donne ici {fmt(marginal_fuel_low_kg / 1000)} a {fmt(marginal_fuel_high_kg / 1000)} t de carburant, soit {fmt(marginal_fuel_low_kg * raw_factor_per_kg_fuel / 1000)} a {fmt(marginal_fuel_high_kg * raw_factor_per_kg_fuel / 1000)} tCO2e d'utilisation avec le facteur Brightway actuel. Cette plage doit etre remplacee par un modele de mission A321 lorsque les donnees seront disponibles.

## Rapprochement de la production

La production et les autres phases hors utilisation valent **{fmt(excel_before_eol / 1000)} tCO2e avant fin de vie** dans STELIA et **{fmt(production_brightway / 1000)} tCO2e** dans la racine production Brightway, soit un ecart de {fmt((production_brightway / excel_before_eol - 1) * 100, 1)} %. Le chiffre Brightway de **{fmt(non_use_brightway / 1000)} tCO2e** ne doit pas etre appele production : il inclut un credit de recyclage de {fmt(recycling_credit_brightway / 1000)} tCO2e.

| Phase | STELIA kgCO2e | Brightway kgCO2e | Lecture |
|---|---:|---:|---|
| Sous-ensembles | {fmt(excel_subassemblies, 1)} | 1 968,7 | Proches a l'echelle du total |
| IFE | {fmt(excel_ife, 1)} | 305,3 | Principal ecart, proxys electroniques non reconciles |
| Assemblage | {fmt(excel_assembly, 1)} | 55,4 | Suspicion d'unite kWh/MJ sur le gaz Excel |
| Livraison et emballage | {fmt(excel_delivery, 1)} | 120,6 | Perimetre et double comptage a verifier |
| Fin de vie | {fmt(excel_landfill, 1)} en enfouissement | {fmt(recycling_credit_brightway, 1)} de credit | Scenarios non comparables |

L'IFE represente {fmt(excel_ife / excel_before_eol * 100, 1)} % du climat de production STELIA, et non 75 %. Le chiffre proche de 75 % concerne la contribution ponderee/normalisee de certains indicateurs ou des composants electroniques, pas la seule categorie changement climatique.

## Erreurs et faiblesses averees

1. **Reconversion des personnes equivalentes corrigee.** Les {fmt(excel_total_pe_reconverted / 1000)} tCO2e venaient de PE arrondies; la reference brute correcte est {fmt(excel_total / 1000)} tCO2e.
2. **Double comptage POC2026 corrige.** Notre import additionnait les colonnes d'utilisation 1 an et 7 ans, ajoutant artificiellement {fmt(excel_annual_use_rounded / 1000)} tCO2e.
3. **Melange de fins de vie POC2026 corrige.** Notre import additionnait aussi l'enfouissement et le scenario alternatif de recyclage ({fmt(recycling_rounded / 1000)} tCO2e). Le total detaille corrige vaut {fmt(detailed_total_rounded / 1000)} tCO2e et reproduit exactement la reference brute.
4. **Branche OPERA brute incorrecte.** Une tkm de siege embarque n'est pas une tkm de charge utile cargo. La quantite tkm n'est pas dix fois trop grande; c'est le facteur carburant qui est environ cinq fois trop grand.
5. **Deux parametres OPERA inutilisables et sans effet.** `fauteuil_cargo_conso_passive=42 700` vaut un dixieme des {fmt(annual_seat_tkm, 0)} tkm/an physiques, mais il peut s'agir soit d'un chiffre incomplet, soit d'une quantite equivalente calibree Sphera; l'unite manque. `fauteuil_kero_conso_passive` ne reproduit que 52,5 % des litres du Masterboard, car il remplace 200 passagers par 105 kg/passager. Aucun des deux ne pilote un echange.
6. **Resultats Sphera figes.** `Master!IA5=8 260` et `IB5=57 600` kgCO2e/an sont des constantes. Le fichier ne contient pas les quantites d'activite Sphera, la version de base, ni les facteurs permettant un recalcul independant.
7. **Ecart de base ACV reel.** A masse de carburant identique, OPERA/eecoinvent donne {fmt(physical_brightway_use / 1000)} tCO2e contre {fmt(excel_use_rounded / 1000)} tCO2e dans STELIA, soit {fmt((physical_brightway_use / excel_use_rounded - 1) * 100, 1)} % de plus.
8. **Configuration avion incoherente.** Le Masterboard nomme un A321XLR mais retient 95 t de MTOW. Airbus publie 101 a 101,5 t. Corriger uniquement ce point reduirait l'allocation massique d'environ {fmt((1 - mtow_kg / 101500) * 100, 1)} %.
9. **Hypotheses insuffisamment documentees.** Les 200 passagers, 2,5 L/100 pkm, 700 vols/an, taux de remplissage et partage passagers/cargo ne sont pas sources. La methode ICAO utilise des consommations par type et distance, des taux de remplissage et un partage passagers-cargo.
10. **Duree 30 ans obsolete mais sans effet numerique.** `Conso Vol!B4` vaut 30 ans, alors que l'etude finale utilise 7 ans. La division en B14 annule cette duree dans le resultat annuel, mais ce reliquat fragilise la tracabilite.
11. **Foreground non parametrique.** {disconnected_parameters} parametres ne pilotent aucun echange. Le paquet Brightway contient surtout des montants figes et des formules descriptives; modifier un parametre ne recalcule pas automatiquement ces montants.
12. **Branche passager mal cablee.** L'activite `consommation passive siege pkm` appelle le processus `1tkm`; elle ne constitue donc pas un calcul passager independant.
13. **Montant acier incoherent.** `Ensemble palette optimisee` stocke {fmt(number(palette_steel_exchange['amount']), 5)} kg pour une formule dont les parametres totalisent {fmt(palette_steel_formula_amount, 5)} kg, soit un facteur {fmt(number(palette_steel_exchange['amount']) / palette_steel_formula_amount, 3)}.
14. **Formules invalides.** {invalid_meal_table_formulas} echanges de `Ensemble tablette repas` contiennent des libelles d'activite dans une expression arithmetique; ils ne peuvent pas etre recalcules tels quels.
15. **Fin de vie a separer.** Avec ecoinvent cut-off, l'enfouissement, le recyclage cut-off et le recyclage avec credits de substitution sont des scenarios differents. Les additionner ou injecter des credits evites dans le scenario cut-off principal rend la comparaison methodologiquement incoherente.
16. **Production Brightway mal nommee dans certaines vues.** {fmt(non_use_brightway / 1000)} tCO2e correspond a la production moins {fmt(abs(recycling_credit_brightway) / 1000)} tCO2e de credit de recyclage. La production avant fin de vie vaut {fmt(production_brightway / 1000)} tCO2e.
17. **Masses non reconciliees.** Le siege vaut 109,967 kg pour l'usage, 114,320 kg dans la fin de vie et 123,309 kg dans la BOM recente hors emballages. L'usage et la production ne portent donc pas necessairement la meme configuration.
18. **IFE non auditable dans Excel et mal apparie dans Brightway.** STELIA contient {fmt(excel_ife, 1)} kgCO2e de resultats electroniques figes; Brightway en retrouve environ 305,3 avec des proxys dont les unites ne sont pas toutes homogenes.
19. **Gaz d'assemblage probablement mal unite.** Le libelle Excel indique 217,352 kWh/siege, mais l'impact ressemble a une saisie en MJ. La sous-estimation potentielle est d'environ 30 kgCO2e.
20. **Electronique Brightway a remapper.** La powerbox, les LED et le support ecran utilisent des unites ecoinvent comme des kilogrammes ou ajoutent un ecran deja modele. Environ 170 kgCO2e du support ecran sont suspects.
21. **Emballage potentiellement compte deux fois.** Les 53 kg d'emballage final apparaissent dans l'activite d'emballage et dans le transport vers Hambourg; le poste Brightway livraison/emballage atteint 120,6 kgCO2e contre {fmt(excel_delivery, 1)} dans STELIA.

## Valeur a retenir aujourd'hui

- **{fmt(excel_total / 1000)} tCO2e** : resultat brut historique STELIA/Sphera, utile pour comparaison mais non independant.
- **{fmt((physical_brightway_use + production_brightway) / 1000)} tCO2e avant fin de vie** : meilleure estimation actuelle compatible avec les facteurs Brightway/eecoinvent disponibles, en conservant la physique de carburant STELIA et sans credit de recyclage indu.
- **{fmt((marginal_fuel_central_kg * raw_factor_per_kg_fuel + production_brightway) / 1000)} tCO2e avant fin de vie**, avec une plage de **{fmt((marginal_fuel_low_kg * raw_factor_per_kg_fuel + production_brightway) / 1000)} a {fmt((marginal_fuel_high_kg * raw_factor_per_kg_fuel + production_brightway) / 1000)} tCO2e** : sensibilite marginale d'ecoconception, a ne pas presenter comme l'empreinte attributionnelle du siege.
- **{fmt((raw_use + non_use_brightway) / 1000)} tCO2e** : a rejeter, car il assimile le siege a du fret cargo.

## Modele correct a construire

Le foreground Brightway doit proposer deux racines explicites. La racine attributionnelle part d'une consommation avion par mission, puis applique le partage passagers/cargo et une regle d'allocation documentee. La racine ecoconception part d'une variation de masse et d'une courbe marginale de consommation. Dans les deux cas, la masse de Jet A-1 alimente une seule fois le marche regional de kerosene et les emissions de combustion. Les effets non-CO2 de l'aviation doivent etre presentes dans un scenario separe, car ils ne sont pas couverts par le simple GWP100 du carburant.
"""
    return scenarios, issues, bridge, report


def main() -> int:
    args = parse_args()
    output_root = args.output_root.resolve(strict=False)
    scenarios, issues, bridge, report = build_audit(output_root)
    data = output_root / "data"
    reports = output_root / "reports"
    reports.mkdir(parents=True, exist_ok=True)

    write_csv(data / "stelia_aircraft_use_scenarios.csv", scenarios)
    write_csv(data / "stelia_aircraft_use_audit_issues.csv", issues)
    write_csv(data / "brightway_raw_aircraft_use_diagnostic.csv", issues)
    write_csv(data / "brightway_raw_aircraft_use_bridge.csv", bridge)
    report_path = reports / "stelia_aircraft_use_audit.md"
    report_path.write_text(report, encoding="utf-8")
    (reports / "brightway_raw_aircraft_use_diagnostic.md").write_text(
        report,
        encoding="utf-8",
    )

    dashboard_path = output_root / "summaries" / "general_kpis.json"
    if dashboard_path.exists():
        dashboard = json.loads(dashboard_path.read_text(encoding="utf-8"))
        dashboard.setdefault("brightway_model", {})["raw_aircraft_use_audit"] = {
            "scenarios": scenarios,
            "issues": issues,
            "bridge": bridge,
            "report": "../reports/stelia_aircraft_use_audit.md",
        }
        dashboard_path.write_text(
            json.dumps(dashboard, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    print(f"Wrote {report_path}")
    print(f"Wrote {data / 'stelia_aircraft_use_scenarios.csv'}")
    print(f"Wrote {data / 'stelia_aircraft_use_audit_issues.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
