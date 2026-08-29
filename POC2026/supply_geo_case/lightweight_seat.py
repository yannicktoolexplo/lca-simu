"""Engineering screening model for the 50 percent lightweight seat concept."""

from __future__ import annotations

import math
import subprocess
import sys
import tempfile
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import yaml


INDICATOR_METHODS = {
    "Acidification": "acidification",
    "Climate Change - total": "climate_change",
    "Ecotoxicity, freshwater - total": "ecotoxicity_freshwater",
    "Resource use, fossils": "resource_use_fossils",
    "Eutrophication, freshwater": "eutrophication_freshwater",
    "Eutrophication, marine": "eutrophication_marine",
    "Eutrophication, terrestrial": "eutrophication_terrestrial",
    "Human toxicity, cancer - total": "human_toxicity_cancer",
    "Human toxicity, non-cancer - total": "human_toxicity_non_cancer",
    "Ionising radiation, human health": "ionising_radiation",
    "Land Use": "land_use",
    "Resource use, mineral and metals": "resource_use_minerals_metals",
    "Ozone depletion": "ozone_depletion",
    "Particulate matter": "particulate_matter",
    "Photochemical ozone formation, human health": "photochemical_ozone",
    "Water use": "water_use",
}

LOCALIZATION_SCENARIO_IDS = (
    "current_export",
    "france_first",
    "europe_first",
    "fully_globalized",
)


def is_exact_brightway_rows(rows: list[dict[str, Any]]) -> bool:
    """Return whether rows are a complete, usable Brightway result set."""
    if len(rows) != len(INDICATOR_METHODS):
        return False
    if {str(row.get("indicator_id") or "") for row in rows} != set(INDICATOR_METHODS):
        return False
    if not all(str(row.get("calculation_status") or "").startswith("brightway_exact") for row in rows):
        return False
    climate = next((row for row in rows if row.get("indicator_id") == "Climate Change - total"), {})
    return float(climate.get("fuel_factor_raw_per_kg") or 0.0) > 0.0


def is_exact_localization_rows(rows: list[dict[str, Any]]) -> bool:
    """Validate a complete four-scenario, 16-indicator result matrix."""
    if len(rows) != len(LOCALIZATION_SCENARIO_IDS) * len(INDICATOR_METHODS):
        return False
    for scenario_id in LOCALIZATION_SCENARIO_IDS:
        scenario_rows = [row for row in rows if row.get("sourcing_scenario_id") == scenario_id]
        if {str(row.get("indicator_id") or "") for row in scenario_rows} != set(INDICATOR_METHODS):
            return False
        if not all(str(row.get("calculation_status") or "").startswith("brightway_exact") for row in scenario_rows):
            return False
    climate = next(
        (
            row
            for row in rows
            if row.get("sourcing_scenario_id") == "current_export"
            and row.get("indicator_id") == "Climate Change - total"
        ),
        {},
    )
    return float(climate.get("fuel_factor_raw_per_kg") or 0.0) > 0.0


def _runtime_python(runtime: dict[str, Any]) -> str:
    external = runtime.get("external_python") if isinstance(runtime.get("external_python"), dict) else {}
    current = runtime.get("current_python") if isinstance(runtime.get("current_python"), dict) else {}
    python = str(runtime.get("python") or "")
    if not python and external.get("available"):
        python = str(external.get("python") or "")
    if not python and current.get("available"):
        python = sys.executable
    return python


def ascii_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join(
        "".join(char for char in text if not unicodedata.combining(char))
        .lower()
        .replace("_", " ")
        .replace("-", " ")
        .split()
    )


def load_scenario_config(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def classify_family(name: str, config: dict[str, Any]) -> str:
    key = ascii_key(name)
    for family_id, spec in config.get("families", {}).items():
        tokens = [ascii_key(token) for token in spec.get("bom_tokens", [])]
        if any(token and token in key for token in tokens):
            return family_id
    return ""


def _fallback_family_masses(config: dict[str, Any]) -> dict[str, float]:
    baseline = float(config.get("baseline_mass_kg") or 0.0)
    return {
        family_id: baseline * float(spec.get("fallback_mass_share") or 0.0)
        for family_id, spec in config.get("families", {}).items()
    }


def extract_reconciled_mass_budget(
    masterboard_path: Path,
    config: dict[str, Any],
) -> tuple[dict[str, float], dict[str, Any]]:
    """Read product masses, remove logistics rows, then reconcile to OPERA mass."""

    baseline_mass = float(config.get("baseline_mass_kg") or 0.0)
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(masterboard_path, read_only=True, data_only=True)
        try:
            sheet = workbook["BOM"]
            raw: dict[str, float] = defaultdict(float)
            unmatched: dict[str, float] = defaultdict(float)
            excluded_prefixes = (
                "consommation",
                "energie",
                "donnees",
                "nettoyage",
                "papier/",
                "packaging",
            )
            for row in sheet.iter_rows(min_row=2, values_only=True):
                equipment = str(row[0] or "").strip()
                part_number = str(row[1] or "").strip()
                amount = row[6]
                if not isinstance(amount, (int, float)) or not math.isfinite(amount) or amount <= 0.0:
                    continue
                equipment_key = ascii_key(equipment)
                if part_number.upper() == "PACKAGING" or equipment_key.startswith(excluded_prefixes):
                    continue
                family_id = classify_family(equipment, config)
                if family_id:
                    raw[family_id] += float(amount)
                else:
                    unmatched[equipment] += float(amount)
        finally:
            workbook.close()
        raw_total = sum(raw.values())
        if raw_total <= 0.0 or unmatched:
            raise ValueError(f"BOM mass classification incomplete: {dict(unmatched)}")
        reconciliation_factor = baseline_mass / raw_total
        masses = {family_id: value * reconciliation_factor for family_id, value in raw.items()}
        return masses, {
            "source": str(masterboard_path),
            "status": "masterboard_reconciled_to_opera_mass",
            "raw_bom_mass_kg": round(raw_total, 9),
            "baseline_mass_kg": round(baseline_mass, 9),
            "reconciliation_factor": round(reconciliation_factor, 12),
            "unmatched_equipment": [],
        }
    except Exception as exc:
        return _fallback_family_masses(config), {
            "source": str(masterboard_path),
            "status": "fallback_configured_mass_shares",
            "raw_bom_mass_kg": "",
            "baseline_mass_kg": round(baseline_mass, 9),
            "reconciliation_factor": "",
            "unmatched_equipment": [],
            "fallback_reason": str(exc),
        }


def build_mass_budget_rows(
    family_masses: dict[str, float],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for family_id, spec in config.get("families", {}).items():
        baseline = float(family_masses.get(family_id) or 0.0)
        reduction = float(spec.get("reduction_fraction") or 0.0)
        target = baseline * (1.0 - reduction)
        process_multiplier = float(spec.get("process_complexity_multiplier") or 1.0)
        rows.append(
            {
                "scenario_id": config.get("scenario_id"),
                "family_id": family_id,
                "family_label": spec.get("label"),
                "baseline_mass_kg": round(baseline, 6),
                "target_mass_kg": round(target, 6),
                "mass_saved_kg": round(baseline - target, 6),
                "reduction_pct": round(100.0 * reduction, 4),
                "retained_mass_fraction": round(1.0 - reduction, 9),
                "process_complexity_multiplier": round(process_multiplier, 6),
                "lca_exchange_scale_factor": round((1.0 - reduction) * process_multiplier, 9),
                "strategy": spec.get("strategy"),
                "candidate_materials": spec.get("candidate_materials"),
                "maturity": spec.get("maturity"),
                "confidence": spec.get("confidence"),
                "verification": spec.get("verification"),
            }
        )
    target_config = float(config.get("target_mass_kg") or 0.0)
    target_computed = sum(float(row["target_mass_kg"]) for row in rows)
    if rows and abs(target_computed - target_config) > 0.0001:
        correction = target_config - target_computed
        rows[-1]["target_mass_kg"] = round(float(rows[-1]["target_mass_kg"]) + correction, 6)
        rows[-1]["mass_saved_kg"] = round(
            float(rows[-1]["baseline_mass_kg"]) - float(rows[-1]["target_mass_kg"]),
            6,
        )
        retained = float(rows[-1]["target_mass_kg"]) / max(float(rows[-1]["baseline_mass_kg"]), 1e-12)
        rows[-1]["retained_mass_fraction"] = round(retained, 9)
        rows[-1]["reduction_pct"] = round(100.0 * (1.0 - retained), 4)
        rows[-1]["lca_exchange_scale_factor"] = round(
            retained * float(rows[-1]["process_complexity_multiplier"]),
            9,
        )
    return rows


def method_indicator_id(method_text: str) -> str:
    key = ascii_key(method_text)
    mapping = (
        ("acidification", "Acidification"),
        ("climate change |", "Climate Change - total"),
        ("ecotoxicity freshwater |", "Ecotoxicity, freshwater - total"),
        ("energy resources non renewable", "Resource use, fossils"),
        ("eutrophication freshwater", "Eutrophication, freshwater"),
        ("eutrophication marine", "Eutrophication, marine"),
        ("eutrophication terrestrial", "Eutrophication, terrestrial"),
        ("human toxicity carcinogenic |", "Human toxicity, cancer - total"),
        ("human toxicity non carcinogenic |", "Human toxicity, non-cancer - total"),
        ("ionising radiation human health", "Ionising radiation, human health"),
        ("land use", "Land Use"),
        ("material resources metals/minerals", "Resource use, mineral and metals"),
        ("ozone depletion", "Ozone depletion"),
        ("particulate matter formation", "Particulate matter"),
        ("photochemical oxidant formation human health", "Photochemical ozone formation, human health"),
        ("water use", "Water use"),
    )
    for token, label in mapping:
        if ascii_key(token) in key:
            return label
    return ""


def _indicator_metadata(
    indicator_unit_views: list[dict[str, Any]],
    reference_person_equivalent_results: list[dict[str, Any]],
    reference_weighting_factors: list[dict[str, Any]],
) -> dict[str, dict[str, float | str]]:
    metadata: dict[str, dict[str, float | str]] = {}
    for row in indicator_unit_views:
        if not row.get("include_in_person_equivalent"):
            continue
        label = str(row.get("short_label") or "")
        metadata[label] = {
            "normalization_factor": float(row.get("normalization_factor_per_person_year") or 0.0),
            "raw_unit": str(row.get("raw_unit") or ""),
            "use_phase_person_equivalent": 0.0,
            "weight_fraction": 0.0,
        }
    for row in reference_person_equivalent_results:
        label = str(row.get("short_label") or "")
        if label in metadata:
            metadata[label]["use_phase_person_equivalent"] = float(row.get("use_phase_person_equivalent") or 0.0)

    category_tokens = {
        "Climate change": "Climate Change - total",
        "Ozone": "Ozone depletion",
        "Cancer": "Human toxicity, cancer - total",
        "Non cancer": "Human toxicity, non-cancer - total",
        "Particulate": "Particulate matter",
        "Ionising": "Ionising radiation, human health",
        "Photochemical": "Photochemical ozone formation, human health",
        "Minerals and metal": "Resource use, mineral and metals",
        "Energy carrier ressource": "Resource use, fossils",
        "Water": "Water use",
        "Freshwater ecotox": "Ecotoxicity, freshwater - total",
        "Land use": "Land Use",
        "Marine Eutroph": "Eutrophication, marine",
        "Freshwater Eutroph": "Eutrophication, freshwater",
        "Terrestrial Eutroph": "Eutrophication, terrestrial",
        "Acidification": "Acidification",
    }
    for row in reference_weighting_factors:
        label = category_tokens.get(str(row.get("category") or ""))
        if label in metadata:
            metadata[label]["weight_fraction"] = float(row.get("ef30_weight_pct") or 0.0) / 100.0
    return metadata


def _fallback_exact_rows(
    impact_rows: list[dict[str, Any]],
    mass_rows: list[dict[str, Any]],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    scales = {row["family_id"]: float(row["lca_exchange_scale_factor"]) for row in mass_rows}
    grouped: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0])
    for row in impact_rows:
        if str(row.get("family")) not in {"Sous-ensemble", "IFE", "Site assemblage", "Livraison"}:
            continue
        indicator = str(row.get("indicator") or "")
        short_label = indicator.removeprefix("EF 3.0 ")
        if " [" in short_label:
            short_label = short_label.rsplit(" [", 1)[0]
        if short_label not in INDICATOR_METHODS:
            continue
        value = float(row.get("value") or 0.0)
        family = str(row.get("family") or "")
        if family == "Site assemblage":
            scale = float(config.get("overhead_scales", {}).get("energie infrastructures", 0.82))
        elif family == "Livraison":
            scale = float(config.get("overhead_scales", {}).get("transport vers hambourg", 0.50))
        else:
            family_id = classify_family(str(row.get("system") or ""), config)
            scale = scales.get(family_id, 0.60)
        grouped[short_label][0] += value
        grouped[short_label][1] += value * scale
    return [
        {
            "indicator_id": label,
            "method": "fallback detailed workbook",
            "raw_unit": "",
            "baseline_production_raw": values[0],
            "lightweight_production_raw": values[1],
            "fuel_factor_raw_per_kg": 0.0,
            "calculation_status": "screening_detailed_workbook_scaled",
        }
        for label, values in grouped.items()
    ]


def run_exact_brightway(
    *,
    runtime: dict[str, Any],
    config_path: Path,
    runner_path: Path,
) -> list[dict[str, Any]]:
    persisted_path = runner_path.parents[1] / "outputs" / "data" / "lightweight_seat_exact_lcia.csv"
    if persisted_path.exists() and persisted_path.stat().st_mtime >= max(
        config_path.stat().st_mtime,
        runner_path.stat().st_mtime,
    ):
        import csv

        with persisted_path.open("r", encoding="utf-8", newline="") as stream:
            persisted_rows = list(csv.DictReader(stream))
        if is_exact_brightway_rows(persisted_rows):
            return persisted_rows
    if not runtime.get("can_execute_brightway") or not runner_path.exists():
        return []
    external = runtime.get("external_python") if isinstance(runtime.get("external_python"), dict) else {}
    current = runtime.get("current_python") if isinstance(runtime.get("current_python"), dict) else {}
    python = str(runtime.get("python") or "")
    if not python and external.get("available"):
        python = str(external.get("python") or "")
    if not python and current.get("available"):
        python = sys.executable
    if not python:
        return []
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as handle:
        output_path = Path(handle.name)
    try:
        command = [python, str(runner_path), "--config", str(config_path), "--output-csv", str(output_path)]
        completed = subprocess.run(command, capture_output=True, text=True, timeout=900, check=False)
        if completed.returncode != 0 or not output_path.exists():
            return [{
                "indicator_id": "runtime_error",
                "calculation_status": "brightway_runtime_error",
                "error": (completed.stderr or completed.stdout or "unknown Brightway error")[-2000:],
            }]
        import csv

        with output_path.open("r", encoding="utf-8", newline="") as stream:
            rows = list(csv.DictReader(stream))
        if not is_exact_brightway_rows(rows):
            return [{
                "indicator_id": "runtime_error",
                "calculation_status": "brightway_runtime_error",
                "error": "Brightway runner returned an incomplete or unusable result set",
            }]
        return rows
    finally:
        output_path.unlink(missing_ok=True)


def run_exact_localization_scenarios(
    *,
    runtime: dict[str, Any],
    config_path: Path,
    runner_path: Path | None,
) -> list[dict[str, Any]]:
    if runner_path is None:
        return []
    persisted_path = runner_path.parents[1] / "outputs" / "data" / "lightweight_seat_localization_exact_lcia.csv"
    if persisted_path.exists() and persisted_path.stat().st_mtime >= max(
        config_path.stat().st_mtime,
        runner_path.stat().st_mtime,
    ):
        import csv

        with persisted_path.open("r", encoding="utf-8", newline="") as stream:
            persisted_rows = list(csv.DictReader(stream))
        if is_exact_localization_rows(persisted_rows):
            return persisted_rows
    if not runtime.get("can_execute_brightway") or not runner_path.exists():
        return []
    python = _runtime_python(runtime)
    if not python:
        return []
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as handle:
        output_path = Path(handle.name)
    try:
        command = [python, str(runner_path), "--config", str(config_path), "--output-csv", str(output_path)]
        completed = subprocess.run(command, capture_output=True, text=True, timeout=1800, check=False)
        if completed.returncode != 0 or not output_path.exists():
            return []
        import csv

        with output_path.open("r", encoding="utf-8", newline="") as stream:
            rows = list(csv.DictReader(stream))
        return rows if is_exact_localization_rows(rows) else []
    finally:
        output_path.unlink(missing_ok=True)


def is_exact_named_supplier_rows(rows: list[dict[str, Any]], scenario_ids: list[str]) -> bool:
    if not scenario_ids or len(rows) != len(scenario_ids) * len(INDICATOR_METHODS):
        return False
    for scenario_id in scenario_ids:
        scenario_rows = [row for row in rows if row.get("sourcing_scenario_id") == scenario_id]
        if {str(row.get("indicator_id") or "") for row in scenario_rows} != set(INDICATOR_METHODS):
            return False
        if not all(str(row.get("calculation_status") or "").startswith("brightway_exact") for row in scenario_rows):
            return False
    return True


def run_exact_named_supplier_scenarios(
    *,
    runtime: dict[str, Any],
    config_path: Path,
    runner_path: Path | None,
    supplier_scenarios: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if runner_path is None or not supplier_scenarios:
        return []
    scenario_ids = [str(row.get("scenario_id") or "") for row in supplier_scenarios if row.get("scenario_id")]
    persisted_path = runner_path.parents[1] / "outputs" / "data" / "lightweight_seat_named_supplier_exact_lcia.csv"
    if persisted_path.exists() and persisted_path.stat().st_mtime >= max(
        config_path.stat().st_mtime,
        runner_path.stat().st_mtime,
    ):
        import csv

        with persisted_path.open("r", encoding="utf-8", newline="") as stream:
            persisted_rows = list(csv.DictReader(stream))
        factors = {
            str(row.get("scenario_id")): round(float(row.get("transport_amount_factor") or 1.0), 9)
            for row in supplier_scenarios
        }
        cache_factors = {
            str(row.get("sourcing_scenario_id")): round(float(row.get("transport_amount_factor") or 1.0), 9)
            for row in persisted_rows
        }
        if is_exact_named_supplier_rows(persisted_rows, scenario_ids) and all(cache_factors.get(key) == value for key, value in factors.items()):
            return persisted_rows
    if not runtime.get("can_execute_brightway") or not runner_path.exists():
        return []
    python = _runtime_python(runtime)
    if not python:
        return []
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as scenario_handle:
        scenario_path = Path(scenario_handle.name)
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as output_handle:
        output_path = Path(output_handle.name)
    try:
        import csv

        fieldnames: list[str] = []
        for row in supplier_scenarios:
            for key in row:
                if key not in fieldnames:
                    fieldnames.append(key)
        with scenario_path.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(supplier_scenarios)
        command = [
            python,
            str(runner_path),
            "--config",
            str(config_path),
            "--supplier-scenarios-csv",
            str(scenario_path),
            "--output-csv",
            str(output_path),
        ]
        completed = subprocess.run(command, capture_output=True, text=True, timeout=1800, check=False)
        if completed.returncode != 0 or not output_path.exists():
            return []
        with output_path.open("r", encoding="utf-8", newline="") as stream:
            rows = list(csv.DictReader(stream))
        return rows if is_exact_named_supplier_rows(rows, scenario_ids) else []
    finally:
        scenario_path.unlink(missing_ok=True)
        output_path.unlink(missing_ok=True)


def build_localization_results(
    *,
    exact_rows: list[dict[str, Any]],
    metadata: dict[str, dict[str, Any]],
    avoided_fuel: dict[str, float],
    regional_scenarios: list[dict[str, Any]],
    scenario_ids: list[str] | tuple[str, ...] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    profiles = {str(row.get("scenario_id") or ""): row for row in regional_scenarios}
    indicator_rows: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []

    ordered_scenario_ids = list(scenario_ids or LOCALIZATION_SCENARIO_IDS)
    for sourcing_id in ordered_scenario_ids:
        source_rows = [row for row in exact_rows if row.get("sourcing_scenario_id") == sourcing_id]
        scenario_indicators: list[dict[str, Any]] = []
        for exact in source_rows:
            indicator_id = str(exact.get("indicator_id") or "")
            if indicator_id not in metadata:
                continue
            meta = metadata[indicator_id]
            norm = float(meta.get("normalization_factor") or 0.0)
            weight = float(meta.get("weight_fraction") or 0.0)
            baseline_production = float(exact.get("baseline_production_raw") or 0.0)
            scenario_production = float(exact.get("lightweight_production_raw") or 0.0)
            baseline_use = float(meta.get("use_phase_person_equivalent") or 0.0) * norm
            fuel_factor = float(exact.get("fuel_factor_raw_per_kg") or 0.0)
            avoided_use = min(baseline_use, avoided_fuel.get("central", 0.0) * fuel_factor)
            scenario_use = baseline_use - avoided_use
            baseline_total = baseline_production + baseline_use
            scenario_total = scenario_production + scenario_use
            row = {
                "scenario_id": exact.get("scenario_id"),
                "sourcing_scenario_id": sourcing_id,
                "label": exact.get("label"),
                "indicator_id": indicator_id,
                "indicator_key": INDICATOR_METHODS[indicator_id],
                "raw_unit": exact.get("raw_unit") or meta.get("raw_unit"),
                "baseline_production_raw": round(baseline_production, 12),
                "localized_lightweight_production_raw": round(scenario_production, 12),
                "localized_lightweight_total_central_raw": round(scenario_total, 12),
                "localized_lightweight_total_central_person_equivalent": round(scenario_total / norm, 12) if norm else "",
                "localized_lightweight_total_central_weighted_point": round(scenario_total / norm * weight, 12) if norm else "",
                "baseline_total_central_weighted_point": round(baseline_total / norm * weight, 12) if norm else "",
                "calculation_status": exact.get("calculation_status"),
            }
            scenario_indicators.append(row)
            indicator_rows.append(row)

        climate = next(
            (row for row in scenario_indicators if row.get("indicator_id") == "Climate Change - total"),
            {},
        )
        weighted_baseline = sum(float(row.get("baseline_total_central_weighted_point") or 0.0) for row in scenario_indicators)
        weighted_scenario = sum(float(row.get("localized_lightweight_total_central_weighted_point") or 0.0) for row in scenario_indicators)
        first = source_rows[0] if source_rows else {}
        profile = profiles.get(sourcing_id, {})
        summaries.append({
            "scenario_id": first.get("scenario_id"),
            "sourcing_scenario_id": sourcing_id,
            "label": first.get("label"),
            "target_scope": first.get("target_scope"),
            "localization_status": first.get("localization_status"),
            "description": profile.get("description", ""),
            "electricity_scope": first.get("elec_switch_param") or "supply actuelle",
            "aluminium_scope": first.get("al_switch_param") or "supply actuelle",
            "transport_amount_factor": round(float(first.get("transport_amount_factor") or 1.0), 6),
            "local_content_objective_pct": profile.get("local_content_target_pct", ""),
            "current_role_mass_already_target_pct": profile.get("current_role_mass_already_target_pct", ""),
            "current_path_mass_already_target_pct": profile.get("current_path_mass_already_target_pct", ""),
            "electricity_replacements": int(float(first.get("electricity_replacements") or 0)),
            "aluminium_replacements": int(float(first.get("aluminium_replacements") or 0)),
            "transport_scaled_exchanges": int(float(first.get("transport_scaled_exchanges") or 0)),
            "production_climate_kgco2e": climate.get("localized_lightweight_production_raw", ""),
            "cycle_climate_central_kgco2e": climate.get("localized_lightweight_total_central_raw", ""),
            "weighted_point": round(weighted_scenario, 9),
            "weighted_reduction_vs_reference_pct": round(
                100.0 * (weighted_baseline - weighted_scenario) / weighted_baseline,
                6,
            ) if weighted_baseline else "",
            "indicator_count": len(scenario_indicators),
            "calculation_status": first.get("calculation_status", ""),
        })

    reference_scenario_id = "current_export" if "current_export" in ordered_scenario_ids else ordered_scenario_ids[0] if ordered_scenario_ids else ""
    current_production = {
        row["indicator_id"]: float(row.get("localized_lightweight_production_raw") or 0.0)
        for row in indicator_rows
        if row.get("sourcing_scenario_id") == reference_scenario_id
    }
    for row in indicator_rows:
        reference = current_production.get(str(row.get("indicator_id") or ""), 0.0)
        value = float(row.get("localized_lightweight_production_raw") or 0.0)
        row["production_delta_vs_lightweight_current_pct"] = round(
            100.0 * (value - reference) / abs(reference),
            6,
        ) if reference else ""
    current_summary = next((row for row in summaries if row.get("sourcing_scenario_id") == reference_scenario_id), {})
    current_climate = float(current_summary.get("production_climate_kgco2e") or 0.0)
    current_weighted = float(current_summary.get("weighted_point") or 0.0)
    for row in summaries:
        production = float(row.get("production_climate_kgco2e") or 0.0)
        weighted = float(row.get("weighted_point") or 0.0)
        row["production_climate_delta_vs_lightweight_current_pct"] = round(
            100.0 * (production - current_climate) / abs(current_climate),
            6,
        ) if current_climate else ""
        row["weighted_delta_vs_lightweight_current_pct"] = round(
            100.0 * (weighted - current_weighted) / abs(current_weighted),
            6,
        ) if current_weighted else ""
    return indicator_rows, summaries


def build_lightweight_scenario(
    *,
    config_path: Path,
    masterboard_path: Path,
    runtime: dict[str, Any],
    runner_path: Path,
    impact_rows: list[dict[str, Any]],
    indicator_unit_views: list[dict[str, Any]],
    reference_person_equivalent_results: list[dict[str, Any]],
    reference_weighting_factors: list[dict[str, Any]],
    localization_runner_path: Path | None = None,
    regional_scenarios: list[dict[str, Any]] | None = None,
    supplier_alternative_payload: dict[str, Any] | None = None,
    supplier_runner_path: Path | None = None,
) -> dict[str, Any]:
    config = load_scenario_config(config_path)
    family_masses, reconciliation = extract_reconciled_mass_budget(masterboard_path, config)
    mass_rows = build_mass_budget_rows(family_masses, config)
    exact_rows = run_exact_brightway(runtime=runtime, config_path=config_path, runner_path=runner_path)
    runtime_error = next((row for row in exact_rows if row.get("indicator_id") == "runtime_error"), None)
    if not exact_rows or runtime_error:
        exact_rows = _fallback_exact_rows(impact_rows, mass_rows, config)
    metadata = _indicator_metadata(
        indicator_unit_views,
        reference_person_equivalent_results,
        reference_weighting_factors,
    )

    flight = config.get("flight_use", {})
    baseline_mass = float(config.get("baseline_mass_kg") or 0.0)
    target_mass = float(config.get("target_mass_kg") or 0.0)
    mass_saved = baseline_mass - target_mass
    lifetime_distance = (
        float(flight.get("average_flight_distance_km") or 0.0)
        * float(flight.get("annual_flight_cycles") or 0.0)
        * float(flight.get("lifetime_years") or 0.0)
    )
    fuel_coefficients = flight.get("marginal_fuel_kg_per_kg_1000km", {})
    avoided_fuel = {
        bound: mass_saved * lifetime_distance / 1000.0 * float(coefficient or 0.0)
        for bound, coefficient in fuel_coefficients.items()
    }
    localization_exact_rows = run_exact_localization_scenarios(
        runtime=runtime,
        config_path=config_path,
        runner_path=localization_runner_path,
    )
    localization_indicator_rows, localization_summaries = build_localization_results(
        exact_rows=localization_exact_rows,
        metadata=metadata,
        avoided_fuel=avoided_fuel,
        regional_scenarios=regional_scenarios or [],
    )
    supplier_payload = supplier_alternative_payload or {}
    supplier_scenario_inputs = supplier_payload.get("scenario_summaries", [])
    supplier_exact_rows = run_exact_named_supplier_scenarios(
        runtime=runtime,
        config_path=config_path,
        runner_path=supplier_runner_path,
        supplier_scenarios=supplier_scenario_inputs,
    )
    supplier_scenario_ids = [str(row.get("scenario_id") or "") for row in supplier_scenario_inputs if row.get("scenario_id")]
    supplier_indicator_rows, supplier_lcia_summaries = build_localization_results(
        exact_rows=supplier_exact_rows,
        metadata=metadata,
        avoided_fuel=avoided_fuel,
        regional_scenarios=supplier_scenario_inputs,
        scenario_ids=supplier_scenario_ids,
    )
    supplier_lcia_by_id = {str(row.get("sourcing_scenario_id") or ""): row for row in supplier_lcia_summaries}
    supplier_summaries = [
        {
            **selection,
            **supplier_lcia_by_id.get(str(selection.get("scenario_id") or ""), {}),
            "scenario_id": selection.get("scenario_id"),
        }
        for selection in supplier_scenario_inputs
    ]
    current_lightweight_production = {
        str(row.get("indicator_id") or ""): float(row.get("localized_lightweight_production_raw") or 0.0)
        for row in localization_indicator_rows
        if row.get("sourcing_scenario_id") == "current_export"
    }
    for row in supplier_indicator_rows:
        reference = current_lightweight_production.get(str(row.get("indicator_id") or ""), 0.0)
        value = float(row.get("localized_lightweight_production_raw") or 0.0)
        row["production_delta_vs_lightweight_current_pct"] = round(
            100.0 * (value - reference) / abs(reference),
            6,
        ) if reference else ""
    current_lightweight_summary = next(
        (row for row in localization_summaries if row.get("sourcing_scenario_id") == "current_export"),
        {},
    )
    current_lightweight_climate = float(current_lightweight_summary.get("production_climate_kgco2e") or 0.0)
    current_lightweight_weighted = float(current_lightweight_summary.get("weighted_point") or 0.0)
    for row in supplier_summaries:
        production = float(row.get("production_climate_kgco2e") or 0.0)
        weighted = float(row.get("weighted_point") or 0.0)
        row["production_climate_delta_vs_lightweight_current_pct"] = round(
            100.0 * (production - current_lightweight_climate) / abs(current_lightweight_climate),
            6,
        ) if current_lightweight_climate else ""
        row["weighted_delta_vs_lightweight_current_pct"] = round(
            100.0 * (weighted - current_lightweight_weighted) / abs(current_lightweight_weighted),
            6,
        ) if current_lightweight_weighted else ""

    indicator_rows: list[dict[str, Any]] = []
    for exact in exact_rows:
        indicator_id = str(exact.get("indicator_id") or "")
        if indicator_id not in metadata:
            continue
        meta = metadata[indicator_id]
        norm = float(meta.get("normalization_factor") or 0.0)
        weight = float(meta.get("weight_fraction") or 0.0)
        baseline_production = float(exact.get("baseline_production_raw") or 0.0)
        scenario_production = float(exact.get("lightweight_production_raw") or 0.0)
        baseline_use = float(meta.get("use_phase_person_equivalent") or 0.0) * norm
        fuel_factor = float(exact.get("fuel_factor_raw_per_kg") or 0.0)
        row: dict[str, Any] = {
            "scenario_id": config.get("scenario_id"),
            "indicator_id": indicator_id,
            "indicator_key": INDICATOR_METHODS[indicator_id],
            "raw_unit": exact.get("raw_unit") or meta.get("raw_unit"),
            "baseline_production_raw": round(baseline_production, 12),
            "lightweight_production_raw": round(scenario_production, 12),
            "production_delta_raw": round(scenario_production - baseline_production, 12),
            "production_reduction_pct": round(
                100.0 * (baseline_production - scenario_production) / baseline_production,
                6,
            ) if baseline_production else "",
            "baseline_use_raw": round(baseline_use, 12),
            "fuel_factor_raw_per_kg": round(fuel_factor, 12),
            "normalization_factor_per_person_year": round(norm, 12),
            "ef30_weight_fraction": round(weight, 9),
            "calculation_status": exact.get("calculation_status"),
        }
        for bound in ("low", "central", "high"):
            avoided_use = min(baseline_use, avoided_fuel.get(bound, 0.0) * fuel_factor)
            scenario_use = baseline_use - avoided_use
            combined_baseline = baseline_production + baseline_use
            combined_scenario = scenario_production + scenario_use
            row[f"avoided_fuel_{bound}_kg"] = round(avoided_fuel.get(bound, 0.0), 6)
            row[f"avoided_use_{bound}_raw"] = round(avoided_use, 12)
            row[f"lightweight_use_{bound}_raw"] = round(scenario_use, 12)
            row[f"baseline_total_{bound}_raw"] = round(combined_baseline, 12)
            row[f"lightweight_total_{bound}_raw"] = round(combined_scenario, 12)
            row[f"total_delta_{bound}_raw"] = round(combined_scenario - combined_baseline, 12)
            row[f"baseline_total_{bound}_person_equivalent"] = round(combined_baseline / norm, 12) if norm else ""
            row[f"lightweight_total_{bound}_person_equivalent"] = round(combined_scenario / norm, 12) if norm else ""
            row[f"baseline_total_{bound}_weighted_point"] = round(combined_baseline / norm * weight, 12) if norm else ""
            row[f"lightweight_total_{bound}_weighted_point"] = round(combined_scenario / norm * weight, 12) if norm else ""
        indicator_rows.append(row)

    weighted_baseline = sum(float(row.get("baseline_total_central_weighted_point") or 0.0) for row in indicator_rows)
    weighted_scenario = sum(float(row.get("lightweight_total_central_weighted_point") or 0.0) for row in indicator_rows)
    climate = next((row for row in indicator_rows if row.get("indicator_id") == "Climate Change - total"), {})
    mass_baseline_computed = sum(float(row.get("baseline_mass_kg") or 0.0) for row in mass_rows)
    mass_target_computed = sum(float(row.get("target_mass_kg") or 0.0) for row in mass_rows)
    gates = [
        {
            "scenario_id": config.get("scenario_id"),
            **gate,
        }
        for gate in config.get("certification_gates", [])
    ]
    return {
        "schema_version": "poc2026.lightweight_seat.results.v1",
        "scenario_id": config.get("scenario_id"),
        "label": config.get("label"),
        "status": config.get("status"),
        "functional_unit": config.get("functional_unit"),
        "mass_reconciliation": reconciliation,
        "mass_budget": mass_rows,
        "certification_gates": gates,
        "evidence_sources": config.get("evidence_sources", []),
        "indicator_results": indicator_rows,
        "exact_runtime_rows": exact_rows,
        "localization_exact_runtime_rows": localization_exact_rows,
        "localization_indicator_results": localization_indicator_rows,
        "localization_scenarios": localization_summaries,
        "named_supplier_exact_runtime_rows": supplier_exact_rows,
        "named_supplier_indicator_results": supplier_indicator_rows,
        "named_supplier_scenarios": supplier_summaries,
        "named_supplier_assignments": supplier_payload.get("assignments", []),
        "named_supplier_routes": supplier_payload.get("routes", []),
        "named_supplier_candidate_audit": supplier_payload.get("candidate_audit", []),
        "named_supplier_loads": supplier_payload.get("supplier_loads", []),
        "summary": {
            "baseline_mass_kg": round(mass_baseline_computed, 6),
            "target_mass_kg": round(mass_target_computed, 6),
            "mass_saved_kg": round(mass_baseline_computed - mass_target_computed, 6),
            "mass_reduction_pct": round(100.0 * (mass_baseline_computed - mass_target_computed) / mass_baseline_computed, 4),
            "lifetime_distance_km": round(lifetime_distance, 3),
            "avoided_fuel_low_kg": round(avoided_fuel.get("low", 0.0), 6),
            "avoided_fuel_central_kg": round(avoided_fuel.get("central", 0.0), 6),
            "avoided_fuel_high_kg": round(avoided_fuel.get("high", 0.0), 6),
            "baseline_production_kgco2e": climate.get("baseline_production_raw", ""),
            "lightweight_production_kgco2e": climate.get("lightweight_production_raw", ""),
            "baseline_use_kgco2e": climate.get("baseline_use_raw", ""),
            "lightweight_use_central_kgco2e": climate.get("lightweight_use_central_raw", ""),
            "avoided_use_low_kgco2e": climate.get("avoided_use_low_raw", ""),
            "avoided_use_central_kgco2e": climate.get("avoided_use_central_raw", ""),
            "avoided_use_high_kgco2e": climate.get("avoided_use_high_raw", ""),
            "baseline_total_central_kgco2e": climate.get("baseline_total_central_raw", ""),
            "lightweight_total_central_kgco2e": climate.get("lightweight_total_central_raw", ""),
            "weighted_baseline_point": round(weighted_baseline, 9),
            "weighted_lightweight_point": round(weighted_scenario, 9),
            "weighted_reduction_pct": round(100.0 * (weighted_baseline - weighted_scenario) / weighted_baseline, 6) if weighted_baseline else "",
            "indicator_count": len(indicator_rows),
            "localization_scenario_count": len(localization_summaries),
            "localization_indicator_count": len(localization_indicator_rows),
            "named_supplier_scenario_count": len(supplier_summaries),
            "named_supplier_indicator_count": len(supplier_indicator_rows),
            "named_supplier_assignment_count": len(supplier_payload.get("assignments", [])),
            "calculation_status": "brightway_exact_foreground_scaled" if exact_rows and not str(exact_rows[0].get("calculation_status", "")).startswith("screening") else "screening_detailed_workbook_scaled",
            "certification_status": "concept_non_certifie",
        },
    }
