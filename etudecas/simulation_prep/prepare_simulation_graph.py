#!/usr/bin/env python3
"""
Prepare a simulation-ready copy of a geocoded supply graph.

Outputs:
- simulation-ready graph JSON
- preparation report JSON
- preparation report Markdown
"""

from __future__ import annotations

import argparse
import json
import math
import re
import subprocess
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from statistics import median
from typing import Any

PRODUCT_SERVICE_TARGETS = {
    "item:268967": 0.80,
    "item:268091": 0.93,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare simulation-ready supply graph.")
    parser.add_argument(
        "--input",
        default="etudecas/result_geocodage/supply_graph_poc_geocoded.json",
        help="Input geocoded graph JSON.",
    )
    parser.add_argument(
        "--output-graph",
        default="etudecas/simulation_prep/result/reference_baseline/supply_graph_reference_baseline_simulation_ready.json",
        help="Output simulation-ready graph JSON.",
    )
    parser.add_argument(
        "--output-report-json",
        default="etudecas/simulation_prep/result/reference_baseline/simulation_prep_report.json",
        help="Output prep report JSON.",
    )
    parser.add_argument(
        "--output-report-md",
        default="etudecas/simulation_prep/result/reference_baseline/simulation_prep_report.md",
        help="Output prep report Markdown.",
    )
    parser.add_argument(
        "--data-poc-xlsx",
        default="etudecas/donnees/Data_poc.xlsx",
        help="Optional Data_poc.xlsx path used to align material prices from Relations_acteurs.",
    )
    parser.add_argument(
        "--demand-pf-xlsx",
        default="etudecas/donnees/demand_PF.xlsx",
        help="Optional demand_PF.xlsx path used to load weekly customer demand.",
    )
    parser.add_argument(
        "--customer-warm-start-days",
        type=float,
        default=0.0,
        help="Optional customer warm-start stock buffer in demand-days (default: 0).",
    )
    parser.add_argument(
        "--upstream-dc-warm-start-days",
        type=float,
        default=0.0,
        help="Optional upstream DC warm-start stock buffer in demand-days (default: 0).",
    )
    parser.add_argument(
        "--simulation-days",
        type=int,
        default=30,
        help="Target simulation horizon in days injected into scenario.horizon.steps_to_run (default: 30).",
    )
    parser.add_argument(
        "--annual-carry-rate",
        type=float,
        default=0.20,
        help="Annual inventory carrying rate applied to item value to derive daily holding cost (default: 0.20).",
    )
    return parser.parse_args()


def to_float(x: Any) -> float | None:
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def normalize_unit(unit: Any) -> str:
    s = str(unit or "").strip().upper()
    aliases = {
        "UNIT": "UN",
        "UNITE": "UN",
        "UNITS": "UN",
    }
    return aliases.get(s, s)


def convert_unit_value(value: float, from_unit: Any, to_unit: Any) -> float:
    source = normalize_unit(from_unit)
    target = normalize_unit(to_unit)
    if not source or not target or source == target:
        return value
    if source == "KG" and target == "G":
        return value / 1000.0
    if source == "G" and target == "KG":
        return value * 1000.0
    return value


def derive_item_unit_value_map(
    edges: list[dict[str, Any]],
    item_unit_map: dict[str, str],
    price_map: dict[tuple[str, str, str], dict[str, Any]],
) -> tuple[dict[str, float], dict[str, Any]]:
    candidates_by_item: dict[str, list[float]] = defaultdict(list)
    global_candidates: list[float] = []
    priced_edge_item_pairs = 0

    for edge in edges:
        src = str(edge.get("from") or "")
        dst = str(edge.get("to") or "")
        ot = edge.get("order_terms") or {}
        for raw_item_id in (edge.get("items") or []):
            item_id = str(raw_item_id)
            rec = price_map.get((src, dst, item_id)) or ot
            sell_price = to_float(rec.get("sell_price"))
            price_base = to_float(rec.get("price_base")) or 1.0
            quantity_unit = normalize_unit(rec.get("quantity_unit"))
            if sell_price is None or sell_price <= 0 or price_base <= 0 or not quantity_unit:
                continue
            target_unit = item_unit_map.get(item_id, quantity_unit)
            unit_value = convert_unit_value(sell_price / price_base, quantity_unit, target_unit)
            if unit_value <= 0:
                continue
            candidates_by_item[item_id].append(unit_value)
            global_candidates.append(unit_value)
            priced_edge_item_pairs += 1

    item_unit_value_map = {
        item_id: float(median(values))
        for item_id, values in candidates_by_item.items()
        if values
    }
    fallback_global_unit_value = float(median(global_candidates)) if global_candidates else 1.0
    stats = {
        "priced_edge_item_pairs": priced_edge_item_pairs,
        "priced_items": len(item_unit_value_map),
        "fallback_global_unit_value_per_item_unit": fallback_global_unit_value,
    }
    return item_unit_value_map, stats


def holding_cost_per_unit_day_from_value(
    *,
    item_id: str,
    unit: Any,
    item_unit_map: dict[str, str],
    item_unit_value_map: dict[str, float],
    fallback_global_unit_value: float,
    annual_carry_rate: float,
) -> tuple[float, str, float]:
    target_unit = normalize_unit(unit) or item_unit_map.get(item_id, "")
    value_unit = item_unit_map.get(item_id, target_unit)
    unit_value = item_unit_value_map.get(item_id)
    source = "item_value_median_from_priced_edges"
    if unit_value is None or unit_value <= 0:
        unit_value = fallback_global_unit_value
        source = "global_value_median_fallback"
    unit_value = convert_unit_value(unit_value, value_unit, target_unit)
    holding_cost = max(0.0, unit_value * max(0.0, annual_carry_rate) / 365.0)
    return holding_cost, source, unit_value


def infer_item_unit_map(nodes: list[dict[str, Any]], edges: list[dict[str, Any]]) -> dict[str, str]:
    votes: dict[str, Counter[str]] = defaultdict(Counter)

    for e in edges:
        u = normalize_unit(((e.get("order_terms") or {}).get("quantity_unit")))
        if not u:
            continue
        for item_id in (e.get("items") or []):
            votes[str(item_id)][u] += 3

    for n in nodes:
        for p in (n.get("processes") or []):
            for inp in (p.get("inputs") or []):
                item_id = str(inp.get("item_id"))
                u = normalize_unit(inp.get("ratio_unit"))
                if item_id and u:
                    votes[item_id][u] += 1

    priority = {"KG": 4, "G": 3, "UN": 2, "M": 1}
    out: dict[str, str] = {}
    for item_id, cnt in votes.items():
        best = sorted(cnt.items(), key=lambda x: (-x[1], -priority.get(x[0], 0), x[0]))[0][0]
        out[item_id] = best
    return out


def detect_unsourced_input_item(
    node: dict[str, Any],
    edges: list[dict[str, Any]],
    *,
    preferred_items: list[str] | None = None,
) -> tuple[str | None, str]:
    inbound_pairs = {
        (str(edge.get("to")), str(item_id))
        for edge in edges
        for item_id in (edge.get("items") or [])
    }
    node_id = str(node.get("id"))
    preferred_rank = {item_id: idx for idx, item_id in enumerate(preferred_items or [])}
    candidates: list[tuple[int, str, str]] = []
    for process in node.get("processes") or []:
        for inp in process.get("inputs") or []:
            item_id = str(inp.get("item_id"))
            if not item_id or (node_id, item_id) in inbound_pairs:
                continue
            unit = str(inp.get("ratio_unit") or "G")
            candidates.append((preferred_rank.get(item_id, len(preferred_rank)), item_id, unit))
    if not candidates:
        return None, "G"
    _, item_id, unit = sorted(candidates)[0]
    return item_id, unit


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlon / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def build_node_maps(nodes: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, tuple[float, float]]]:
    by_id: dict[str, dict[str, Any]] = {}
    coords: dict[str, tuple[float, float]] = {}
    for n in nodes:
        nid = str(n.get("id"))
        by_id[nid] = n
        geo = n.get("geo") or {}
        lat = to_float(geo.get("lat"))
        lon = to_float(geo.get("lon"))
        if lat is not None and lon is not None:
            coords[nid] = (lat, lon)
    return by_id, coords


def estimate_lead_time_days(distance_km: float, src_type: str, dst_type: str) -> float:
    base_days = max(1.0, distance_km / 650.0)
    lane_buffer = 1.0
    if src_type == "supplier_dc" and dst_type == "factory":
        lane_buffer = 2.0
    elif src_type == "factory" and dst_type == "distribution_center":
        lane_buffer = 1.5
    elif src_type == "distribution_center" and dst_type == "customer":
        lane_buffer = 1.0
    return round(base_days + lane_buffer, 2)


def estimate_transport_cost_per_unit(distance_km: float) -> float:
    # Simple placeholder estimate for early simulation setup.
    return round(max(0.05, 0.03 + 0.0011 * distance_km), 4)


def canonical_item_id(value: Any) -> str:
    s = str(value or "").strip()
    if not s:
        return ""
    if s.lower().startswith("item:"):
        raw = s.split(":", 1)[1].strip()
        if raw.isdigit():
            return f"item:{raw.zfill(6)}"
        return f"item:{raw}"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        ival = int(float(value))
        return f"item:{ival:06d}"
    s_clean = s.replace(" ", "")
    if re.fullmatch(r"[0-9]+(?:\\.0+)?", s_clean):
        ival = int(float(s_clean))
        return f"item:{ival:06d}"
    return f"item:{s_clean}"


def canonical_actor_id(value: Any) -> str:
    s = str(value or "").strip()
    if not s:
        return ""
    m = re.match(r"^([A-Za-z]+)\s*-\s*(.+)$", s)
    if not m:
        return s.replace(" ", "")
    prefix = m.group(1).upper()
    code = m.group(2).strip().upper().replace(" ", "")

    if prefix == "M":
        if code.startswith("D") and code[1:].isdigit():
            code = code[1:]
        return f"M-{code}"
    if prefix == "DC":
        if code.startswith("D") and code[1:].isdigit():
            code = code[1:]
        if code == "1910":
            code = "1920"
        return f"DC-{code}"
    if prefix == "SDC":
        if code.startswith("D") and code[1:].isdigit():
            code = code[1:]
        return f"SDC-{code}"
    if prefix == "C":
        return f"C-{code}"
    return s.replace(" ", "")


def profile_day0_value(profile: list[dict[str, Any]]) -> float:
    if not profile:
        return 0.0
    step_candidates: list[tuple[int, float]] = []
    for p in profile:
        if not isinstance(p, dict):
            continue
        ptype = str(p.get("type", "constant")).lower()
        if ptype == "constant":
            return max(0.0, to_float(p.get("value")) or 0.0)
        if ptype == "step":
            start = int(to_float(p.get("start_day")) or 0)
            val = max(0.0, to_float(p.get("value")) or 0.0)
            if start <= 0:
                step_candidates.append((start, val))
        if ptype == "piecewise":
            for pt in (p.get("points") or []):
                if not isinstance(pt, dict):
                    continue
                t = int(to_float(pt.get("t")) or 0)
                v = max(0.0, to_float(pt.get("value")) or 0.0)
                if t <= 0:
                    step_candidates.append((t, v))
    if step_candidates:
        step_candidates.sort(key=lambda x: x[0])
        return step_candidates[-1][1]
    return 0.0


def weekly_fluctuating_profile(base: float, horizon_days: int) -> list[dict[str, Any]]:
    """
    Deterministic weekly pattern centered on the same mean as `base`.
    """
    factors = [0.8, 0.95, 1.05, 1.2, 1.1, 1.0, 0.9]  # mean = 1.0
    days = max(1, int(horizon_days))
    points = [{"t": d, "value": round(max(0.0, base * factors[d % 7]), 6)} for d in range(days)]
    return [
        {
            "type": "piecewise",
            "points": points,
            "uom": "unit/day",
            "is_default": False,
            "source": "simulation_prep_assumption_weekly_fluctuation",
        }
    ]


def weekly_demand_piecewise_profile(
    weekly_values: list[float],
    horizon_days: int,
    *,
    source: str,
) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    for idx, weekly_value in enumerate(weekly_values):
        day = idx * 7
        if day >= horizon_days:
            break
        points.append(
            {
                "t": day,
                "value": round(max(0.0, weekly_value) / 7.0, 6),
            }
        )
    covered_days = len(points) * 7
    if points and covered_days < horizon_days:
        # Keep XLSX weekly totals exact: do not extend the last weekly daily rate
        # beyond the provided source periods.
        points.append({"t": covered_days, "value": 0.0})
    if not points:
        points = [{"t": 0, "value": 0.0}]
    return [
        {
            "type": "piecewise",
            "points": points,
            "uom": "unit/day",
            "is_default": False,
            "source": source,
            "source_period_unit": "week",
            "daily_distribution": "uniform_over_7_days",
            "repeat_period_days": 365,
            "repeat_mode": "annual_cycle",
        }
    ]


def load_xlsx_sheet_rows_via_powershell(
    xlsx_path: Path,
    sheet_name: str,
) -> tuple[list[dict[str, Any]], str | None]:
    script = r"""
param(
    [string]$WorkbookPath,
    [string]$SheetName
)
$excel = $null
$workbook = $null
$worksheet = $null
try {
    $resolved = (Resolve-Path -LiteralPath $WorkbookPath).Path
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($resolved)
    try {
        $worksheet = $workbook.Worksheets.Item($SheetName)
    } catch {
        Write-Output '{"error":"sheet_not_found"}'
        exit 0
    }
    $used = $worksheet.UsedRange.Value2
    if ($null -eq $used) {
        Write-Output '[]'
        exit 0
    }
    $rowCount = $used.GetLength(0)
    $colCount = $used.GetLength(1)
    if ($rowCount -lt 1 -or $colCount -lt 1) {
        Write-Output '[]'
        exit 0
    }
    $headers = @()
    for ($c = 1; $c -le $colCount; $c++) {
        $headers += [string]($used[1, $c])
    }
    $rows = New-Object System.Collections.Generic.List[object]
    for ($r = 2; $r -le $rowCount; $r++) {
        $obj = [ordered]@{}
        for ($c = 1; $c -le $colCount; $c++) {
            $header = $headers[$c - 1]
            if ([string]::IsNullOrWhiteSpace($header)) {
                continue
            }
            $obj[$header.Trim()] = $used[$r, $c]
        }
        $rows.Add([pscustomobject]$obj)
    }
    $rows | ConvertTo-Json -Depth 4 -Compress
} finally {
    if ($workbook) { $workbook.Close($false) }
    if ($excel) { $excel.Quit() }
    foreach ($obj in @($worksheet, $workbook, $excel)) {
        if ($null -ne $obj) {
            try { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($obj) } catch {}
        }
    }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""
    wrapped_script = f"& {{ {script} }}"
    cmd = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-Command",
        wrapped_script,
        str(xlsx_path),
        sheet_name,
    ]
    try:
        completed = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=False,
        )
    except Exception as exc:
        return [], f"powershell_excel_com_failed:{exc}"
    if completed.returncode != 0:
        err = (completed.stderr or completed.stdout or "").strip()
        return [], f"powershell_excel_com_failed:{err or 'unknown_error'}"
    payload = (completed.stdout or "").strip()
    if not payload:
        return [], None
    try:
        parsed = json.loads(payload)
    except Exception as exc:
        return [], f"powershell_excel_json_parse_failed:{exc}"
    if isinstance(parsed, dict) and parsed.get("error") == "sheet_not_found":
        return [], "sheet_not_found"
    if isinstance(parsed, dict):
        return [parsed], None
    if isinstance(parsed, list):
        return parsed, None
    return [], None


def rebalance_weekly_demand_rows(
    weekly_rows: list[tuple[int, float]],
) -> tuple[list[tuple[int, float]], dict[str, Any]]:
    """Keep a non-negative weekly demand profile while preserving the workbook annual total.

    The source workbook can contain negative weekly demand corrections. The simulator only
    supports non-negative customer withdrawals, so we net negative weeks against subsequent
    positive weeks, and as a last resort against previous positive weeks, while keeping the
    total yearly demand equal to the workbook total whenever that total stays non-negative.
    """

    ordered_rows = sorted((int(step), float(value)) for step, value in weekly_rows)
    adjusted_rows: list[tuple[int, float]] = []
    carry_qty = 0.0
    negative_steps: list[dict[str, float]] = []
    forward_adjustments: list[dict[str, float]] = []
    backward_adjustments: list[dict[str, float]] = []

    for step_num, raw_qty in ordered_rows:
        if raw_qty < 0.0:
            negative_steps.append({"step": int(step_num), "raw_qty": round(raw_qty, 6)})
        carry_in_qty = carry_qty
        net_qty = raw_qty + carry_qty
        if net_qty <= 0.0:
            adjusted_qty = 0.0
            carry_qty = net_qty
        else:
            adjusted_qty = net_qty
            carry_qty = 0.0
        adjusted_rows.append((step_num, adjusted_qty))
        if raw_qty < 0.0 or abs(carry_in_qty) > 1e-9:
            forward_adjustments.append(
                {
                    "step": int(step_num),
                    "raw_qty": round(raw_qty, 6),
                    "carry_in_qty": round(carry_in_qty, 6),
                    "adjusted_qty": round(adjusted_qty, 6),
                    "carry_out_qty": round(carry_qty, 6),
                }
            )

    if carry_qty < -1e-9:
        debt_qty = -carry_qty
        for idx in range(len(adjusted_rows) - 1, -1, -1):
            step_num, adjusted_qty = adjusted_rows[idx]
            if adjusted_qty <= 1e-9:
                continue
            reduction_qty = min(adjusted_qty, debt_qty)
            adjusted_rows[idx] = (step_num, adjusted_qty - reduction_qty)
            debt_qty -= reduction_qty
            backward_adjustments.append(
                {
                    "step": int(step_num),
                    "reduction_qty": round(reduction_qty, 6),
                    "adjusted_qty_after_reduction": round(adjusted_rows[idx][1], 6),
                }
            )
            if debt_qty <= 1e-9:
                break
        carry_qty = -debt_qty

    raw_total_qty = sum(value for _step, value in ordered_rows)
    adjusted_total_qty = sum(value for _step, value in adjusted_rows)
    return adjusted_rows, {
        "raw_total_qty": round(raw_total_qty, 6),
        "adjusted_total_qty": round(adjusted_total_qty, 6),
        "negative_steps": negative_steps,
        "forward_adjustments": forward_adjustments,
        "backward_adjustments": list(reversed(backward_adjustments)),
        "unresolved_negative_balance_qty": round(max(0.0, -carry_qty), 6),
    }


def load_weekly_demand_from_pf_xlsx(
    xlsx_path: Path,
) -> tuple[dict[tuple[str, str], list[float]], dict[str, Any]]:
    demand_map: dict[tuple[str, str], list[float]] = {}
    stats: dict[str, Any] = {
        "enabled": False,
        "xlsx_path": str(xlsx_path),
        "sheet_found": False,
        "rows_read": 0,
        "rows_mapped": 0,
        "pairs_loaded": 0,
        "annual_total_per_pair": {},
        "simulation_profile_total_per_pair": {},
        "negative_demand_adjustments": {},
    }
    if not xlsx_path.exists():
        stats["error"] = "xlsx_not_found"
        return demand_map, stats

    try:
        import openpyxl  # type: ignore
    except Exception:
        openpyxl = None  # type: ignore

    raw_rows: dict[tuple[str, str], list[tuple[int, float]]] = defaultdict(list)
    if openpyxl is not None:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Demande" not in wb.sheetnames:
            stats["error"] = "demande_sheet_not_found"
            return demand_map, stats
        stats["enabled"] = True
        stats["sheet_found"] = True

        ws = wb["Demande"]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(v).strip() if v is not None else "" for v in next(rows)]
        except StopIteration:
            stats["error"] = "empty_sheet"
            return demand_map, stats
        idx = {name: i for i, name in enumerate(header)}
        required = {"product", "customer_id", "step"}
        if not required.issubset(idx):
            stats["error"] = "missing_required_columns"
            stats["columns"] = header
            return demand_map, stats

        for rec in rows:
            stats["rows_read"] += 1
            if not rec:
                continue
            item_id = canonical_item_id(rec[idx["product"]])
            customer_id = canonical_actor_id(rec[idx["customer_id"]])
            step_num = int(to_float(rec[idx["step"]]) or 0)
            real_idx = idx.get("real demand")
            forecast_idx = idx.get("forecast_demand")
            demand_value = None
            if real_idx is not None:
                demand_value = to_float(rec[real_idx])
            if demand_value is None and forecast_idx is not None:
                demand_value = to_float(rec[forecast_idx])
            if not item_id or not customer_id or step_num <= 0 or demand_value is None:
                continue
            raw_rows[(customer_id, item_id)].append((step_num, demand_value))
            stats["rows_mapped"] += 1
    else:
        stats["excel_reader"] = "powershell_excel_com"
        rows, err = load_xlsx_sheet_rows_via_powershell(xlsx_path, "Demande")
        if err == "sheet_not_found":
            stats["error"] = "demande_sheet_not_found"
            return demand_map, stats
        if err:
            stats["error"] = err
            return demand_map, stats
        stats["enabled"] = True
        stats["sheet_found"] = True
        for rec in rows:
            stats["rows_read"] += 1
            if not isinstance(rec, dict):
                continue
            item_id = canonical_item_id(rec.get("product"))
            customer_id = canonical_actor_id(rec.get("customer_id"))
            step_num = int(to_float(rec.get("step")) or 0)
            demand_value = to_float(rec.get("real demand"))
            if demand_value is None:
                demand_value = to_float(rec.get("forecast_demand"))
            if not item_id or not customer_id or step_num <= 0 or demand_value is None:
                continue
            raw_rows[(customer_id, item_id)].append((step_num, demand_value))
            stats["rows_mapped"] += 1

    for pair, values in raw_rows.items():
        adjusted_rows, adjustment_info = rebalance_weekly_demand_rows(values)
        demand_map[pair] = [float(v) for _, v in adjusted_rows]
        pair_key = f"{pair[0]}::{pair[1]}"
        if adjustment_info["negative_steps"]:
            stats["negative_demand_adjustments"][pair_key] = adjustment_info
    stats["pairs_loaded"] = len(demand_map)
    stats["periods_per_pair"] = {
        f"{pair[0]}::{pair[1]}": len(values)
        for pair, values in sorted(demand_map.items())
    }
    stats["annual_total_per_pair"] = {
        f"{pair[0]}::{pair[1]}": round(sum(value for _step, value in raw_rows[pair]), 6)
        for pair in sorted(demand_map)
    }
    stats["simulation_profile_total_per_pair"] = {
        f"{pair[0]}::{pair[1]}": round(sum(values), 6)
        for pair, values in sorted(demand_map.items())
    }
    return demand_map, stats


def load_prices_from_data_poc(xlsx_path: Path) -> tuple[dict[tuple[str, str, str], dict[str, Any]], dict[str, Any]]:
    price_map: dict[tuple[str, str, str], dict[str, Any]] = {}
    stats: dict[str, Any] = {
        "enabled": False,
        "xlsx_path": str(xlsx_path),
        "sheet_found": False,
        "rows_read": 0,
        "rows_mapped": 0,
    }
    if not xlsx_path.exists():
        stats["error"] = "xlsx_not_found"
        return price_map, stats

    try:
        import openpyxl  # type: ignore
    except Exception:
        openpyxl = None  # type: ignore

    if openpyxl is not None:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Relations_acteurs" not in wb.sheetnames:
            stats["error"] = "relations_acteurs_sheet_missing"
            return price_map, stats

        stats["enabled"] = True
        stats["sheet_found"] = True
        ws = wb["Relations_acteurs"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return price_map, stats
        headers = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(headers)}
        required = {"customer", "supplier", "product", "sell_price", "price_base", "quantity_unit"}
        if not required.issubset(set(idx.keys())):
            stats["error"] = "missing_required_columns"
            stats["columns"] = headers
            return price_map, stats

        for r in rows[1:]:
            stats["rows_read"] += 1
            customer = canonical_actor_id(r[idx["customer"]])
            supplier = canonical_actor_id(r[idx["supplier"]])
            item_id = canonical_item_id(r[idx["product"]])
            if not customer or not supplier or not item_id:
                continue
            sell_price = to_float(r[idx["sell_price"]])
            price_base = to_float(r[idx["price_base"]])
            unit = normalize_unit(r[idx["quantity_unit"]])
            key = (supplier, customer, item_id)
            rec = {
                "sell_price": sell_price,
                "price_base": price_base if price_base and price_base > 0 else 1.0,
                "quantity_unit": unit,
                "source": "data_poc_relations_acteurs",
            }
            existing = price_map.get(key)
            if existing is None:
                price_map[key] = rec
            else:
                ex_sp = to_float(existing.get("sell_price"))
                if ex_sp is None and sell_price is not None:
                    price_map[key] = rec
    else:
        stats["excel_reader"] = "powershell_excel_com"
        rows, err = load_xlsx_sheet_rows_via_powershell(xlsx_path, "Relations_acteurs")
        if err == "sheet_not_found":
            stats["error"] = "relations_acteurs_sheet_missing"
            return price_map, stats
        if err:
            stats["error"] = err
            return price_map, stats
        stats["enabled"] = True
        stats["sheet_found"] = True
        for rec in rows:
            stats["rows_read"] += 1
            if not isinstance(rec, dict):
                continue
            customer = canonical_actor_id(rec.get("customer"))
            supplier = canonical_actor_id(rec.get("supplier"))
            item_id = canonical_item_id(rec.get("product"))
            if not customer or not supplier or not item_id:
                continue
            sell_price = to_float(rec.get("sell_price"))
            price_base = to_float(rec.get("price_base"))
            unit = normalize_unit(rec.get("quantity_unit"))
            key = (supplier, customer, item_id)
            row = {
                "sell_price": sell_price,
                "price_base": price_base if price_base and price_base > 0 else 1.0,
                "quantity_unit": unit,
                "source": "data_poc_relations_acteurs",
            }
            existing = price_map.get(key)
            if existing is None:
                price_map[key] = row
            else:
                ex_sp = to_float(existing.get("sell_price"))
                if ex_sp is None and sell_price is not None:
                    price_map[key] = row

    stats["rows_mapped"] = len(price_map)
    return price_map, stats


def actor_type_from_role(role: Any, actor_id: str) -> str:
    role_s = str(role or "").strip().lower()
    if "supplier distribution center" in role_s:
        return "supplier_dc"
    if role_s == "manufacturer" or "manufacturer" in role_s:
        return "factory"
    if role_s == "distribution center" or "distribution center" in role_s:
        return "distribution_center"
    if role_s == "customer" or "customer" in role_s:
        return "customer"
    if actor_id.startswith("SDC-"):
        return "supplier_dc"
    if actor_id.startswith("M-"):
        return "factory"
    if actor_id.startswith("DC-"):
        return "distribution_center"
    if actor_id.startswith("C-"):
        return "customer"
    return "unknown"


def actor_id_from_acteurs_row(rec: dict[str, Any]) -> str:
    role = str(rec.get("role") or "").strip()
    description = str(rec.get("description") or "").strip()
    match = re.search(r"([A-Z]{1,3}[0-9A-Z]+)\s*$", description)
    code = match.group(1) if match else ""
    role_s = role.lower()
    if not code:
        if "customer" in role_s:
            return "C-XXXXX"
        return ""
    if "supplier distribution center" in role_s:
        return canonical_actor_id(f"SDC - {code}")
    if role_s == "manufacturer" or "manufacturer" in role_s:
        return canonical_actor_id(f"M - {code}")
    if role_s == "distribution center" or "distribution center" in role_s:
        return canonical_actor_id(f"DC - {code}")
    if "customer" in role_s:
        return canonical_actor_id(f"C - {code}")
    return canonical_actor_id(code)


def load_actor_context_from_demand_pf_xlsx(
    xlsx_path: Path,
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    actor_map: dict[str, dict[str, Any]] = {}
    stats: dict[str, Any] = {
        "enabled": False,
        "xlsx_path": str(xlsx_path),
        "sheet_found": False,
        "rows_read": 0,
        "rows_mapped": 0,
    }
    if not xlsx_path.exists():
        stats["error"] = "xlsx_not_found"
        return actor_map, stats

    try:
        import openpyxl  # type: ignore
    except Exception:
        openpyxl = None  # type: ignore

    if openpyxl is not None:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Acteurs" not in wb.sheetnames:
            stats["error"] = "acteurs_sheet_not_found"
            return actor_map, stats
        stats["enabled"] = True
        stats["sheet_found"] = True
        ws = wb["Acteurs"]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(v).strip() if v is not None else "" for v in next(rows)]
        except StopIteration:
            stats["error"] = "empty_sheet"
            return actor_map, stats
        for rec_values in rows:
            stats["rows_read"] += 1
            rec = {
                header[i]: rec_values[i] if i < len(rec_values) else None
                for i in range(len(header))
                if header[i]
            }
            actor_id = actor_id_from_acteurs_row(rec)
            if not actor_id:
                continue
            actor_map[actor_id] = {
                "actor_id": actor_id,
                "role": str(rec.get("role") or "").strip(),
                "description": str(rec.get("description") or "").strip(),
                "location_ID": str(rec.get("location_ID") or "").strip(),
                "manufactured_products": str(rec.get("manufactured_products") or "").strip(),
                "procured_product": str(rec.get("procured_product") or "").strip(),
            }
            stats["rows_mapped"] += 1
    else:
        stats["excel_reader"] = "powershell_excel_com"
        rows, err = load_xlsx_sheet_rows_via_powershell(xlsx_path, "Acteurs")
        if err == "sheet_not_found":
            stats["error"] = "acteurs_sheet_not_found"
            return actor_map, stats
        if err:
            stats["error"] = err
            return actor_map, stats
        stats["enabled"] = True
        stats["sheet_found"] = True
        for rec in rows:
            stats["rows_read"] += 1
            if not isinstance(rec, dict):
                continue
            actor_id = actor_id_from_acteurs_row(rec)
            if not actor_id:
                continue
            actor_map[actor_id] = {
                "actor_id": actor_id,
                "role": str(rec.get("role") or "").strip(),
                "description": str(rec.get("description") or "").strip(),
                "location_ID": str(rec.get("location_ID") or "").strip(),
                "manufactured_products": str(rec.get("manufactured_products") or "").strip(),
                "procured_product": str(rec.get("procured_product") or "").strip(),
            }
            stats["rows_mapped"] += 1
    return actor_map, stats


def load_relations_from_demand_pf_xlsx(
    xlsx_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    relation_rows: list[dict[str, Any]] = []
    stats: dict[str, Any] = {
        "enabled": False,
        "xlsx_path": str(xlsx_path),
        "sheet_found": False,
        "rows_read": 0,
        "rows_mapped": 0,
    }
    if not xlsx_path.exists():
        stats["error"] = "xlsx_not_found"
        return relation_rows, stats

    try:
        import openpyxl  # type: ignore
    except Exception:
        openpyxl = None  # type: ignore

    if openpyxl is not None:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Relations_acteurs" not in wb.sheetnames:
            stats["error"] = "relations_acteurs_sheet_missing"
            return relation_rows, stats
        stats["enabled"] = True
        stats["sheet_found"] = True
        ws = wb["Relations_acteurs"]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(v).strip() if v is not None else "" for v in next(rows)]
        except StopIteration:
            stats["error"] = "empty_sheet"
            return relation_rows, stats
        for rec_values in rows:
            stats["rows_read"] += 1
            rec = {
                header[i]: rec_values[i] if i < len(rec_values) else None
                for i in range(len(header))
                if header[i]
            }
            supplier = canonical_actor_id(rec.get("supplier"))
            customer = canonical_actor_id(rec.get("customer"))
            item_id = canonical_item_id(rec.get("product"))
            if not supplier or not customer or not item_id:
                continue
            relation_rows.append(
                {
                    "supplier": supplier,
                    "customer": customer,
                    "item_id": item_id,
                    "sell_price": to_float(rec.get("sell_price")),
                    "price_base": to_float(rec.get("price_base")) or 1.0,
                    "quantity_unit": normalize_unit(rec.get("quantity_unit")),
                    "supply_order_frequency": to_float(rec.get("supply_order_frequency")),
                    "customer_priority_rank": to_float(rec.get("customer_priority_rank")),
                    "supplier_priority_rank": to_float(rec.get("supplier_priority_rank")),
                    "delay_step_limit": to_float(rec.get("delay_step_limit")),
                    "transport_cost": to_float(rec.get("transport_cost")),
                    "source": "demand_pf_relations_acteurs",
                }
            )
            stats["rows_mapped"] += 1
    else:
        stats["excel_reader"] = "powershell_excel_com"
        rows, err = load_xlsx_sheet_rows_via_powershell(xlsx_path, "Relations_acteurs")
        if err == "sheet_not_found":
            stats["error"] = "relations_acteurs_sheet_missing"
            return relation_rows, stats
        if err:
            stats["error"] = err
            return relation_rows, stats
        stats["enabled"] = True
        stats["sheet_found"] = True
        for rec in rows:
            stats["rows_read"] += 1
            if not isinstance(rec, dict):
                continue
            supplier = canonical_actor_id(rec.get("supplier"))
            customer = canonical_actor_id(rec.get("customer"))
            item_id = canonical_item_id(rec.get("product"))
            if not supplier or not customer or not item_id:
                continue
            relation_rows.append(
                {
                    "supplier": supplier,
                    "customer": customer,
                    "item_id": item_id,
                    "sell_price": to_float(rec.get("sell_price")),
                    "price_base": to_float(rec.get("price_base")) or 1.0,
                    "quantity_unit": normalize_unit(rec.get("quantity_unit")),
                    "supply_order_frequency": to_float(rec.get("supply_order_frequency")),
                    "customer_priority_rank": to_float(rec.get("customer_priority_rank")),
                    "supplier_priority_rank": to_float(rec.get("supplier_priority_rank")),
                    "delay_step_limit": to_float(rec.get("delay_step_limit")),
                    "transport_cost": to_float(rec.get("transport_cost")),
                    "source": "demand_pf_relations_acteurs",
                }
            )
            stats["rows_mapped"] += 1

    return relation_rows, stats


def ensure_sim_meta(graph: dict[str, Any]) -> None:
    meta = graph.setdefault("meta", {})
    meta.pop("baseline_rebuild", None)
    prep = {
        "prepared_for_simulation": True,
        "prepared_at_utc": datetime.now(timezone.utc).isoformat(),
        "assumption_version": "v1",
        "notes": [
            "Missing/default values were replaced with explicit assumptions for first-pass simulation.",
            "Review simulation_prep_report before running production scenarios.",
        ],
    }
    meta["simulation_prep"] = prep


def prepare_graph(
    graph: dict[str, Any],
    data_poc_xlsx: Path | None = None,
    demand_pf_xlsx: Path | None = None,
    customer_warm_start_days: float = 0.0,
    upstream_dc_warm_start_days: float = 0.0,
    simulation_days: int = 30,
    annual_carry_rate: float = 0.20,
) -> tuple[dict[str, Any], dict[str, Any]]:
    g = deepcopy(graph)
    ensure_sim_meta(g)

    nodes = g.get("nodes", []) or []
    edges = g.get("edges", []) or []
    scenarios = g.get("scenarios", []) or []
    node_by_id, coords = build_node_maps(nodes)

    change_counts = defaultdict(int)
    changed_edge_ids: list[str] = []
    changed_node_ids: list[str] = []
    changed_demand_rows: list[dict[str, str]] = []
    invented_entities: list[dict[str, str]] = []

    # Reconcile known alias mismatch: DC-1910 (relations) == DC-1920 (actors).
    if "DC-1910" in node_by_id and "DC-1920" in node_by_id:
        for e in edges:
            edge_changed = False
            if str(e.get("from")) == "DC-1910":
                e["from"] = "DC-1920"
                edge_changed = True
            if str(e.get("to")) == "DC-1910":
                e["to"] = "DC-1920"
                edge_changed = True
            if edge_changed:
                change_counts["edge_endpoint_reconciled_dc1910_to_dc1920"] += 1
                changed_edge_ids.append(str(e.get("id")))

        nodes[:] = [n for n in nodes if str(n.get("id")) != "DC-1910"]
        change_counts["node_alias_merged_dc1910_into_dc1920"] += 1
        changed_node_ids.extend(["DC-1910", "DC-1920"])
        node_by_id, coords = build_node_maps(nodes)

    # Recover customer location from Acteurs context.
    customer = node_by_id.get("C-XXXXX")
    if isinstance(customer, dict) and not str(customer.get("location_ID") or "").strip():
        customer["location_ID"] = "Paris"
        geo = customer.get("geo")
        if not isinstance(geo, dict):
            geo = {}
            customer["geo"] = geo
        if not str(geo.get("country") or "").strip():
            geo["country"] = "France"
        raw_geo = geo.get("raw")
        if not isinstance(raw_geo, dict):
            raw_geo = {}
            geo["raw"] = raw_geo
        raw_geo["location_ID"] = "Paris"
        raw_geo["source"] = "simulation_prep_acteurs_recovery"
        change_counts["customer_location_recovered_from_acteurs"] += 1
        changed_node_ids.append("C-XXXXX")
        node_by_id, coords = build_node_maps(nodes)

    item_unit_map = infer_item_unit_map(nodes, edges)
    price_map: dict[tuple[str, str, str], dict[str, Any]] = {}
    price_import_stats: dict[str, Any] = {"enabled": False}
    if data_poc_xlsx is not None:
        price_map, price_import_stats = load_prices_from_data_poc(data_poc_xlsx)
    demand_pf_map: dict[tuple[str, str], list[float]] = {}
    demand_pf_stats: dict[str, Any] = {"enabled": False}
    demand_pf_actor_map: dict[str, dict[str, Any]] = {}
    demand_pf_actor_stats: dict[str, Any] = {"enabled": False}
    demand_pf_relation_rows: list[dict[str, Any]] = []
    demand_pf_relation_stats: dict[str, Any] = {"enabled": False}
    if demand_pf_xlsx is not None:
        demand_pf_map, demand_pf_stats = load_weekly_demand_from_pf_xlsx(demand_pf_xlsx)
        demand_pf_actor_map, demand_pf_actor_stats = load_actor_context_from_demand_pf_xlsx(demand_pf_xlsx)
        demand_pf_relation_rows, demand_pf_relation_stats = load_relations_from_demand_pf_xlsx(demand_pf_xlsx)

    invented_dst_node = "M-1810"

    for rel in demand_pf_relation_rows:
        key = (str(rel.get("supplier") or ""), str(rel.get("customer") or ""), str(rel.get("item_id") or ""))
        if not all(key):
            continue
        if key not in price_map:
            price_map[key] = {
                "sell_price": rel.get("sell_price"),
                "price_base": rel.get("price_base"),
                "quantity_unit": rel.get("quantity_unit"),
                "source": "demand_pf_relations_acteurs",
            }

    def ensure_actor_node(actor_id: str) -> dict[str, Any] | None:
        nonlocal node_by_id, coords
        actor = demand_pf_actor_map.get(actor_id) or {}
        node = node_by_id.get(actor_id)
        role = str(actor.get("role") or "")
        description = str(actor.get("description") or actor_id)
        location_id = str(actor.get("location_ID") or "")
        ntype = actor_type_from_role(role, actor_id)
        legacy_prefix = actor_id.split("-", 1)[0] if "-" in actor_id else actor_id
        legacy_code = actor_id.split("-", 1)[1] if "-" in actor_id else actor_id
        legacy_key = f"{legacy_prefix} - {legacy_code}"

        if not isinstance(node, dict):
            node = {
                "id": actor_id,
                "type": ntype,
                "name": description or actor_id,
                "role_raw": role,
                "location_ID": location_id,
                "geo": {
                    "lat": None,
                    "lon": None,
                    "country": "",
                    "raw": {
                        "location_ID": location_id,
                        "source": "demand_pf_acteurs",
                    },
                },
                "attrs": {
                    "legacy_key": legacy_key,
                    "location_ID": location_id,
                },
                "metadata": {
                    "description": description,
                    "source_sheet": "Acteurs",
                },
                "inventory": {
                    "states": [],
                    "backlogs": [],
                    "wip": [],
                },
                "processes": [],
                "policies": {},
            }
            nodes.append(node)
            change_counts["node_added_from_demand_pf_acteurs"] += 1
            changed_node_ids.append(actor_id)
            node_by_id, coords = build_node_maps(nodes)
            node = node_by_id.get(actor_id)
        if not isinstance(node, dict):
            return None

        if location_id and not str(node.get("location_ID") or "").strip():
            node["location_ID"] = location_id
            change_counts["node_location_filled_from_demand_pf_acteurs"] += 1
            changed_node_ids.append(actor_id)
        attrs = node.get("attrs")
        if not isinstance(attrs, dict):
            attrs = {}
            node["attrs"] = attrs
        if location_id and not str(attrs.get("location_ID") or "").strip():
            attrs["location_ID"] = location_id
        if not str(attrs.get("legacy_key") or "").strip():
            attrs["legacy_key"] = legacy_key
        if description and not str(node.get("name") or "").strip():
            node["name"] = description
        if role and not str(node.get("role_raw") or "").strip():
            node["role_raw"] = role
        metadata = node.get("metadata")
        if not isinstance(metadata, dict):
            metadata = {}
            node["metadata"] = metadata
        if description and not str(metadata.get("description") or "").strip():
            metadata["description"] = description
            metadata["source_sheet"] = "Acteurs"
        return node

    def ensure_node_inventory_state(node_id: str, item_id: str, uom: str, initial_source: str) -> None:
        nonlocal change_counts, changed_node_ids
        n = node_by_id.get(node_id)
        if not isinstance(n, dict):
            return
        inv = n.get("inventory")
        if not isinstance(inv, dict):
            inv = {"states": [], "backlogs": [], "wip": []}
            n["inventory"] = inv
        states = inv.get("states")
        if not isinstance(states, list):
            states = []
            inv["states"] = states
        state = next((s for s in states if str(s.get("item_id")) == item_id), None)
        if isinstance(state, dict):
            return
        states.append(
            {
                "item_id": item_id,
                "state_id": f"I_{item_id.replace('item:', '')}_{node_id.replace('-', '_')}",
                "initial": 0.0,
                "uom": uom or item_unit_map.get(item_id, "UN"),
                "holding_cost": {
                    "value": 0.0,
                    "per": "unit*day",
                    "is_default": True,
                    "source": "simulation_prep_assumption_pending_value_based",
                },
                "is_default_initial": False,
                "initial_source": initial_source,
                "uom_source": "simulation_prep_inferred_from_relations_bom",
            }
        )
        change_counts["inventory_state_added_from_demand_pf_relations"] += 1
        changed_node_ids.append(node_id)

    def ensure_relation_edge(rel: dict[str, Any]) -> None:
        nonlocal node_by_id, coords
        supplier = str(rel.get("supplier") or "")
        customer = str(rel.get("customer") or "")
        item_id = str(rel.get("item_id") or "")
        if not supplier or not customer or not item_id:
            return
        ensure_actor_node(supplier)
        ensure_actor_node(customer)
        supplier_node = node_by_id.get(supplier)
        customer_node = node_by_id.get(customer)
        if not isinstance(supplier_node, dict) or not isinstance(customer_node, dict):
            return
        item_code = item_id.split(":", 1)[1]
        existing = next(
            (
                e
                for e in edges
                if str(e.get("from")) == supplier and str(e.get("to")) == customer and item_id in (e.get("items") or [])
            ),
            None,
        )
        if not isinstance(existing, dict):
            edge_id = f"edge:{supplier}_TO_{customer}_{item_code}"
            existing = {
                "id": edge_id,
                "type": "transport",
                "from": supplier,
                "to": customer,
                "items": [item_id],
                "order_terms": {},
                "lead_time": {"is_default": True},
                "transport_cost": {"is_default": True},
                "delay_step_limit": {"is_default": True},
                "source": "demand_pf_relations_acteurs",
            }
            edges.append(existing)
            change_counts["edge_added_from_demand_pf_relations"] += 1
            changed_edge_ids.append(edge_id)
            node_by_id, coords = build_node_maps(nodes)
            item_unit_map[item_id] = rel.get("quantity_unit") or item_unit_map.get(item_id, "UN")

        ot = existing.get("order_terms")
        if not isinstance(ot, dict):
            ot = {}
        rel_sell_price = rel.get("sell_price")
        rel_price_base = rel.get("price_base")
        rel_quantity_unit = rel.get("quantity_unit")
        if rel_sell_price is not None:
            ot["sell_price"] = rel_sell_price
        if rel_price_base is not None:
            ot["price_base"] = rel_price_base
        if rel_quantity_unit:
            ot["quantity_unit"] = rel_quantity_unit
        sof = rel.get("supply_order_frequency")
        if sof is not None and sof > 0:
            ot["supply_order_frequency"] = {"value": sof, "time_unit": "day", "is_default": False}
        cpr = rel.get("customer_priority_rank")
        if cpr is not None:
            ot["customer_priority_rank"] = cpr
        spr = rel.get("supplier_priority_rank")
        if spr is not None:
            ot["supplier_priority_rank"] = spr
        ot["source"] = "demand_pf_relations_acteurs"
        existing["order_terms"] = ot

        dsl = existing.get("delay_step_limit")
        if not isinstance(dsl, dict):
            dsl = {}
        if rel.get("delay_step_limit") is not None and rel.get("delay_step_limit") > 0:
            dsl["value"] = int(rel["delay_step_limit"])
            dsl["is_default"] = False
            dsl["source"] = "demand_pf_relations_acteurs"
            existing["delay_step_limit"] = dsl

        tc = existing.get("transport_cost")
        if not isinstance(tc, dict):
            tc = {}
        if rel.get("transport_cost") is not None and rel.get("transport_cost") > 0:
            tc["value"] = rel["transport_cost"]
            tc["per"] = "unit"
            tc["is_default"] = False
            tc["source"] = "demand_pf_relations_acteurs"
            existing["transport_cost"] = tc

        ensure_node_inventory_state(supplier, item_id, str(rel.get("quantity_unit") or ""), "demand_pf_relations_acteurs")
        ensure_node_inventory_state(customer, item_id, str(rel.get("quantity_unit") or ""), "demand_pf_relations_acteurs")
        change_counts["edge_order_terms_aligned_from_demand_pf_relations"] += 1
        if str(existing.get("id")) not in changed_edge_ids:
            changed_edge_ids.append(str(existing.get("id")))

    for rel in demand_pf_relation_rows:
        ensure_relation_edge(rel)

    actual_inbound_relations: dict[tuple[str, str], int] = defaultdict(int)
    for e in edges:
        if str(e.get("source") or "") == "simulation_prep_gaillac_question_mark_assumption":
            continue
        if bool(e.get("is_assumed")):
            continue
        dst = str(e.get("to"))
        for item_id in (e.get("items") or []):
            actual_inbound_relations[(dst, str(item_id))] += 1

    obsolete_assumed_edges = []
    for e in edges:
        if str(e.get("source") or "") != "simulation_prep_gaillac_question_mark_assumption" and not bool(e.get("is_assumed")):
            continue
        dst = str(e.get("to"))
        items = [str(item_id) for item_id in (e.get("items") or [])]
        if any(actual_inbound_relations.get((dst, item_id), 0) > 0 for item_id in items):
            obsolete_assumed_edges.append(str(e.get("id")))
    if obsolete_assumed_edges:
        edges[:] = [e for e in edges if str(e.get("id")) not in set(obsolete_assumed_edges)]
        change_counts["assumed_gaillac_supplier_edge_removed_due_to_real_relation"] += len(obsolete_assumed_edges)
        changed_edge_ids.extend(obsolete_assumed_edges)
        node_by_id, coords = build_node_maps(nodes)
        item_unit_map = infer_item_unit_map(nodes, edges)

    for node_id, node in list(node_by_id.items()):
        assumptions = node.get("assumptions")
        if not isinstance(assumptions, dict):
            continue
        remove_keys = []
        for key in list(assumptions.keys()):
            if not key.startswith("item_") or not key.endswith("_supplier_mapping"):
                continue
            item_code = key[len("item_") : -len("_supplier_mapping")]
            item_id = canonical_item_id(item_code)
            if actual_inbound_relations.get((invented_dst_node, item_id), 0) > 0:
                remove_keys.append(key)
        for key in remove_keys:
            assumptions.pop(key, None)
            change_counts["assumed_gaillac_supplier_mapping_removed_due_to_real_relation"] += 1
            changed_node_ids.append(node_id)

        inv = node.get("inventory")
        if not isinstance(inv, dict):
            continue
        states = inv.get("states")
        if not isinstance(states, list):
            continue
        kept_states = []
        removed_any = False
        for state in states:
            item_id = str(state.get("item_id") or "")
            initial_source = str(state.get("initial_source") or "")
            assumption_label = str(state.get("assumption_label") or "")
            has_real_relation = actual_inbound_relations.get((invented_dst_node, item_id), 0) > 0
            if node_id != invented_dst_node and has_real_relation and (
                initial_source == "simulation_prep_gaillac_question_mark_assumption" or assumption_label == "GAILLAC?"
            ):
                removed_any = True
                continue
            kept_states.append(state)
        if removed_any:
            inv["states"] = kept_states
            change_counts["assumed_gaillac_inventory_state_removed_due_to_real_relation"] += 1
            changed_node_ids.append(node_id)

    # Assumed supplier mapping for the currently modeled but unsourced M-1810 input.
    invented_item_id = None
    invented_item_code = ""
    invented_unit = "G"
    has_process_input = False
    if invented_dst_node in node_by_id:
        invented_item_id, invented_unit = detect_unsourced_input_item(
            node_by_id[invented_dst_node],
            edges,
            preferred_items=["item:007923", "item:693710"],
        )
        has_process_input = invented_item_id is not None
        invented_item_code = invented_item_id.split(":", 1)[1] if invented_item_id else ""

    has_existing_lane = any(
        str(e.get("to")) == invented_dst_node and invented_item_id in (e.get("items") or [])
        for e in edges
    ) if invented_item_id else False
    invented_source_tag = "simulation_prep_gaillac_question_mark_assumption"
    assumption_label = "GAILLAC?"
    supplier_candidates = ["SDC-1450", "DC-1450"]
    invented_supplier_id = next((sid for sid in supplier_candidates if sid in node_by_id), supplier_candidates[0])
    invented_edge_id = f"edge:{invented_supplier_id}_TO_M-1810_{invented_item_code}_Q" if invented_item_code else ""

    if has_process_input and invented_item_id and not has_existing_lane and invented_dst_node in node_by_id:
        dst_node = node_by_id[invented_dst_node]
        dst_geo = (dst_node.get("geo") or {}) if isinstance(dst_node.get("geo"), dict) else {}
        unit = item_unit_map.get(invented_item_id, invented_unit or "G")
        initial_qty = 3000.0

        inv_node = node_by_id.get(invented_supplier_id)
        if not isinstance(inv_node, dict):
            inv_node = {
                "id": invented_supplier_id,
                "type": "supplier_dc",
                "name": "Supplier of Raw Materials - D1450 ?",
                "location_ID": "France - GAILLAC - 81600",
                "geo": {
                    "lat": to_float(dst_geo.get("lat")),
                    "lon": to_float(dst_geo.get("lon")),
                    "country": dst_geo.get("country") or "France",
                    "raw": {
                        "method": "simulation_prep_assumed_supplier",
                        "note": f"Assumed Gaillac supplier mapping (uncertain) for {invented_item_id}",
                    },
                },
                "inventory": {
                    "states": [],
                    "backlogs": [],
                    "wip": [],
                },
                "processes": [],
                "policies": {},
                "assumptions": {
                    "is_assumed": True,
                    "label": assumption_label,
                    "source": invented_source_tag,
                },
            }
            nodes.append(inv_node)
            change_counts["assumed_gaillac_supplier_node_added"] += 1
            changed_node_ids.append(invented_supplier_id)
            invented_entities.append(
                {
                    "entity_type": "node",
                    "id": invented_supplier_id,
                    "label": assumption_label,
                }
            )
            node_by_id, coords = build_node_maps(nodes)

        assumptions = inv_node.get("assumptions")
        if not isinstance(assumptions, dict):
            assumptions = {}
            inv_node["assumptions"] = assumptions
        assumptions["is_assumed"] = True
        assumptions["label"] = assumption_label
        assumptions["source"] = invented_source_tag
        assumptions[f"item_{invented_item_code}_supplier_mapping"] = {
            "is_assumed": True,
            "label": assumption_label,
            "source": invented_source_tag,
        }
        change_counts["assumed_gaillac_supplier_node_tagged"] += 1

        inv_states = ((inv_node.get("inventory") or {}).get("states") or [])
        has_inv_state = any(str(s.get("item_id")) == invented_item_id for s in inv_states)
        if not has_inv_state:
            inv_states.append(
                {
                    "item_id": invented_item_id,
                    "state_id": f"I_{invented_item_code}_Q",
                    "initial": initial_qty,
                    "uom": unit,
                    "holding_cost": {
                        "value": 0.0,
                        "per": "unit*day",
                        "is_default": True,
                        "source": "simulation_prep_assumption_pending_value_based",
                    },
                    "is_default_initial": False,
                    "initial_source": invented_source_tag,
                    "uom_source": invented_source_tag,
                    "assumption_label": assumption_label,
                }
            )
            inv_node.setdefault("inventory", {})["states"] = inv_states
            change_counts["assumed_gaillac_supplier_inventory_state_added"] += 1
            changed_node_ids.append(invented_supplier_id)

        if not any(str(e.get("id")) == invented_edge_id for e in edges):
            edges.append(
                {
                    "id": invented_edge_id,
                    "type": "transport",
                    "from": invented_supplier_id,
                    "to": invented_dst_node,
                    "items": [invented_item_id],
                    "order_terms": {
                        "sell_price": None,
                        "price_base": 1.0,
                        "quantity_unit": unit,
                        "supply_order_frequency": {"value": 1, "time_unit": "day", "is_default": True},
                        "customer_priority_rank": None,
                        "supplier_priority_rank": None,
                    },
                    "lead_time": {
                        "type": "erlang_pipeline",
                        "mean": 1.0,
                        "stages": 4,
                        "time_unit": "day",
                        "is_default": False,
                        "source": invented_source_tag,
                    },
                    "distance_km": 0.0,
                    "transport_cost": {
                        "value": 0.05,
                        "per": "unit",
                        "is_default": False,
                        "source": invented_source_tag,
                    },
                    "delay_step_limit": {
                        "value": 21,
                        "is_default": False,
                        "source": invented_source_tag,
                    },
                    "is_assumed": True,
                    "assumption_label": assumption_label,
                    "source": invented_source_tag,
                }
            )
            change_counts["assumed_gaillac_supplier_edge_added"] += 1
            changed_edge_ids.append(invented_edge_id)
            invented_entities.append(
                {
                    "entity_type": "edge",
                    "id": invented_edge_id,
                    "label": f"SUPPLY_LINK_{assumption_label}_MISSING_INPUT_{invented_item_code}",
                }
            )
            node_by_id, coords = build_node_maps(nodes)
            item_unit_map = infer_item_unit_map(nodes, edges)

        # Ensure destination factory tracks modeled inventory state for this input.
        dst_inv = dst_node.get("inventory")
        if not isinstance(dst_inv, dict):
            dst_inv = {"states": [], "backlogs": [], "wip": []}
            dst_node["inventory"] = dst_inv
        dst_states = dst_inv.get("states")
        if not isinstance(dst_states, list):
            dst_states = []
            dst_inv["states"] = dst_states
        has_dst_state = any(str(s.get("item_id")) == invented_item_id for s in dst_states)
        if not has_dst_state:
            dst_states.append(
                {
                    "item_id": invented_item_id,
                    "state_id": f"I_{invented_item_code}_ASSUMED_Q",
                    "initial": 700.0,
                    "uom": unit,
                    "holding_cost": {
                        "value": 0.0,
                        "per": "unit*day",
                        "is_default": True,
                        "source": "simulation_prep_assumption_pending_value_based",
                    },
                    "is_default_initial": False,
                    "initial_source": invented_source_tag,
                    "uom_source": invented_source_tag,
                    "assumption_label": assumption_label,
                }
            )
            change_counts["assumed_gaillac_destination_inventory_state_added"] += 1
            changed_node_ids.append(invented_dst_node)

    customer_ids = {
        str(n.get("id"))
        for n in nodes
        if str(n.get("type") or "") == "customer"
    }
    customer_item_pairs: set[tuple[str, str]] = set()
    for e in edges:
        dst = str(e.get("to"))
        if dst not in customer_ids:
            continue
        for item_id in (e.get("items") or []):
            customer_item_pairs.add((dst, str(item_id)))

    # Edge-level enrichment
    for e in edges:
        eid = str(e.get("id"))
        src = str(e.get("from"))
        dst = str(e.get("to"))
        src_type = str((node_by_id.get(src) or {}).get("type") or "")
        dst_type = str((node_by_id.get(dst) or {}).get("type") or "")

        src_c = coords.get(src)
        dst_c = coords.get(dst)
        computed_dist = None
        if src_c and dst_c:
            computed_dist = round(haversine_km(src_c[0], src_c[1], dst_c[0], dst_c[1]), 1)

        if e.get("distance_km") is None and computed_dist is not None:
            e["distance_km"] = computed_dist
            change_counts["edge_distance_filled"] += 1
            changed_edge_ids.append(eid)

        effective_dist = to_float(e.get("distance_km")) or computed_dist or 500.0

        # Try to align pricing terms from Data_poc Relations_acteurs.
        ot = e.get("order_terms") or {}
        order_terms_changed = False
        for item_id in (e.get("items") or []):
            key = (src, dst, str(item_id))
            rec = price_map.get(key)
            if not rec:
                continue
            target_price = to_float(rec.get("sell_price"))
            target_base = to_float(rec.get("price_base")) or 1.0
            target_unit = normalize_unit(rec.get("quantity_unit"))
            cur_price = to_float(ot.get("sell_price"))
            cur_base = to_float(ot.get("price_base"))
            cur_unit = normalize_unit(ot.get("quantity_unit"))

            if target_price is not None and (cur_price is None or abs(cur_price - target_price) > 1e-12):
                ot["sell_price"] = target_price
                order_terms_changed = True
            if cur_base is None or abs(cur_base - target_base) > 1e-12:
                ot["price_base"] = target_base
                order_terms_changed = True
            if target_unit and cur_unit != target_unit:
                ot["quantity_unit"] = target_unit
                order_terms_changed = True

        if order_terms_changed:
            ot["source"] = "data_poc_relations_acteurs"
            e["order_terms"] = ot
            change_counts["edge_order_terms_pricing_aligned_from_data_poc"] += 1
            if eid not in changed_edge_ids:
                changed_edge_ids.append(eid)

    item_unit_value_map, holding_cost_value_stats = derive_item_unit_value_map(edges, item_unit_map, price_map)
    fallback_global_unit_value = float(
        holding_cost_value_stats.get("fallback_global_unit_value_per_item_unit") or 1.0
    )

    # Normalize scenario horizons to a consistent simulation window.
    target_sim_days = max(1, int(simulation_days))
    for scn in scenarios:
        h = scn.get("horizon")
        if not isinstance(h, dict):
            h = {}
        current_steps = int(to_float(h.get("steps_to_run")) or 0)
        current_unit = str(h.get("time_unit") or "").strip()
        if current_steps != target_sim_days or current_unit.lower() != "day":
            h["steps_to_run"] = target_sim_days
            h["time_unit"] = "Day"
            scn["horizon"] = h
            change_counts["scenario_horizon_updated"] += 1

    for e in edges:
        eid = str(e.get("id"))
        src = str(e.get("from"))
        dst = str(e.get("to"))
        src_type = str((node_by_id.get(src) or {}).get("type") or "")
        dst_type = str((node_by_id.get(dst) or {}).get("type") or "")

        src_c = coords.get(src)
        dst_c = coords.get(dst)
        computed_dist = None
        if src_c and dst_c:
            computed_dist = round(haversine_km(src_c[0], src_c[1], dst_c[0], dst_c[1]), 1)
        effective_dist = to_float(e.get("distance_km")) or computed_dist or 500.0

        lead = e.get("lead_time") or {}
        if not isinstance(lead, dict):
            lead = {}
        if lead.get("is_default") is True or to_float(lead.get("mean")) is None:
            lead["type"] = "erlang_pipeline"
            lead["mean"] = estimate_lead_time_days(effective_dist, src_type, dst_type)
            lead["stages"] = int(lead.get("stages") or 4)
            lead["time_unit"] = "day"
            lead["is_default"] = False
            lead["source"] = "simulation_prep_assumption"
            e["lead_time"] = lead
            change_counts["edge_lead_time_updated"] += 1
            if eid not in changed_edge_ids:
                changed_edge_ids.append(eid)

        tc = e.get("transport_cost") or {}
        if not isinstance(tc, dict):
            tc = {}
        tc_value = to_float(tc.get("value"))
        if tc.get("is_default") is True or tc_value is None or tc_value == 0:
            tc["value"] = estimate_transport_cost_per_unit(effective_dist)
            tc["per"] = "unit"
            tc["is_default"] = False
            tc["source"] = "simulation_prep_assumption"
            e["transport_cost"] = tc
            change_counts["edge_transport_cost_updated"] += 1
            if eid not in changed_edge_ids:
                changed_edge_ids.append(eid)

        dsl = e.get("delay_step_limit") or {}
        if not isinstance(dsl, dict):
            dsl = {}
        if dsl.get("is_default") is True or dsl.get("value") in (None, 999):
            dsl["value"] = 21
            dsl["is_default"] = False
            dsl["source"] = "simulation_prep_assumption"
            e["delay_step_limit"] = dsl
            change_counts["edge_delay_limit_updated"] += 1
            if eid not in changed_edge_ids:
                changed_edge_ids.append(eid)

    # Node-level enrichment
    node_base_stock = {
        "supplier_dc": 180.0,
        "factory": 0.0,
        "distribution_center": 120.0,
        "customer": 0.0,
    }
    for n in nodes:
        nid = str(n.get("id"))
        ntype = str(n.get("type") or "unknown")
        inv = n.get("inventory") or {}
        states = inv.get("states") or []
        factory_output_items = {
            str(out.get("item_id"))
            for p in (n.get("processes") or [])
            for out in (p.get("outputs") or [])
            if out.get("item_id")
        }
        node_changed = False

        for st in states:
            item_id = str(st.get("item_id"))
            target_uom = item_unit_map.get(item_id, "")
            current_uom = normalize_unit(st.get("uom"))
            if target_uom and current_uom and target_uom != current_uom:
                st["uom"] = target_uom
                st["uom_source"] = "simulation_prep_inferred_from_relations_bom"
                change_counts["inventory_uom_harmonized"] += 1
                node_changed = True

            initial = to_float(st.get("initial"))
            if initial is None or initial <= 0:
                if ntype == "factory" and item_id in factory_output_items:
                    # If no explicit data exists, finished-goods start at zero to observe production dynamics.
                    new_initial = 0.0
                    st["initial_source"] = "simulation_prep_assumption_factory_output_zero"
                else:
                    base = node_base_stock.get(ntype, 500.0)
                    new_initial = round(base, 2)
                    if ntype == "customer":
                        new_initial = 0.0
                    st["initial_source"] = "simulation_prep_assumption"
                st["initial"] = new_initial
                st["is_default_initial"] = False
                change_counts["inventory_initial_updated"] += 1
                node_changed = True

            hc = st.get("holding_cost") or {}
            if isinstance(hc, dict):
                hc_val = to_float(hc.get("value"))
                source = str(hc.get("source") or "")
                unit_for_cost = target_uom or current_uom
                item_id = str(st.get("item_id") or "")
                target_hc, target_hc_source, unit_value_used = holding_cost_per_unit_day_from_value(
                    item_id=item_id,
                    unit=unit_for_cost,
                    item_unit_map=item_unit_map,
                    item_unit_value_map=item_unit_value_map,
                    fallback_global_unit_value=fallback_global_unit_value,
                    annual_carry_rate=annual_carry_rate,
                )
                should_fill = hc.get("is_default") is True or hc_val is None or hc_val == 0
                should_rescale_assumed = source.startswith("simulation_prep_") and (
                    hc_val is None or abs(hc_val - target_hc) > 1e-12
                )
                if should_fill or should_rescale_assumed:
                    hc["value"] = target_hc
                    hc["per"] = "unit*day"
                    hc["is_default"] = False
                    hc["source"] = f"simulation_prep_{target_hc_source}"
                    hc["annual_carry_rate"] = round(max(0.0, annual_carry_rate), 6)
                    hc["unit_value_basis"] = round(max(0.0, unit_value_used), 6)
                    st["holding_cost"] = hc
                    change_counts["inventory_holding_cost_updated"] += 1
                    change_counts[f"inventory_holding_cost_source::{target_hc_source}"] += 1
                    node_changed = True

        policies = n.get("policies")
        if not isinstance(policies, dict):
            policies = {}
        sim_policy = policies.get("simulation_policy")
        if not isinstance(sim_policy, dict):
            sim_policy = {}
        if not sim_policy:
            sim_policy.update(
                {
                    "service_level_target": 0.95,
                    "review_period_days": 7,
                    "reorder_mode": "base_stock",
                    "source": "simulation_prep_assumption",
                }
            )
            policies["simulation_policy"] = sim_policy
            n["policies"] = policies
            change_counts["node_policy_added"] += 1
            node_changed = True

        processes = n.get("processes") or []
        for p in processes:
            for inp in (p.get("inputs") or []):
                item_id = str(inp.get("item_id"))
                target_uom = item_unit_map.get(item_id, "")
                current_uom = normalize_unit(inp.get("ratio_unit"))
                if target_uom and not current_uom:
                    inp["ratio_unit"] = target_uom
                    change_counts["process_input_uom_filled"] += 1
                    node_changed = True

            cap = p.get("capacity") or {}
            if isinstance(cap, dict):
                max_rate = to_float(cap.get("max_rate")) or 0.0
                if cap.get("is_default") is True:
                    batch_size = to_float(p.get("batch_size")) or 1000.0
                    cap["max_rate"] = round(max(max_rate, batch_size / 40.0), 2)
                    cap["uom"] = cap.get("uom") or "unit/day"
                    cap["is_default"] = False
                    cap["source"] = "simulation_prep_assumption"
                    p["capacity"] = cap
                    change_counts["process_capacity_updated"] += 1
                    node_changed = True

            cost = p.get("cost") or {}
            if isinstance(cost, dict):
                cval = to_float(cost.get("value"))
                if cost.get("is_default") is True or cval is None or cval == 0:
                    cost["value"] = 0.35
                    cost["per"] = "unit"
                    cost["is_default"] = False
                    cost["source"] = "simulation_prep_assumption"
                    p["cost"] = cost
                    change_counts["process_cost_updated"] += 1
                    node_changed = True

        if node_changed:
            changed_node_ids.append(nid)

    # Capacity envelope by finished item after process-enrichment
    # (used to build realistic default demand assumptions).
    item_production_capacity: dict[str, float] = defaultdict(float)
    for n in nodes:
        for p in (n.get("processes") or []):
            cap = to_float(((p.get("capacity") or {}).get("max_rate"))) or 0.0
            if cap <= 0:
                continue
            for out in (p.get("outputs") or []):
                item_id = str(out.get("item_id"))
                if item_id:
                    item_production_capacity[item_id] += cap

    # Demand enrichment
    for scn in scenarios:
        sid = str(scn.get("id") or "unknown_scn")
        demands = scn.get("demand") or []
        if not isinstance(demands, list):
            demands = []
        existing_pairs = {
            (str(d.get("node_id")), str(d.get("item_id")))
            for d in demands
            if isinstance(d, dict)
        }

        for pair in sorted(customer_item_pairs):
            if pair in existing_pairs:
                continue
            node_id, item_id = pair
            demands.append(
                {
                    "node_id": node_id,
                    "item_id": item_id,
                    "profile": [
                        {
                            "type": "constant",
                            "value": 0.0,
                            "uom": "unit/day",
                            "is_default": True,
                        }
                    ],
                    "defaults": {"demand": True},
                }
            )
            change_counts["demand_rows_added"] += 1
            changed_demand_rows.append(
                {
                    "scenario_id": sid,
                    "node_id": node_id,
                    "item_id": item_id,
                    "action": "added_missing_customer_demand",
                }
            )
        scn["demand"] = demands

        for d in demands:
            if not isinstance(d, dict):
                continue
            node_id = str(d.get("node_id") or "")
            item_id = canonical_item_id(d.get("item_id"))
            profile = d.get("profile") or []
            if not isinstance(profile, list) or not profile:
                profile = [{"type": "constant", "value": 0.0, "uom": "unit/day", "is_default": True}]
                d["profile"] = profile

            if item_id in PRODUCT_SERVICE_TARGETS:
                d["service_level_target"] = PRODUCT_SERVICE_TARGETS[item_id]
                d["service_level_target_source"] = "simulation_prep_business_rule_by_product"

            demand_pair = (node_id, item_id)
            if demand_pair in demand_pf_map:
                annual_total = round(sum(demand_pf_map[demand_pair]), 6)
                d["profile"] = weekly_demand_piecewise_profile(
                    demand_pf_map[demand_pair],
                    target_sim_days,
                    source="demand_pf_real_demand_weekly_uniform_daily",
                )
                d.setdefault("defaults", {})
                d["defaults"]["demand"] = False
                d["source_truth"] = {
                    "type": "demand_pf_weekly_real_demand",
                    "xlsx_path": str(demand_pf_xlsx) if demand_pf_xlsx is not None else "",
                    "sheet": "Demande",
                    "annual_total_qty": annual_total,
                    "period_count_weeks": len(demand_pf_map[demand_pair]),
                }
                change_counts["demand_profile_loaded_from_demand_pf"] += 1
                changed_demand_rows.append(
                    {
                        "scenario_id": sid,
                        "node_id": node_id,
                        "item_id": item_id,
                        "action": "loaded_from_demand_pf_weekly_real_demand",
                        "annual_total_qty": annual_total,
                    }
                )
                continue

            is_all_zero = True
            for p in profile:
                val = to_float((p or {}).get("value"))
                if val is not None and abs(val) > 1e-9:
                    is_all_zero = False
                    break

            if is_all_zero:
                if str((node_by_id.get(node_id) or {}).get("type")) == "customer":
                    # Realistic default: align demand to nominal producible envelope when not provided.
                    cap = to_float(item_production_capacity.get(item_id)) or 0.0
                    base = round(max(10.0, 1.0 * cap), 2) if cap > 0 else 25.0
                    d["profile"] = weekly_fluctuating_profile(base, target_sim_days)
                else:
                    base = 40.0
                    for p in profile:
                        if isinstance(p, dict):
                            p["type"] = p.get("type") or "constant"
                            p["value"] = base
                            p["uom"] = p.get("uom") or "unit/day"
                            p["is_default"] = False
                            p["source"] = "simulation_prep_assumption"
                change_counts["demand_profile_updated"] += 1
                changed_demand_rows.append(
                    {
                        "scenario_id": sid,
                        "node_id": node_id,
                        "item_id": str(d.get("item_id")),
                        "action": "filled_zero_or_missing_profile",
                    }
                )

    for n in nodes:
        if str(n.get("type") or "") != "customer":
            continue
        nid = str(n.get("id") or "")
        item_targets = {
            item_id: PRODUCT_SERVICE_TARGETS[item_id]
            for customer_id, item_id in sorted(customer_item_pairs)
            if customer_id == nid and item_id in PRODUCT_SERVICE_TARGETS
        }
        if not item_targets:
            continue
        policies = n.get("policies")
        if not isinstance(policies, dict):
            policies = {}
            n["policies"] = policies
        sim_policy = policies.get("simulation_policy")
        if not isinstance(sim_policy, dict):
            sim_policy = {}
            policies["simulation_policy"] = sim_policy
        if sim_policy.get("service_level_target_by_item") != item_targets:
            sim_policy["service_level_target_by_item"] = item_targets
            sim_policy["service_level_target_source"] = "simulation_prep_business_rule_by_product"
            change_counts["customer_service_target_by_item_updated"] += 1
            changed_node_ids.append(nid)

    # Warm-start service buffers on demand-serving lanes to avoid cold-start bias:
    # - customer buffer: 1 day of demand
    # - immediate upstream DC buffer: 2 days of demand
    customer_buffer_days = max(0.0, float(customer_warm_start_days))
    upstream_dc_buffer_days = max(0.0, float(upstream_dc_warm_start_days))
    demand_day0_by_pair: dict[tuple[str, str], float] = defaultdict(float)
    for scn in scenarios:
        for d in (scn.get("demand") or []):
            if not isinstance(d, dict):
                continue
            pair = (str(d.get("node_id")), str(d.get("item_id")))
            v = profile_day0_value(d.get("profile") or [])
            if v > demand_day0_by_pair[pair]:
                demand_day0_by_pair[pair] = v

    inbound_to_customer_item: dict[tuple[str, str], list[str]] = defaultdict(list)
    for e in edges:
        dst = str(e.get("to"))
        if str((node_by_id.get(dst) or {}).get("type") or "") != "customer":
            continue
        src = str(e.get("from"))
        for item_id in (e.get("items") or []):
            inbound_to_customer_item[(dst, str(item_id))].append(src)

    def ensure_inventory_state(node_id: str, item_id: str, target_initial: float, initial_source: str) -> None:
        nonlocal change_counts, changed_node_ids
        n = node_by_id.get(node_id)
        if not isinstance(n, dict):
            return
        inv = n.get("inventory")
        if not isinstance(inv, dict):
            inv = {"states": [], "backlogs": [], "wip": []}
            n["inventory"] = inv
        states = inv.get("states")
        if not isinstance(states, list):
            states = []
            inv["states"] = states
        state = next((s for s in states if str(s.get("item_id")) == item_id), None)
        target_initial = round(max(0.0, target_initial), 6)
        unit = item_unit_map.get(item_id, "UN")
        if not isinstance(state, dict):
            state = {
                "item_id": item_id,
                "state_id": f"I_{item_id.replace('item:', '')}_{node_id.replace('-', '_')}",
                "initial": target_initial,
                "uom": unit,
                "holding_cost": {
                    "value": 0.0,
                    "per": "unit*day",
                    "is_default": True,
                    "source": "simulation_prep_assumption_pending_value_based",
                },
                "is_default_initial": False,
                "initial_source": initial_source,
                "uom_source": "simulation_prep_inferred_from_relations_bom",
            }
            states.append(state)
            change_counts["inventory_state_added_for_service_warm_start"] += 1
            changed_node_ids.append(node_id)
        else:
            current = to_float(state.get("initial")) or 0.0
            if target_initial > current + 1e-9:
                state["initial"] = target_initial
                state["is_default_initial"] = False
                state["initial_source"] = initial_source
                if not normalize_unit(state.get("uom")) and unit:
                    state["uom"] = unit
                    state["uom_source"] = "simulation_prep_inferred_from_relations_bom"
                change_counts["inventory_initial_updated_for_service_warm_start"] += 1
                changed_node_ids.append(node_id)

    for (customer_id, item_id), d0 in sorted(demand_day0_by_pair.items()):
        if d0 <= 0:
            continue
        if customer_buffer_days > 0:
            customer_target = d0 * customer_buffer_days
            ensure_inventory_state(
                customer_id,
                item_id,
                customer_target,
                "simulation_prep_assumption_customer_warm_start_buffer",
            )
        for src in inbound_to_customer_item.get((customer_id, item_id), []):
            src_type = str((node_by_id.get(src) or {}).get("type") or "")
            if src_type != "distribution_center":
                continue
            if upstream_dc_buffer_days > 0:
                upstream_target = d0 * upstream_dc_buffer_days
                ensure_inventory_state(
                    src,
                    item_id,
                    upstream_target,
                    "simulation_prep_assumption_upstream_dc_warm_start_buffer",
                )

    # Validation snapshot after enrichment
    missing_geo_nodes = []
    for n in nodes:
        geo = n.get("geo") or {}
        if to_float(geo.get("lat")) is None or to_float(geo.get("lon")) is None:
            missing_geo_nodes.append(str(n.get("id")))

    edges_missing_distance = [str(e.get("id")) for e in edges if to_float(e.get("distance_km")) is None]
    edges_zero_transport_cost = []
    edges_factory_input_missing_sell_price = []
    for e in edges:
        tc = e.get("transport_cost") or {}
        if to_float((tc or {}).get("value")) in (None, 0.0):
            edges_zero_transport_cost.append(str(e.get("id")))
        dst = str(e.get("to"))
        dst_type = str((node_by_id.get(dst) or {}).get("type") or "")
        if dst_type == "factory":
            ot = e.get("order_terms") or {}
            if to_float((ot or {}).get("sell_price")) is None:
                edges_factory_input_missing_sell_price.append(str(e.get("id")))
    zero_demand_rows = []
    for scn in scenarios:
        sid = str(scn.get("id") or "")
        for d in (scn.get("demand") or []):
            vals = []
            for p in (d.get("profile") or []):
                if not isinstance(p, dict):
                    continue
                ptype = str(p.get("type", "constant")).lower()
                if ptype == "piecewise":
                    for pt in (p.get("points") or []):
                        if not isinstance(pt, dict):
                            continue
                        v = to_float(pt.get("value"))
                        if v is not None:
                            vals.append(v)
                    continue
                v = to_float(p.get("value"))
                if v is not None:
                    vals.append(v)
            if not vals or all(v == 0 for v in vals):
                zero_demand_rows.append({"scenario_id": sid, "node_id": str(d.get("node_id")), "item_id": str(d.get("item_id"))})

    report = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "changes": dict(sorted(change_counts.items())),
        "invented_entities": invented_entities,
        "item_unit_map_sample": dict(sorted(item_unit_map.items())[:20]),
        "changed_entities": {
            "edge_count": len(set(changed_edge_ids)),
            "node_count": len(set(changed_node_ids)),
            "demand_rows_count": len(changed_demand_rows),
            "demand_rows": changed_demand_rows,
        },
        "validation_after_prep": {
            "missing_geo_nodes_count": len(missing_geo_nodes),
            "missing_geo_nodes": sorted(missing_geo_nodes),
            "edges_missing_distance_count": len(edges_missing_distance),
            "edges_zero_transport_cost_count": len(edges_zero_transport_cost),
            "edges_factory_input_missing_sell_price_count": len(edges_factory_input_missing_sell_price),
            "edges_factory_input_missing_sell_price": sorted(edges_factory_input_missing_sell_price),
            "zero_demand_rows_count": len(zero_demand_rows),
            "zero_demand_rows": zero_demand_rows,
        },
        "data_poc_price_integration": price_import_stats,
        "demand_pf_integration": demand_pf_stats,
        "demand_pf_actor_integration": demand_pf_actor_stats,
        "demand_pf_relation_integration": demand_pf_relation_stats,
        "assumptions": {
            "inventory_base_stock_by_node_type": node_base_stock,
            "delay_step_limit_assumed": 21,
            "holding_cost_model_assumed": {
                "formula": "item_unit_value * annual_carry_rate / 365",
                "annual_carry_rate": round(max(0.0, annual_carry_rate), 6),
                "item_unit_value_basis": "median(sell_price / price_base) per item after Data_poc pricing alignment",
                "fallback_unit_value_basis": "global median priced item-unit value",
                "priced_items": holding_cost_value_stats.get("priced_items", 0),
                "priced_edge_item_pairs": holding_cost_value_stats.get("priced_edge_item_pairs", 0),
                "fallback_global_unit_value_per_item_unit": round(fallback_global_unit_value, 6),
            },
            "process_cost_per_unit_assumed": 0.35,
            "demand_default_customer_rule": "weekly_fluctuating_profile(base=max(10, 1.0 * total_production_capacity_of_item))",
            "demand_default_customer_weekly_pattern": [0.8, 0.95, 1.05, 1.2, 1.1, 1.0, 0.9],
            "demand_constant_assumed_customer_fallback": 25.0,
            "demand_constant_assumed_other": 40.0,
            "service_warm_start_customer_buffer_days": customer_buffer_days,
            "service_warm_start_upstream_dc_buffer_days": upstream_dc_buffer_days,
            "dc_alias_reconciliation": "DC-1910 merged into DC-1920",
            "customer_location_recovery": "C-XXXXX location_ID set to Paris",
            "assumed_supplier_label_for_unsourced_m1810_item": (
                f"{invented_item_id}: GAILLAC?" if invented_item_id else None
            ),
            "simulation_horizon_days_default": target_sim_days,
            "product_service_targets": PRODUCT_SERVICE_TARGETS,
            "demand_pf_mapping_rule": "weekly demand values from demand_PF.xlsx converted to daily rates via uniform division by 7",
            "demand_pf_tail_rule": "no artificial tail extension beyond the provided weekly source periods",
        },
    }
    return g, report


def report_markdown(report: dict[str, Any], input_path: str, output_graph_path: str) -> str:
    c = report["changes"]
    v = report["validation_after_prep"]
    ch = report["changed_entities"]
    return f"""# Simulation prep report

## Inputs / outputs
- Input graph: {input_path}
- Output graph: {output_graph_path}
- Generated at (UTC): {report['generated_at_utc']}

## What was enriched
- Edge distances filled: {c.get('edge_distance_filled', 0)}
- Edge lead times updated: {c.get('edge_lead_time_updated', 0)}
- Edge transport costs updated: {c.get('edge_transport_cost_updated', 0)}
- Edge delay limits updated: {c.get('edge_delay_limit_updated', 0)}
- Edge pricing aligned from Data_poc Relations_acteurs: {c.get('edge_order_terms_pricing_aligned_from_data_poc', 0)}
- Edge pricing aligned from demand_PF Relations_acteurs: {c.get('edge_order_terms_aligned_from_demand_pf_relations', 0)}
- Nodes added from demand_PF Acteurs: {c.get('node_added_from_demand_pf_acteurs', 0)}
- Node locations filled from demand_PF Acteurs: {c.get('node_location_filled_from_demand_pf_acteurs', 0)}
- Edges added from demand_PF Relations_acteurs: {c.get('edge_added_from_demand_pf_relations', 0)}
- Inventory states added from demand_PF Relations_acteurs: {c.get('inventory_state_added_from_demand_pf_relations', 0)}
- Inventory initials updated: {c.get('inventory_initial_updated', 0)}
- Inventory holding costs updated: {c.get('inventory_holding_cost_updated', 0)}
- Holding-cost source item-value median: {c.get('inventory_holding_cost_source::item_value_median_from_priced_edges', 0)}
- Holding-cost source global fallback: {c.get('inventory_holding_cost_source::global_value_median_fallback', 0)}
- Inventory UOM harmonized: {c.get('inventory_uom_harmonized', 0)}
- Node policies added: {c.get('node_policy_added', 0)}
- Process capacities updated: {c.get('process_capacity_updated', 0)}
- Process costs updated: {c.get('process_cost_updated', 0)}
- DC alias reconciliations (1910->1920): {c.get('node_alias_merged_dc1910_into_dc1920', 0)}
- Customer location recovered: {c.get('customer_location_recovered_from_acteurs', 0)}
- Assumed Gaillac supplier nodes added: {c.get('assumed_gaillac_supplier_node_added', 0)}
- Assumed Gaillac supplier node tags updated: {c.get('assumed_gaillac_supplier_node_tagged', 0)}
- Assumed Gaillac supplier edges added: {c.get('assumed_gaillac_supplier_edge_added', 0)}
- Assumed Gaillac supplier inventory states added: {c.get('assumed_gaillac_supplier_inventory_state_added', 0)}
- Assumed destination inventory states added (M-1810 unsourced input): {c.get('assumed_gaillac_destination_inventory_state_added', 0)}
- Demand rows added: {c.get('demand_rows_added', 0)}
- Demand rows updated: {c.get('demand_profile_updated', 0)}
- Demand rows loaded from demand_PF.xlsx: {c.get('demand_profile_loaded_from_demand_pf', 0)}
- Scenario horizons updated to default simulation days: {c.get('scenario_horizon_updated', 0)}

## Changed entities
- Changed edges: {ch.get('edge_count', 0)}
- Changed nodes: {ch.get('node_count', 0)}
- Changed demand rows: {ch.get('demand_rows_count', 0)}

## Validation after prep
- Missing geo nodes: {v.get('missing_geo_nodes_count', 0)}
- Edges still missing distance: {v.get('edges_missing_distance_count', 0)}
- Edges still zero transport cost: {v.get('edges_zero_transport_cost_count', 0)}
- Factory inbound edges missing sell_price: {v.get('edges_factory_input_missing_sell_price_count', 0)}
- Zero-demand rows remaining: {v.get('zero_demand_rows_count', 0)}

## Data_poc pricing import
- Enabled: {report.get('data_poc_price_integration', {}).get('enabled', False)}
- XLSX path: {report.get('data_poc_price_integration', {}).get('xlsx_path', '')}
- Rows read: {report.get('data_poc_price_integration', {}).get('rows_read', 0)}
- Rows mapped: {report.get('data_poc_price_integration', {}).get('rows_mapped', 0)}
- Error: {report.get('data_poc_price_integration', {}).get('error', 'none')}

## demand_PF import
- Enabled: {report.get('demand_pf_integration', {}).get('enabled', False)}
- XLSX path: {report.get('demand_pf_integration', {}).get('xlsx_path', '')}
- Sheet found: {report.get('demand_pf_integration', {}).get('sheet_found', False)}
- Rows read: {report.get('demand_pf_integration', {}).get('rows_read', 0)}
- Rows mapped: {report.get('demand_pf_integration', {}).get('rows_mapped', 0)}
- Pairs loaded: {report.get('demand_pf_integration', {}).get('pairs_loaded', 0)}
- Annual totals by pair: `{report.get('demand_pf_integration', {}).get('annual_total_per_pair', {})}`
- Error: {report.get('demand_pf_integration', {}).get('error', 'none')}

## demand_PF Acteurs import
- Enabled: {report.get('demand_pf_actor_integration', {}).get('enabled', False)}
- Rows read: {report.get('demand_pf_actor_integration', {}).get('rows_read', 0)}
- Rows mapped: {report.get('demand_pf_actor_integration', {}).get('rows_mapped', 0)}
- Error: {report.get('demand_pf_actor_integration', {}).get('error', 'none')}

## demand_PF Relations_acteurs import
- Enabled: {report.get('demand_pf_relation_integration', {}).get('enabled', False)}
- Rows read: {report.get('demand_pf_relation_integration', {}).get('rows_read', 0)}
- Rows mapped: {report.get('demand_pf_relation_integration', {}).get('rows_mapped', 0)}
- Error: {report.get('demand_pf_relation_integration', {}).get('error', 'none')}

## Holding cost model
- Formula: {report.get('assumptions', {}).get('holding_cost_model_assumed', {}).get('formula', 'n/a')}
- Annual carry rate: {report.get('assumptions', {}).get('holding_cost_model_assumed', {}).get('annual_carry_rate', 'n/a')}
- Item value basis: {report.get('assumptions', {}).get('holding_cost_model_assumed', {}).get('item_unit_value_basis', 'n/a')}
- Fallback unit value basis: {report.get('assumptions', {}).get('holding_cost_model_assumed', {}).get('fallback_unit_value_basis', 'n/a')}
- Priced items used: {report.get('assumptions', {}).get('holding_cost_model_assumed', {}).get('priced_items', 0)}
- Priced edge-item pairs used: {report.get('assumptions', {}).get('holding_cost_model_assumed', {}).get('priced_edge_item_pairs', 0)}
- Fallback global unit value: {report.get('assumptions', {}).get('holding_cost_model_assumed', {}).get('fallback_global_unit_value_per_item_unit', 'n/a')}

## Review reminder
This graph is assumption-based and intended for pre-simulation validation.
Review the assumptions in simulation_prep_report.json before scenario studies.
"""


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    out_graph = Path(args.output_graph)
    out_report_json = Path(args.output_report_json)
    out_report_md = Path(args.output_report_md)
    data_poc_xlsx = Path(args.data_poc_xlsx) if str(args.data_poc_xlsx).strip() else None
    demand_pf_xlsx = Path(args.demand_pf_xlsx) if str(args.demand_pf_xlsx).strip() else None

    out_graph.parent.mkdir(parents=True, exist_ok=True)

    raw = json.loads(input_path.read_text(encoding="utf-8"))
    prepared, report = prepare_graph(
        raw,
        data_poc_xlsx,
        demand_pf_xlsx,
        customer_warm_start_days=args.customer_warm_start_days,
        upstream_dc_warm_start_days=args.upstream_dc_warm_start_days,
        simulation_days=args.simulation_days,
        annual_carry_rate=args.annual_carry_rate,
    )

    out_graph.write_text(json.dumps(prepared, indent=2, ensure_ascii=False), encoding="utf-8")
    out_report_json.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    out_report_md.write_text(
        report_markdown(report, str(input_path), str(out_graph)),
        encoding="utf-8",
    )
    print(f"[OK] Simulation-ready graph written: {out_graph.resolve()}")
    print(f"[OK] Prep report written: {out_report_json.resolve()}")
    print(f"[OK] Prep report (md) written: {out_report_md.resolve()}")


if __name__ == "__main__":
    main()
