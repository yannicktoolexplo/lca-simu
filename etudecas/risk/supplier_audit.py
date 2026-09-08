"""Read supplier-audit workbooks and expose their criticality to the map.

The source workbook is an operational audit template rather than a flat data
table. Its cached Excel results are read from each criterion block, while the
raw answers are checked so an empty copy cannot be scored as a completed audit.
"""

from __future__ import annotations

import html
import csv
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SUPPLIER_AUDIT_XLSX = REPO_ROOT / "etudecas/data/source/Trame d'audit fournisseur finalisé.xlsx"
DEFAULT_SUPPLIER_AUDIT_SOURCE = DEFAULT_SUPPLIER_AUDIT_XLSX.parent
DEFAULT_SUPPLIER_CONTEXT_PROXIES = REPO_ROOT / "etudecas/data/source/supplier_context_proxies.csv"
DEFAULT_SUPPLIER_PUBLIC_EVIDENCE = REPO_ROOT / "etudecas/data/source/supplier_public_evidence.csv"
AUDIT_SHEET_INDEX = 3
CRITERIA_FIRST_ROW = 31
CRITERIA_LAST_ROW = 58
RESILIENCE_THRESHOLD_WEEKS = 70.0
MATURITY_THRESHOLD = 0.75
CRITICALITY_THRESHOLD = 15.0
AUDIT_SCORE_WEIGHT = 0.30

# The audit workbook contains observations that the simulation cannot directly
# observe.  These criterion-level coefficients are therefore explicit proxies,
# never audit answers.  The last three values are respectively a maturity
# offset, an exposure multiplier and a recovery-time multiplier.
CRITERION_PROXY_CONFIG: dict[int, tuple[str, float, float, float]] = {
    31: ("operations", 0.04, 0.82, 0.78),
    32: ("operations", 0.06, 0.94, 0.88),
    33: ("operations", 0.02, 1.00, 0.95),
    34: ("safety", 0.03, 0.92, 0.92),
    35: ("data", 0.05, 0.78, 0.72),
    36: ("operations", 0.02, 1.08, 1.10),
    37: ("operations", 0.00, 0.88, 0.94),
    38: ("operations", -0.02, 1.10, 1.18),
    39: ("data", 0.04, 0.86, 0.82),
    40: ("safety", 0.01, 0.82, 0.82),
    41: ("safety", 0.02, 0.88, 0.86),
    42: ("people", -0.02, 0.92, 0.96),
    43: ("safety", 0.00, 0.98, 1.02),
    44: ("safety", 0.00, 0.94, 0.98),
    45: ("safety", 0.00, 0.90, 0.94),
    46: ("safety", 0.03, 0.88, 0.86),
    47: ("climate", -0.03, 1.06, 1.12),
    48: ("industrial", -0.02, 1.12, 1.18),
    49: ("data", 0.03, 0.82, 0.78),
    50: ("people", -0.02, 0.96, 1.00),
    51: ("market", -0.04, 1.08, 1.08),
    52: ("suppliers", -0.01, 1.12, 1.15),
    53: ("suppliers", 0.00, 0.92, 0.94),
    54: ("suppliers", 0.01, 0.90, 0.90),
    55: ("suppliers", -0.02, 1.04, 1.06),
    56: ("suppliers", -0.01, 0.98, 1.00),
    57: ("continuity", -0.01, 1.18, 1.22),
    58: ("transport", 0.00, 1.08, 1.14),
}

FAMILY_ROWS: tuple[tuple[str, tuple[int, ...]], ...] = (
    ("Technique & production", (31, 32, 33, 36, 37, 38)),
    ("Organisation & données", (35, 39, 49, 57)),
    ("RH & sécurité", (34, 40, 41, 42, 43, 44, 45, 46)),
    ("Risques externes", (47, 48, 58)),
    ("Fournisseurs", (52, 53, 54, 55, 56)),
    ("Activité & marché", (51,)),
)

QUALITY_WARNINGS = (
    "La formule Excel de priorité mélange une maturité stockée entre 0 et 1 avec une échelle sur 100.",
    "L'indice carte recalcule donc les trois dimensions sur une échelle commune et conserve séparément les valeurs Excel.",
    "Neuf agrégats de criticité omettent des lignes renseignées et quatre valeurs C saisies en dur ne correspondent pas à P × I ; elles restent signalées comme données source à valider.",
    "La synthèse H31:H58 est exprimée en semaines malgré un libellé intermédiaire en jours ; H59 ne doit pas être utilisé comme total hebdomadaire.",
    "Les résultats proviennent du cache de calcul enregistré dans le classeur ; ouvrir puis enregistrer le fichier dans Excel après toute modification.",
)


def _as_float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _finite_float(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _mean(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _is_answered(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _display_source_path(source: Path) -> str:
    try:
        return source.resolve().relative_to(REPO_ROOT).as_posix()
    except ValueError:
        return source.as_posix()


def _node_has_coordinates(node: dict[str, Any]) -> bool:
    geo = node.get("geo") if isinstance(node.get("geo"), dict) else {}
    lat = _finite_float(node.get("lat"))
    lon = _finite_float(node.get("lon"))
    if lat is None:
        lat = _finite_float(geo.get("lat"))
    if lon is None:
        lon = _finite_float(geo.get("lon"))
    return lat is not None and lon is not None


def _supplier_id(*values: Any) -> str:
    for value in values:
        match = re.search(r"VD\s*[-_ ]?\s*(\d+[A-Z]?)", str(value or ""), flags=re.IGNORECASE)
        if match:
            return f"SDC-VD{match.group(1).upper()}"
    return ""


def _criticality_level(value: float) -> int:
    if value > 12.0:
        return 5
    if value >= 9.0:
        return 4
    if value >= 5.0:
        return 3
    if value >= 3.0:
        return 2
    return 1


def _resilience_level(weeks: float) -> int:
    if weeks > 300.0:
        return 5
    if weeks > 150.0:
        return 4
    if weeks >= 60.0:
        return 3
    if weeks >= 15.0:
        return 2
    return 1


ACTION_MATRIX: tuple[tuple[str, ...], ...] = (
    ("Pas d'action", "Pas d'action", "Pas d'action", "Action moyen terme", "Action moyen terme"),
    ("Pas d'action", "Pas d'action", "Action moyen terme", "Action moyen terme", "Action court terme"),
    ("Pas d'action", "Action moyen terme", "Action court terme", "Action court terme", "Action immédiate"),
    ("Action moyen terme", "Action moyen terme", "Action court terme", "Action immédiate", "Mesure conservatoire"),
    ("Action moyen terme", "Action court terme", "Action immédiate", "Mesure conservatoire", "Mesure conservatoire"),
)


def audit_risk_index(maturity: float, criticality: float, resilience_weeks: float) -> float:
    """Return a normalized 0..1 audit risk index.

    The 20/40/40 weights come from the workbook dashboard.  Maturity is first
    converted from a 0..1 capability score to a risk deficit; criticality is
    already designed on a 0..100 scale; recovery time is capped at the explicit
    70-week dashboard threshold.
    """

    maturity_risk = 1.0 - max(0.0, min(1.0, maturity))
    criticality_risk = max(0.0, min(1.0, criticality / 100.0))
    resilience_risk = max(0.0, min(1.0, resilience_weeks / RESILIENCE_THRESHOLD_WEEKS))
    return max(0.0, min(1.0, 0.20 * maturity_risk + 0.40 * criticality_risk + 0.40 * resilience_risk))


def audit_action(criticality: float, resilience_weeks: float) -> str:
    return ACTION_MATRIX[_criticality_level(criticality) - 1][_resilience_level(resilience_weeks) - 1]


def load_supplier_audits(path: Path | str) -> dict[str, dict[str, Any]]:
    """Load the audit currently stored in ``path``, keyed by simulation node id."""

    source = Path(path)
    if not source.exists():
        return {}
    if source.is_dir():
        candidates = sorted(
            (
                candidate
                for candidate in source.glob("*.xlsx")
                if not candidate.name.startswith("~$")
                if "audit" in candidate.name.casefold() and "fournisseur" in candidate.name.casefold()
            ),
            key=lambda candidate: (
                "finalis" in candidate.name.casefold(),
                candidate.stat().st_mtime,
                candidate.name.casefold(),
            ),
        )
        merged: dict[str, dict[str, Any]] = {}
        for candidate in candidates:
            merged.update(load_supplier_audits(candidate))
        return merged
    # The criteria sheet is wide and accessed by coordinates. A normal workbook
    # load avoids the repeated XML scans caused by random access in read-only mode.
    workbook = load_workbook(source, data_only=True, read_only=False)
    if len(workbook.worksheets) <= AUDIT_SHEET_INDEX:
        raise ValueError(f"Supplier audit workbook has no criteria sheet: {source}")
    sheet = workbook.worksheets[AUDIT_SHEET_INDEX]
    dashboard = workbook.worksheets[0]
    supplier_field_id = _supplier_id(sheet["K2"].value)
    product_field_id = _supplier_id(sheet["R2"].value)
    supplier_id = supplier_field_id or product_field_id
    if not supplier_id:
        raise ValueError(f"Unable to identify a VD supplier in audit workbook: {source}")

    criteria_by_row: dict[int, dict[str, Any]] = {}
    calculation_corrections: list[dict[str, Any]] = []
    working_days_per_week = _as_float(sheet["K31"].value) or 5.0
    for criterion_index, row_index in enumerate(range(CRITERIA_FIRST_ROW, CRITERIA_LAST_ROW + 1)):
        label = str(sheet.cell(row_index, 3).value or "").strip()
        if not label:
            continue
        block_start_column = 3 + 7 * criterion_index
        cached_maturity = _as_float(sheet.cell(row_index, 6).value)
        cached_criticality = _as_float(sheet.cell(row_index, 7).value)
        cached_resilience = _as_float(sheet.cell(row_index, 8).value)
        maturity_raw = sheet.cell(27, block_start_column + 1).value
        criticality_raw = sheet.cell(26, block_start_column + 4).value
        resilience_raw = sheet.cell(27, block_start_column + 6).value
        maturity_value = _finite_float(maturity_raw)
        criticality_value = _finite_float(criticality_raw)
        resilience_value = _finite_float(resilience_raw)
        calculation_cache_ready = all(
            value is not None for value in (maturity_value, criticality_value, resilience_value)
        )
        maturity = maturity_value if maturity_value is not None else 0.0
        criticality = criticality_value if criticality_value is not None else 0.0
        resilience = (resilience_value if resilience_value is not None else 0.0) / working_days_per_week
        question_rows = [
            question_row
            for question_row in range(13, 25)
            if _is_answered(sheet.cell(question_row, block_start_column).value)
        ]
        answered_question_count = sum(
            all(
                _is_answered(sheet.cell(question_row, block_start_column + offset).value)
                for offset in (1, 2, 3, 5)
            )
            for question_row in question_rows
        )
        any_answer = any(
            _is_answered(sheet.cell(question_row, block_start_column + offset).value)
            for question_row in question_rows
            for offset in (1, 2, 3, 5)
        )
        response_status = (
            "complete"
            if question_rows and answered_question_count == len(question_rows)
            else "partial"
            if any_answer
            else "not_assessed"
        )
        cached_values = (cached_maturity, cached_criticality, cached_resilience)
        corrected_values = (maturity, criticality, resilience)
        if any(abs(cached - corrected) > 1e-9 for cached, corrected in zip(cached_values, corrected_values)):
            calculation_corrections.append(
                {
                    "criterion": label,
                    "summary_row": row_index,
                    "cached": {
                        "maturity": round(cached_maturity, 6),
                        "criticality": round(cached_criticality, 6),
                        "resilience_weeks": round(cached_resilience, 6),
                    },
                    "corrected": {
                        "maturity": round(maturity, 6),
                        "criticality": round(criticality, 6),
                        "resilience_weeks": round(resilience, 6),
                    },
                }
            )
        criteria_by_row[row_index] = {
            "row": row_index,
            "label": label,
            "response_status": response_status,
            "calculation_cache_ready": calculation_cache_ready,
            "question_count": len(question_rows),
            "answered_question_count": answered_question_count,
            "maturity": round(maturity, 6),
            "criticality": round(criticality, 6),
            "resilience_weeks": round(resilience, 6),
            "risk_index": round(audit_risk_index(maturity, criticality, resilience), 6),
            "action": audit_action(criticality, resilience),
            "cached_summary": {
                "maturity": round(cached_maturity, 6),
                "criticality": round(cached_criticality, 6),
                "resilience_weeks": round(cached_resilience, 6),
            },
        }

    families: list[dict[str, Any]] = []
    for family, rows in FAMILY_ROWS:
        selected = [criteria_by_row[row] for row in rows if row in criteria_by_row]
        maturity = _mean([float(row["maturity"]) for row in selected])
        criticality = _mean([float(row["criticality"]) for row in selected])
        resilience = _mean([float(row["resilience_weeks"]) for row in selected])
        families.append(
            {
                "family": family,
                "criterion_count": len(selected),
                "maturity": round(maturity, 6),
                "criticality": round(criticality, 6),
                "resilience_weeks": round(resilience, 6),
                "risk_index": round(audit_risk_index(maturity, criticality, resilience), 6),
                "action": audit_action(criticality, resilience),
            }
        )

    assigned_rows = {row for _family, rows in FAMILY_ROWS for row in rows}
    unassigned_criteria = [
        criteria_by_row[row]["label"]
        for row in sorted(set(criteria_by_row) - assigned_rows)
    ]
    completed_criterion_count = sum(
        1 for criterion in criteria_by_row.values() if criterion["response_status"] == "complete"
    )
    partial_criterion_count = sum(
        1 for criterion in criteria_by_row.values() if criterion["response_status"] == "partial"
    )
    question_count = sum(int(criterion["question_count"]) for criterion in criteria_by_row.values())
    answered_question_count = sum(
        int(criterion["answered_question_count"]) for criterion in criteria_by_row.values()
    )
    calculated_criterion_count = sum(
        1 for criterion in criteria_by_row.values() if criterion["calculation_cache_ready"]
    )
    is_complete = (
        bool(criteria_by_row)
        and completed_criterion_count == len(criteria_by_row)
        and calculated_criterion_count == len(criteria_by_row)
    )
    maturity = _mean([float(row["maturity"]) for row in families])
    criticality = _mean([float(row["criticality"]) for row in families])
    resilience = _mean([float(row["resilience_weeks"]) for row in families])
    saved_priority = _as_float(dashboard["I18"].value)
    quality_warnings = list(QUALITY_WARNINGS)
    if product_field_id and not supplier_field_id:
        quality_warnings.append(
            "L'identifiant fournisseur VD est renseigné dans le champ Produit et non dans le champ Fournisseur."
        )
    if unassigned_criteria:
        quality_warnings.append(
            "Critère(s) absent(s) des six familles du Dashboard et donc hors moyenne globale : "
            + ", ".join(unassigned_criteria)
            + "."
        )
    if not is_complete:
        quality_warnings.append(
            f"Audit incomplet : {completed_criterion_count}/{len(criteria_by_row)} critères et "
            f"{answered_question_count}/{question_count} lignes-question sont entièrement renseignés ; "
            f"{calculated_criterion_count}/{len(criteria_by_row)} synthèses ont un résultat Excel disponible. "
            "Aucun score d'audit n'est injecté dans la criticité simulée."
        )
    for correction in calculation_corrections:
        quality_warnings.append(
            "Référence de synthèse corrigée pour "
            + str(correction["criterion"])
            + " : lecture directe de son bloc source utilisée dans la carte."
        )
    result = {
        supplier_id: {
            "audit_status": "audited" if is_complete else "in_progress",
            "supplier_id": supplier_id,
            "supplier_label": str(sheet["K2"].value or supplier_id),
            "product": str(sheet["R2"].value or ""),
            "audited_company": str(sheet["D2"].value or ""),
            "source_file": _display_source_path(source),
            "criterion_count": len(criteria_by_row),
            "answered_criterion_count": completed_criterion_count,
            "partial_criterion_count": partial_criterion_count,
            "question_count": question_count,
            "answered_question_count": answered_question_count,
            "calculated_criterion_count": calculated_criterion_count,
            "aggregated_criterion_count": len(assigned_rows & set(criteria_by_row)),
            "unassigned_criteria": unassigned_criteria,
            "calculation_corrections": calculation_corrections,
            "maturity": round(maturity, 6),
            "criticality": round(criticality, 6),
            "resilience_weeks": round(resilience, 6),
            "audit_risk_index": round(audit_risk_index(maturity, criticality, resilience), 6),
            "saved_excel_priority": round(saved_priority, 6),
            "recommended_action": audit_action(criticality, resilience),
            "families": families,
            "criteria": list(criteria_by_row.values()),
            "quality_warnings": quality_warnings,
            "methodology": {
                "formula": "20% x (1 - maturite) + 40% x criticite/100 + 40% x min(resilience/70 semaines, 1)",
                "weights": {"maturity_deficit": 0.20, "criticality": 0.40, "resilience": 0.40},
                "resilience_threshold_weeks": RESILIENCE_THRESHOLD_WEEKS,
                "simulation_blend": {
                    "structural_criticality": 1.0 - AUDIT_SCORE_WEIGHT,
                    "supplier_audit": AUDIT_SCORE_WEIGHT,
                },
            },
        }
    }
    workbook.close()
    return result


def _clamp(value: float, low: float = 0.0, high: float = 1.0) -> float:
    return max(low, min(high, value))


def _load_context_proxies(path: Path | str = DEFAULT_SUPPLIER_CONTEXT_PROXIES) -> dict[str, dict[str, Any]]:
    source = Path(path)
    if not source.exists():
        return {}
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        return {
            str(row.get("supplier_id") or "").strip(): dict(row)
            for row in csv.DictReader(handle)
            if str(row.get("supplier_id") or "").strip()
        }


def load_supplier_public_evidence(
    path: Path | str = DEFAULT_SUPPLIER_PUBLIC_EVIDENCE,
) -> dict[str, list[dict[str, Any]]]:
    """Load reviewed public facts keyed only by supplier matricule."""

    source = Path(path)
    if not source.exists():
        return {}
    evidence: dict[str, list[dict[str, Any]]] = {}
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        for raw in csv.DictReader(handle):
            supplier_id = str(raw.get("supplier_id") or "").strip()
            source_url = str(raw.get("source_url") or "").strip()
            if not supplier_id or not source_url.startswith(("https://", "http://")):
                continue
            row = {key: str(value or "").strip() for key, value in raw.items()}
            row["confidence"] = round(_clamp(_as_float(raw.get("confidence"))), 4)
            evidence.setdefault(supplier_id, []).append(row)
    for rows in evidence.values():
        rows.sort(
            key=lambda row: (str(row.get("event_date") or ""), str(row.get("evidence_kind") or "")),
            reverse=True,
        )
    return evidence


def _proxy_drivers(
    row: dict[str, Any],
    context: dict[str, Any],
    maxima: dict[str, float],
) -> dict[str, float]:
    def number(key: str) -> float:
        return max(0.0, _as_float(row.get(key)))

    confidence = _clamp(_as_float(context.get("confidence")))
    public_maturity = 0.5 + confidence * (_clamp(_as_float(context.get("public_maturity_signal")), 0.0, 1.0) - 0.5)
    public_resilience = 0.5 + confidence * (_clamp(_as_float(context.get("public_resilience_signal")), 0.0, 1.0) - 0.5)
    geo = 0.30 + confidence * (_clamp(_as_float(context.get("geographic_hazard_signal"))) - 0.30)
    industrial = 0.45 + confidence * (_clamp(_as_float(context.get("industrial_hazard_signal"))) - 0.45)
    active = number("active_days") / max(1.0, maxima["active_days"])
    volume = math.log1p(number("total_shipped_qty")) / max(1.0, maxima["log_volume"])
    shortage = math.log1p(number("shortage_supported_qty")) / max(1.0, maxima["log_shortage"])
    structural = number("structural_criticality_score") / max(0.01, maxima["structural"])
    local = number("local_criticality_score") / max(0.01, maxima["local"])
    system = number("system_criticality_score") / max(0.01, maxima["system"])
    lead = _clamp(number("avg_procurement_lead_days") / 180.0)
    pair_count = number("sole_source_pairs") + number("shared_source_pairs")
    sole = number("sole_source_pairs") / pair_count if pair_count else 0.35
    shared = 1.0 - sole
    no_flow = 1.0 if number("active_days") <= 0.0 else 0.0
    capacity = _clamp(number("max_capacity_utilization"))
    data_quality = _clamp(0.20 + 0.55 * active + 0.15 * (1.0 - no_flow) + 0.10 * min(pair_count, 1.0))
    core = _clamp(
        0.26 * structural
        + 0.14 * local
        + 0.12 * system
        + 0.14 * volume
        + 0.14 * sole
        + 0.10 * lead
        + 0.05 * shortage
        + 0.03 * capacity
        + 0.02 * no_flow
    )
    return {
        "confidence": confidence,
        "public_maturity": public_maturity,
        "public_resilience": public_resilience,
        "geo": _clamp(geo),
        "industrial": _clamp(industrial),
        "active": _clamp(active),
        "volume": _clamp(volume),
        "shortage": _clamp(shortage),
        "structural": _clamp(structural),
        "local": _clamp(local),
        "system": _clamp(system),
        "lead": lead,
        "lead_days": number("avg_procurement_lead_days"),
        "sole": _clamp(sole),
        "shared": _clamp(shared),
        "no_flow": no_flow,
        "capacity": capacity,
        "data_quality": data_quality,
        "core": core,
    }


def _criterion_proxy_values(
    criterion_row: int,
    drivers: dict[str, float],
) -> tuple[float, float, float, float]:
    group, maturity_offset, exposure_factor, recovery_factor = CRITERION_PROXY_CONFIG.get(
        criterion_row, ("operations", 0.0, 1.0, 1.0)
    )
    pub_maturity = drivers["public_maturity"]
    pub_resilience = drivers["public_resilience"]
    active = drivers["active"]
    shared = drivers["shared"]
    data_quality = drivers["data_quality"]
    core = drivers["core"]

    maturity_by_group = {
        "operations": 0.48 * pub_maturity + 0.22 * active + 0.18 * pub_resilience + 0.12 * data_quality,
        "data": 0.42 * pub_maturity + 0.30 * data_quality + 0.18 * active + 0.10 * pub_resilience,
        "safety": 0.58 * pub_maturity + 0.20 * pub_resilience + 0.12 * active + 0.10 * data_quality,
        "people": 0.46 * pub_maturity + 0.22 * active + 0.18 * data_quality + 0.14 * pub_resilience,
        "climate": 0.35 * pub_maturity + 0.35 * pub_resilience + 0.15 * data_quality + 0.15 * active,
        "industrial": 0.45 * pub_maturity + 0.28 * pub_resilience + 0.15 * data_quality + 0.12 * active,
        "market": 0.34 * pub_resilience + 0.24 * shared + 0.22 * active + 0.20 * pub_maturity,
        "suppliers": 0.38 * pub_maturity + 0.24 * shared + 0.22 * pub_resilience + 0.16 * data_quality,
        "continuity": 0.34 * pub_maturity + 0.34 * pub_resilience + 0.18 * shared + 0.14 * data_quality,
        "transport": 0.32 * pub_resilience + 0.24 * active + 0.22 * shared + 0.22 * pub_maturity,
    }
    exposure_by_group = {
        "operations": 0.58 * core + 0.18 * drivers["capacity"] + 0.14 * drivers["shortage"] + 0.10 * drivers["sole"],
        "data": 0.52 * core + 0.28 * (1.0 - data_quality) + 0.20 * drivers["no_flow"],
        "safety": 0.45 * core + 0.35 * drivers["industrial"] + 0.20 * drivers["geo"],
        "people": 0.52 * core + 0.25 * drivers["industrial"] + 0.23 * (1.0 - active),
        "climate": 0.52 * drivers["geo"] + 0.23 * drivers["lead"] + 0.25 * drivers["structural"],
        "industrial": 0.48 * drivers["industrial"] + 0.22 * drivers["capacity"] + 0.30 * core,
        "market": 0.40 * drivers["volume"] + 0.32 * drivers["structural"] + 0.28 * drivers["sole"],
        "suppliers": 0.42 * drivers["sole"] + 0.23 * drivers["lead"] + 0.25 * drivers["structural"] + 0.10 * core,
        "continuity": 0.38 * drivers["structural"] + 0.24 * drivers["lead"] + 0.18 * drivers["geo"] + 0.20 * drivers["industrial"],
        "transport": 0.38 * drivers["lead"] + 0.24 * drivers["geo"] + 0.20 * drivers["sole"] + 0.18 * drivers["structural"],
    }
    maturity = _clamp(maturity_by_group[group] + maturity_offset, 0.20, 0.95)
    exposure = _clamp(exposure_by_group[group] * exposure_factor)
    criticality = max(1.0, min(100.0, 4.0 + 46.0 * exposure))
    recovery_pressure = _clamp(
        0.30 * exposure
        + 0.20 * drivers["sole"]
        + 0.15 * drivers["geo"]
        + 0.15 * drivers["industrial"]
        + 0.10 * drivers["no_flow"]
        + 0.10 * (1.0 - pub_resilience)
    )
    resilience_weeks = max(
        2.0,
        min(
            350.0,
            recovery_factor
            * (2.0 + 0.75 * drivers["lead_days"] / 7.0 + 82.0 * recovery_pressure),
        ),
    )
    estimate_confidence = _clamp(0.30 + 0.38 * drivers["confidence"] + 0.22 * data_quality + 0.10 * (1.0 - drivers["no_flow"]))
    return maturity, criticality, resilience_weeks, estimate_confidence


def estimate_supplier_audit_profiles(
    audits: dict[str, dict[str, Any]],
    ranking_rows: list[dict[str, Any]],
    context_path: Path | str = DEFAULT_SUPPLIER_CONTEXT_PROXIES,
    evidence_path: Path | str = DEFAULT_SUPPLIER_PUBLIC_EVIDENCE,
) -> dict[str, dict[str, Any]]:
    """Fill unaudited profiles with clearly-labelled simulation proxies.

    Estimates use the existing operational exposure metrics, lead times,
    sourcing concentration and conservative public-context signals.  They do
    not count as questionnaire answers and are never returned by
    :func:`supplier_audit_score`.
    """

    contexts = _load_context_proxies(context_path)
    evidence_by_id = load_supplier_public_evidence(evidence_path)
    maxima = {
        "active_days": max((_as_float(row.get("active_days")) for row in ranking_rows), default=1.0),
        "log_volume": max((math.log1p(max(0.0, _as_float(row.get("total_shipped_qty")))) for row in ranking_rows), default=1.0),
        "log_shortage": max((math.log1p(max(0.0, _as_float(row.get("shortage_supported_qty")))) for row in ranking_rows), default=1.0),
        "structural": max((_as_float(row.get("structural_criticality_score")) for row in ranking_rows), default=1.0),
        "local": max((_as_float(row.get("local_criticality_score")) for row in ranking_rows), default=1.0),
        "system": max((_as_float(row.get("system_criticality_score")) for row in ranking_rows), default=1.0),
    }
    rows_by_id = {str(row.get("supplier_id") or ""): row for row in ranking_rows}
    for supplier_id, profile in list(audits.items()):
        # URLs remain in the internal evidence register but are deliberately
        # excluded from the standalone map payload and its visible panels.
        public_evidence = [
            {key: value for key, value in row.items() if key != "source_url"}
            for row in evidence_by_id.get(supplier_id, [])
        ]
        profile["public_evidence"] = public_evidence
        profile["public_evidence_count"] = len(public_evidence)
        profile["public_context_status"] = (
            "identity_conflict"
            if supplier_id == "SDC-VD0993480A"
            else "documented"
            if public_evidence
            else "no_verified_evidence"
        )
        # Preserve any workbook-backed profile, whether complete or partial.
        if str(profile.get("audit_status") or "") in {"audited", "in_progress"}:
            profile["graph_supplier_name"] = supplier_id
            continue
        ranking = rows_by_id.get(supplier_id)
        if not ranking or not profile.get("criteria"):
            continue
        context = contexts.get(supplier_id, {})
        drivers = _proxy_drivers(ranking, context, maxima)
        criteria: list[dict[str, Any]] = []
        for template_row in profile.get("criteria", []):
            row_number = int(template_row.get("row") or 0)
            maturity, criticality, resilience, confidence = _criterion_proxy_values(row_number, drivers)
            criteria.append(
                {
                    **template_row,
                    "response_status": "estimated",
                    "calculation_cache_ready": False,
                    "answered_question_count": 0,
                    "maturity": round(maturity, 6),
                    "criticality": round(criticality, 6),
                    "resilience_weeks": round(resilience, 6),
                    "risk_index": round(audit_risk_index(maturity, criticality, resilience), 6),
                    "action": audit_action(criticality, resilience),
                    "estimate_confidence": round(confidence, 6),
                    "estimate_basis": "proxy simulation + contexte public pondéré",
                }
            )
        criteria_by_row = {int(row["row"]): row for row in criteria}
        families: list[dict[str, Any]] = []
        for family, family_rows in FAMILY_ROWS:
            selected = [criteria_by_row[row] for row in family_rows if row in criteria_by_row]
            maturity = _mean([float(row["maturity"]) for row in selected])
            criticality = _mean([float(row["criticality"]) for row in selected])
            resilience = _mean([float(row["resilience_weeks"]) for row in selected])
            confidence = _mean([float(row["estimate_confidence"]) for row in selected])
            families.append(
                {
                    "family": family,
                    "criterion_count": len(selected),
                    "maturity": round(maturity, 6),
                    "criticality": round(criticality, 6),
                    "resilience_weeks": round(resilience, 6),
                    "risk_index": round(audit_risk_index(maturity, criticality, resilience), 6),
                    "action": audit_action(criticality, resilience),
                    "estimate_confidence": round(confidence, 6),
                }
            )
        maturity = _mean([float(row["maturity"]) for row in families])
        criticality = _mean([float(row["criticality"]) for row in families])
        resilience = _mean([float(row["resilience_weeks"]) for row in families])
        estimate_confidence = _mean([float(row["estimate_confidence"]) for row in criteria])
        profile.update(
            {
                "audit_status": "estimated",
                "estimation_status": "proxy_estimate",
                "graph_supplier_name": supplier_id,
                "supplier_label": supplier_id,
                "product": "",
                "audited_company": "",
                "estimated_criterion_count": len(criteria),
                "maturity": round(maturity, 6),
                "criticality": round(criticality, 6),
                "resilience_weeks": round(resilience, 6),
                "audit_risk_index": round(audit_risk_index(maturity, criticality, resilience), 6),
                "estimate_confidence": round(estimate_confidence, 6),
                "recommended_action": audit_action(criticality, resilience),
                "families": families,
                "criteria": criteria,
                "public_context": {
                    "confidence": round(drivers["confidence"], 6),
                    "source_count": int(_as_float(context.get("source_count"))),
                    "evidence_basis": str(context.get("evidence_basis") or "simulation_only"),
                },
                "quality_warnings": [
                    "Estimation indicative : aucune réponse de la trame n'a été inventée ni marquée comme auditée.",
                    "Les valeurs utilisent des proxys de flux, délai, mono-source, exposition réseau et contexte public ; elles doivent être remplacées par un audit terrain.",
                    "L'absence de résultat web n'est jamais interprétée comme une absence de risque.",
                ],
                "methodology": {
                    **dict(profile.get("methodology") or {}),
                    "estimate_formula": "proxys simulation normalisés + signaux publics rétractés vers un prior neutre selon leur confiance",
                    "estimate_does_not_change_structural_rank": True,
                },
            }
        )
    return audits


def supplier_estimated_score(audit: dict[str, Any] | None) -> float | None:
    if not audit or str(audit.get("audit_status") or "") != "estimated":
        return None
    value = _finite_float(audit.get("audit_risk_index"))
    return _clamp(value) if value is not None else None


def supplier_audit_score(audit: dict[str, Any] | None) -> float | None:
    """Return the normalized audit score only for a genuinely completed audit."""

    if not audit or str(audit.get("audit_status") or "") != "audited":
        return None
    value = audit.get("audit_risk_index")
    if value in (None, ""):
        return None
    try:
        score = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(score):
        return None
    return max(0.0, min(1.0, score))


def expand_supplier_audit_coverage(
    nodes: list[dict[str, Any]],
    audits: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """Return an audit/profile entry for every supplier node in the graph.

    A completed workbook remains supplier-specific.  Other suppliers receive
    the same criterion catalogue with explicit missing values, never copied
    scores.  This makes coverage visible without inventing audit evidence.
    """

    template = next((audit for audit in audits.values() if audit.get("criteria")), None)
    covered: dict[str, dict[str, Any]] = {}
    for node in nodes:
        if str(node.get("type") or "") != "supplier_dc":
            continue
        supplier_id = str(node.get("id") or "").strip()
        if not supplier_id:
            continue
        completed = audits.get(supplier_id)
        if completed:
            completed_copy = dict(completed)
            completed_copy["graph_supplier_name"] = supplier_id
            completed_copy["map_marker_available"] = _node_has_coordinates(node)
            covered[supplier_id] = completed_copy
            continue

        families = []
        criteria = []
        if template:
            families = [
                {
                    "family": row.get("family"),
                    "criterion_count": row.get("criterion_count", 0),
                    "maturity": None,
                    "criticality": None,
                    "resilience_weeks": None,
                    "risk_index": None,
                    "action": "À renseigner",
                }
                for row in template.get("families", [])
            ]
            criteria = [
                {
                    "row": row.get("row"),
                    "label": row.get("label"),
                    "response_status": "not_assessed",
                    "calculation_cache_ready": False,
                    "question_count": row.get("question_count", 0),
                    "answered_question_count": 0,
                    "maturity": None,
                    "criticality": None,
                    "resilience_weeks": None,
                    "risk_index": None,
                    "action": "À renseigner",
                }
                for row in template.get("criteria", [])
            ]
        covered[supplier_id] = {
            "audit_status": "not_assessed",
            "supplier_id": supplier_id,
            "supplier_label": "",
            "graph_supplier_name": supplier_id,
            "map_marker_available": _node_has_coordinates(node),
            "product": "",
            "audited_company": "",
            "source_file": str((template or {}).get("source_file") or ""),
            "criterion_count": len(criteria),
            "answered_criterion_count": 0,
            "partial_criterion_count": 0,
            "question_count": int((template or {}).get("question_count") or 0),
            "answered_question_count": 0,
            "calculated_criterion_count": 0,
            "aggregated_criterion_count": 0,
            "reference_aggregated_criterion_count": int(
                (template or {}).get("aggregated_criterion_count") or 0
            ),
            "unassigned_criteria": list((template or {}).get("unassigned_criteria") or []),
            "calculation_corrections": [],
            "maturity": None,
            "criticality": None,
            "resilience_weeks": None,
            "audit_risk_index": None,
            "saved_excel_priority": None,
            "recommended_action": "Audit à renseigner",
            "families": families,
            "criteria": criteria,
            "quality_warnings": [
                "Aucune réponse d'audit propre à ce fournisseur n'est disponible.",
                "Le référentiel est affiché pour préparation ; aucun score RAJA n'est recopié et la criticité simulée reste structurelle.",
            ],
            "methodology": dict((template or {}).get("methodology") or {}),
        }
    return covered


def supplier_audit_coverage_summary(audits: dict[str, dict[str, Any]]) -> dict[str, Any]:
    audited_ids = sorted(node_id for node_id, audit in audits.items() if supplier_audit_score(audit) is not None)
    estimated_ids = sorted(
        node_id for node_id, audit in audits.items() if supplier_estimated_score(audit) is not None
    )
    pending_ids = sorted(node_id for node_id in audits if node_id not in {*audited_ids, *estimated_ids})
    mapped_ids = sorted(node_id for node_id, audit in audits.items() if audit.get("map_marker_available"))
    unlocated_ids = sorted(node_id for node_id in audits if node_id not in mapped_ids)
    return {
        "supplier_count": len(audits),
        "audited_supplier_count": len(audited_ids),
        "estimated_supplier_count": len(estimated_ids),
        "pending_supplier_count": len(pending_ids),
        "audited_supplier_ids": audited_ids,
        "estimated_supplier_ids": estimated_ids,
        "pending_supplier_ids": pending_ids,
        "map_marker_supplier_count": len(mapped_ids),
        "unlocated_supplier_count": len(unlocated_ids),
        "unlocated_supplier_ids": unlocated_ids,
    }


def blend_criticality_with_audit(structural_score: float, audit: dict[str, Any] | None) -> float:
    audit_score = supplier_audit_score(audit)
    if audit_score is None:
        return max(0.0, min(1.0, structural_score))
    return max(
        0.0,
        min(1.0, (1.0 - AUDIT_SCORE_WEIGHT) * structural_score + AUDIT_SCORE_WEIGHT * audit_score),
    )


def _fmt_pct(value: Any) -> str:
    if value in (None, ""):
        return "À renseigner"
    return f"{100.0 * _as_float(value):.1f} %"


def _fmt_num(value: Any) -> str:
    if value in (None, ""):
        return "À renseigner"
    return f"{_as_float(value):.1f}"


def build_supplier_audit_radar_figures(audit: dict[str, Any]) -> list[dict[str, Any]]:
    """Build the three family radars defined by the workbook dashboard."""

    families = list(audit.get("families") or [])
    categories = [str(row.get("family") or "") for row in families]
    is_audited = supplier_audit_score(audit) is not None
    is_estimated = supplier_estimated_score(audit) is not None
    has_values = is_audited or is_estimated

    def values_for(key: str, multiplier: float = 1.0) -> list[float]:
        if not has_values:
            return []
        values: list[float] = []
        for row in families:
            value = _finite_float(row.get(key))
            if value is None:
                return []
            values.append(round(value * multiplier, 6))
        return values

    maturity_values = values_for("maturity", 100.0)
    criticality_values = values_for("criticality")
    resilience_values = values_for("resilience_weeks")
    resilience_peak = max([RESILIENCE_THRESHOLD_WEEKS, *resilience_values])
    resilience_max = max(100.0, 25.0 * math.ceil(resilience_peak / 25.0))
    common = {
        "categories": categories,
        "value_label": "Valeur auditée" if is_audited else "Estimation par proxy",
        "has_audit_values": has_values,
        "value_status": "audited" if is_audited else "estimated" if is_estimated else "missing",
    }
    return [
        {
            **common,
            "kind": "radar",
            "title": "Maturité par famille",
            "values": maturity_values,
            "threshold": 100.0 * MATURITY_THRESHOLD,
            "threshold_label": "Seuil 75 %",
            "radial_max": 100.0,
            "unit": " %",
            "color": "#0f766e",
        },
        {
            **common,
            "kind": "radar",
            "title": "Criticité par famille",
            "values": criticality_values,
            "threshold": CRITICALITY_THRESHOLD,
            "threshold_label": "Seuil 15 / 100",
            "radial_max": 100.0,
            "unit": " / 100",
            "color": "#be123c",
        },
        {
            **common,
            "kind": "radar",
            "title": "Résilience par famille",
            "values": resilience_values,
            "threshold": RESILIENCE_THRESHOLD_WEEKS,
            "threshold_label": "Seuil 70 semaines",
            "radial_max": resilience_max,
            "unit": " sem.",
            "color": "#7c3aed",
        },
    ]


def render_supplier_audit_html(audit: dict[str, Any]) -> str:
    """Render the immediately visible audit summary and family table."""

    def e(value: Any) -> str:
        return html.escape(str(value if value not in (None, "") else "n/a"))

    is_audited = supplier_audit_score(audit) is not None
    is_estimated = supplier_estimated_score(audit) is not None
    family_rows = "".join(
        "<tr>"
        f"<td>{e(row.get('family'))}</td>"
        f"<td>{e(_fmt_pct(row.get('maturity')))}</td>"
        f"<td>{e(_fmt_num(row.get('criticality')))}{(' / 100' if row.get('criticality') not in (None, '') else '')}</td>"
        f"<td>{e(_fmt_num(row.get('resilience_weeks')))}{(' sem.' if row.get('resilience_weeks') not in (None, '') else '')}</td>"
        f"<td>{e(_fmt_pct(row.get('risk_index')))}</td>"
        f"<td>{e(row.get('action'))}</td>"
        "</tr>"
        for row in audit.get("families", [])
    )
    warnings = "".join(f"<li>{e(value)}</li>" for value in audit.get("quality_warnings", []))
    if is_audited:
        status_text = (
            f"Réponses d'audit complètes — {int(audit.get('answered_criterion_count') or 0)} critères ; provenance à confirmer"
        )
    elif is_estimated:
        status_text = (
            f"Estimation par proxy — {int(audit.get('estimated_criterion_count') or 0)} critères ; non auditée"
        )
    elif audit.get("audit_status") == "in_progress":
        status_text = (
            "Audit partiellement renseigné — "
            f"{int(audit.get('answered_criterion_count') or 0)}/"
            f"{int(audit.get('criterion_count') or 0)} critères complets"
        )
    else:
        status_text = (
            f"Audit à renseigner — référentiel de {int(audit.get('criterion_count') or 0)} critères prêt"
        )
    criticality_suffix = " / 100" if audit.get("criticality") not in (None, "") else ""
    resilience_suffix = " semaines" if audit.get("resilience_weeks") not in (None, "") else ""
    confidence_html = (
        f'<div class="dataKvLabel">Confiance estimation</div><div class="dataKvValue">{e(_fmt_pct(audit.get("estimate_confidence")))}</div>'
        if is_estimated
        else ""
    )
    return (
        '<div class="factoryHtmlPanelContent dataSummaryPanelContent supplierAuditPanel">'
        '<div class="orderLedgerTextHeader">Critères fournisseur — audit et estimations</div>'
        f'<div class="orderLedgerStatus">{e(status_text)}</div>'
        '<div class="dataSummaryScroll">'
        '<section class="dataSummarySection">'
        '<div class="dataSummarySectionTitle">Synthèse normalisée</div>'
        '<div class="dataKvGrid">'
        f'<div class="dataKvLabel">Matricule</div><div class="dataKvValue">{e(audit.get("supplier_id"))}</div>'
        f'<div class="dataKvLabel">Maturité</div><div class="dataKvValue">{e(_fmt_pct(audit.get("maturity")))}</div>'
        f'<div class="dataKvLabel">Statut</div><div class="dataKvValue">{e(status_text)}</div>'
        f'<div class="dataKvLabel">Criticité {"audit" if is_audited else "estimée"}</div><div class="dataKvValue">{e(_fmt_num(audit.get("criticality")))}{criticality_suffix}</div>'
        f'<div class="dataKvLabel">Résilience</div><div class="dataKvValue">{e(_fmt_num(audit.get("resilience_weeks")))}{resilience_suffix}</div>'
        f'<div class="dataKvLabel">Indice de risque {"audit" if is_audited else "estimé"}</div><div class="dataKvValue">{e(_fmt_pct(audit.get("audit_risk_index")))}</div>'
        f'{confidence_html}'
        f'<div class="dataKvLabel">Action matrice</div><div class="dataKvValue">{e(audit.get("recommended_action"))}</div>'
        '</div></section>'
        '<div class="orderLedgerStatus">Les trois radars et le détail des critères sont accessibles directement dans les onglets du panneau.</div>'
        '<section class="dataSummarySection"><div class="dataSummarySectionTitle">Familles de criticité</div>'
        '<div class="dataSummaryTableWrap"><table class="dataSummaryTable"><thead><tr>'
        '<th>Famille</th><th>Maturité</th><th>Criticité</th><th>Résilience</th><th>Indice</th><th>Action</th>'
        f'</tr></thead><tbody>{family_rows}</tbody></table></div></section>'
        '<details class="riskMethodDetails"><summary>Méthode et points de vigilance</summary>'
        f'<div class="orderLedgerStatus">{e((audit.get("methodology") or {}).get("formula"))}</div><ul>{warnings}</ul>'
        '</details></div></div>'
    )


def render_supplier_audit_criteria_html(audit: dict[str, Any]) -> str:
    """Render all criteria in a dedicated, scrollable panel."""

    def e(value: Any) -> str:
        return html.escape(str(value if value not in (None, "") else "n/a"))

    criterion_rows = "".join(
        "<tr>"
        f"<td>{e(row.get('label'))}</td>"
        f"<td>{e(_fmt_pct(row.get('maturity')))}</td>"
        f"<td>{e(_fmt_num(row.get('criticality')))}</td>"
        f"<td>{e(_fmt_num(row.get('resilience_weeks')))}</td>"
        f"<td>{e(_fmt_pct(row.get('risk_index')))}</td>"
        f"<td>{e(row.get('action'))}</td>"
        "</tr>"
        for row in sorted(audit.get("criteria", []), key=lambda row: int(row.get("row") or 0))
    )
    count = int(audit.get("criterion_count") or len(audit.get("criteria") or []))
    answered = int(audit.get("answered_criterion_count") or 0)
    is_estimated = supplier_estimated_score(audit) is not None
    coverage_text = (
        f"{count}/{count} critères estimés par proxy — aucune réponse d'audit inventée."
        if is_estimated
        else f"{answered}/{count} critères complètement renseignés."
    )
    return (
        '<div class="factoryHtmlPanelContent dataSummaryPanelContent supplierAuditCriteriaPanel">'
        f'<div class="orderLedgerTextHeader">Détail des {count} critères</div>'
        f'<div class="orderLedgerStatus">{e(coverage_text)}</div>'
        '<div class="dataSummaryScroll"><div class="dataSummaryTableWrap">'
        '<table class="dataSummaryTable"><thead><tr>'
        '<th>Critère</th><th>Maturité</th><th>Criticité</th><th>Semaines</th><th>Indice</th><th>Action</th>'
        f'</tr></thead><tbody>{criterion_rows}</tbody></table></div></div></div>'
    )


def render_supplier_public_context_html(audit: dict[str, Any]) -> str:
    """Render reviewed public evidence without exposing supplier names."""

    def e(value: Any) -> str:
        return html.escape(str(value if value not in (None, "") else "n/a"))

    kind_labels = {
        "success": "Réussite / capacité",
        "certification": "Certification",
        "incident": "Incident",
        "financial_risk": "Fragilité financière",
        "regulatory": "Contexte réglementaire",
        "identity": "Identité / activité",
    }
    rows = audit.get("public_evidence") if isinstance(audit.get("public_evidence"), list) else []
    evidence_html = "".join(
        '<section class="dataSummarySection supplierPublicEvidenceItem">'
        f'<div class="dataSummarySectionTitle">{e(kind_labels.get(str(row.get("evidence_kind") or ""), row.get("evidence_kind")))}</div>'
        '<div class="dataKvGrid">'
        f'<div class="dataKvLabel">Date</div><div class="dataKvValue">{e(row.get("event_date"))}</div>'
        f'<div class="dataKvLabel">Portée</div><div class="dataKvValue">{e(row.get("scope"))}</div>'
        f'<div class="dataKvLabel">Vérification</div><div class="dataKvValue">{e(row.get("verification_status"))}</div>'
        f'<div class="dataKvLabel">Confiance</div><div class="dataKvValue">{e(_fmt_pct(row.get("confidence")))}</div>'
        '</div>'
        f'<div class="orderLedgerStatus">{e(row.get("summary"))}</div>'
        '</section>'
        for row in rows
    )
    if not rows:
        if audit.get("public_context_status") == "identity_conflict":
            empty_text = (
                "Contexte public non activé : l'identité du classeur et celle de la liste fournisseur se contredisent. "
                "Aucun fait externe n'est attribué tant que le rapprochement du matricule n'est pas validé."
            )
        else:
            empty_text = (
                "Aucun fait public suffisamment vérifié n'a été retenu pour ce matricule. "
                "Cela ne signifie ni absence d'incident ni absence de risque."
            )
        evidence_html = f'<section class="dataSummarySection"><div class="orderLedgerStatus">{e(empty_text)}</div></section>'
    return (
        '<div class="factoryHtmlPanelContent dataSummaryPanelContent supplierPublicContextPanel">'
        '<div class="orderLedgerTextHeader">Contexte public documenté</div>'
        f'<div class="orderLedgerStatus">Matricule {e(audit.get("supplier_id"))} — {len(rows)} fait(s) retenu(s).</div>'
        '<div class="orderLedgerStatus">Recherche non exhaustive. Les faits groupe ou réseau ne valent pas preuve pour le site et ne remplacent pas un audit terrain.</div>'
        f'<div class="dataSummaryScroll">{evidence_html}</div></div>'
    )


def build_supplier_audit_panel_asset(audit: dict[str, Any]) -> dict[str, Any]:
    """Return a tabbed audit asset with summary, radars and full criteria."""

    radar_figures = build_supplier_audit_radar_figures(audit)
    criterion_count = int(audit.get("criterion_count") or 0)
    return {
        "bundle": [
            {"label": "Synthèse", "asset": {"html": render_supplier_audit_html(audit)}},
            {"label": "Contexte public", "asset": {"html": render_supplier_public_context_html(audit)}},
            {"label": "Radar maturité", "asset": {"figure": radar_figures[0]}},
            {"label": "Radar criticité", "asset": {"figure": radar_figures[1]}},
            {"label": "Radar résilience", "asset": {"figure": radar_figures[2]}},
            {
                "label": f"{criterion_count} critères",
                "asset": {"html": render_supplier_audit_criteria_html(audit)},
            },
        ]
    }


def attach_supplier_audit_panels(
    supplier_hover_images: dict[str, dict[str, Any]],
    audits: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """Expose audit and simulated criticality as distinct supplier tabs."""

    merged = {node_id: dict(panels) for node_id, panels in supplier_hover_images.items()}
    for supplier_id, audit in audits.items():
        panels = merged.setdefault(supplier_id, {})
        simulated_criticality = dict(panels.get("incoming") or {})
        bundle = [
            {
                "label": "Audit fournisseur",
                "asset": build_supplier_audit_panel_asset(audit),
                "role": "supplier_audit",
            }
        ]
        if simulated_criticality:
            bundle.append(
                {
                    "label": "Criticité simulée",
                    "asset": simulated_criticality,
                    "role": "simulated_criticality",
                }
            )
        panels["incoming"] = {"bundle": bundle}
    return merged
