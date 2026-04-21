#!/usr/bin/env python3
"""
Integrate auxiliary case-study Excel workbooks into supply_graph_poc.json.

This script enriches the base graph with:
- lane metadata from FIA sheets (lead times, pricing, order hints)
- missing supplier locations from Fournisseur.xlsx
- a second-tier upstream layer for item 021081 -> 773474 -> M-1430
- an analysis report describing what changed
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
from collections import defaultdict
from pathlib import Path
from typing import Any


PRODUCT_WORKBOOKS = ("021081.xlsx", "268191.xlsx", "268967.xlsx")
LOCATION_WORKBOOK = "Fournisseur.xlsx"
BASE_WORKBOOK = "Data_poc.xlsx"
UPSTREAM_OUTPUT_ITEM = "item:773474"
UPSTREAM_INPUT_ITEM = "item:021081"
UPSTREAM_PRODUCER_ID = "SDC-1450"
UPSTREAM_CAPACITY_G_PER_DAY = 500000.0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update supply_graph_poc.json from case-data Excel files.")
    parser.add_argument(
        "--input-json",
        default="etudecas/donnees/supply_graph_poc.json",
        help="Base graph JSON to update in place.",
    )
    parser.add_argument(
        "--data-dir",
        default="etudecas/donnees",
        help="Directory containing Excel workbooks.",
    )
    parser.add_argument(
        "--output-json",
        default="etudecas/donnees/supply_graph_poc.json",
        help="Output path for the updated graph JSON.",
    )
    parser.add_argument(
        "--report-json",
        default="etudecas/donnees/case_data_update_report.json",
        help="Output path for the machine-readable update report.",
    )
    parser.add_argument(
        "--report-md",
        default="etudecas/donnees/case_data_update_report.md",
        help="Output path for the Markdown update report.",
    )
    return parser.parse_args()


def load_workbook_rows(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        openpyxl = None  # type: ignore

    if openpyxl is not None:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            return [], []
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            return [], []
        headers = [str(value).strip() if value is not None else "" for value in raw_rows[0]]
        rows: list[dict[str, Any]] = []
        for raw in raw_rows[1:]:
            if not any(value not in (None, "") for value in raw):
                continue
            rows.append({headers[idx]: raw[idx] for idx in range(min(len(headers), len(raw)))})
        return headers, rows

    script = r"""
param(
    [string]$WorkbookPath,
    [string]$SheetName
)
$excel = $null
$workbook = $null
$worksheet = $null
try {
    $resolved = (Resolve-Path -LiteralPath $WorkbookPath).Path
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($resolved)
    try {
        $worksheet = $workbook.Worksheets.Item($SheetName)
    } catch {
        Write-Output '{"error":"sheet_not_found"}'
        exit 0
    }
    $used = $worksheet.UsedRange.Value2
    if ($null -eq $used) {
        Write-Output '[]'
        exit 0
    }
    $rowCount = $used.GetLength(0)
    $colCount = $used.GetLength(1)
    if ($rowCount -lt 1 -or $colCount -lt 1) {
        Write-Output '[]'
        exit 0
    }
    $headers = @()
    for ($c = 1; $c -le $colCount; $c++) {
        $headers += [string]($used[1, $c])
    }
    $rows = New-Object System.Collections.Generic.List[object]
    for ($r = 2; $r -le $rowCount; $r++) {
        $obj = [ordered]@{}
        for ($c = 1; $c -le $colCount; $c++) {
            $header = $headers[$c - 1]
            if ([string]::IsNullOrWhiteSpace($header)) {
                continue
            }
            $obj[$header.Trim()] = $used[$r, $c]
        }
        $rows.Add([pscustomobject]$obj)
    }
    $rows | ConvertTo-Json -Depth 4 -Compress
} finally {
    if ($workbook) { $workbook.Close($false) }
    if ($excel) { $excel.Quit() }
    foreach ($obj in @($worksheet, $workbook, $excel)) {
        if ($null -ne $obj) {
            try { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($obj) } catch {}
        }
    }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""
    cmd = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-Command",
        f"& {{ {script} }}",
        str(path),
        sheet_name,
    ]
    completed = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if completed.returncode != 0:
        raise RuntimeError((completed.stderr or completed.stdout or "").strip() or "Excel COM fallback failed")
    payload = (completed.stdout or "").strip()
    if not payload:
        return [], []
    parsed = json.loads(payload)
    if isinstance(parsed, dict) and parsed.get("error") == "sheet_not_found":
        return [], []
    if isinstance(parsed, dict):
        parsed_rows = [parsed]
    elif isinstance(parsed, list):
        parsed_rows = [row for row in parsed if isinstance(row, dict)]
    else:
        parsed_rows = []
    headers: list[str] = []
    for row in parsed_rows:
        for key in row.keys():
            key_text = str(key).strip()
            if key_text and key_text not in headers:
                headers.append(key_text)
    rows = [{header: row.get(header) for header in headers} for row in parsed_rows]
    return headers, rows


def to_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_unit(value: Any) -> str | None:
    raw = str(value or "").strip().upper().replace(".", "")
    aliases = {
        "UNITE": "UN",
        "UNIT": "UN",
        "UNITS": "UN",
        "ZUN": "UN",
    }
    raw = aliases.get(raw, raw)
    return raw or None


def canonical_item_id(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.lower().startswith("item:"):
        text = text.split(":", 1)[1].strip()
    if re.fullmatch(r"[0-9]+(?:\.0+)?", text):
        return f"item:{int(float(text)):06d}"
    return f"item:{text.replace(' ', '')}"


def canonical_actor_id(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    match = re.match(r"^([A-Za-z]+)\s*-\s*(.+)$", text)
    if not match:
        compact = text.replace(" ", "").upper()
        if compact.startswith("VD"):
            return f"SDC-{compact}"
        if compact.startswith("D") and compact[1:].isdigit():
            return f"SDC-{compact}"
        return compact

    prefix = match.group(1).upper()
    code = match.group(2).strip().upper().replace(" ", "")
    if prefix == "M":
        if code.startswith("D") and code[1:].isdigit():
            code = code[1:]
        return f"M-{code}"
    if prefix in {"DC", "SDC"}:
        if code.startswith("D") and code[1:].isdigit():
            code = code[1:]
        return f"SDC-{code}" if prefix == "SDC" or code.startswith("VD") else f"{prefix}-{code}"
    if prefix == "C":
        return f"C-{code}"
    return text.replace(" ", "")


def actor_code_from_node_id(node_id: str) -> str:
    if "-" not in node_id:
        return node_id
    prefix, code = node_id.split("-", 1)
    if prefix in {"M", "DC", "SDC"} and code.isdigit():
        return f"D{code}"
    return code


def item_code_from_item_id(item_id: str) -> str:
    return item_id.split(":", 1)[1] if ":" in item_id else item_id


def ensure_item(items: list[dict[str, Any]], item_id: str, default_uom: str | None, report: dict[str, Any]) -> None:
    if any(str(item.get("id")) == item_id for item in items):
        return
    code = item_code_from_item_id(item_id)
    items.append(
        {
            "id": item_id,
            "code": code,
            "name": f"Product {code}",
            "kind": "unknown",
            "uom_default": default_uom,
            "defaults": {
                "kind": True,
                "name": True,
                "uom_default": default_uom is None,
            },
        }
    )
    report["created_items"].append(item_id)


def ensure_inventory_state(
    node: dict[str, Any],
    item_id: str,
    uom: str | None,
    report: dict[str, Any],
    initial: float = 0.0,
) -> None:
    inventory = node.setdefault("inventory", {})
    states = inventory.setdefault("states", [])
    target_uom = uom or "UN"
    for state in states:
        if str(state.get("item_id")) == item_id:
            if normalize_unit(state.get("uom")) != normalize_unit(target_uom):
                state["uom"] = target_uom
                report["updated_inventory_states"].append(
                    {
                        "node_id": str(node.get("id")),
                        "item_id": item_id,
                        "action": "uom_updated",
                        "uom": target_uom,
                    }
                )
            return
    states.append(
        {
            "item_id": item_id,
            "state_id": f"I_{item_code_from_item_id(item_id)}",
            "initial": initial,
            "uom": target_uom,
            "holding_cost": {
                "value": 0.0,
                "per": "unit*day",
                "is_default": True,
            },
            "is_default_initial": True,
        }
    )
    report["updated_inventory_states"].append(
        {
            "node_id": str(node.get("id")),
            "item_id": item_id,
            "action": "created",
            "uom": target_uom,
        }
    )


def make_supplier_name(actor_code: str) -> str:
    return f"Supplier of Raw Materials - {actor_code}"


def make_supplier_node(actor_code: str, supplier_id: str, location_id: str | None) -> dict[str, Any]:
    attrs: dict[str, Any] = {"source_sheet": "FIA"}
    if location_id:
        attrs["location_source"] = LOCATION_WORKBOOK
    return {
        "id": supplier_id,
        "type": "supplier_dc",
        "name": make_supplier_name(actor_code),
        "location_ID": location_id,
        "geo": {"lat": None, "lon": None, "country": None, "raw": None},
        "inventory": {"states": [], "backlogs": [], "wip": []},
        "processes": [],
        "policies": {"ordering": [], "production": []},
        "accounting": {"defaults_used": True},
        "lca": {"defaults_used": True},
        "attrs": attrs,
        "defaults": {"geo": True},
        "role_raw": "Supplier Distribution Center",
    }


def ensure_supplier_node(
    nodes: list[dict[str, Any]],
    node_by_id: dict[str, dict[str, Any]],
    supplier_id: str,
    location_id: str | None,
    report: dict[str, Any],
) -> dict[str, Any]:
    node = node_by_id.get(supplier_id)
    actor_code = actor_code_from_node_id(supplier_id)
    if node is None:
        node = make_supplier_node(actor_code, supplier_id, location_id)
        nodes.append(node)
        node_by_id[supplier_id] = node
        report["created_nodes"].append(supplier_id)
    if not str(node.get("name") or "").strip() or str(node.get("name")) == supplier_id:
        node["name"] = make_supplier_name(actor_code)
    if location_id and not str(node.get("location_ID") or "").strip():
        node["location_ID"] = location_id
        node.setdefault("attrs", {})["location_source"] = LOCATION_WORKBOOK
        report["updated_node_locations"].append(
            {"node_id": supplier_id, "location_ID": location_id, "action": "filled_missing"}
        )
    return node


def ensure_edge(
    edges: list[dict[str, Any]],
    supplier_id: str,
    destination_id: str,
    item_id: str,
    report: dict[str, Any],
) -> dict[str, Any]:
    for edge in edges:
        if (
            str(edge.get("from")) == supplier_id
            and str(edge.get("to")) == destination_id
            and item_id in (edge.get("items") or [])
        ):
            return edge

    edge = {
        "id": f"edge:{supplier_id}_TO_{destination_id}_{item_code_from_item_id(item_id)}",
        "type": "transport",
        "from": supplier_id,
        "to": destination_id,
        "items": [item_id],
        "mode": "truck",
        "distance_km": None,
        "lead_time": {
            "type": "erlang_pipeline",
            "mean": 8.0,
            "stages": 4,
            "time_unit": "day",
            "is_default": True,
        },
        "order_terms": {
            "sell_price": None,
            "price_base": None,
            "quantity_unit": None,
            "is_default": True,
        },
        "transport_cost": {"value": 0.0, "per": "unit", "is_default": True},
        "delay_step_limit": {"value": 999, "is_default": True},
        "attrs": {"source_sheet": "FIA"},
    }
    edges.append(edge)
    report["created_edges"].append(edge["id"])
    return edge


def update_edge_from_fia(
    edge: dict[str, Any],
    row: dict[str, Any],
    workbook_name: str,
    product_code: str,
    report: dict[str, Any],
) -> None:
    price = to_float(
        row.get("Montant")
        if "Montant" in row
        else row.get("Montant/Pourcentage de condition en l'absence de barème")
    )
    price_base = to_float(
        row.get("Base de prix")
        if "Base de prix" in row
        else row.get("Base de prix de la condition")
    )
    lead_days = to_float(row.get("Délai prévisionnel de livraison en jours"))
    order_qty = to_float(row.get("Quantité standard de commande"))
    unit = normalize_unit(
        row.get("Unité de quantité")
        if "Unité de quantité" in row
        else row.get("Unité de quantité de la condition")
    )
    currency = str(
        row.get("Devise")
        if "Devise" in row
        else row.get("Unité de condition (devise ou pourcentage)")
        or ""
    ).strip()

    order_terms = edge.setdefault("order_terms", {})
    order_terms["sell_price"] = price
    order_terms["price_base"] = price_base if price_base is not None else 1.0
    order_terms["quantity_unit"] = unit
    order_terms["is_default"] = price is None
    order_terms["source"] = "case_data_fia"

    if lead_days is not None and lead_days > 0:
        edge["lead_time"] = {
            "type": "erlang_pipeline",
            "mean": float(lead_days),
            "stages": 4,
            "time_unit": "day",
            "is_default": False,
            "source": "case_data_fia",
        }
        edge["delay_step_limit"] = {
            "value": int(max(lead_days + 14, lead_days * 2)),
            "is_default": False,
            "source": "case_data_fia_assumption_from_lead_time",
        }

    attrs = edge.setdefault("attrs", {})
    attrs.update(
        {
            "source_sheet": "FIA",
            "source_workbook": workbook_name,
            "product_code": product_code,
            "supplier_account": row.get("Numéro de compte fournisseur"),
            "standard_order_qty": order_qty,
            "standard_order_uom": unit,
            "pricing_currency": currency or None,
        }
    )
    report["updated_edges"].append(
        {
            "edge_id": str(edge.get("id")),
            "workbook": workbook_name,
            "product_code": product_code,
            "lead_time_days": lead_days,
            "sell_price": price,
            "price_base": price_base,
            "quantity_unit": unit,
            "standard_order_qty": order_qty,
        }
    )


def parse_product_workbook(path: Path) -> dict[str, Any]:
    bom_headers, bom_rows = load_workbook_rows(path, "BOM")
    _, fia_rows = load_workbook_rows(path, "FIA")

    output_products = []
    for row in bom_rows:
        value = row.get("Produit Fini")
        if value in (None, ""):
            continue
        output_products.append(canonical_item_id(value))

    unique_outputs = sorted(set(output_products))
    output_item_id = unique_outputs[0] if unique_outputs else ""
    product_code = item_code_from_item_id(output_item_id) if output_item_id else ""

    output_qty = None
    output_uom = None
    for header in bom_headers:
        match = re.search(r"Qt[ée]\s*\(([^)]+)\)", header, re.IGNORECASE)
        if match:
            output_uom = normalize_unit(match.group(1))
            if bom_rows:
                output_qty = to_float(bom_rows[0].get(header))
            break

    normalized_bom = []
    for row in bom_rows:
        component_id = canonical_item_id(row.get("N° composante"))
        qty = to_float(row.get("Qté composants (UQB)"))
        qty_uom = normalize_unit(row.get("Unité de quantité"))
        if not component_id or qty is None or qty_uom is None:
            continue
        normalized_bom.append(
            {
                "item_id": component_id,
                "ratio_per_batch": qty,
                "ratio_unit": qty_uom,
                "defaults": {"quantity": False},
            }
        )

    return {
        "path": path,
        "file_name": path.name,
        "file_stem": path.stem,
        "output_item_id": output_item_id,
        "output_product_code": product_code,
        "output_batch_size": output_qty,
        "output_batch_uom": output_uom,
        "bom_inputs": normalized_bom,
        "fia_rows": fia_rows,
    }


def sync_process_inputs(
    node: dict[str, Any],
    workbook: dict[str, Any],
    report: dict[str, Any],
) -> None:
    output_item_id = workbook["output_item_id"]
    bom_inputs = workbook["bom_inputs"]
    if not output_item_id or not bom_inputs:
        return

    processes = node.setdefault("processes", [])
    process = None
    for candidate in processes:
        outputs = candidate.get("outputs") or []
        if any(str(out.get("item_id")) == output_item_id for out in outputs):
            process = candidate
            break

    if process is None and str(node.get("id")) == UPSTREAM_PRODUCER_ID and output_item_id == UPSTREAM_OUTPUT_ITEM:
        process = {
            "id": "proc:MAKE_773474",
            "type": "transform",
            "outputs": [{"item_id": UPSTREAM_OUTPUT_ITEM, "rate_id": "Q_MAKE_773474", "uom": "G/day"}],
            "inputs": [],
            "batch_size": workbook.get("output_batch_size") or 1000.0,
            "batch_size_unit": workbook.get("output_batch_uom") or "G",
            "wip": {
                "state_id": "WIP_773474",
                "tau_process": 3.0,
                "time_unit": "day",
                "is_default": True,
            },
            "capacity": {
                "max_rate": UPSTREAM_CAPACITY_G_PER_DAY,
                "uom": "G/day",
                "is_default": False,
                "source": "case_data_upstream_bom_assumption",
            },
            "cost": {"value": 0.0, "per": "unit", "is_default": True},
            "lca": {
                "factors": [{"impact": "GWP100", "value": 0.0, "per": "unit", "is_default": True}]
            },
            "attrs": {
                "source_workbook": workbook["file_name"],
                "source_sheet": "BOM",
            },
        }
        processes.append(process)
        report["created_processes"].append({"node_id": str(node.get("id")), "process_id": process["id"]})

    if process is None:
        return

    process["inputs"] = bom_inputs
    if workbook.get("output_batch_size") is not None:
        process["batch_size"] = workbook["output_batch_size"]
    if workbook.get("output_batch_uom") is not None:
        process["batch_size_unit"] = workbook["output_batch_uom"]
    process.setdefault("attrs", {})["source_workbook"] = workbook["file_name"]
    process.setdefault("attrs", {})["source_sheet"] = "BOM"
    report["synced_processes"].append(
        {
            "node_id": str(node.get("id")),
            "process_id": str(process.get("id")),
            "output_item_id": output_item_id,
            "input_count": len(bom_inputs),
        }
    )


def markdown_report(report: dict[str, Any]) -> str:
    lines = [
        "# Case data update report",
        "",
        "## Summary",
        "",
        f"- Updated graph: `{report['paths']['output_json']}`",
        f"- Created items: {len(report['created_items'])}",
        f"- Removed orphan items: {len(report['removed_items'])}",
        f"- Created nodes: {len(report['created_nodes'])}",
        f"- Created edges: {len(report['created_edges'])}",
        f"- Created processes: {len(report['created_processes'])}",
        f"- Synced processes: {len(report['synced_processes'])}",
        f"- Updated edges from FIA: {len(report['updated_edges'])}",
        f"- Updated node locations: {len(report['updated_node_locations'])}",
        "",
        "## Workbook findings",
        "",
    ]
    for finding in report["workbook_findings"]:
        lines.append(
            f"- `{finding['file_name']}` -> output `{finding['output_product_code'] or 'n/a'}`"
            f" (BOM rows={finding['bom_rows']}, FIA rows={finding['fia_rows']},"
            f" file/product mismatch={finding['file_product_mismatch']})"
        )
    lines.extend(
        [
            "",
            "## Important assumptions",
            "",
            "- `268191.xlsx` is interpreted as product `268091` because the BOM sheet explicitly points to `268091`.",
            "- `021081.xlsx` is modeled as an upstream component feeding supplier `SDC-1450`, which now transforms `021081` into `773474` before delivery to `M-1430`.",
            f"- The new `SDC-1450` transformation capacity is set to {UPSTREAM_CAPACITY_G_PER_DAY:.0f} G/day to avoid creating an artificial bottleneck.",
            "- FIA lead times are applied directly to lanes, and delay limits are set to `max(lead + 14, 2 * lead)` as a simulation cap assumption.",
            "- Component `007923` is kept in the 268091 BOM but left unconstrained because no supplier lane is provided in the new FIA data.",
            "",
            "## Unresolved points",
            "",
        ]
    )
    if report["unresolved"]:
        for line in report["unresolved"]:
            lines.append(f"- {line}")
    else:
        lines.append("- None.")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    args = parse_args()
    input_json = Path(args.input_json)
    data_dir = Path(args.data_dir)
    output_json = Path(args.output_json)
    report_json = Path(args.report_json)
    report_md = Path(args.report_md)

    graph = json.loads(input_json.read_text(encoding="utf-8"))
    items = graph.setdefault("items", [])
    nodes = graph.setdefault("nodes", [])
    edges = graph.setdefault("edges", [])
    meta = graph.setdefault("meta", {})

    node_by_id = {str(node.get("id")): node for node in nodes if node.get("id")}

    report: dict[str, Any] = {
        "paths": {
            "input_json": str(input_json),
            "output_json": str(output_json),
            "report_json": str(report_json),
            "report_md": str(report_md),
        },
        "created_items": [],
        "removed_items": [],
        "created_nodes": [],
        "created_edges": [],
        "created_processes": [],
        "synced_processes": [],
        "updated_edges": [],
        "updated_node_locations": [],
        "updated_inventory_states": [],
        "workbook_findings": [],
        "unresolved": [],
    }

    _, location_rows = load_workbook_rows(data_dir / LOCATION_WORKBOOK, "LOCATION")
    location_map: dict[str, str] = {}
    for row in location_rows:
        supplier_code = str(row.get("Numéro de compte fournisseur") or "").strip().upper()
        location_id = str(row.get("location") or "").strip()
        if supplier_code and location_id:
            location_map[supplier_code] = location_id

    for node_id, node in node_by_id.items():
        actor_code = actor_code_from_node_id(node_id)
        location_id = location_map.get(actor_code)
        if location_id and not str(node.get("location_ID") or "").strip():
            node["location_ID"] = location_id
            node.setdefault("attrs", {})["location_source"] = LOCATION_WORKBOOK
            report["updated_node_locations"].append(
                {"node_id": node_id, "location_ID": location_id, "action": "filled_missing"}
            )
        if node_id == UPSTREAM_PRODUCER_ID and location_id:
            node["location_ID"] = location_id
            if str(node.get("name") or "").strip() == node_id:
                node["name"] = make_supplier_name(actor_code)

    workbooks = [parse_product_workbook(data_dir / name) for name in PRODUCT_WORKBOOKS]
    for workbook in workbooks:
        output_product_code = workbook["output_product_code"]
        report["workbook_findings"].append(
            {
                "file_name": workbook["file_name"],
                "output_product_code": output_product_code,
                "bom_rows": len(workbook["bom_inputs"]),
                "fia_rows": len(workbook["fia_rows"]),
                "file_product_mismatch": bool(
                    output_product_code and workbook["file_stem"] not in {output_product_code, "268191"}
                )
                or (workbook["file_stem"] == "268191" and output_product_code != "268091"),
            }
        )

    destination_by_output: dict[str, str] = {}
    for node in nodes:
        for process in node.get("processes") or []:
            for output in process.get("outputs") or []:
                item_id = str(output.get("item_id"))
                if item_id:
                    destination_by_output[item_id] = str(node.get("id"))
    destination_by_output[UPSTREAM_OUTPUT_ITEM] = UPSTREAM_PRODUCER_ID

    for workbook in workbooks:
        output_item_id = workbook["output_item_id"]
        output_uom = workbook.get("output_batch_uom")
        if output_item_id:
            ensure_item(items, output_item_id, output_uom, report)
        for bom_input in workbook["bom_inputs"]:
            ensure_item(items, str(bom_input["item_id"]), normalize_unit(bom_input.get("ratio_unit")), report)

        destination_id = destination_by_output.get(output_item_id)
        if not destination_id:
            report["unresolved"].append(
                f"No destination node found for workbook {workbook['file_name']} / output {output_item_id}."
            )
            continue

        destination_node = node_by_id.get(destination_id)
        if destination_node is None:
            report["unresolved"].append(f"Destination node {destination_id} is missing from the graph.")
            continue

        sync_process_inputs(destination_node, workbook, report)
        if output_item_id:
            ensure_inventory_state(destination_node, output_item_id, output_uom, report)
        for bom_input in workbook["bom_inputs"]:
            ensure_inventory_state(
                destination_node,
                str(bom_input["item_id"]),
                normalize_unit(bom_input.get("ratio_unit")),
                report,
            )

        for row in workbook["fia_rows"]:
            item_id = canonical_item_id(row.get("Numéro d'article"))
            supplier_account = str(row.get("Numéro de compte fournisseur") or "").strip().upper()
            if not item_id or not supplier_account:
                continue
            supplier_id = canonical_actor_id(f"SDC - {supplier_account}")
            supplier_location = location_map.get(supplier_account)
            ensure_item(items, item_id, normalize_unit(row.get("Unité de quantité") or row.get("Unité de quantité de la condition")), report)
            supplier_node = ensure_supplier_node(nodes, node_by_id, supplier_id, supplier_location, report)
            ensure_inventory_state(
                supplier_node,
                item_id,
                normalize_unit(row.get("Unité de quantité") or row.get("Unité de quantité de la condition")),
                report,
            )
            edge = ensure_edge(edges, supplier_id, destination_id, item_id, report)
            update_edge_from_fia(edge, row, workbook["file_name"], workbook["output_product_code"], report)

    inbound_pairs = {
        (str(edge.get("to")), str(item_id))
        for edge in edges
        for item_id in (edge.get("items") or [])
    }
    outbound_pairs = {
        (str(edge.get("from")), str(item_id))
        for edge in edges
        for item_id in (edge.get("items") or [])
    }
    demand_pairs = {
        (str(d.get("node_id")), str(d.get("item_id")))
        for scenario in (graph.get("scenarios") or [])
        for d in (scenario.get("demand") or [])
        if d.get("node_id") and d.get("item_id")
    }
    for node in nodes:
        inventory = node.get("inventory") or {}
        states = inventory.get("states") or []
        if not states:
            continue
        node_id = str(node.get("id"))
        process_outputs = {
            str(out.get("item_id"))
            for process in (node.get("processes") or [])
            for out in (process.get("outputs") or [])
            if out.get("item_id")
        }
        kept_states = []
        for state in states:
            item_id = str(state.get("item_id"))
            keep = (
                (node_id, item_id) in inbound_pairs
                or (node_id, item_id) in outbound_pairs
                or (node_id, item_id) in demand_pairs
                or item_id in process_outputs
            )
            if keep:
                kept_states.append(state)
                continue
            report["unresolved"].append(
                f"Removed orphan inventory state {node_id}/{item_id} because no inbound lane is provided."
            )
        inventory["states"] = kept_states

    referenced_items = {
        str(item_id)
        for edge in edges
        for item_id in (edge.get("items") or [])
    }
    referenced_items.update(
        str(spec.get("item_id"))
        for node in nodes
        for inventory_state in ((node.get("inventory") or {}).get("states") or [])
        for spec in [inventory_state]
        if spec.get("item_id")
    )
    referenced_items.update(
        str(spec.get("item_id"))
        for node in nodes
        for process in (node.get("processes") or [])
        for spec in (process.get("inputs") or [])
        if spec.get("item_id")
    )
    referenced_items.update(
        str(spec.get("item_id"))
        for node in nodes
        for process in (node.get("processes") or [])
        for spec in (process.get("outputs") or [])
        if spec.get("item_id")
    )
    referenced_items.update(
        str(demand.get("item_id"))
        for scenario in (graph.get("scenarios") or [])
        for demand in (scenario.get("demand") or [])
        if demand.get("item_id")
    )
    kept_items = []
    for item in items:
        item_id = str(item.get("id"))
        if item_id in referenced_items:
            kept_items.append(item)
            continue
        report["removed_items"].append(item_id)
    items[:] = kept_items

    meta["source_file"] = BASE_WORKBOOK
    meta["source_files"] = [BASE_WORKBOOK, LOCATION_WORKBOOK, *PRODUCT_WORKBOOKS]
    notes = meta.setdefault("notes", [])
    case_note = (
        "Graph enriched with auxiliary case-study Excel workbooks "
        "(021081.xlsx, 268191.xlsx, 268967.xlsx, Fournisseur.xlsx)."
    )
    if case_note not in notes:
        notes.append(case_note)
    meta["case_data_refresh"] = {
        "enabled": True,
        "source_files": [LOCATION_WORKBOOK, *PRODUCT_WORKBOOKS],
        "upstream_extension": {
            "input_item": UPSTREAM_INPUT_ITEM,
            "producer_node_id": UPSTREAM_PRODUCER_ID,
            "output_item": UPSTREAM_OUTPUT_ITEM,
        },
    }

    output_json.write_text(json.dumps(graph, indent=2, ensure_ascii=False), encoding="utf-8")
    report_json.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    report_md.write_text(markdown_report(report), encoding="utf-8")

    print(f"[OK] Updated graph: {output_json.resolve()}")
    print(f"[OK] Report JSON: {report_json.resolve()}")
    print(f"[OK] Report MD: {report_md.resolve()}")


if __name__ == "__main__":
    main()
